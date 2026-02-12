# -*- coding: utf-8 -*-
from __future__ import unicode_literals
import csv
import xlwt
from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.models import User
from django.http import HttpResponseRedirect, HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.mail import EmailMessage,EmailMultiAlternatives
from django.http import JsonResponse
from interlocutorc.forms import *
from interlocutorc.models import *
from configuracion.models import *
from itertools import chain
from django.contrib import messages
from django.conf import settings
from datetime import date, datetime, timedelta
from datetime import datetime, time
from collections import defaultdict
from django.db.models import Q,Count,Sum,F
import pandas as pd
import os
import sys
import xlrd
import requests
import ast
import random
import string
import json
import holidays_co
import base64
import pytz
from django.core.files import File
import ssl
import re
import openpyxl
import zipfile
from django.template.loader import get_template
import uuid
from django.shortcuts import redirect
from PIL import Image
from django.core.files.base import ContentFile
from django.views.decorators.clickjacking import xframe_options_exempt
from xhtml2pdf import pisa
from django.contrib.staticfiles import finders
from django.shortcuts import get_object_or_404

try:
    from cStringIO import StringIO as BytesIO  # más rápido si disponible
except ImportError:
    from StringIO import StringIO as BytesIO   # fallback puro Python



PAGINADOR = 9999999999999999999999999999999

# ************* IP PRODUCCION ******************
#IP_SAP = 'https://192.168.1.2:50000/b1s/v1/'
#IP_SERVIDOR = 'https://192.155.95.186'

# ************* IP CALIDAD ******************
IP_SAP = 'https://172.16.100.20:50000/b1s/v1/'
IP_SERVIDOR = 'https://160.153.178.159'


def link_callback(uri, rel):
    """
    Convierte las rutas relativas de imágenes/estáticos en rutas absolutas en disco
    para que xhtml2pdf pueda encontrarlas.
    """

    # Si la URI empieza con MEDIA_URL (/bodega/...)
    if uri.startswith(settings.MEDIA_URL):
        # quitamos el prefijo /bodega/ y construimos la ruta real en MEDIA_ROOT
        path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
        return path

    # Si la URI empieza con STATIC_URL (por si usas imágenes estáticas)
    if uri.startswith(settings.STATIC_URL):
        path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
        return path

    # Si no es ni media ni static, lo devolvemos tal cual
    return uri

# Configuración SAP
SAP_URL = IP_SAP
SAP_CRED = {
    "CompanyDB": "PCS",
    "UserName": "manager1",
    "Password": "HYC909"
}

# Variables globales para sesión
SESSION = None
ROUTEID = None


def sap_login():
    """Hace login en SAP y guarda la sesión global."""
    global SESSION, ROUTEID

    url = IP_SAP + "Login"
    payload = json.dumps(SAP_CRED)  # dict → JSON
    headers = {"Content-Type": "application/json"}

    response = requests.post(url, data=payload, headers=headers, verify=False)

    if response.status_code != 200:
        raise Exception("Error al hacer login SAP: %s" % response.text)

    respuesta = json.loads(response.text)
    SESSION = respuesta["SessionId"]
    ROUTEID = response.cookies.get("ROUTEID")
    return respuesta

def sap_request(url, metodo="GET", data=None):
    """Hace una request al Service Layer con manejo automático de login/401."""
    global SESSION, ROUTEID

    # Si no hay sesión → login
    if not SESSION:
        sap_login()

    # Armar cookie
    cookie = "B1SESSION=" + SESSION
    if ROUTEID:
        cookie += "; ROUTEID=" + ROUTEID

    headers = {
        'Prefer': 'odata.maxpagesize=999999',
        'Cookie': cookie,
        'Content-Type': 'application/json'
    }

    # Primer intento
    response = requests.request(metodo, url, headers=headers, data=data, verify=False)

    # Si la sesión caducó → relogin y reintentar
    if response.status_code == 401:
        sap_login()
        cookie = "B1SESSION=" + SESSION
        if ROUTEID:
            cookie += "; ROUTEID=" + ROUTEID
        headers['Cookie'] = cookie
        response = requests.request(metodo, url, headers=headers, data=data, verify=False)

    return response

#Justificacion
def config_justificacion(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        lista_justificacion = Justificacion.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.consultar_localizacion == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            #return HttpResponseRedirect('/administracion/')

        return render(request, "config_justificacion.html", {'user': current_user,
                                                          'lista_justificacion': lista_justificacion,
                                                           'permiso_usuario': usuario_datos
                                                           })
    else:
        pass

def config_perfiles_pcs(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        lista_perfiles = Perfiles_PCS.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.consultar_localizacion == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            #return HttpResponseRedirect('/administracion/')

        return render(request, "config_perfiles_pcs.html", {'user': current_user,
                                                          'lista_perfiles': lista_perfiles,
                                                           'permiso_usuario': usuario_datos
                                                           })
    else:
        pass

def config_perfiles_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.registrar_localizacion == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            #return HttpResponseRedirect('/configuracion/continentes/')

        return render(request, "config_perfiles_pcs_registrar.html", {'user': current_user,
                                                                     'permiso_usuario': usuario_datos
                                                                     })

    elif request.method == 'POST':
        current_user = request.user

        id = request.POST['id']
        nombre = request.POST['descripcion']
        codigo_conecta = request.POST['codigo_conecta']
        comercial_direccion = request.POST['comercial_direccion']
        exportaciones_direccion = request.POST['exportaciones_direccion']
        desarrollo_empresarial_direccion = request.POST['desarrollo_empresarial_direccion']
        administrativo_financiero_direccion = request.POST['administrativo_financiero_direccion']
        mercadeo_innovacion_direccion = request.POST['mercadeo_innovacion_direccion']
        operaciones_logistica_direccion = request.POST['operaciones_logistica_direccion']
        comercial_coordinacion_1 = request.POST['comercial_coordinacion_1']
        comercial_coordinacion_2 = request.POST['comercial_coordinacion_2']
        exportaciones_Coordinacion_1 = request.POST['exportaciones_Coordinacion_1']
        desarrollo_empresarial_coordinacion_1 = request.POST['desarrollo_empresarial_coordinacion_1']
        desarrollo_empresarial_coordinacion_2 = request.POST['desarrollo_empresarial_coordinacion_2']
        administrativo_financiero_coordinacion_1 = request.POST['administrativo_financiero_coordinacion_1']
        administrativo_financiero_coordinacion_2 = request.POST['administrativo_financiero_coordinacion_2']
        administrativo_financiero_coordinacion_3 = request.POST['administrativo_financiero_coordinacion_3']
        mercadeo_innovacion_coordinacion_1 = request.POST['mercadeo_innovacion_coordinacion_1']
        mercadeo_innovacion_coordinacion_2 = request.POST['mercadeo_innovacion_coordinacion_2']
        operaciones_logistica_coordinacion_1 = request.POST['operaciones_logistica_coordinacion_1']
        operaciones_logistica_coordinacion_2 = request.POST['operaciones_logistica_coordinacion_2']
        comercial_analista_senior_1 = request.POST['comercial_analista_senior_1']
        comercial_analista_senior_2 = request.POST['comercial_analista_senior_2']
        comercial_analista_senior_3 = request.POST['comercial_analista_senior_3']
        comercial_analista_senior_4 = request.POST['comercial_analista_senior_4']
        comercial_analista_senior_5 = request.POST['comercial_analista_senior_5']
        comercial_analista_senior_6 = request.POST['comercial_analista_senior_6']
        exportaciones_analista_senior_1 = request.POST['exportaciones_analista_senior_1']
        exportaciones_analista_senior_2 = request.POST['exportaciones_analista_senior_2']
        exportaciones_analista_senior_3 = request.POST['exportaciones_analista_senior_3']
        desarrollo_empresarial_analista_senior_1 = request.POST['desarrollo_empresarial_analista_senior_1']
        administrativo_financiero_analista_senior_1 = request.POST['administrativo_financiero_analista_senior_1']
        administrativo_financiero_analista_senior_2 = request.POST['administrativo_financiero_analista_senior_2']
        administrativo_financiero_analista_senior_3 = request.POST['administrativo_financiero_analista_senior_3']
        administrativo_financiero_analista_senior_4 = request.POST['administrativo_financiero_analista_senior_4']
        administrativo_financiero_analista_senior_5 = request.POST['administrativo_financiero_analista_senior_5']
        mercadeo_innovacion_analista_senior_1 = request.POST['mercadeo_innovacion_analista_senior_1']
        operaciones_logistica_analista_senior_1 = request.POST['operaciones_logistica_analista_senior_1']
        operaciones_logistica_analista_senior_2 = request.POST['operaciones_logistica_analista_senior_2']
        operaciones_logistica_analista_senior_3 = request.POST['operaciones_logistica_analista_senior_3']
        operaciones_logistica_analista_senior_4 = request.POST['operaciones_logistica_analista_senior_4']
        operaciones_logistica_analista_senior_5 = request.POST['operaciones_logistica_analista_senior_5']
        comercial_analista_junior_1 = request.POST['comercial_analista_junior_1']
        comercial_analista_junior_2 = request.POST['comercial_analista_junior_2']
        comercial_analista_junior_3 = request.POST['comercial_analista_junior_3']
        comercial_analista_junior_4 = request.POST['comercial_analista_junior_4']
        comercial_analista_junior_5 = request.POST['comercial_analista_junior_5']
        comercial_analista_junior_6 = request.POST['comercial_analista_junior_6']
        comercial_analista_junior_7 = request.POST['comercial_analista_junior_7']
        comercial_analista_junior_8 = request.POST['comercial_analista_junior_8']
        exportaciones_analista_junior_1 = request.POST['exportaciones_analista_junior_1']
        exportaciones_analista_junior_2 = request.POST['exportaciones_analista_junior_2']
        exportaciones_analista_junior_3 = request.POST['exportaciones_analista_junior_3']
        exportaciones_analista_junior_4 = request.POST['exportaciones_analista_junior_4']
        exportaciones_analista_junior_5 = request.POST['exportaciones_analista_junior_5']
        desarrollo_empresarial_analista_junior_1 = request.POST['desarrollo_empresarial_analista_junior_1']
        administrativo_financiero_analista_junior_1 = request.POST['administrativo_financiero_analista_junior_1']
        mercadeo_innovacion_analista_junior_1 = request.POST['mercadeo_innovacion_analista_junior_1']
        operaciones_logistica_analista_junior_1 = request.POST['operaciones_logistica_analista_junior_1']
        analista_senior_2 = request.POST['analista_senior_2']

        if id =='':
            perfil = Perfiles_PCS(
                nombre=nombre,
                codigo_conecta=codigo_conecta,
                comercial_direccion=comercial_direccion,
                exportaciones_direccion=exportaciones_direccion,
                desarrollo_empresarial_direccion=desarrollo_empresarial_direccion,
                administrativo_financiero_direccion=administrativo_financiero_direccion,
                mercadeo_innovacion_direccion=mercadeo_innovacion_direccion,
                operaciones_logistica_direccion=operaciones_logistica_direccion,
                comercial_coordinacion_1=comercial_coordinacion_1,
                comercial_coordinacion_2=comercial_coordinacion_2,
                exportaciones_Coordinacion_1=exportaciones_Coordinacion_1,
                desarrollo_empresarial_coordinacion_1=desarrollo_empresarial_coordinacion_1,
                desarrollo_empresarial_coordinacion_2=desarrollo_empresarial_coordinacion_2,
                administrativo_financiero_coordinacion_1=administrativo_financiero_coordinacion_1,
                administrativo_financiero_coordinacion_2=administrativo_financiero_coordinacion_2,
                administrativo_financiero_coordinacion_3=administrativo_financiero_coordinacion_3,
                mercadeo_innovacion_coordinacion_1=mercadeo_innovacion_coordinacion_1,
                mercadeo_innovacion_coordinacion_2=mercadeo_innovacion_coordinacion_2,
                operaciones_logistica_coordinacion_1=operaciones_logistica_coordinacion_1,
                operaciones_logistica_coordinacion_2=operaciones_logistica_coordinacion_2,
                comercial_analista_senior_1=comercial_analista_senior_1,
                comercial_analista_senior_2=comercial_analista_senior_2,
                comercial_analista_senior_3=comercial_analista_senior_3,
                comercial_analista_senior_4=comercial_analista_senior_4,
                comercial_analista_senior_5=comercial_analista_senior_5,
                comercial_analista_senior_6=comercial_analista_senior_6,
                exportaciones_analista_senior_1=exportaciones_analista_senior_1,
                exportaciones_analista_senior_2=exportaciones_analista_senior_2,
                exportaciones_analista_senior_3=exportaciones_analista_senior_3,
                desarrollo_empresarial_analista_senior_1=desarrollo_empresarial_analista_senior_1,
                administrativo_financiero_analista_senior_1=administrativo_financiero_analista_senior_1,
                administrativo_financiero_analista_senior_2=administrativo_financiero_analista_senior_2,
                administrativo_financiero_analista_senior_3=administrativo_financiero_analista_senior_3,
                administrativo_financiero_analista_senior_4=administrativo_financiero_analista_senior_4,
                administrativo_financiero_analista_senior_5=administrativo_financiero_analista_senior_5,
                mercadeo_innovacion_analista_senior_1=mercadeo_innovacion_analista_senior_1,
                operaciones_logistica_analista_senior_1=operaciones_logistica_analista_senior_1,
                operaciones_logistica_analista_senior_2=operaciones_logistica_analista_senior_2,
                operaciones_logistica_analista_senior_3=operaciones_logistica_analista_senior_3,
                operaciones_logistica_analista_senior_4=operaciones_logistica_analista_senior_4,
                operaciones_logistica_analista_senior_5=operaciones_logistica_analista_senior_5,
                comercial_analista_junior_1=comercial_analista_junior_1,
                comercial_analista_junior_2=comercial_analista_junior_2,
                comercial_analista_junior_3=comercial_analista_junior_3,
                comercial_analista_junior_4=comercial_analista_junior_4,
                comercial_analista_junior_5=comercial_analista_junior_5,
                comercial_analista_junior_6=comercial_analista_junior_6,
                comercial_analista_junior_7=comercial_analista_junior_7,
                comercial_analista_junior_8=comercial_analista_junior_8,
                exportaciones_analista_junior_1=exportaciones_analista_junior_1,
                exportaciones_analista_junior_2=exportaciones_analista_junior_2,
                exportaciones_analista_junior_3=exportaciones_analista_junior_3,
                exportaciones_analista_junior_4=exportaciones_analista_junior_4,
                exportaciones_analista_junior_5=exportaciones_analista_junior_5,
                desarrollo_empresarial_analista_junior_1=desarrollo_empresarial_analista_junior_1,
                administrativo_financiero_analista_junior_1=administrativo_financiero_analista_junior_1,
                mercadeo_innovacion_analista_junior_1=mercadeo_innovacion_analista_junior_1,
                operaciones_logistica_analista_junior_1=operaciones_logistica_analista_junior_1,
                analista_senior_2=analista_senior_2,

            )
            perfil.save()
        else:
            perfil = Perfiles_PCS(
                pk=int(id),
                nombre=nombre,
                codigo_conecta=codigo_conecta,
                comercial_direccion=comercial_direccion,
                exportaciones_direccion=exportaciones_direccion,
                desarrollo_empresarial_direccion=desarrollo_empresarial_direccion,
                administrativo_financiero_direccion=administrativo_financiero_direccion,
                mercadeo_innovacion_direccion=mercadeo_innovacion_direccion,
                operaciones_logistica_direccion=operaciones_logistica_direccion,
                comercial_coordinacion_1=comercial_coordinacion_1,
                comercial_coordinacion_2=comercial_coordinacion_2,
                exportaciones_Coordinacion_1=exportaciones_Coordinacion_1,
                desarrollo_empresarial_coordinacion_1=desarrollo_empresarial_coordinacion_1,
                desarrollo_empresarial_coordinacion_2=desarrollo_empresarial_coordinacion_2,
                administrativo_financiero_coordinacion_1=administrativo_financiero_coordinacion_1,
                administrativo_financiero_coordinacion_2=administrativo_financiero_coordinacion_2,
                administrativo_financiero_coordinacion_3=administrativo_financiero_coordinacion_3,
                mercadeo_innovacion_coordinacion_1=mercadeo_innovacion_coordinacion_1,
                mercadeo_innovacion_coordinacion_2=mercadeo_innovacion_coordinacion_2,
                operaciones_logistica_coordinacion_1=operaciones_logistica_coordinacion_1,
                operaciones_logistica_coordinacion_2=operaciones_logistica_coordinacion_2,
                comercial_analista_senior_1=comercial_analista_senior_1,
                comercial_analista_senior_2=comercial_analista_senior_2,
                comercial_analista_senior_3=comercial_analista_senior_3,
                comercial_analista_senior_4=comercial_analista_senior_4,
                comercial_analista_senior_5=comercial_analista_senior_5,
                comercial_analista_senior_6=comercial_analista_senior_6,
                exportaciones_analista_senior_1=exportaciones_analista_senior_1,
                exportaciones_analista_senior_2=exportaciones_analista_senior_2,
                exportaciones_analista_senior_3=exportaciones_analista_senior_3,
                desarrollo_empresarial_analista_senior_1=desarrollo_empresarial_analista_senior_1,
                administrativo_financiero_analista_senior_1=administrativo_financiero_analista_senior_1,
                administrativo_financiero_analista_senior_2=administrativo_financiero_analista_senior_2,
                administrativo_financiero_analista_senior_3=administrativo_financiero_analista_senior_3,
                administrativo_financiero_analista_senior_4=administrativo_financiero_analista_senior_4,
                administrativo_financiero_analista_senior_5=administrativo_financiero_analista_senior_5,
                mercadeo_innovacion_analista_senior_1=mercadeo_innovacion_analista_senior_1,
                operaciones_logistica_analista_senior_1=operaciones_logistica_analista_senior_1,
                operaciones_logistica_analista_senior_2=operaciones_logistica_analista_senior_2,
                operaciones_logistica_analista_senior_3=operaciones_logistica_analista_senior_3,
                operaciones_logistica_analista_senior_4=operaciones_logistica_analista_senior_4,
                operaciones_logistica_analista_senior_5=operaciones_logistica_analista_senior_5,
                comercial_analista_junior_1=comercial_analista_junior_1,
                comercial_analista_junior_2=comercial_analista_junior_2,
                comercial_analista_junior_3=comercial_analista_junior_3,
                comercial_analista_junior_4=comercial_analista_junior_4,
                comercial_analista_junior_5=comercial_analista_junior_5,
                comercial_analista_junior_6=comercial_analista_junior_6,
                comercial_analista_junior_7=comercial_analista_junior_7,
                comercial_analista_junior_8=comercial_analista_junior_8,
                exportaciones_analista_junior_1=exportaciones_analista_junior_1,
                exportaciones_analista_junior_2=exportaciones_analista_junior_2,
                exportaciones_analista_junior_3=exportaciones_analista_junior_3,
                exportaciones_analista_junior_4=exportaciones_analista_junior_4,
                exportaciones_analista_junior_5=exportaciones_analista_junior_5,
                desarrollo_empresarial_analista_junior_1=desarrollo_empresarial_analista_junior_1,
                administrativo_financiero_analista_junior_1=administrativo_financiero_analista_junior_1,
                mercadeo_innovacion_analista_junior_1=mercadeo_innovacion_analista_junior_1,
                operaciones_logistica_analista_junior_1=operaciones_logistica_analista_junior_1,
                analista_senior_2=analista_senior_2,

            )
            perfil.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el perfil ' + (nombre).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/perfiles_pcs/')

def config_perfiles_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        perfiles = Perfiles_PCS.objects.get(pk=id)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "config_perfiles_pcs_registrar.html", {'user': current_user,
                                                                         'permisos': permisos,
                                                                         'perfiles': perfiles,
                                                                        'permiso_usuario': usuario_datos,})

def config_perfiles_borrar(request, id):
    if request.method == 'GET':

        perfiles = Perfiles_PCS.objects.get(pk=id)
        usuario_actual = request.user
        current_user = request.user

        if perfiles.usuarios_datos_set.all().exists():
            messages.add_message(request, messages.ERROR,
                                 'No se puede borrar el perfil ' + str(
                                     id) + ' porque tiene un usuario asociado')

            return HttpResponseRedirect('/configuracion/perfiles_pcs/')

        else:

            perfiles.delete()
            messages.add_message(request, messages.WARNING,
                                 'Se ha borrado el perfil ' + str(id) + ' satisfactoriamente')

            return HttpResponseRedirect('/configuracion/perfiles_pcs/')

def config_justificacion_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        #usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.registrar_localizacion == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            #return HttpResponseRedirect('/configuracion/continentes/')

        return render(request, "config_justificacion_registrar.html", {'user': current_user,
                                                                     #'permiso_usuario': usuario_datos
                                                                     })

    elif request.method == 'POST':
        current_user = request.user
        #usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        #if usuario_datos.perfiles.registrar_localizacion == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            #return HttpResponseRedirect('/configuracion/continentes/')

        descripcion = request.POST['descripcion']

        justificacion = Justificacion(
            descripcion=descripcion,

        )
        justificacion.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado la justificacion ' + (descripcion).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/justificacion/')


def config_justificacion_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        justificacion = Justificacion.objects.get(pk=id)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "config_justificacion_editar.html", {'user': current_user,
                                                                         'permisos': permisos,
                                                                         'justificacion': justificacion,
                                                                        'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


        id = request.POST['id']
        descripcion = request.POST['descripcion']


        justificacion = Justificacion(
            id=id,
            descripcion=descripcion,

        )
        justificacion.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado la justificacion ' + (descripcion).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/justificacion/')


def config_justificacion_borrar(request, id):
    if request.method == 'GET':

        justificacion = Justificacion.objects.get(pk=id)
        usuario_actual = request.user
        current_user = request.user
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        justificacion.delete()
        messages.add_message(request, messages.WARNING,
                                'Se ha borrado la justificacion ' + str(id) + ' satisfactoriamente')

        return HttpResponseRedirect('/configuracion/justificacion/')


# Continentes
def config_continentes(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        lista_continentes = Continentes.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.consultar_localizacion == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            #return HttpResponseRedirect('/administracion/')

        return render(request, "config_continentes.html", {'user': current_user,
                                                          'lista_continentes': lista_continentes,
                                                           'permiso_usuario': usuario_datos
                                                           })
    else:
        pass

def config_continentes_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        #usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.registrar_localizacion == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            #return HttpResponseRedirect('/configuracion/continentes/')

        return render(request, "config_continentes_registrar.html", {'user': current_user,
                                                                     #'permiso_usuario': usuario_datos
                                                                     })

    elif request.method == 'POST':
        current_user = request.user
        #usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        #if usuario_datos.perfiles.registrar_localizacion == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            #return HttpResponseRedirect('/configuracion/continentes/')

        codigo = request.POST['codigo']
        descripcion = request.POST['descripcion']

        lista_continentes = Continentes.objects.all()
        lista_continentes = lista_continentes.filter(id=id)
        if lista_continentes.exists():
            messages.add_message(request, messages.ERROR,
                                 'Registro ya existe con el codigo ' + (codigo).encode(
                                     'utf-8').strip())

            return HttpResponseRedirect('/configuracion/continentes/')

        continentes = Continentes(

            codigo=codigo,
            descripcion=descripcion,

        )
        continentes.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el Continente ' + (descripcion).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/continentes/')

def config_continentes_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        continentes = Continentes.objects.get(pk=id)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "config_continentes_editar.html", {'user': current_user,
                                                                         'permisos': permisos,
                                                                         'continentes': continentes,
                                                                        'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


        id = request.POST['id']
        codigo = request.POST['codigo']
        descripcion = request.POST['descripcion']


        continentes = Continentes(
            id=id,
            codigo=codigo,
            descripcion=descripcion,

        )
        continentes.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado el Continente ' + (descripcion).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/continentes/')

def config_continentes_borrar(request, id):

        if request.method == 'GET':

            continente = Continentes.objects.get(pk=id)
            usuario_actual = request.user
            current_user = request.user
            permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
            usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


            if continente.paises_set.all().exists():
                messages.add_message(request, messages.ERROR,
                                     'No se puede borrar el continente ' + str(
                                         id) + ' porque tiene un pais asociado')

                return HttpResponseRedirect('/configuracion/continentes/')

            else:


                continente.delete()
                messages.add_message(request, messages.WARNING,
                                     'Se ha borrado el Continente ' + str(id) + ' satisfactoriamente')

                return HttpResponseRedirect('/configuracion/continentes/')



# Paises
def config_paises(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        #permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_paises = Paises.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.consultar_localizacion == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            #return HttpResponseRedirect('/administracion/')

        return render(request, "config_paises.html", {'user': current_user,
                                                          'lista_paises': lista_paises,
                                                          #'permisos': permisos,
                                                      'permiso_usuario': usuario_datos
                                                      })
    else:
        pass

def config_paises_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        #usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        #permisos = Permisos.objects.filter(usuario_id=current_user.id).first()

        lista_continentes = Continentes.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.registrar_localizacion == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            #return HttpResponseRedirect('/configuracion/paises/')

        return render(request, "config_paises_registrar.html", {'user': current_user,
                                                                   'lista_continentes': lista_continentes,
                                                                #'permisos': permisos,
                                                                #'permiso_usuario': usuario_datos,
                                                                })
    elif request.method == 'POST':
        current_user = request.user
        #usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        #if usuario_datos.perfiles.registrar_localizacion == False:
          #  messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
           # return HttpResponseRedirect('/configuracion/paises/')

        continentes = request.POST['continentes']
        codigo = request.POST['codigo']
        descripcion = request.POST['descripcion']


        paises = Paises(
            continentes_id=int(continentes),
            codigo=codigo,
            descripcion=descripcion,

        )
        paises.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el Pais ' + (descripcion).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/paises/')

def config_paises_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        paises = Paises.objects.get(pk=id)

        lista_continentes = Continentes.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "config_paises_editar.html", {'user': current_user,
                                                                         'permisos': permisos,
                                                                         'lista_continentes': lista_continentes,
                                                                         'paises': paises,
                                                                        'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        id = request.POST['id']
        continentes = request.POST['continentes']
        codigo = request.POST['codigo']
        descripcion = request.POST['descripcion']


        paises = Paises(
            id=None if not id else id,
            continentes_id=int (continentes),
            codigo=codigo,
            descripcion=descripcion,

        )
        paises.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado el Pais ' + (descripcion).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/paises/')

def config_paises_borrar(request, id):

        if request.method == 'GET':
            current_user = request.user
            permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
            usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
            pais = Paises.objects.get(pk=id)
            usuario_actual = request.user


            if pais.departamentos_set.all().exists():
                messages.add_message(request, messages.ERROR,
                                     'No se puede borrar el Pais ' + str(
                                         id) + ' porque tiene un Departamento asociado')

                return HttpResponseRedirect('/configuracion/paises/')

            else:

                pais.delete()
                messages.add_message(request, messages.WARNING,
                                     'Se ha borrado el Pais ' + str(id) + ' satisfactoriamente')

                return HttpResponseRedirect('/configuracion/paises/')

def config_paises_datos(request):
    if request.method == 'GET':

        lista_pais = Paises.objects.all()

        cuenta = lista_pais.count()
        paginador = Paginator(lista_pais, PAGINADOR)
        pagina = request.GET.get('page')

        try:
            paises = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            paises = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            paises = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        paises_dict = []

        for pais in paises:
            pais_dict = {
                'id': pais.id,
                'codigo': pais.codigo,
                'descripcion': pais.descripcion,
                'continentes': str(pais.continentes_id),

            }
            paises_dict.append(pais_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': paises_dict
        }

        return JsonResponse(response_dict)



# Departamentos
def config_departamentos(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        #permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_departamentos = Departamentos.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.consultar_localizacion == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            #return HttpResponseRedirect('/administracion/')

        return render(request, "config_departamentos.html", {'user': current_user,
                                                          'lista_departamentos': lista_departamentos,
                                                          #'permisos': permisos,
                                                          'permiso_usuario': usuario_datos,
                                                             })
    else:
        pass

def config_departamentos_registrar(request):
        # Render  administracion.html
        if request.method == 'GET':
            current_user = request.user
            #usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
            #permisos = Permisos.objects.filter(usuario_id=current_user.id).first()

            lista_paises = Paises.objects.all()
            if not current_user.is_staff:
                return HttpResponseRedirect('/login/')

            #if usuario_datos.perfiles.registrar_localizacion == False:
                #messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
                #return HttpResponseRedirect('/configuracion/departamentos/')

            return render(request, "config_departamentos_registrar.html", {'user': current_user,
                                                                    'lista_paises': lista_paises,
                                                                    #'permisos': permisos,
                                                                    #'permiso_usuario': usuario_datos,
                                                                           })
        elif request.method == 'POST':

            current_user = request.user
            #usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

            #if usuario_datos.perfiles.registrar_localizacion == False:
                #messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
                #return HttpResponseRedirect('/configuracion/departamentos/')


            paises = request.POST['paises']
            codigo = request.POST['codigo']
            descripcion = request.POST['descripcion']

            departamentos = Departamentos(
                paises_id=int(paises),
                codigo=codigo,
                descripcion=descripcion,

            )
            departamentos.save()

            messages.add_message(request, messages.INFO,
                                 'Se ha registrado el Departamento ' + (descripcion).encode(
                                     'utf-8').strip() + ' satisfactoriamente.')

            return HttpResponseRedirect('/configuracion/departamentos/')

def config_departamentos_editar(request, id):
            # Render  administracion.html
            if request.method == 'GET':
                current_user = request.user
                usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
                permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
                departamentos = Departamentos.objects.get(pk=id)

                lista_paises = Paises.objects.all()
                if not current_user.is_staff:
                    return HttpResponseRedirect('/login/')

                return render(request, "config_departamentos_editar.html", {'user': current_user,
                                                                     'permisos': permisos,
                                                                     'lista_paises': lista_paises,
                                                                     'departamentos': departamentos,
                                                                    'permiso_usuario': usuario_datos,})
            elif request.method == 'POST':
                current_user = request.user
                usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()



                id = request.POST['id']
                paises = request.POST['paises']
                codigo = request.POST['codigo']
                descripcion = request.POST['descripcion']

                departamentos = Departamentos(
                    id=None if not id else id,
                    paises_id=int(paises),
                    codigo=codigo,
                    descripcion=descripcion,

                )
                departamentos.save()

                messages.add_message(request, messages.SUCCESS,
                                     'Se ha editado el Departamento ' + (descripcion).encode(
                                         'utf-8').strip() + ' satisfactoriamente.')

                return HttpResponseRedirect('/configuracion/departamentos/')

def config_departamentos_borrar(request, id):

                if request.method == 'GET':
                    current_user = request.user
                    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
                    departamento = Departamentos.objects.get(pk=id)
                    usuario_actual = request.user

                    if departamento.municipios_set.all().exists():
                        messages.add_message(request, messages.ERROR,
                                             'No se puede borrar el Departamento ' + str(
                                                 id) + ' porque tiene un Municipio asociado')

                        return HttpResponseRedirect('/configuracion/departamentos/')

                    else:

                        departamento.delete()
                        messages.add_message(request, messages.WARNING,
                                             'Se ha borrado el Departamento ' + str(id) + ' satisfactoriamente')

                        return HttpResponseRedirect('/configuracion/departamentos/')

def config_departamentos_datos(request):
    if request.method == 'GET':

        lista_departamento = Departamentos.objects.all()

        cuenta = lista_departamento.count()
        paginador = Paginator(lista_departamento, PAGINADOR)
        pagina = request.GET.get('page')

        try:
            departamentos = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            departamentos = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            departamentos = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada departamento
        departamentos_dict = []

        for departamento in departamentos:
            departamento_dict = {
                'id': departamento.id,
                'codigo': departamento.codigo,
                'descripcion': departamento.descripcion,
                'paises': str(departamento.paises_id),

            }
            departamentos_dict.append(departamento_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': departamentos_dict
        }

        return JsonResponse(response_dict)

# Municipios
def config_municipios(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        #permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_municipios = Municipios.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.consultar_localizacion == False:
            #messages.add_message(request, messages.ERROR,
              #                   'No tienes permitido el acceso a ese modulo')
            #return HttpResponseRedirect('/administracion/')

        return render(request, "config_municipios.html", {'user': current_user,
                                                             'lista_municipios': lista_municipios,
                                                           #  'permisos': permisos,
                                                          'permiso_usuario': usuario_datos,
                                                          })
    else:
        pass

def config_municipios_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        #usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        #permisos = Permisos.objects.filter(usuario_id=current_user.id).first()

        lista_departamentos = Departamentos.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.registrar_localizacion == False:
            #messages.add_message(request, messages.ERROR,
             #                    'No tienes permisos para registrar en este modulo')
            #return HttpResponseRedirect('/configuracion/municipios/')

        return render(request, "config_municipios_registrar.html", {'user': current_user,
                                                                       'lista_departamentos': lista_departamentos,
                                                                     #  'permisos': permisos,
                                                                    #'permiso_usuario': usuario_datos,
                                                                    })
    elif request.method == 'POST':

        current_user = request.user
        #usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        #if usuario_datos.perfiles.registrar_localizacion == False:
         #   messages.add_message(request, messages.ERROR,
          #                       'No tienes permisos para registrar en este modulo')
           # return HttpResponseRedirect('/configuracion/municipios/')

        departamentos = request.POST['departamentos']
        codigo = request.POST['codigo']
        descripcion = request.POST['descripcion']

        municipios = Municipios(
            departamentos_id=int(departamentos),
            codigo=codigo,
            descripcion=descripcion,

        )
        municipios.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el Municipio ' + (descripcion).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/municipios/')

def config_municipios_editar(request, id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        municipios = Municipios.objects.get(pk=id)

        lista_departamentos = Departamentos.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "config_municipios_editar.html", {'user': current_user,
                                                                    'permisos': permisos,
                                                                    'lista_departamentos': lista_departamentos,
                                                                    'municipios': municipios,
                                                                 'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


        id = request.POST['id']
        departamentos = request.POST['departamentos']
        codigo = request.POST['codigo']
        descripcion = request.POST['descripcion']

        municipios = Municipios(
            id=None if not id else id,
            departamentos_id=int(departamentos),
            codigo=codigo,
            descripcion=descripcion,

        )
        municipios.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado el Municipio ' + (descripcion).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/municipios/')

def config_municipios_borrar(request, id):

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        municipio = Municipios.objects.get(pk=id)
        usuario_actual = request.user

        

        municipio.delete()
        messages.add_message(request, messages.WARNING,
                             'Se ha borrado el Departamento ' + str(id) + ' satisfactoriamente')

        return HttpResponseRedirect('/configuracion/municipios/')

def config_municipios_datos(request):

    if request.method == 'GET':

        lista_municipio = Municipios.objects.all()

        cuenta = lista_municipio.count()
        paginador = Paginator(lista_municipio, PAGINADOR)
        pagina = request.GET.get('page')

        try:
            municipios = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            municipios = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            municipios = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada municipio
        municipios_dict = []

        for municipio in municipios:
            municipio_dict = {
                'id': municipio.id,
                'codigo': municipio.codigo,
                'descripcion': municipio.descripcion,
                'departamentos': str(municipio.departamentos_id),

            }
            municipios_dict.append(municipio_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': municipios_dict
        }

        return JsonResponse(response_dict)




# Areas de Atencion
def config_areas_atencion(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        areas = AreasAtencion.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_areas_atencion.html", {'user': current_user,
                                                             'areas': areas,
                                                          'permiso_usuario': usuario_datos,
                                                          })
    else:
        pass

def config_areas_atencion_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_areas_atencion_registrar.html", {'user': current_user,
                                                                    'permiso_usuario': usuario_datos,
                                                                    })
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        nombre = request.POST['nombre']

        areas = AreasAtencion(
            descripcion=nombre,
        )
        areas.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el area ' + nombre + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/areas_atencion/')

def config_areas_atencion_editar(request, id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        areas = AreasAtencion.objects.get(pk=id)
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "config_areas_atencion_editar.html", {'user': current_user,
                                                                    'permisos': permisos,
                                                                    'areas': areas,
                                                                 'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


        nombre = request.POST['nombre']
        id = request.POST['id']

        areas = AreasAtencion(
            id=None if not id else id,
            descripcion=nombre,

        )
        areas.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado el grupo de atencion ' + nombre + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/areas_atencion/')

def config_areas_atencion_editar_integrantes(request, id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        areas = AreasAtencion.objects.get(pk=id)
        usuarios=Usuarios_datos.objects.filter(~Q(atencion=1))
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "config_areas_atencion_editar_integrantes.html", {'user': current_user,
                                                                    'permisos': permisos,
                                                                    'areas': areas,
                                                                    'usuarios': usuarios,
                                                                 'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


        integrante = int(request.POST['integrante'])
        id = request.POST['id']
        persona=Usuarios_datos.objects.filter(usuario_id=integrante).first()
        personas = PersonasAtencion(
            area_id=int(id),
            nombre=persona.usuario.first_name + ' ' + persona.usuario.last_name ,
            telefono=persona.telefono,
            email=persona.usuario.email,
            usuario=persona.usuario.username,
        )
        personas.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha agregado al usuario' + persona.usuario.username + ' al grupo satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/areas_atencion/')

def config_areas_atencion_borrar(request, id):
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        areas = AreasAtencion.objects.get(pk=id)
        usuario_actual = request.user
        if areas.peticiones_set.all().exists():
            messages.add_message(request, messages.ERROR,
                                 'No se puede borrar el grupo ' + str(
                                     id) + ' porque tiene una peticion asociado')

            return HttpResponseRedirect('/configuracion/areas_atencion/')

        else:

            areas.delete()
            messages.add_message(request, messages.WARNING,
                                 'Se ha borrado el grupo de atencion ' + str(id) + ' satisfactoriamente')

            return HttpResponseRedirect('/configuracion/areas_atencion/')


#Peticiones

def config_peticiones(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        peticiones = Peticiones.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_peticiones.html", {'user': current_user,
                                                             'peticiones': peticiones,
                                                          'permiso_usuario': usuario_datos,
                                                          })
    else:
        pass

def config_peticiones_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        lista_areas = AreasAtencion.objects.all()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_peticiones_registrar.html", {'user': current_user,
                                                                    'permiso_usuario': usuario_datos,
                                                                    'lista_areas': lista_areas,
                                                                    })
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        area = request.POST['area']
        nombre = request.POST['nombre']

        peticion = Peticiones(
            area_id=area,
            descripcion=nombre,

        )
        peticion.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado la peticion ' + nombre + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/peticion/')

def config_peticiones_editar(request, id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        peticiones = Peticiones.objects.get(pk=id)

        lista_areas = AreasAtencion.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "config_peticiones_editar.html", {'user': current_user,
                                                                    'permisos': permisos,
                                                                    'peticiones': peticiones,
                                                                    'lista_areas': lista_areas,
                                                                 'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


        id = request.POST['id']
        area = request.POST['area']
        nombre = request.POST['nombre']

        peticiones = Peticiones(
            id=None if not id else id,
            area_id=area,
            descripcion=nombre,

        )
        peticiones.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado la peticion ' + nombre + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/peticion/')

def config_peticiones_borrar(request, id):
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        peticion = Peticiones.objects.get(pk=id)
        usuario_actual = request.user

        peticion.delete()
        messages.add_message(request, messages.WARNING,
                             'Se ha borrado la peticion ' + str(id) + ' satisfactoriamente')

        return HttpResponseRedirect('/configuracion/peticion/')

#Personas Atencion
def config_personas_atencion(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        areas = PersonasAtencion.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_persona_atencion.html", {'user': current_user,
                                                             'areas': areas,
                                                          'permiso_usuario': usuario_datos,
                                                          })
    else:
        pass

def config_personas_aten_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        lista_areas = AreasAtencion.objects.all()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_personas_aten_registrar.html", {'user': current_user,
                                                                    'permiso_usuario': usuario_datos,
                                                                    'lista_areas': lista_areas,
                                                                    })
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        area = request.POST['area']
        nombre = request.POST['nombre']
        telefono = request.POST['telefono']
        email = request.POST['email']

        personas = PersonasAtencion(
            area_id=area,
            nombre=nombre,
            telefono=telefono,
            email=email,
        )
        personas.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado la peticion ' + nombre + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/personas_aten/')

def config_personas_aten_editar(request, id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        personas = PersonasAtencion.objects.get(pk=id)

        lista_areas = AreasAtencion.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "config_personas_atencion_editar.html", {'user': current_user,
                                                                    'permisos': permisos,
                                                                    'personas': personas,
                                                                    'lista_areas': lista_areas,
                                                                 'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


        id = request.POST['id']
        area = request.POST['area']
        nombre = request.POST['nombre']
        telefono = request.POST['telefono']
        email = request.POST['email']

        personas = PersonasAtencion(
            id=None if not id else id,
            area_id=area,
            nombre=nombre,
            telefono=telefono,
            email=email,
        )
        personas.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado el integrante ' + nombre + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/personas_aten/')

def config_personas_aten_borrar(request, id):
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        personas = PersonasAtencion.objects.get(pk=id)
        usuario_actual = request.user

        personas.delete()
        messages.add_message(request, messages.WARNING,
                             'Se ha borrado el integrante ' + str(id) + ' satisfactoriamente')

        return HttpResponseRedirect('/configuracion/personas_aten/')

def config_solicitudes(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        personas = PersonasAtencion.objects.filter(usuario=usuario_datos.usuario.username)
        condicion_facturar_sin_novedades = Q(area__descripcion='Facturar pedido sin novedades')
        condicion_facturar_con_novedades = Q(area__descripcion='Facturar pedido con novedades')
        condiciones = condicion_facturar_sin_novedades | condicion_facturar_con_novedades
        personas_con_descripcion_especifica = PersonasAtencion.objects.filter(
            usuario=usuario_datos.usuario.username
        ).filter(condiciones)
        hay_personas_con_descripcion_especifica = personas_con_descripcion_especifica.exists()
        resultado_booleano = hay_personas_con_descripcion_especifica
        dato_lista = []
        for persona in personas:
            peticiones=Peticiones.objects.filter(area_id=persona.area_id)
            for peticion in peticiones:
                solicitudes=RespuestaPedido.objects.filter(peticion_id=peticion.id,estado='pendiente').order_by('-fecha')
                for solicitud in solicitudes:
                    data_prueba = {
                        'id': solicitud.id,
                        'entry': solicitud.entry_pedido,
                        'empresa': solicitud.empresa,
                        'pedido': solicitud.num_pedido,
                        'peticion': solicitud.peticion.descripcion,
                        'fecha':solicitud.fecha ,
                        'estado': solicitud.estado,
                        'respuesta': solicitud.respuesta,
                    }
                    dato_lista.append(data_prueba)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_solicitudes.html", {'user': current_user,
                                                             'solicitudes': dato_lista,
                                                             'resultado_booleano': resultado_booleano,
                                                          'permiso_usuario': usuario_datos,
                                                          })
    else:
        pass

def config_respuesta_factura_pedido(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        pedidos=RespuestaPedido.objects.filter(
            Q(peticion__descripcion='Facturar pedido sin novedades') |
            Q(peticion__descripcion='Facturar Pedido con novedades'),
            estado='pendiente'
        )
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_respuesta_factura_pedido.html", {'user': current_user,
                                                          'permiso_usuario': usuario_datos,
                                                          'pedidos': pedidos,
                                                          })
    else:
        hoy = date.today()
        hora = datetime.now().time()
        pedidos = request.POST.getlist('pedido[]')
        respuestas = request.POST.getlist('respuesta[]')
        adjuntos = request.FILES.getlist('adjunto[]')
        for pedido, respuesta, adjunto in zip(pedidos, respuestas, adjuntos):
            info_pedido=RespuestaPedido.objects.filter(id=int(pedido)).first()
            if adjunto:
                fs = FileSystemStorage()
                filename = fs.save(adjunto.name, adjunto)
                uploaded_file_url = fs.url(filename)
                RespuestaPedido.objects.filter(id=int(pedido)).update(respuesta=respuesta, estado='respondido',
                                                                   doc_respuesta=uploaded_file_url)
                email = EmailMessage('RESPUESTA PETICION ' + str(info_pedido.num_pedido),
                                     str(respuesta) + ', Documento adjunto: ' + IP_SERVIDOR + str(
                                         uploaded_file_url),
                                     to=[info_pedido.email])
                email.send()
                log = LogRespuestaPedido(
                    peticion_id=info_pedido.peticion.id,
                    num_pedido=info_pedido.num_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=info_pedido.email,
                    empresa=str(info_pedido.empresa),
                    accion='Respuesta a la solicitud'
                )
                log.save()
        messages.add_message(request, messages.INFO,
                             'Se han subido las facturas satisfactoriamente.')
        return HttpResponseRedirect('/configuracion/respuesta_factura/')

def config_ordenes_otroscanales(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.codigo
        url2 = IP_SAP + "SQLQueries('ProductosOtrosPedido9')/List?cliente='" + empresa +"'"

        response = sap_request(url2)
        response = response.text
        response = response.replace('null', ' " " ')
        response = ast.literal_eval(response)
        response = response['value']
        dato_lista=[]
        for datos in response:
            data_prueba = {
                'n_producto': datos['CodeBars'],
                'nombre': datos['ItemName'],
                'plu': datos['U_PLU'],
                'codigo': datos['ItemCode'],

            }
            dato_lista.append(data_prueba)
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_pedidos_otroscanales.html", {'user': current_user,
                                                          'permiso_usuario': usuario_datos,
                                                          'pedidos': dato_lista,
                                                          })
    else:

        hoy = date.today()
        hora = datetime.now().time()
        descripcion = request.POST.getlist('nombre[]')
        necesidad = request.POST.getlist('necesidad[]')
        observaciones = request.POST.getlist('observaciones[]')
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        ultimo_registro = PedidosOtrosCanales.objects.all().last()
        if ultimo_registro == None:
            consecutivo = 1
        else:
            consecutivo = ultimo_registro.num_pedido + 1

        pedidootros = PedidosOtrosCanales(
            num_pedido=consecutivo,
            empresa=usuario_datos.empresa,
            fecha=hoy,
            hora=hora,
        )
        pedidootros.save()
        for descripcion, necesidad, observaciones in zip(descripcion, necesidad, observaciones):
            descripcion, nombre = descripcion.split('_')

            detallepedidootros = DetallesPedidosOtrosCanales(
                num_pedido_id=int(consecutivo),
                cantidad=necesidad,
                referencia=descripcion,
                nombre=nombre,
                observaciones=observaciones,
            )
            detallepedidootros.save()

        usuarios_correo = Usuarios_datos.objects.filter(perfil_pcs_id__in=[8, 21, 23])

        for correos in usuarios_correo:
            email = EmailMessage(' Nuevo Pedido Recibido de Otros Canales – Pedido  '+str(pedidootros),
                                 'Te informamos que has recibido un nuevo pedido a través de otros canales de la empresa . '+str(usuario_datos.empresa.nombre),
                                 to=[correos.usuario.email])
            email.send()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el pedido satisfactoriamente')
        return HttpResponseRedirect('/configuracion/orden_cliente_otroscanales/')

def obtener_imagen_producto(request):
    numero_producto = request.GET.get('numero')
    try:
        numero_producto = numero_producto.split('_')[0]
        # Realiza la consulta a la base de datos para obtener la URL de la imagen
        imagen_url = ImagenesOtrosCanales.objects.get(referencia=numero_producto).imagen
        return JsonResponse({'imagen': imagen_url})
    except Imagenes.DoesNotExist:
        return JsonResponse({'imagen': ''})

def config_ordenes_otroscanales_pcs(request):

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        usuarios= Empresas.objects.filter(tipo='Cadena')
        pedidos = PedidosOtrosCanales.objects.filter(
            Q(estado='en proceso') | Q(estado='predistribuido')
        )



        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_ordenes_otroscanales_pcs.html", {'user': current_user,
                                                        'usuarios': usuarios,
                                                        'pedidos': pedidos,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    else:
        pass

def config_ordenes_otroscanales_pcs_cliente(request):

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        usuarios= Empresas.objects.filter(tipo='Cadena')
        pedidos = PedidosOtrosCanales.objects.filter(estado='en proceso',empresa_id=usuario_datos.empresa.pk)



        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_ordenes_otroscanales_cliente.html", {'user': current_user,
                                                        'usuarios': usuarios,
                                                        'pedidos': pedidos,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    else:
        pass

def config_ordenes_otroscanales_empresario(request):

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        usuarios = Usuarios_datos.objects.filter(empresa__tipo='OtrosPedidos')
        empresas=Empresas.objects.filter(tipo='Empresario')



        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_ordenes_otroscanales_empresarios.html", {'user': current_user,
                                                        'usuarios': usuarios,
                                                        'permiso_usuario': usuario_datos,
                                                        'empresas': empresas,
                                                        })
    else:
        pass

def generar_pdf_bytes(contexto, template_name='pedido_asignacion_pdf.html'):
    """
    Renderiza un template como PDF y devuelve el contenido en bytes.
    """
    template = get_template(template_name)
    html = template.render(contexto)

    result = BytesIO()
    pdf = pisa.pisaDocument(
        BytesIO(html.encode('utf-8')),
        dest=result,
        link_callback=link_callback   # 👈 clave para resolver imágenes locales
    )

    if pdf.err:
        return None
    return result.getvalue()

def descargar_pedidos_zip(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            ids = data.get('ids', [])

            zip_buffer = BytesIO()
            zip_file = zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED)

            # Agrupar asignaciones por (pedido, empresa)
            grupos = defaultdict(list)
            asignaciones = AsignacionPedidosOtrosCanales.objects.filter(pk__in=ids)

            for asignacion in asignaciones:
                clave = (asignacion.num_detalle.num_pedido.num_pedido, asignacion.empresa.id)
                grupos[clave].append(asignacion)

            # Generar un PDF por grupo
            for (pedido_id, empresa_id), asignaciones_grupo in grupos.items():
                try:
                    asignacion_titulos = asignaciones_grupo[0]

                    preciocompra = 1000  # tu lógica real de precio unitario

                    # Preparamos filas con sus subtotales
                    filas = []
                    subtotal = 0
                    for a in asignaciones_grupo:
                        total_fila = a.cantidad * preciocompra
                        imagen_obj = ImagenesOtrosCanales.objects.filter(referencia=a.num_detalle.referencia).first()
                        if imagen_obj:
                            if imagen_obj.imagen.startswith("/"):  # ya viene con /
                                imagen_url = imagen_obj.imagen
                            else:
                                imagen_url = settings.MEDIA_URL + imagen_obj.imagen
                        else:
                            imagen_url = None
                        filas.append({
                            'referencia': a.num_detalle.referencia,
                            'nombre': a.num_detalle.nombre,
                            'cantidad': a.cantidad,
                            'preciocompra': preciocompra,
                            'total_fila': total_fila,
                            'imagen': imagen_url,
                        })
                        subtotal += total_fila

                    iva = int(subtotal * 0.19)  # ejemplo: 19% de IVA
                    total = subtotal + iva

                    contexto = {
                        'form_id': pedido_id,
                        'filas': filas,  # 👈 ya no pasamos asignacion cruda
                        'asignacion_titulos': asignacion_titulos,
                        'subtotal': subtotal,
                        'iva': iva,
                        'total': total,
                        'clave': 'prueba'
                    }

                    pdf_bytes = generar_pdf_bytes(contexto)
                    nombre_archivo = "{}_{}".format(
                        asignacion_titulos.num_detalle.num_pedido.num_pedido,
                        asignacion_titulos.empresa.nombre
                    )
                    if pdf_bytes:
                        filename = 'pedido_%s.pdf' % nombre_archivo
                        zip_file.writestr(filename, pdf_bytes)
                except Exception:
                    continue  # si falla uno, seguimos

            zip_file.close()
            zip_buffer.seek(0)

            response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename=pedidos.zip'
            return response

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

def config_ordenes_otroscanales_empresario_facturacion(request):

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        usuarios = Usuarios_datos.objects.filter(empresa__tipo='OtrosPedidos')
        empresas=Empresas.objects.all()



        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_ordenes_otroscanales_empresarios_facturar.html", {'user': current_user,
                                                        'usuarios': usuarios,
                                                        'permiso_usuario': usuario_datos,
                                                        'empresas': empresas,
                                                        })
    else:
        pass

def config_ordenes_otroscanales_empresario_recibo(request):

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        usuarios = Usuarios_datos.objects.filter(empresa__tipo='OtrosPedidos')
        empresas=Empresas.objects.filter(tipo='Empresario')



        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_ordenes_otroscanales_empresarios_recibo.html", {'user': current_user,
                                                        'usuarios': usuarios,
                                                        'permiso_usuario': usuario_datos,
                                                        'empresas': empresas,
                                                        })
    else:
        pass

def config_imagen_otroscanales_empresario(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        lista_imagenes = ImagenesOtrosCanales.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.consultar_localizacion == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            #return HttpResponseRedirect('/administracion/')

        return render(request, "config_imagen_otroscanales_empresario.html", {'user': current_user,
                                                          'lista_imagenes': lista_imagenes,
                                                           'permiso_usuario': usuario_datos
                                                           })
    else:
        pass

def config_imagen_otroscanales_empresario_borrar(request, id):

        if request.method == 'GET':

            imagen = ImagenesOtrosCanales.objects.get(pk=id)
            usuario_actual = request.user
            current_user = request.user
            permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
            usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()



            imagen.delete()
            messages.add_message(request, messages.WARNING,
                                     'Se ha borradola imagen ' + str(id) + ' satisfactoriamente')

            return HttpResponseRedirect('/configuracion/imagen_empresiario_otroscanales/')

def config_imagen_otroscanales_empresario_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.registrar_localizacion == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            #return HttpResponseRedirect('/configuracion/continentes/')

        return render(request, "config_imagen_otroscanales_empresario_resgistrar.html", {'user': current_user,
                                                                     'permiso_usuario': usuario_datos
                                                                     })

    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        referencia = request.POST['referencia']

        try:
            adjunto = request.POST['adjunto']
        except:
            adjunto = request.FILES['adjunto']

        if adjunto != '':
            myfile = request.FILES['adjunto']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = fs.url(filename)
        else:
            uploaded_file_url = None

        url3 = IP_SAP + "SQLQueries('Consultasdescripcionporref')/List?referencia='" + str(
            referencia) + "'"

        response1 = sap_request(url3)
        response1 = response1.text
        response1 = response1.replace('null', ' " " ')
        response1 = ast.literal_eval(response1)
        response1 = response1['value'][0]
        descripcion = response1['ItemName']

        imagenes = ImagenesOtrosCanales(

            referencia=referencia,
            descripcion=descripcion,
            imagen=uploaded_file_url,

        )
        imagenes.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado la imagen para la referencia ' + (referencia).encode('utf-8').strip() + ' satisfactoriamente.')



        return HttpResponseRedirect('/configuracion/imagen_empresiario_otroscanales/')

def config_imagen_otroscanales_empresario_registrar_masivo(request):
    # Render  administracion.html
    carpeta = os.path.join(settings.MEDIA_ROOT, 'imagenes')

    if not os.path.exists(carpeta):
        os.makedirs(carpeta)

    for archivo in os.listdir(carpeta):
        if archivo.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):

            # La referencia será el nombre del archivo sin extensión
            referencia = os.path.splitext(archivo)[0]

            # Construir la URL accesible
            ruta_url = os.path.join(settings.MEDIA_URL, 'imagenes', archivo)

            # Evitar duplicados (mismo nombre de archivo ya guardado)
            if ImagenesOtrosCanales.objects.filter(referencia=referencia).exists():
                continue

            # Crear registro
            ImagenesOtrosCanales.objects.create(
                referencia=referencia,
                imagen=ruta_url,
                descripcion=""  # vacío por ahora
            )

    return HttpResponseRedirect('/configuracion/imagen_empresiario_otroscanales/' )

def config_ordenes_otroscanales_pcs_detalle(request, id):

    if request.method == 'GET':
        identificacion_asignacion = request.GET.get('asignacion')
        asignacion_editar=AsignacionPedidosOtrosCanales.objects.filter(pk=identificacion_asignacion).first()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        pedido = PedidosOtrosCanales.objects.filter(num_pedido=id).first()
        pedidodetalle = DetallesPedidosOtrosCanales.objects.filter(num_pedido=id,multiple=False)
        dato_lista=[]
        pedidodetalle_multiple = DetallesPedidosOtrosCanales.objects.filter(num_pedido=id,multiple=True)
        dato_lista_multiple = []
        pedidodetalle_plu = DetallesPedidosOtrosCanales_plus.objects.filter(num_pedido=id)
        dato_lista_plu = []

        for detalles_plu in pedidodetalle_plu:
            articulo_lista_plu = []
            asignaciones_plu = AsignacionPedidosOtrosCanales.objects.filter(num_detalle_id=detalles_plu.id)
            articulos_plu = MaestroArticulo.objects.filter(u_plu=detalles_plu.u_plu)
            for articulo_plu in articulos_plu:
                articulo_lista_plu.append(articulo_plu.proveedorCodigo)
            data_detalle_plu = {
                'id': detalles_plu.id,
                'u_plu': detalles_plu.u_plu,
                'cantidad': detalles_plu.cantidad,
                'observaciones': detalles_plu.observaciones,
                'asignaciones': asignaciones_plu,
                'articulos': articulo_lista_plu
            }
            dato_lista_plu.append(data_detalle_plu)

        for detalles in pedidodetalle:
            articulo_lista = []
            asignaciones = AsignacionPedidosOtrosCanales.objects.filter(num_detalle_id=detalles.id)
            #articulos = MaestroArticulo.objects.filter(u_plu=detalles.u_plu)
            articulos = MaestroArticulo.objects.filter(itemCode=detalles.referencia)
            for articulo in articulos:
                articulo_lista.append(articulo.proveedorCodigo)
            data_detalle = {
                'id': detalles.id,
                'u_plu': detalles.u_plu,
                'referencia': detalles.referencia,
                'nombre': detalles.nombre,
                'cantidad': detalles.cantidad,
                'observaciones': detalles.observaciones,
                'asignaciones': asignaciones,
                'articulos': articulo_lista
            }
            dato_lista.append(data_detalle)

        for detalles_multiple in pedidodetalle_multiple:
            articulo_lista_multiple = []
            asignaciones = AsignacionPedidosOtrosCanales.objects.filter(num_detalle_id=detalles_multiple.id)
            #articulos = MaestroArticulo.objects.filter(u_plu=detalles.u_plu)
            articulos = MaestroArticulo.objects.filter(itemCode=detalles_multiple.referencia)
            for articulo in articulos:
                articulo_lista_multiple.append(articulo.proveedorCodigo)
            data_detalle_multiple = {
                'id': detalles_multiple.id,
                'u_plu': detalles_multiple.u_plu,
                'referencia': detalles_multiple.referencia,
                'nombre': detalles_multiple.nombre,
                'cantidad': detalles_multiple.cantidad,
                'observaciones': detalles_multiple.observaciones,
                'asignaciones': asignaciones,
                'articulos': articulo_lista_multiple
            }
            dato_lista_multiple.append(data_detalle_multiple)

        empresas=Empresas.objects.filter(tipo='Empresario')

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_ordenes_otroscanales_pcs_detalle.html", {'user': current_user,
                                                        'pedido': pedido,
                                                        'pedidodetalle_plu': dato_lista_plu,
                                                        'pedidodetalle': dato_lista,
                                                        'pedidodetalle_multiple': dato_lista_multiple,
                                                        'asignacion_editar': asignacion_editar,
                                                        'empresas': empresas,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    else:
        pass

def config_ordenes_otroscanales_pcs_eliminar(request, id):

    if request.method == 'GET':
        identificacion_asignacion = request.GET.get('asignacion')
        asignacion_editar=AsignacionPedidosOtrosCanales.objects.get(pk=identificacion_asignacion)
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


        asignacion_editar.delete()
        messages.add_message(request, messages.WARNING,
                             'Se ha borrado la asignacion ' + str(id) + ' satisfactoriamente')


        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/'+id)

    else:
        pass

def config_ordenes_otroscanales_pcs_producto_eliminar(request, id):

    if request.method == 'GET':
        current_user = request.user
        pedido_id = request.GET.get('pedido')

        asignaciones=AsignacionPedidosOtrosCanales.objects.filter(num_detalle_id=int(id))
        for asignacion in asignaciones:
            asignacion.delete()

        detalle=DetallesPedidosOtrosCanales.objects.get(pk=id)
        detalle.delete()

        messages.add_message(request, messages.WARNING,
                             'Se ha borrado la linea de producto ' + str(id) + ' satisfactoriamente')


        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/'+pedido_id)

    else:
        pass

def config_ordenes_otroscanales_pcs_detalle_edi(request):

    if request.method == 'GET':
        pedido_pk = request.GET.get('pedido_pk')
        pedido=PedidosOtrosCanales.objects.filter(pk=pedido_pk).first()
        usuario=Usuarios_datos.objects.filter(empresa_id=pedido.empresa).first()
        empresa=Empresas.objects.filter(id=usuario.empresa_id).first()
        fecha = pedido.fecha
        fechas = fecha.strftime('%Y-%m-%d')
        fecha = fechas[2:4] + fechas[5:7] + fechas[8:10]
        fecha2 = fechas[0:4] + fechas[5:7] + fechas[8:10]
        fecha_maxima = pedido.fecha_maxima
        fechas_maxima = fecha_maxima.strftime('%Y-%m-%d')
        fechas_maxima = fechas_maxima[0:4] + fechas_maxima[5:7] + fechas_maxima[8:10]
        fecha_minima = pedido.fecha_minima
        fecha_minima = fecha_minima.strftime('%Y-%m-%d')
        fecha_minima = fecha_minima[0:4] + fecha_minima[5:7] + fecha_minima[8:10]
        hora = pedido.hora
        hora = hora.strftime('%H%M')
        detalle=DetallesPedidosOtrosCanales.objects.filter(num_pedido=pedido_pk).first()
        detalle_num=detalle.referencia
        asignaciones=AsignacionPedidosOtrosCanales.objects.filter(num_detalle__num_pedido_id=pedido_pk)
        detallesasignaciones=DetallesPedidosOtrosCanales.objects.filter(num_pedido_id=pedido_pk)

        url3 = IP_SAP + "SQLQueries('consultaediempresariosdia')/List?Cliente='" + str(
            empresa.nombre) + "'"

        response1 = sap_request(url3)
        response1 = response1.text
        response1 = response1.replace('null', ' " " ')
        response1 = ast.literal_eval(response1)
        response1 = response1['value'][0]
        EanCliente=response1['U_EAN']
        condiciones_pago=response1['ExtraDays']

        if pedido.numero_pedido_cliente:
            num_ped_cliente=pedido.numero_pedido_cliente
        else:
            num_ped_cliente=pedido.num_pedido

        edi_data = "UNB+UNOA:2+"+str(EanCliente)+"+7701081000011+"+fecha+":"+hora+"+0000"+str(pedido.num_pedido)+"_4_09++ORDERS" \
                    "\nUNH+1+ORDERS:D:96A:UN:EAN008\nBGM+YB1+0000000"+str(num_ped_cliente)+"\nDTM+137:" \
                    +fecha2+"0000:203\nDTM+63:"+fechas_maxima+"0000:203\nDTM+64:"+fecha_minima+"0000:203\n" \
                    "FTX+PUR+++PRENDA DOBLADA\nNAD+BY+"+str(EanCliente)+"::9\nNAD+SN+"+str(EanCliente)+"::9\n" \
                    "NAD+SG+"+'7701001994062'+"::9\nNAD+SU+7701081000011::9\nNAD+DP+"+'7701002222222'+"::9\n" \
                    "NAD+IV+"+str(EanCliente)+"::9\nNAD+ITO+"+'7701002222222'+"::9\nPAT+1++9:3:D:"+str(condiciones_pago)
        contador=0
        for detalleasignacion in detallesasignaciones:
            contador=contador+1
            url3 = IP_SAP + "SQLQueries('ProductosEDI3')/List?producto='" + str(
                detalleasignacion.referencia) + "'"
            response1 = sap_request(url3)
            response1 = response1.text
            response1 = response1.replace('null', ' " " ')
            response1 = ast.literal_eval(response1)
            response1 = response1['value'][-1]
            if response1['U_PLU']==" ":
                plu=response1['ItemCode']
            else:
                plu=response1['U_PLU']
            edi_complement="\nLIN+"+str(contador)+"++"+response1['CodeBars']+":EN" \
                    "\nPIA+1+"+plu+":AT\nQTY+21:"+str(detalleasignacion.cantidad)+":NAR\nPRI+AAB:"+str(int(response1['Price']))+"" \
                    "\nPRI+AAA:"+str(int(response1['Price']))
            edi_data = edi_data + edi_complement
            asignacionesdetalle = AsignacionPedidosOtrosCanales.objects.filter(num_detalle=detalleasignacion.pk)
            for asignaciondetalle in asignacionesdetalle:
                nombre_emp=asignaciondetalle.empresa.nombre
                url4 = IP_SAP + "SQLQueries('consultaediempresariootros1')/List?cliente='" + str(
                    empresa.nombre) + "'&Empresario='" + str(nombre_emp) + "'"

                response1 = sap_request(url4)
                response1 = response1.text
                response1 = response1.replace('null', ' " " ')
                response1 = ast.literal_eval(response1)
                response1 = response1['value']
                if response1==[]:
                    edi_emp=str(EanCliente)
                else:
                    edi_emp=response1[0]['U_EAN']
                edi_complement = "\nLOC+7+"+str(edi_emp)+"::9\nQTY+11:" + str(asignaciondetalle.cantidad) + ":NAR"
                edi_data = edi_data + edi_complement
        edi_final="\nUNS+S\nCNT+2:2"
        edi_data = edi_data + edi_final
        edi_count=edi_data.count("\n")
        edi_finales="\nUNT+"+str(edi_count)+"+1\nUNZ+1+0000"+str(pedido.num_pedido)+"_4_09"
        edi_data = edi_data + edi_finales


        # Crear la respuesta HTTP con el contenido del archivo EDI
        response = HttpResponse(edi_data, content_type='application/edi')
        response['Content-Disposition'] = 'attachment; filename="archivo.edi"'
        return response
    else:
        pass

def config_ordenes_aviso_despacho_detalle_edi(request):

    if request.method == 'GET':
        pedido_pk = request.GET.get('pedido_pk')
        fecha = datetime.now()
        fechas = fecha.strftime('%Y-%m-%d')
        fecha = fechas[2:4] + fechas[5:7] + fechas[8:10]
        fecha_actual=fechas[0:4] + fechas[5:7] + fechas[8:10]
        hora = timezone.now()
        hora = hora.strftime('%H%M')
        url = IP_SAP + "Login"
        consecutivo_edi=Consecutivo.objects.all().first()
        consuc_edi="000"+str(consecutivo_edi.valor)


        url3 = IP_SAP + "SQLQueries('consultaavisodespacho6')/List?pedido='" + str(
            pedido_pk) + "'"

        response1 = sap_request(url3)
        response1 = response1.text
        response1 = response1.replace('null', ' " " ')
        response1 = ast.literal_eval(response1)
        response1 = response1['value']
        dato_cambiante_cantidad_cajas="esteesundatocambiantedecajas"
        for datos in response1:
            fecha_minima = datos['DocDate']
            fechas_maxima = datos['DocDueDate']
            numero_pedido=datos['DocNum']
            GLNCliente=datos['U_EAN']
            NumeroOrden=datos['NumAtCard']
            NumeroFactura=datos['NumeroFactura']
            edi_data = "UNB+UNOA:2+7701081000011+"+datos['U_EAN']+"+"+fecha+":"+hora+'+PCS'+str(consuc_edi)+"+PASSWORD+DESADV" \
                        "\nUNH+1+DESADV:D:96A:UN:EAN005\nBGM+351+"+str(consuc_edi)+"+9\nDTM+137:" \
                        +fecha_minima+"0000:203\nDTM+11:"+fecha_actual+"0000:203\nDTM+17:"+fechas_maxima+"0000:203\n" \
                        "RFF+ON:"+datos['NumAtCard']+"\nRFF+IV:"+str(datos['NumeroFactura']) +"\nNAD+DP+"+datos['EanTienda']+ "::9\nRFF+VA:"+datos['LicTradNum'][:-2]+"\nNAD+BY+"+datos['U_EAN']+"::9\n" \
                        "RFF+VA:"+datos['LicTradNum'][:-2]+"\nNAD+SU+7701081000011::9\nRFF+VA:890985438\nNAD+CA+7701081000011::9\n" \
                        "TDT+20++30+31++++:::123456\nEQD+BX\nCPS+1\nPAC+"+dato_cambiante_cantidad_cajas+"++BX\nCPS+2+1\nPAC+"+dato_cambiante_cantidad_cajas+"++BX"
        contador=0
        url3 = IP_SAP + "SQLQueries('consultaavisodespachodetalle1')/List?pedido='" + str(
            pedido_pk) + "'"

        response1 = sap_request(url3)
        response1 = response1.text
        response1 = response1.replace('null', ' " " ')
        response1 = ast.literal_eval(response1)
        response1 = response1['value']
        cantidad_embalada=0
        for detallepedido in response1:
            cantidad_embalada=cantidad_embalada+int(detallepedido['PackQty'])
            contador=contador+1
            edi_complement="\nLIN+"+str(contador)+"++"+detallepedido['CodeBars']+":EN" \
                    "\nQTY+12:"+str(detallepedido['Quantity'])+":NAR"
            edi_data = edi_data + edi_complement

        edi_final="\nCNT+2:2"
        edi_data = edi_data + edi_final
        edi_data=edi_data.replace(dato_cambiante_cantidad_cajas,str(cantidad_embalada))
        edi_count=edi_data.count("\n")
        edi_finales="\nUNT+"+str(edi_count)+"+1\nUNZ+1+PCS"+str(consuc_edi)
        edi_data = edi_data + edi_finales

        name_file=str(GLNCliente)+'_'+str(NumeroOrden)+'_'+str(NumeroFactura)
        # Crear la respuesta HTTP con el contenido del archivo EDI
        response = HttpResponse(edi_data, content_type='application/edi')
        response['Content-Disposition'] = 'attachment; filename="'+name_file+'".edi'
        valor_edi=1+consecutivo_edi.valor
        Consecutivo.objects.filter(pk=consecutivo_edi.pk).update(valor=valor_edi)
        return response
    else:
        pass

def config_ordenes_otroscanales_pcs_detalle_cliente(request, id):

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        pedido = PedidosOtrosCanales.objects.filter(num_pedido=id).first()
        pedidodetalle = DetallesPedidosOtrosCanales.objects.filter(num_pedido=id)
        dato_lista=[]
        for detalles in pedidodetalle:
            variable= AsignacionPedidosOtrosCanales.objects.filter(num_detalle_id=detalles.id)
            data_detalle = {
                'id': detalles.id,
                'u_plu': detalles.u_plu,
                'referencia': detalles.referencia,
                'cantidad': detalles.cantidad,
                'nombre': detalles.nombre,
                'observaciones': detalles.observaciones,
                'asignaciones': AsignacionPedidosOtrosCanales.objects.filter(num_detalle_id=detalles.id)
            }
            dato_lista.append(data_detalle)
        empresas=Empresas.objects.all()


        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_ordenes_otroscanales_pcs_detalle_cliente.html", {'user': current_user,
                                                        'pedido': pedido,
                                                        'pedidodetalle': dato_lista,
                                                        'empresas': empresas,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    else:
        pass

def config_ordenes_otroscanales_pcs_eliminar_cliente(request, id):

    if request.method == 'GET':

        identificacion_asignacion = request.GET.get('asignacion')
        asignacion_editar = PedidosOtrosCanales.objects.get(num_pedido=id)
        i=asignacion_editar.estado
        if i !='en proceso':
            messages.add_message(request, messages.ERROR,
                                 'No se puede borrar el pedido, no se encuentra en proceso')
            return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales_cliente/')
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        asignacion_editar.delete()
        messages.add_message(request, messages.WARNING,
                             'Se ha borrado el pedido ' + str(id) + ' satisfactoriamente')

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales_cliente/' )

    else:
        pass

def config_ordenes_otroscanales_pcs_detalles(request):
    if request.method == 'POST':
        hoy = date.today()
        current_user = request.user
        empresa = request.POST['empresa']
        id = request.POST['id']
        pedido = request.POST['pedido']
        cantidad = request.POST['cantidad']

        # detallepedido_id = DetallesPedidosOtrosCanales.objects.filter(id=id).first()
        # empresario = Empresas.objects.filter(id=empresa).first()
        # articulo = MaestroArticulo.objects.filter(u_plu=detallepedido_id.u_plu,
        #                                           proveedorCodigo=empresario.codigo).first()
        #
        # detallepedidootros = DetallesPedidosOtrosCanales(
        #     num_pedido_id=pedido,
        #     cantidad=cantidad,
        #     nombre=articulo.itemName,
        #     referencia=articulo.itemCode,
        #     u_plu=articulo.u_plu,
        #     observaciones=detallepedido_id.observaciones,
        #     empresa=empresario,
        #     # articulo=articulo,
        # )
        # detallepedidootros.save()

        if int(cantidad) <= 0:
            if request.is_ajax():
                return JsonResponse({'success': False, 'message': 'No es posible asignar valores <= 0.'})
            messages.error(request, 'No es posible asignar valores que sean iguales o inferiores a cero.')
            return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/' + pedido + '/')

        detalle = DetallesPedidosOtrosCanales.objects.filter(pk=id).first()
        cantidad_pedido = detalle.cantidad
        asignaciones = AsignacionPedidosOtrosCanales.objects.filter(num_detalle=id)
        suma_cantidad = asignaciones.aggregate(total_cantidad=Sum('cantidad'))['total_cantidad'] or 0

        try:
            id_asignacion = request.POST['id_asignacion']
            asignacion_editar = AsignacionPedidosOtrosCanales.objects.filter(pk=int(id_asignacion)).first()
            cantidad_total = int(cantidad) + suma_cantidad - int(asignacion_editar.cantidad)

            if cantidad_total > cantidad_pedido:
                if request.is_ajax():
                    return JsonResponse({'success': False, 'message': 'Cantidad máxima superada.'})
                messages.error(request, 'No se puede asignar. La cantidad máxima de la orden fue superada')
                return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/' + pedido + '/')

            if cantidad_total == cantidad_pedido:
                PedidosOtrosCanales.objects.filter(pk=pedido).update(estado='asignado')

            asignacion = AsignacionPedidosOtrosCanales(
                pk=int(id_asignacion),
                num_detalle_id=id,
                cantidad=cantidad,
                empresa_id=empresa,
                fecha=hoy,
            )
            asignacion.save()
        except:
            cantidad_total = int(cantidad) + suma_cantidad

            if cantidad_total > cantidad_pedido:
                if request.is_ajax():
                    return JsonResponse({'success': False, 'message': 'Cantidad máxima superada.'})
                messages.error(request, 'No se puede asignar. La cantidad máxima de la orden fue superada')
                return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/' + pedido + '/')

            if cantidad_total == cantidad_pedido:
                PedidosOtrosCanales.objects.filter(pk=pedido).update(estado='asignado')

            asignacion = AsignacionPedidosOtrosCanales(
                num_detalle_id=id,
                cantidad=cantidad,
                empresa_id=empresa,
                fecha=hoy,
            )
            asignacion.save()

        if request.is_ajax():
            asignaciones_data = []
            asignaciones = AsignacionPedidosOtrosCanales.objects.filter(num_detalle=id).select_related('empresa')
            for a in asignaciones:
                asignaciones_data.append({
                    'id': a.pk,
                    'empresa': a.empresa.nombre,
                    'cantidad': a.cantidad,
                })

            return JsonResponse({
                'success': True,
                'message': 'Asignación guardada correctamente',
                'detalle_id': id,
                'pedido_id': pedido,
                'asignaciones': asignaciones_data
            })

        return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/' + pedido + '/')

def config_ordenes_otroscanales_pcs_plus_detalles(request):
    if request.method == 'POST':
        hoy = date.today()
        current_user = request.user
        empresa = request.POST['empresa']
        id = request.POST['id']
        pedido = request.POST['pedido']
        cantidad = request.POST['cantidad_plu']

        if int(cantidad) <= 0:
            if request.is_ajax():
                return JsonResponse({'success': False, 'message': 'No es posible asignar valores <= 0.'})
            messages.error(request, 'No es posible asignar valores que sean iguales o inferiores a cero.')
            return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/' + pedido + '/')

        detalle_plu = DetallesPedidosOtrosCanales_plus.objects.filter(pk=id).first()
        observacion = detalle_plu.observaciones
        plu_pedido = detalle_plu.u_plu
        plu_cantidad = detalle_plu.cantidad
        empresario = Empresas.objects.filter(pk=int(empresa)).first()

        articulo_plu = MaestroArticulo.objects.filter(u_plu=plu_pedido, proveedorCodigo=empresario.codigo).first()
        detalle_all = DetallesPedidosOtrosCanales.objects.filter(num_pedido=int(pedido),u_plu=plu_pedido)
        detalle_cantidad_total = 0
        for d in detalle_all:
            detalle_cantidad_total = d.cantidad

        cantidad_total = int(cantidad) + detalle_cantidad_total

        if plu_cantidad < cantidad_total:
            if request.is_ajax():
                return JsonResponse({'success': False, 'message': 'Cantidad máxima superada.'})
            messages.error(request, 'No se puede asignar. La cantidad máxima de la articulo '+plu_pedido+' fue superada')
            return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/' + pedido + '/')

        detalle = detalle_all.filter(referencia=articulo_plu.itemCode)
        if detalle:
            if request.is_ajax():
                return JsonResponse({'success': False, 'message': 'El producto: ' + articulo_plu.itemName +' - '+ plu_pedido + ' ya fue registrado'})
            messages.error(request,
                           'El producto: ' + articulo_plu.itemName +' - '+ plu_pedido + ' ya fue registrado')
            return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/' + pedido + '/')

        # Crear detalle completo
        detallepedidootros = DetallesPedidosOtrosCanales(
            num_pedido_id=int(pedido),
            cantidad=cantidad,
            nombre=articulo_plu.itemName,
            referencia=articulo_plu.itemCode,
            u_plu=articulo_plu.u_plu,
            observaciones=observacion,
            empresa=empresario,
            multiple=True,
        )
        detallepedidootros.save()

        asignacion = AsignacionPedidosOtrosCanales(
            num_detalle=detallepedidootros,
            cantidad=cantidad,
            empresa=empresario,
            fecha=hoy,
        )
        asignacion.save()

        return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/' + pedido + '/')

def config_ordenes_otroscanales_pcs_detalles_todos(request):
    if request.method == 'POST':
        hoy = date.today()
        empresa_id = request.POST.get('empresa')   # empresa destino
        pedido_id = request.POST.get('pedido')     # pedido a asignar

        if not empresa_id or not pedido_id:
            messages.add_message(request, messages.ERROR, "Faltan datos de empresa o pedido")
            return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/')

        # Obtener pedido
        pedido = get_object_or_404(PedidosOtrosCanales, pk=pedido_id)

        # Traer todos los detalles del pedido
        detalles = DetallesPedidosOtrosCanales.objects.filter(num_pedido=pedido)

        if not detalles.exists():
            messages.add_message(request, messages.ERROR, "El pedido no tiene detalles para asignar")
            return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/')

        for detalle in detalles:
            # Verificar si ya existen asignaciones para este detalle
            asignaciones = AsignacionPedidosOtrosCanales.objects.filter(num_detalle=detalle)
            suma_asignada = asignaciones.aggregate(total=Sum('cantidad'))['total']
            if suma_asignada is None:
                suma_asignada = 0

            # Si ya está completamente asignado, lo saltamos
            if suma_asignada >= detalle.cantidad:
                continue

            # Cantidad pendiente por asignar en este detalle
            cantidad_pendiente = detalle.cantidad - suma_asignada

            # Crear asignación
            asignacion = AsignacionPedidosOtrosCanales(
                num_detalle=detalle,
                cantidad=cantidad_pendiente,
                empresa_id=empresa_id,
                fecha=hoy
            )
            asignacion.save()

        # Revisar si quedó todo asignado
        detalles_sin_asignar = DetallesPedidosOtrosCanales.objects.filter(
            num_pedido=pedido
        ).exclude(
            asignacionpedidosotroscanales__cantidad__gte=F('cantidad')
        )

        if not detalles_sin_asignar.exists():
            pedido.estado = 'asignado'
            pedido.save()

        messages.add_message(request, messages.SUCCESS,
                             "Se asignaron todos los detalles del pedido a la empresa seleccionada")
        return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/' + str(pedido_id) + '/')

def config_ordenes_otroscanales_pcs_estado(request):
    if request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        estado = request.POST['estado']
        pedido = request.POST['pedido']
        id = request.POST['id']

        if estado == 'asignado':
            detalle = DetallesPedidosOtrosCanales.objects.filter(num_pedido=id)
            for d in detalle:
                asignacion_cantidad = 0
                asignaciones = AsignacionPedidosOtrosCanales.objects.filter(num_detalle_id=d.id)
                for asignacion in asignaciones:
                    asignacion_cantidad += asignacion.cantidad
                if not asignacion_cantidad == d.cantidad:
                    messages.add_message(request, messages.ERROR,
                                         "Al producto: "+d.nombre+' (REF: '+d.referencia+') no tiene todas las unidades asignadas.')
                    return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/' + pedido + '/')

            detalle_plus = DetallesPedidosOtrosCanales_plus.objects.filter(num_pedido=id)
            for d_plus in detalle_plus:
                detalle = DetallesPedidosOtrosCanales.objects.filter(num_pedido=id,u_plu=d_plus.u_plu)
                asignacion_cantidad = 0
                for d in detalle:
                    asignaciones = AsignacionPedidosOtrosCanales.objects.filter(num_detalle_id=d.id)
                    for asignacion in asignaciones:
                        asignacion_cantidad += asignacion.cantidad
                if not asignacion_cantidad == d_plus.cantidad:
                    messages.add_message(request, messages.ERROR,
                                         "los productos con U_PLU: " + d_plus.u_plu + ' no tiene todas las unidades asignadas.')
                    return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/' + pedido + '/')

            PedidosOtrosCanales.objects.filter(pk=id).update(estado=estado)
            messages.add_message(request, messages.SUCCESS,
                                 "el pedido: " + pedido + ' se ha asignado correctamente.')
            return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/' + pedido + '/')

        PedidosOtrosCanales.objects.filter(pk=id).update(estado=estado)

        return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/'+pedido+'/')

def config_ordenes_otroscanales_pcs_fecha_minima(request):
    if request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        fecha_minima = request.POST['fecha_minima']
        pedido = request.POST['pedido']
        id = request.POST['id']
        PedidosOtrosCanales.objects.filter(pk=id).update(fecha_minima=fecha_minima)

        return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/'+pedido+'/')

def config_ordenes_otroscanales_pcs_fecha_maxima(request):
    if request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        fecha_maxima = request.POST['fecha_maxima']
        fecha_objeto = datetime.strptime(fecha_maxima, '%Y-%m-%d').date()
        pedido = request.POST['pedido']
        id = request.POST['id']
        variable=PedidosOtrosCanales.objects.filter(pk=id).first()
        if variable.fecha_minima < fecha_objeto:
            PedidosOtrosCanales.objects.filter(pk=id).update(fecha_maxima=fecha_maxima)
        else:
            messages.add_message(request, messages.ERROR,
                                 'La fecha maxima debe ser mayor a la fecha minima' )
        return HttpResponseRedirect('/configuracion/orden_pcs_otroscanales/detalle/'+pedido+'/')

def config_documentos_creditos(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        documentos_pendientes = DocumentosCredito.objects.filter(estado='pendiente').order_by('-FechaEmision')
        documentos_otro_estado = DocumentosCredito.objects.exclude(estado='pendiente').order_by('-FechaEmision')
        documentos = list(documentos_pendientes) + list(documentos_otro_estado)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "config_documentos_creditas.html", {'user': current_user,
                                                          'documentos': documentos,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        hoy = date.today()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        fecha_hoy = date.today()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.nit
        numerosoli = request.POST['numerosoli']
        respuesta = request.POST['respuesta']

        if respuesta == 'Aprobado':
            DocumentosCredito.objects.filter(pk=numerosoli).update(estado='aprobado')
            consulta = DocumentosCredito.objects.filter(pk=numerosoli).first()
            historiales = HistorialDocumentosCredito.objects.filter(documento=numerosoli).first()
            if historiales == None:
                estado='REGISTRO'
            else:
                estado='ACTUALIZACION'
            historial = HistorialDocumentosCredito(

                documento_id=consulta.pk,
                pagare=consulta.pagare,
                contrato=consulta.contrato,
                carta=consulta.carta,
                ficha=consulta.ficha,
                FechaRespuesta=hoy,
                estado=estado,
                UsuarioRespuesta_id=current_user.pk,

            )
            historial.save()
            email = EmailMessage('DOCUMENTOS APROBADOS',
                                     'Felicidades, tus documentos fueron aprobados, ya podras solicitar creditos en nuestro sistema  : \n',
                                     to=[consulta.Correo])
            email.send()

            messages.add_message(request, messages.INFO,
                                    'Se ha aprobado los documentos satisfactoriamente.')
        else:
            razon = request.POST['razon']
            DocumentosCredito.objects.filter(pk=numerosoli).update(estado='Negado',Razon=razon)
            consulta = DocumentosCredito.objects.filter(pk=numerosoli).first()
            historial = HistorialDocumentosCredito(
                documento_id=consulta.pk,
                pagare=consulta.pagare,
                contrato=consulta.contrato,
                carta=consulta.carta,
                ficha=consulta.ficha,
                FechaRespuesta=hoy,
                estado='NEGADO DE DOCUMENTOS',
                UsuarioRespuesta_id=current_user.pk,

            )
            historial.save()
            email = EmailMessage('DOCUMENTOS DENEGADOS',
                                 'Lamentablemente, sus documentos han sido negados por lo siguiente : \n'
                                 + str(consulta.Razon) ,
                                 to=[consulta.Correo])
            email.send()
            messages.add_message(request, messages.ERROR,
                                 'Se ha rechazado los documentos  satisfactoriamente.')
        return HttpResponseRedirect('/configuracion/documentos_creditos/')

def config_historial_documentos_creditos(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        documentos = HistorialDocumentosCredito.objects.filter(documento_id=id).order_by('-FechaRespuesta')

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "config_historial_documentos_creditos.html", {'user': current_user,
                                                          'documentos': documentos,
                                                        'permiso_usuario': usuario_datos,
                                                        })

def config_servicio_crediya_registro(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        documentos = DocumentosCredito.objects.filter(Empresa_id=empresa.empresa.pk).first()
        empresa = empresa.empresa.nombre
        hoy = date.today()
        hoy = hoy + timedelta(days=0)
        hoy = hoy.strftime("%Y-%m-%d")

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')



        url2 = IP_SAP + "SQLQueries('PedidosHabiles5')/List?fecha='" + hoy +  "'&empresa='" + str(
                empresa)+"'"

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        pedidos=[]
        for pedido in response:
            url2 = IP_SAP + "SQLQueries('PedidosHabilesEntry')/List?DocEntry='" + str(pedido['DocEntry']) + "'"

            responseentry = sap_request(url2)
            responseentry = ast.literal_eval(responseentry.text)
            responseentry = responseentry['value'][0]['WhsCode']

            crediya_preaprobados = CrediyaPreaprobado.objects.filter(NumeroOrdenCompra=pedido['DocNum']).first()
            estado = crediya_preaprobados.estado if crediya_preaprobados else None
            fecha_objeto = datetime.strptime(pedido['TaxDate'], "%Y%m%d")
            fecha_formateada = fecha_objeto.strftime("%Y-%m-%d")
            fecha_ven = datetime.strptime(pedido['DocDueDate'], "%Y%m%d")
            fecha_formateada_ven = fecha_ven.strftime("%Y-%m-%d")
            pedido_det={
                'pedido':pedido['DocNum'],
                'fecha_vencimiento':pedido['DocDueDate'],
                'fecha_formateada_ven':fecha_formateada_ven,
                'pedido':pedido['DocNum'],
                'total':format(int(pedido['DocTotal']) - int(pedido['VatSum']), '0,.0f'),
                'fecha_pedido': pedido['TaxDate'],
                'fecha_formato_pedido':fecha_formateada,
                'estado':estado,
                'almacen':responseentry,
            }
            pedidos.append(pedido_det)


        return render(request, "config_servicio_crediya_registro.html", {'user': current_user,
                                                           'permiso_usuario': usuario_datos,
                                                           'pedidos': pedidos,
                                                           'empresa': empresa,
                                                           'documentos': documentos,
                                                           })
    elif request.method == 'POST':
        current_user = request.user
        hoy = date.today()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        correo = empresa.email
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.pk
        numero_pedido = request.POST['numero_pedido']
        intereses = request.POST['intereses']
        total_desembolso = request.POST['total_desembolso']
        fecha_pedido = request.POST['fecha_pedido']
        fecha_vencimiento = request.POST['fecha_vencimiento']
        valor_orden = request.POST['valor_orden']
        pedido_almacen = request.POST['pedido_almacen']
        fecha_pedido = datetime.strptime(fecha_pedido, "%Y%m%d").date()
        fecha_vencimiento = datetime.strptime(fecha_vencimiento, "%Y%m%d").date()
        if CrediyaPreaprobado.objects.filter(NumeroOrdenCompra=numero_pedido).exists():
            messages.add_message(request, messages.ERROR,
                                 'La solicitud de crédito para el pedido ' + str(numero_pedido) + ' ya fue realizada previamente. ')

            return HttpResponseRedirect('/configuracion/servicio_crediya_registro/')

        preaprobado = CrediyaPreaprobado(
            Empresa_id=empresa,
            UsuarioSolicitado_id=current_user.id,
            TipoIdentificacion='NIT',
            ValorAprobado=total_desembolso,
            NumeroOrdenCompra=numero_pedido,
            Interes=intereses,
            estado='Preaprobado',
            FechaEmision=fecha_pedido,
            FechaVencimiento=fecha_vencimiento,
            FechaSolicitud=hoy,
            ValorOrden=valor_orden,
            Correo=correo,
            Almacen=pedido_almacen,
        )
        preaprobado.save()
        messages.add_message(request, messages.INFO,
                             'El credito para el pedido ' + str(numero_pedido) + ' se encuentra preaprobado.')

        return HttpResponseRedirect('/configuracion/servicio_crediya_registro/')

def config_servicio_crediya_preaprobado(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        doc = empresa.empresa.pk
        empresa = empresa.empresa.nombre

        return render(request, "config_servicio_crediya_preaprobado.html", {'user': current_user,
                                                           'permiso_usuario': usuario_datos,
                                                           'empresa': empresa
                                                           })
    else:
        return HttpResponseRedirect('/configuracion/servicio_crediya_registro/')

def config_registro_documentos_creditos(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        doc = empresa.empresa.pk
        empresa = empresa.empresa.nombre
        lista_empresas = Empresas.objects.all()

        return render(request, "config_servicio_registro_documentos.html", {'user': current_user,
                                                           'permiso_usuario': usuario_datos,
                                                           'lista_empresas': lista_empresas,
                                                           'empresa': empresa
                                                           })
    elif request.method == 'POST':
        current_user = request.user
        hoy = date.today()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        correo = empresa.email
        fecha_hoy = date.today()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.pk
        pagare = request.FILES['pagare'] if 'pagare' in request.FILES else ''
        contrato = request.FILES['contrato'] if 'contrato' in request.FILES else ''
        ficha = request.FILES['ficha'] if 'ficha' in request.FILES else ''
        carta = request.FILES['carta'] if 'carta' in request.FILES else ''
        empresa_docs = request.POST['empresas']
        documentos_empresa=DocumentosCredito.objects.filter(Empresa_id=empresa_docs).first()
        if documentos_empresa != None:
            messages.add_message(request, messages.ERROR,
                                 'No se pueden registrar los documentos. Existe un registro para esta empresa')
            return HttpResponseRedirect('/configuracion/documentos_creditos/')
        if pagare != '':
            myfile = request.FILES['pagare']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            pagare = fs.url(filename)
        else:
            pagare = None
        if contrato != '':
            myfile = request.FILES['contrato']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            contrato = fs.url(filename)
        else:
            contrato = None
        if carta != '':
            myfile = request.FILES['carta']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            carta = fs.url(filename)
        else:
            carta = None
        if ficha != '':
            myfile = request.FILES['ficha']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            ficha = fs.url(filename)
        else:
            ficha = None


        preaprobado = DocumentosCredito(
                Empresa_id=empresa_docs,
                pagare=pagare,
                contrato=contrato,
                carta=carta,
                ficha=ficha,
                estado='aprobado',
                FechaEmision=fecha_hoy,
                Correo=correo,
            )
        preaprobado.save()
        historial = HistorialDocumentosCredito(
            documento_id=preaprobado.pk,
            pagare=preaprobado.pagare,
            contrato=preaprobado.contrato,
            carta=preaprobado.carta,
            ficha=preaprobado.ficha,
            FechaRespuesta=hoy,
            estado='REGISTRO POR ADMINISTRADOR',
            UsuarioRespuesta_id=current_user.pk,

        )
        historial.save()

        return HttpResponseRedirect('/configuracion/documentos_creditos/')

def config_editar_documentos_creditos(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        doc = empresa.empresa.pk
        empresa = empresa.empresa.nombre
        lista_empresas = Empresas.objects.all()
        documentos=DocumentosCredito.objects.filter(pk=id).first()

        return render(request, "config_servicio_editar_documentos.html", {'user': current_user,
                                                           'permiso_usuario': usuario_datos,
                                                           'lista_empresas': lista_empresas,
                                                           'documentos': documentos,
                                                           'empresa': empresa
                                                           })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        hoy = date.today()
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        correo = empresa.email
        fecha_hoy = date.today()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.pk
        pagare = request.FILES['pagare']
        contrato = request.FILES['contrato']
        carta = request.FILES['carta']
        ficha = request.FILES['ficha']
        empresa_docs = request.POST['empresas']
        if pagare != '':
            myfile = request.FILES['pagare']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            pagare = fs.url(filename)
        else:
            pagare = None
        if contrato != '':
            myfile = request.FILES['contrato']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            contrato = fs.url(filename)
        else:
            contrato = None
        if carta != '':
            myfile = request.FILES['carta']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            carta = fs.url(filename)
        else:
            carta = None
        if ficha != '':
            myfile = request.FILES['ficha']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            ficha = fs.url(filename)
        else:
            ficha = None


        preaprobado = DocumentosCredito(
                pk=id,
                Empresa_id=empresa_docs,
                pagare=pagare,
                contrato=contrato,
                carta=carta,
                ficha=ficha,
                estado='aprobado',
                FechaEmision=fecha_hoy,
                Correo=correo,
            )
        preaprobado.save()
        historial = HistorialDocumentosCredito(
            documento_id=preaprobado.pk,
            pagare=preaprobado.pagare,
            contrato=preaprobado.contrato,
            carta=preaprobado.carta,
            ficha=preaprobado.ficha,
            FechaRespuesta=hoy,
            estado='ACTUALIZACION POR ADMINISTRADOR',
            UsuarioRespuesta_id=current_user.pk,

        )
        historial.save()

        return HttpResponseRedirect('/configuracion/documentos_creditos/')

def config_servicio_crediya_preaprobados(request):
    # Render  administracion.html
    if request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        correo=empresa.email
        fecha_hoy = date.today()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.pk
        documentos = DocumentosCredito.objects.filter(Empresa_id=empresa).first()
        pagare = request.FILES['pagare']
        contrato = request.FILES['contrato']
        carta = request.FILES['carta']
        ficha = request.FILES['ficha']
        if pagare != '':
            myfile = request.FILES['pagare']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            pagare = fs.url(filename)
        else:
            pagare = None
        if contrato != '':
            myfile = request.FILES['contrato']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            contrato = fs.url(filename)
        else:
            contrato = None
        if carta != '':
            myfile = request.FILES['carta']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            carta = fs.url(filename)
        else:
            carta = None
        if ficha != '':
            myfile = request.FILES['ficha']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            ficha = fs.url(filename)
        else:
            ficha = None
        if documentos:
            DocumentosCredito.objects.filter(Empresa_id=empresa).update(
                Empresa_id=empresa,
                pagare=pagare,
                contrato=contrato,
                carta=carta,
                ficha=ficha,
                estado='Pendiente',
                FechaEmision=fecha_hoy,
                Correo=correo,
            )
        else:
            preaprobado = DocumentosCredito(
                Empresa_id=empresa,
                pagare=pagare,
                contrato=contrato,
                carta=carta,
                ficha=ficha,
                estado='Pendiente',
                FechaEmision=fecha_hoy,
                Correo=correo,
            )
            preaprobado.save()

        return HttpResponseRedirect('/configuracion/servicio_crediya_registro/')

def config_crediya_consulta_generales(request):

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_empresas = Empresas.objects.all()
        creditos_preaprobados=CrediyaPreaprobado.objects.filter(estado='Preaprobado').order_by('-FechaSolicitud')

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_solicitudes_crediya.html", {'user': current_user,
                                                        'lista_empresas': lista_empresas,
                                                        'permiso_usuario': usuario_datos,
                                                        'creditos_preaprobados': creditos_preaprobados,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        fecha_hoy = date.today()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.nit
        numeropedido = request.POST['numeropedido']
        respuesta = request.POST['respuesta']

        if respuesta=='Aprobado':
            estado_pedido=respuesta
            # Usa la variable global SESSION
            cookie = "B1SESSION=" + SESSION
            if ROUTEID:  # por si tu SAP devuelve ROUTEID
                cookie += "; ROUTEID=" + ROUTEID

            consulta=CrediyaPreaprobado.objects.filter(NumeroOrdenCompra=numeropedido).first()
            if consulta is not None and consulta.FechaVencimiento > fecha_hoy:
                codigo_empresa = str(consulta.Empresa.codigo)
                fecha_actual = datetime.utcnow()
                fecha_vencimient = fecha_actual + timedelta(days=30)
                fecha_formateada_ven = fecha_vencimient.strftime('%Y-%m-%dT%H:%M:%SZ')
                fecha_formateada = fecha_actual.strftime('%Y-%m-%dT%H:%M:%SZ')
                valor_sin_comas = consulta.ValorAprobado.replace(",", "")
                valor_aprobado=float(valor_sin_comas)
                interes_sin_comas = consulta.Interes.replace(",", "")
                intereses=float(interes_sin_comas)

                # URL del endpoint de VendorPayments
                url_pagos_efectuados = IP_SAP + "VendorPayments"

                # Datos del nuevo pago efectuado que quieres enviar
                nuevo_pago_efectuado = {
                    "DocType": "rSupplier",
                    "HandWritten": "tNO",
                    "Printed": "tNO",
                    "DocDate": fecha_formateada,
                    "CardCode": codigo_empresa,
                    "CashAccount": "42100516",
                    "DocCurrency": "$",
                    "CashSum": intereses,
                    "CheckAccount": None,
                    "TransferAccount": "11201005",
                    "TransferSum": valor_aprobado,
                    "TransferDate": fecha_formateada,
                    "TransferReference": None,
                    "LocalCurrency": "tNO",
                    "DocRate": 0.0,
                    "Reference2": None,
                    "CounterReference": None,
                    "Remarks": None,
                    "JournalRemarks": "ANTICIPO OC "+str(consulta.NumeroOrdenCompra),
                    "SplitTransaction": "tNO",
                    "ApplyVAT": "tNO",
                    "TaxDate": fecha_formateada,
                    "BankCode": None,
                    "BankAccount": None,
                    "DiscountPercent": 0.0,
                    "ProjectCode": None,
                    "CurrencyIsLocal": "tNO",
                    "DeductionPercent": 0.0,
                    "DeductionSum": 0.0,
                    "CashSumFC": 0.0,
                    "BoeAccount": None,
                    "BillOfExchangeAmount": 0.0,
                    "BillofExchangeStatus": None,
                    "BillOfExchangeAmountFC": 0.0,
                    "BillOfExchangeAmountSC": 0.0,
                    "BillOfExchangeAgent": None,
                    "WTCode": None,
                    "WTAmount": 0.0,
                    "WTAmountFC": 0.0,
                    "WTAmountSC": 0.0,
                    "WTAccount": None,
                    "WTTaxableAmount": 0.0,
                    "Proforma": "tNO",
                    "PayToBankCode": None,
                    "PayToBankBranch": None,
                    "PayToBankAccountNo": None,
                    "PayToCode": "PRINCIPAL",
                    "PayToBankCountry": None,
                    "IsPayToBank": "tNO",
                    "PaymentPriority": "bopp_Priority_6",
                    "TaxGroup": None,
                    "BankChargeAmount": 0.0,
                    "BankChargeAmountInFC": 0.0,
                    "BankChargeAmountInSC": 0.0,
                    "UnderOverpaymentdifference": 0.0,
                    "UnderOverpaymentdiffSC": 0.0,
                    "WtBaseSum": 0.0,
                    "WtBaseSumFC": 0.0,
                    "WtBaseSumSC": 0.0,
                    "VatDate": fecha_formateada,
                    "TransactionCode": "",
                    "PaymentType": "bopt_None",
                    "TransferRealAmount": 0.0,
                    "DocObjectCode": "bopot_OutgoingPayments",
                    "DocTypte": "rSupplier",
                    "DueDate": fecha_formateada_ven,
                    "LocationCode": None,
                    "Cancelled": "tNO",
                    "ControlAccount": "13300612",
                    "UnderOverpaymentdiffFC": 0.0,
                    "AuthorizationStatus": "pasWithout",
                    "BPLID": None,
                    "BPLName": None,
                    "VATRegNum": None,
                    "BlanketAgreement": None,
                    "PaymentByWTCertif": "tNO",
                    "Cig": None,
                    "Cup": None,
                    "AttachmentEntry": None,
                    "SignatureInputMessage": None,
                    "SignatureDigest": None,
                    "CertificationNumber": None,
                    "PrivateKeyVersion": None,
                    "U_HBT_AreVal": "Comun",
                    "U_HBT_Tercero": None,
                    "U_GSP_TPVREF": None,
                    "U_GSP_TPVLINE": None,
                    "U_GSP_TPVPAYMETHOD": None,
                    "U_HBT1_CPTO": "Pago",
                    "U_HBT_PagoAnticipo": "N",
                    "U_BP_Confd": "N",
                    "U_BP_DocNr": None,
                    "U_BP_Seque": None,
                    "PaymentChecks": [],
                    "PaymentInvoices": [],
                    "PaymentCreditCards": [],
                    "PaymentAccounts": [],
                    "PaymentDocumentReferencesCollection": [],
                    "BillOfExchange": {},
                    "WithholdingTaxCertificatesCollection": [],
                    "ElectronicProtocols": [],
                    "CashFlowAssignments": [],
                    "Payments_ApprovalRequests": [],
                    "WithholdingTaxDataWTXCollection": []
                }

                # Convierte los datos a formato JSON
                payload_pagos_efectuados = json.dumps(nuevo_pago_efectuado)

                # Encabezados de la solicitud HTTP
                headers_pagos_efectuados = {
                    'Content-Type': 'application/json',
                    # Agrega aquí tu cookie de sesión
                    'Cookie': cookie
                }

                # Realiza la solicitud POST al Service Layer para crear un nuevo pago efectuado
                response_pagos_efectuados = requests.post(url_pagos_efectuados, headers=headers_pagos_efectuados,
                                                          data=payload_pagos_efectuados, verify=False)


                # Verifica el estado de la respuesta
                if response_pagos_efectuados.status_code == 201:
                    egreso_creado=response_pagos_efectuados.content
                    egreso_creado = json.loads(egreso_creado)
                    egreso_creado = str(egreso_creado.get("DocNum"))

                    print("Pago efectuado creado exitosamente.")
                else:
                    now = datetime.now(pytz.timezone('America/Bogota'))
                    hoy = now.date()
                    hora = now.time()
                    errores = HistorialErrorEnvioSap(
                        accion='Error al enviar informacion a SAP',
                        fecha=hoy,
                        hora=hora,
                        tipo='crediya',
                        empresa=str(consulta.Empresa.nombre),
                        pedido=consulta.NumeroOrdenCompra,
                        mensaje_sistema=response_pagos_efectuados.text
                    )
                    errores.save()
                    messages.add_message(request, messages.ERROR,
                                         'No se ha podido registrar el pedido  ' + numeropedido + ' .Por la siguiente razon:'+response_pagos_efectuados.text)
                    print("Error al crear el pago efectuado:", response_pagos_efectuados.text)
                    return HttpResponseRedirect('/configuracion/servicio_crediya_consulta/')

                CrediyaPreaprobado.objects.filter(NumeroOrdenCompra=numeropedido).update(estado=estado_pedido,FechaRespuesta=fecha_hoy,UsuarioRespuesta=current_user.id,EgresoCreado=egreso_creado)
                aprobado = RespuestaOrdenCompraApi(
                    Identificacion=consulta.Empresa.nit,
                    TipoIdentificacion='NIT',
                    ValorAprobado=consulta.ValorAprobado.replace(",", ""),
                    FechaEmision=consulta.FechaEmision,
                    NumeroOrdenCompra=consulta.NumeroOrdenCompra,
                    Interes=consulta.Interes.replace(",", ""),
                )
                aprobado.save()
                now = datetime.now(pytz.timezone('America/Bogota'))
                hoy = now.date()
                hora = now.time()
                errores = HistorialErrorEnvioSap(
                    accion='Registrado Con Éxito',
                    fecha=hoy,
                    hora=hora,
                    tipo='crediya',
                    empresa=str(consulta.Empresa.nombre),
                    pedido=consulta.NumeroOrdenCompra,
                )
                errores.save()
                email = EmailMessage('SERVICIO FINANCIERO CREDIYA APROBADO',
                                     'Tu Servicio Financiero para el pedido : \n'
                                     + str(consulta.NumeroOrdenCompra)+"\nha sido aprobado. Podrás consultarlo en el siguiente sitio web:"
                                                    + IP_SERVIDOR+ "/configuracion/servicio_crediya_historial/",
                                     to=[consulta.Correo])
                email.send()

                messages.add_message(request, messages.INFO,
                                     'Se ha aprobado el credito  ' + numeropedido + ' satisfactoriamente.')
            else:
                messages.add_message(request, messages.WARNING,
                                     'No se puede aprobar el crédito ' + numeropedido + ' se encuentra vencido.')
        else:
            razon = request.POST['razon']
            CrediyaPreaprobado.objects.filter(NumeroOrdenCompra=numeropedido).update(estado=respuesta,FechaRespuesta=fecha_hoy,UsuarioRespuesta=current_user.id,Razon=razon)
            consulta = CrediyaPreaprobado.objects.filter(NumeroOrdenCompra=numeropedido).first()
            email = EmailMessage('SERVICIO FINANCIERO CREDIYA DENEGADO',
                                 'Lamentablemente, su Servicio Financiero para el pedido : \n'
                                 + str(consulta.NumeroOrdenCompra) + "\nha sido rechazado por la siguiente razon:\n"
                                ""+razon,
                                 to=[consulta.Correo])
            email.send()
            messages.add_message(request, messages.ERROR,
                                 'Se ha rechazado el credito  ' + numeropedido + ' satisfactoriamente.')

        
        return HttpResponseRedirect('/configuracion/servicio_crediya_consulta/')

def config_credilisto_consulta_generales(request):

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_empresas = Empresas.objects.all()
        creditos_preaprobados=CrediListoPreaprobado.objects.filter(estado='Preaprobado').order_by('-FechaEmision')

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_solicitudes_credilisto.html", {'user': current_user,
                                                        'lista_empresas': lista_empresas,
                                                        'permiso_usuario': usuario_datos,
                                                        'creditos_preaprobados': creditos_preaprobados,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        fecha_hoy = date.today()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.nit
        numeropedido = request.POST['numeropedido']
        respuesta = request.POST['respuesta']
        estado = request.POST['estado']

        if respuesta == 'Aprobado':

            estado_pedido = respuesta

            consulta = CrediListoPreaprobado.objects.filter(NumeroFactura=numeropedido).first()
            if consulta is not None and consulta.FechaVencimiento > fecha_hoy:
                fecha_actual = datetime.now().date()
                factura_id = consulta.ValorFactura.replace(",", "")
                factura_vencimiento = consulta.FechaVencimiento
                fecha_vencimiento = factura_vencimiento
                diferencia = (fecha_vencimiento - fecha_actual).days
                entero = int(factura_id.replace(',', ''))
                intereses_minimo = 20000
                intereses_segundo = 40000
                mitad = entero * 0.8
                interes = diferencia * 0.000533333 * mitad
                if intereses_minimo > interes:
                    interes = intereses_minimo
                elif intereses_segundo > interes:
                    interes = intereses_segundo
                desembolso = mitad - interes

                codigo_empresa = str(consulta.Empresa.codigo)
                fecha_actual = datetime.utcnow()
                fecha_formateada = fecha_actual.strftime('%Y-%m-%dT%H:%M:%SZ')
                valor_aprobado = float(desembolso)
                intereses = float(interes)
                fecha_formateada_ven = consulta.FechaVencimiento
                fecha_formateada_ven = fecha_formateada_ven.strftime('%Y-%m-%dT%H:%M:%SZ')
                CrediListoPreaprobado.objects.filter(NumeroFactura=numeropedido).update(
                    Interes_aplicado=str(int(intereses)), ValorAprobado_aplicado=str(int(valor_aprobado)))

                # Usa la variable global SESSION
                cookie = "B1SESSION=" + SESSION
                if ROUTEID:  # por si tu SAP devuelve ROUTEID
                    cookie += "; ROUTEID=" + ROUTEID
                # URL del endpoint de VendorPayments
                url_pagos_efectuados = IP_SAP + "VendorPayments"
                if estado != 'con pedidos':
                    # Datos del nuevo pago efectuado que quieres enviar
                    nuevo_pago_efectuado = {
                        "DocType": "rSupplier",
                        "HandWritten": "tNO",
                        "Printed": "tNO",
                        "DocDate": fecha_formateada,
                        "CardCode": codigo_empresa,
                        "CashAccount": "42100517",
                        "DocCurrency": "$",
                        "CashSum": intereses,
                        "CheckAccount": None,
                        "TransferAccount": "11201005",
                        "TransferSum": valor_aprobado,
                        "TransferDate": fecha_formateada,
                        "TransferReference": None,
                        "LocalCurrency": "tNO",
                        "DocRate": 0.0,
                        "Reference2": None,
                        "CounterReference": None,
                        "Remarks": None,
                        "JournalRemarks": str(consulta.NumeroFacturaCliente),
                        "SplitTransaction": "tNO",
                        "ApplyVAT": "tNO",
                        "TaxDate": fecha_formateada,
                        "BankCode": None,
                        "BankAccount": None,
                        "DiscountPercent": 0.0,
                        "ProjectCode": None,
                        "CurrencyIsLocal": "tNO",
                        "DeductionPercent": 0.0,
                        "DeductionSum": 0.0,
                        "CashSumFC": 0.0,
                        "BoeAccount": None,
                        "BillOfExchangeAmount": 0.0,
                        "BillofExchangeStatus": None,
                        "BillOfExchangeAmountFC": 0.0,
                        "BillOfExchangeAmountSC": 0.0,
                        "BillOfExchangeAgent": None,
                        "WTCode": None,
                        "WTAmount": 0.0,
                        "WTAmountFC": 0.0,
                        "WTAmountSC": 0.0,
                        "WTAccount": None,
                        "WTTaxableAmount": 0.0,
                        "Proforma": "tNO",
                        "PayToBankCode": None,
                        "PayToBankBranch": None,
                        "PayToBankAccountNo": None,
                        "PayToCode": "PRINCIPAL",
                        "PayToBankCountry": None,
                        "IsPayToBank": "tNO",
                        "PaymentPriority": "bopp_Priority_6",
                        "TaxGroup": None,
                        "BankChargeAmount": 0.0,
                        "BankChargeAmountInFC": 0.0,
                        "BankChargeAmountInSC": 0.0,
                        "UnderOverpaymentdifference": 0.0,
                        "UnderOverpaymentdiffSC": 0.0,
                        "WtBaseSum": 0.0,
                        "WtBaseSumFC": 0.0,
                        "WtBaseSumSC": 0.0,
                        "VatDate": fecha_formateada,
                        "TransactionCode": "",
                        "PaymentType": "bopt_None",
                        "TransferRealAmount": 0.0,
                        "DocObjectCode": "bopot_OutgoingPayments",
                        "DocTypte": "rSupplier",
                        "DueDate": fecha_formateada_ven,
                        "LocationCode": None,
                        "Cancelled": "tNO",
                        "ControlAccount": "13300614",
                        "UnderOverpaymentdiffFC": 0.0,
                        "AuthorizationStatus": "pasWithout",
                        "BPLID": None,
                        "BPLName": None,
                        "VATRegNum": None,
                        "BlanketAgreement": None,
                        "PaymentByWTCertif": "tNO",
                        "Cig": None,
                        "Cup": None,
                        "AttachmentEntry": None,
                        "SignatureInputMessage": None,
                        "SignatureDigest": None,
                        "CertificationNumber": None,
                        "PrivateKeyVersion": None,
                        "U_HBT_AreVal": "Comun",
                        "U_HBT_Tercero": None,
                        "U_GSP_TPVREF": None,
                        "U_GSP_TPVLINE": None,
                        "U_GSP_TPVPAYMETHOD": None,
                        "U_HBT1_CPTO": "Pago",
                        "U_HBT_PagoAnticipo": "N",
                        "U_BP_Confd": "N",
                        "U_BP_DocNr": None,
                        "U_BP_Seque": None,
                        "PaymentChecks": [],
                        "PaymentInvoices": [],
                        "PaymentCreditCards": [],
                        "PaymentAccounts": [],
                        "PaymentDocumentReferencesCollection": [],
                        "BillOfExchange": {},
                        "WithholdingTaxCertificatesCollection": [],
                        "ElectronicProtocols": [],
                        "CashFlowAssignments": [],
                        "Payments_ApprovalRequests": [],
                        "WithholdingTaxDataWTXCollection": []
                    }
                else:
                    pedido_sap = int(request.POST['pedido_sap'])
                    valor_sap = float(request.POST['valor_sap'])
                    valor_aprob = valor_aprobado - valor_sap
                    valor_negativo = valor_sap * -1

                    nuevo_pago_efectuado = {
                        "DocType": "rSupplier",
                        "HandWritten": "tNO",
                        "Printed": "tNO",
                        "DocDate": fecha_formateada,
                        "CardCode": codigo_empresa,
                        "CashAccount": "42100517",
                        "DocCurrency": "$",
                        "CashSum": intereses,
                        "CheckAccount": None,
                        "TransferAccount": "11201005",
                        "TransferSum": valor_aprob,
                        "TransferDate": fecha_formateada,
                        "TransferReference": None,
                        "LocalCurrency": "tNO",
                        "DocRate": 0.0,
                        "Reference2": None,
                        "CounterReference": None,
                        "Remarks": None,
                        "JournalRemarks": str(consulta.NumeroFacturaCliente),
                        "SplitTransaction": "tNO",
                        "ApplyVAT": "tNO",
                        "TaxDate": fecha_formateada,
                        "BankCode": None,
                        "BankAccount": None,
                        "DiscountPercent": 0.0,
                        "ProjectCode": None,
                        "CurrencyIsLocal": "tNO",
                        "DeductionPercent": 0.0,
                        "DeductionSum": 0.0,
                        "CashSumFC": 0.0,
                        "BoeAccount": None,
                        "BillOfExchangeAmount": 0.0,
                        "BillofExchangeStatus": None,
                        "BillOfExchangeAmountFC": 0.0,
                        "BillOfExchangeAmountSC": 0.0,
                        "BillOfExchangeAgent": None,
                        "WTCode": None,
                        "WTAmount": 0.0,
                        "WTAmountFC": 0.0,
                        "WTAmountSC": 0.0,
                        "WTAccount": None,
                        "WTTaxableAmount": 0.0,
                        "Proforma": "tNO",
                        "PayToBankCode": None,
                        "PayToBankBranch": None,
                        "PayToBankAccountNo": None,
                        "PayToCode": "PRINCIPAL",
                        "PayToBankCountry": None,
                        "IsPayToBank": "tNO",
                        "PaymentPriority": "bopp_Priority_6",
                        "TaxGroup": None,
                        "BankChargeAmount": 0.0,
                        "BankChargeAmountInFC": 0.0,
                        "BankChargeAmountInSC": 0.0,
                        "UnderOverpaymentdifference": 0.0,
                        "UnderOverpaymentdiffSC": 0.0,
                        "WtBaseSum": 0.0,
                        "WtBaseSumFC": 0.0,
                        "WtBaseSumSC": 0.0,
                        "VatDate": fecha_formateada,
                        "TransactionCode": "",
                        "PaymentType": "bopt_None",
                        "TransferRealAmount": 0.0,
                        "DocObjectCode": "bopot_OutgoingPayments",
                        "DocTypte": "rSupplier",
                        "DueDate": fecha_formateada_ven,
                        "LocationCode": None,
                        "Cancelled": "tNO",
                        "ControlAccount": "13300614",
                        "UnderOverpaymentdiffFC": 0.0,
                        "AuthorizationStatus": "pasWithout",
                        "BPLID": None,
                        "BPLName": None,
                        "VATRegNum": None,
                        "BlanketAgreement": None,
                        "PaymentByWTCertif": "tNO",
                        "Cig": None,
                        "Cup": None,
                        "AttachmentEntry": None,
                        "SignatureInputMessage": None,
                        "SignatureDigest": None,
                        "CertificationNumber": None,
                        "PrivateKeyVersion": None,
                        "U_HBT_AreVal": "Comun",
                        "U_HBT_Tercero": None,
                        "U_GSP_TPVREF": None,
                        "U_GSP_TPVLINE": None,
                        "U_GSP_TPVPAYMETHOD": None,
                        "U_HBT1_CPTO": "Pago",
                        "U_HBT_PagoAnticipo": "N",
                        "U_BP_Confd": "N",
                        "U_BP_DocNr": None,
                        "U_BP_Seque": None,
                        "PaymentChecks": [],
                        "PaymentInvoices": [
                            {
                                "DocEntry": pedido_sap,
                                "SumApplied": valor_negativo,
                                "AppliedFC": 0.0,
                                "DocRate": 0.0,
                                "DocLine": 2,
                                "InvoiceType": "it_PaymentAdvice",
                                "DiscountPercent": 0.0,
                                "PaidSum": 0.0,
                                "InstallmentId": 1,
                                "WitholdingTaxApplied": 0.0,
                                "WitholdingTaxAppliedFC": 0.0,
                                "WitholdingTaxAppliedSC": 0.0,
                                "LinkDate": None,
                                "DistributionRule": None,
                                "DistributionRule2": None,
                                "DistributionRule3": None,
                                "DistributionRule4": None,
                                "DistributionRule5": None,
                                "TotalDiscount": 0.0,
                                "TotalDiscountFC": 0.0,
                                "TotalDiscountSC": 0.0
                            }
                        ],
                        "PaymentCreditCards": [],
                        "PaymentAccounts": [],
                        "PaymentDocumentReferencesCollection": [],
                        "BillOfExchange": {},
                        "WithholdingTaxCertificatesCollection": [],
                        "ElectronicProtocols": [],
                        "CashFlowAssignments": [],
                        "Payments_ApprovalRequests": [],
                        "WithholdingTaxDataWTXCollection": []
                    }

                pendientes_seleccionados = request.POST.getlist('pendientes_seleccionados[]')
                cruces=''
                for dato in pendientes_seleccionados:

                    variable = dato
                    variable = str(variable).split(",")
                    cruces=cruces +';'+  str(variable[3])
                    valor_negativo_sub = int(variable[2]) * -1
                    if variable[0] == 'Nota':
                        desembolso = desembolso - int(variable[2])
                        facturas_a_agregar = [
                            {
                                    "DocEntry": variable[4],
                                    "SumApplied": valor_negativo_sub,
                                    "AppliedFC": 0.0,
                                    "AppliedSys": 103.5,
                                    "DocRate": 0.0,
                                    "DocLine": 0,
                                    "InvoiceType": "it_PurchaseCreditNote",
                                    "PaidSum": 0.0,
                                    "InstallmentId": 1,
                                    "WitholdingTaxApplied": 0.0,
                                    "WitholdingTaxAppliedFC": 0.0,
                                    "WitholdingTaxAppliedSC": 0.0,
                                    "LinkDate": None,
                                    "DistributionRule": None,
                                    "DistributionRule2": None,
                                    "DistributionRule3": None,
                                    "DistributionRule4": None,
                                    "DistributionRule5": None,
                                    "TotalDiscount": 0.0,
                            }
                        ]
                        nuevo_pago_efectuado["TransferSum"] = float(desembolso)
                    elif variable[0] == 'Anticipo':
                        facturas_a_agregar = [
                            {
                                "DocEntry": variable[4],
                                "SumApplied": valor_negativo_sub,
                                "AppliedFC": 0.0,
                                "DocRate": 0.0,
                                "InvoiceType": "it_PaymentAdvice",
                                "DiscountPercent": 0.0,
                                "PaidSum": 0.0,
                                "InstallmentId": 1,
                                "WitholdingTaxApplied": 0.0,
                                "WitholdingTaxAppliedFC": 0.0,
                                "WitholdingTaxAppliedSC": 0.0,
                                "LinkDate": None,
                                "DistributionRule": None,
                                "DistributionRule2": None,
                                "DistributionRule3": None,
                                "DistributionRule4": None,
                                "DistributionRule5": None,
                                "TotalDiscount": 0.0,
                                "TotalDiscountFC": 0.0,
                                "TotalDiscountSC": 0.0
                            }
                        ]
                    nuevo_pago_efectuado["PaymentInvoices"].extend(facturas_a_agregar)



                # Convierte los datos a formato JSON
                payload_pagos_efectuados = json.dumps(nuevo_pago_efectuado)

                # Encabezados de la solicitud HTTP
                headers_pagos_efectuados = {
                    'Content-Type': 'application/json',
                    # Agrega aquí tu cookie de sesión
                    'Cookie': cookie
                }

                # Realiza la solicitud POST al Service Layer para crear un nuevo pago efectuado
                response_pagos_efectuados = requests.post(url_pagos_efectuados, headers=headers_pagos_efectuados,
                                                          data=payload_pagos_efectuados, verify=False)

                # Verifica el estado de la respuesta
                if response_pagos_efectuados.status_code == 201:
                    egreso_creado = response_pagos_efectuados.content
                    egreso_creado = json.loads(egreso_creado)
                    egreso_creado = str(egreso_creado.get("DocNum"))
                    print("Pago efectuado creado exitosamente.")
                else:
                    now = datetime.now(pytz.timezone('America/Bogota'))
                    hoy = now.date()
                    hora = now.time()
                    errores = HistorialErrorEnvioSap(
                        accion='Error al enviar informacion a SAP',
                        fecha=hoy,
                        hora=hora,
                        tipo='credilisto',
                        empresa=str(consulta.Empresa.nombre),
                        pedido=consulta.NumeroFactura,
                        mensaje_sistema=response_pagos_efectuados.text
                    )
                    errores.save()
                    messages.add_message(request, messages.ERROR,
                                         'No se ha podido registrar la factura  ' + numeropedido + ' .Por la siguiente razon' + response_pagos_efectuados.text)
                    print("Error al crear el pago efectuado:", response_pagos_efectuados.text)
                    return HttpResponseRedirect('/configuracion/servicio_credilisto_consulta/')
                CrediListoPreaprobado.objects.filter(NumeroFactura=numeropedido).update(estado=estado_pedido,
                                                                                        FechaRespuesta=fecha_hoy,
                                                                                        UsuarioRespuesta=current_user.id,
                                                                                        EgresoCreado=egreso_creado,
                                                                                        CrucesAplicados=cruces)
                aprobado = RespuestaFacturaApi(
                    Identificacion=consulta.Empresa.nit,
                    TipoIdentificacion='NIT',
                    ValorAprobado=consulta.ValorAprobado.replace(",", ""),
                    FechaEmision=consulta.FechaEmision,
                    NumeroFactura=consulta.NumeroFactura,
                    Interes=consulta.Interes.replace(",", ""),
                )
                aprobado.save()
                now = datetime.now(pytz.timezone('America/Bogota'))
                hoy = now.date()
                hora = now.time()
                errores = HistorialErrorEnvioSap(
                    accion='Registrado Con Éxito',
                    fecha=hoy,
                    hora=hora,
                    tipo='credilisto',
                    empresa=str(consulta.Empresa.nombre),
                    pedido=consulta.NumeroFactura,
                )
                errores.save()
                email = EmailMessage('SERVICIO FINANCIERO CREDILISTO APROBADO',
                                     'Tu Servicio Financiero para la factura : \n'
                                     + str(
                                         consulta.NumeroFactura) + "\nha sido aprobado. Podrás consultarlo en el siguiente sitio web:"
                                                                   + IP_SERVIDOR +"/configuracion/servicio_credilisto_historial/",
                                     to=[consulta.Correo])
                email.send()

                messages.add_message(request, messages.INFO,
                                     'Se ha aprobado el credito  ' + numeropedido + ' satisfactoriamente.')
            else:
                messages.add_message(request, messages.WARNING,
                                     'No se puede aprobar el crédito ' + numeropedido + ' se encuentra vencido.')
        else:
            razon = request.POST['razon']
            CrediListoPreaprobado.objects.filter(NumeroFactura=numeropedido).update(estado=respuesta,FechaRespuesta=fecha_hoy,UsuarioRespuesta=current_user.id,Razon=razon)
            consulta = CrediListoPreaprobado.objects.filter(NumeroFactura=numeropedido).first()
            email = EmailMessage('SERVICIO FINANCIERO CREDILISTO DENEGADO',
                                 'Lamentablemente, su Servicio Financiero para la factura : \n'
                                 + str(consulta.NumeroFactura) + "\nha sido rechazado.",
                                 to=[consulta.Correo])
            email.send()
            messages.add_message(request, messages.ERROR,
                                 'Se ha rechazado el credito  ' + numeropedido + ' satisfactoriamente.')


        return HttpResponseRedirect('/configuracion/servicio_credilisto_consulta/')

def config_crediya_consulta_historial(request):

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_crediya.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    else:
        pass

def config_credilisto_consulta_historial(request):

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_credilisto.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    else:
        pass

def informacion_complementaria_consulta_crediya(request, ):
    if request.method == 'GET':

        lista_infoc = CrediyaPreaprobado.objects.all()


        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        empresa = request.GET.get('empresa')
        estado = request.GET.get('estado')
        if not fecha_inicio == '' and not fecha_fin == '':
            lista_infoc = lista_infoc.filter(FechaEmision__range=[fecha_inicio, fecha_fin])

        if not empresa == 'T':
            lista_infoc = lista_infoc.filter(Empresa_id=empresa)

        if not estado == '4':
            lista_infoc = lista_infoc.filter(estado=estado)





        cuenta = lista_infoc.count()
        paginador = Paginator(lista_infoc, PAGINADOR)
        pagina = request.GET.get('page')

        try:
            infos = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            infos = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            infos = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        infos_dict = []
        fecha_comparacion=date(1, 1, 1)
        for info in infos:
            fecha_aprobacion = info.FechaRespuesta if info.FechaRespuesta != fecha_comparacion else ''
            usuario_aprobacion = info.UsuarioRespuesta.first_name +' ' + info.UsuarioRespuesta.last_name  if info.UsuarioRespuesta.id != 1 else ''
            almacen= info.Almacen if info.Almacen != None else ''
            info_dict = {

                'empresa': info.Empresa.nombre,
                'pedido': info.NumeroOrdenCompra,
                'almacen': almacen,
                'valor_preaprobado': info.ValorAprobado,
                'interes': info.Interes,
                'fecha': info.FechaEmision,
                'fecha_solicitud': info.FechaSolicitud,
                'vencimiento': info.FechaVencimiento,
                'estado': info.estado,
                'fecha_aprobacion': fecha_aprobacion,
                'usuario_aprobacion': usuario_aprobacion,

            }
            infos_dict.append(info_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': infos_dict
        }

        return JsonResponse(response_dict)

def informacion_complementaria_consulta_cruces_credilisto(request, ):
    if request.method == 'GET' and request.is_ajax():
        estado='sin pedidos'
        factura = request.GET.get('pedido', None)
        informacion_factura= CrediListoPreaprobado.objects.filter(NumeroFactura=factura).first()
        nombre_empresa= informacion_factura.Empresa.codigo


        url2 = IP_SAP + "SQLQueries('consultanotascredito3')/List?ShortName='" + nombre_empresa + "'"

        response = sap_request(url2)
        response = response.text
        response = response.replace('null', ' " " ')
        response = ast.literal_eval(response)
        response = response['value']

        infos_dict = []
        for info in response:
            info_dict = {
                'descripcion': info['LineMemo'],
                'entry': info['DocEntry'],
                'origen': info['BaseRef'],
                'fecha': info['RefDate'],
                'tipo': 'Nota',
                'saldo': int(info['BalDueDeb'])+int(info['BalDueCred']),
            }
            infos_dict.append(info_dict)



        url2 = IP_SAP + "SQLQueries('consultaanticipos5')/List?ShortName='" + nombre_empresa + "'"

        response = sap_request(url2)
        response = response.text
        response = response.replace('null', ' " " ')
        response = ast.literal_eval(response)
        response = response['value']

        for info in response:
            info_dict = {
                'descripcion': info['LineMemo'],
                'entry': info['TransId'],
                'origen': info['BaseRef'],
                'fecha': info['RefDate'],
                'tipo': 'Anticipo',
                'saldo': int(info['BalDueDeb']) + int(info['BalDueCred']),
            }
            infos_dict.append(info_dict)



        resultado = "Este es el resultado para el pedido "


        url2 = IP_SAP + "SQLQueries('ConsultarPedidosCruce')/List?Fatura='" + factura +  "'"

        response = sap_request(url2)

        try:
            response = response.text
            response = response.replace('null', ' " " ')
            response = ast.literal_eval(response)
            response = response['value'][0]['Pedido']
            pedido=response



            url2 = IP_SAP + "SQLQueries('ConsultaPagoEfectuadosc1')/List?pedido='ANTICIPO OC " + str(pedido) + "'"

            response = sap_request(url2)
            response = ast.literal_eval(response.text)
            response = response['value']
            if response ==[]:
                response='No tiene pagos que afecten esta factura. el pedido relacionado es '+str(pedido)
                valor=''
            else:
                pedidos = response[0]['TransId']
                valor = response[0]['DocTotal']
                response='Se encontro un pago para cruzar con el pedido '+str(pedido)+' con valor de '+str(valor)
                estado='con pedidos'
                pedido=pedidos
        except:
            response='No se encontro un pedido asociado a la factura'
            pedido=''
            valor = ''

        return JsonResponse({'resultado': response,'estado':estado,'pedido':pedido, 'valor':valor,'consulta_pendientes':infos_dict})
    else:
        return JsonResponse({'error': 'Bad request'}, status=400)

def informacion_complementaria_consulta_credilisto(request, ):
    if request.method == 'GET':

        lista_infoc = CrediListoPreaprobado.objects.all()


        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        empresa = request.GET.get('empresa')
        estado = request.GET.get('estado')
        if not fecha_inicio == '' and not fecha_fin == '':
            lista_infoc = lista_infoc.filter(FechaSolicitud__range=[fecha_inicio, fecha_fin])

        if not empresa == 'T':
            lista_infoc = lista_infoc.filter(Empresa_id=empresa)

        if not estado == '4':
            lista_infoc = lista_infoc.filter(estado=estado)





        cuenta = lista_infoc.count()
        paginador = Paginator(lista_infoc, PAGINADOR)
        pagina = request.GET.get('page')

        try:
            infos = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            infos = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            infos = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        infos_dict = []
        fecha_comparacion=date(1, 1, 1)
        for info in infos:
            fecha_aprobacion = info.FechaRespuesta if info.FechaRespuesta != fecha_comparacion else ''
            usuario_aprobacion = info.UsuarioRespuesta.first_name +' ' + info.UsuarioRespuesta.last_name  if info.UsuarioRespuesta.id != 1 else ''
            info_dict = {

                'empresa': info.Empresa.nombre,
                'factura': info.NumeroFactura,
                'facturacliente': info.NumeroFacturaCliente,
                'valor_preaprobado': info.ValorAprobado,
                'interes': info.Interes,
                'interes_aplicado': info.Interes_aplicado,
                'fecha': info.FechaEmision,
                'vencimiento': info.FechaVencimiento,
                'fechafactura': info.FechaSolicitud,
                'estado': info.estado,
                'fecha_aprobacion': fecha_aprobacion,
                'usuario_aprobacion': usuario_aprobacion,

            }
            infos_dict.append(info_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': infos_dict
        }

        return JsonResponse(response_dict)

def informacion_complementaria_historial_crediya(request, ):
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_infoc = CrediyaPreaprobado.objects.filter(Empresa_id=usuario_datos.empresa).order_by('-FechaSolicitud')


        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        estado = request.GET.get('estado')
        if not fecha_inicio == '' and not fecha_fin == '':
            lista_infoc = lista_infoc.filter(FechaEmision__range=[fecha_inicio, fecha_fin])


        if not estado == '4':
            lista_infoc = lista_infoc.filter(estado=estado)





        cuenta = lista_infoc.count()
        paginador = Paginator(lista_infoc, PAGINADOR)
        pagina = request.GET.get('page')

        try:
            infos = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            infos = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            infos = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        infos_dict = []

        for info in infos:
            if info.Razon == None:
                razon=''
            else:
                razon=info.Razon
            info_dict = {

                'empresa': info.Empresa.nombre,
                'pedido': info.NumeroOrdenCompra,
                'valor_preaprobado': info.ValorAprobado,
                'interes': info.Interes,
                'fecha': info.FechaEmision,
                'estado': info.estado,
                'fecha_solicitud': info.FechaSolicitud,
                'fecha_vencimiento': info.FechaVencimiento,
                'razon':razon,
                'fecha_aprobado':info.FechaRespuesta,

            }
            infos_dict.append(info_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': infos_dict
        }

        return JsonResponse(response_dict)

def informacion_complementaria_historial_credilisto(request, ):
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_infoc = CrediListoPreaprobado.objects.filter(Empresa_id=usuario_datos.empresa).order_by('-FechaSolicitud')


        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        estado = request.GET.get('estado')
        if not fecha_inicio == '' and not fecha_fin == '':
            lista_infoc = lista_infoc.filter(FechaEmision__range=[fecha_inicio, fecha_fin])


        if not estado == '4':
            lista_infoc = lista_infoc.filter(estado=estado)





        cuenta = lista_infoc.count()
        paginador = Paginator(lista_infoc, PAGINADOR)
        pagina = request.GET.get('page')

        try:
            infos = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            infos = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            infos = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        infos_dict = []

        for info in infos:
            if info.Razon == None:
                razon=''
            else:
                razon=info.Razon
            info_dict = {

                'empresa': info.Empresa.nombre,
                'pedido': info.NumeroFactura,
                'factura_cliente': info.NumeroFacturaCliente,
                'valor_preaprobado': info.ValorAprobado,
                'interes': info.Interes,
                'interesapl': info.Interes_aplicado,
                'fecha': info.FechaEmision,
                'fecha_vencimiento': info.FechaVencimiento,
                'estado': info.estado,
                'fecha_solicitud': info.FechaSolicitud,
                'razon':razon,
                'fecha_aprobado':info.FechaRespuesta,

            }
            infos_dict.append(info_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': infos_dict
        }

        return JsonResponse(response_dict)

def config_servicio_crediya_lista(request):
    # Render  administracion.html
    if request.method == 'GET' and request.is_ajax():
        pedido_id = request.GET.get('pedido_id')
        entero = int(pedido_id.replace(',', ''))
        intereses_minimo=20000
        intereses_segundo=40000
        mitad=entero/2
        interes=0.018*mitad
        if intereses_minimo>interes:
            interes=intereses_minimo
        elif intereses_segundo>interes:
            interes=intereses_segundo
        desembolso=mitad-interes
        desembolso = "{:,.0f}".format(desembolso)
        interes = "{:,.0f}".format(interes)
        mitad = "{:,.0f}".format(mitad)
        data = {'valor1': desembolso, 'valor2': interes, 'valor3':mitad}
        return JsonResponse(data)



def config_servicio_credilisto_lista(request):
    # Render  administracion.html
    if request.method == 'GET' and request.is_ajax():
        fecha_actual = datetime.now().date()
        factura_id = request.GET.get('factura_id')
        factura_vencimiento = request.GET.get('factura_vencimiento')
        fecha_vencimiento = datetime.strptime(factura_vencimiento, '%Y%m%d').date()
        diferencia = (fecha_vencimiento - fecha_actual).days
        entero = int(factura_id.replace(',', ''))
        intereses_minimo=20000
        intereses_segundo=40000
        mitad=entero*0.8
        interes=diferencia*0.000533333*mitad
        if intereses_minimo>interes:
            interes=intereses_minimo
        elif intereses_segundo>interes:
            interes=intereses_segundo
        desembolso=mitad-interes
        desembolso = "{:,.0f}".format(desembolso)
        interes = "{:,.0f}".format(interes)
        mitad = "{:,.0f}".format(mitad)
        diferencia=diferencia
        data = {'valor1': desembolso, 'valor2': interes, 'valor3':mitad,'diferencia':diferencia}
        return JsonResponse(data)



def config_servicio_codigoregistro_lista(request):
    # Render  administracion.html
    if request.method == 'GET' and request.is_ajax():
        codigo=str(uuid.uuid4())
        if CodigosRegistros.objects.filter(codigo=codigo).exists():
            config_servicio_codigoregistro_lista(request)
        data = {'valor1': codigo}
        return JsonResponse(data)



def config_servicio_credilisto_registro(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.nombre
        hoy = date.today()
        hoy = hoy + timedelta(days=9)
        hoy = hoy.strftime("%Y-%m-%d")

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')




        url2 = IP_SAP + "SQLQueries('FacturasHabiles3')/List?fecha='" + hoy +  "'&empresa='" + str(
                empresa)+"'"

        response = sap_request(url2)
        response = response.text
        response = response.replace('null', ' " " ')
        response = ast.literal_eval(response)
        response = response['value']
        facturas=[]
        for factura in response:
            credilisto_preaprobados = CrediListoPreaprobado.objects.filter(NumeroFactura=factura['DocNum']).first()
            estado = credilisto_preaprobados.estado if credilisto_preaprobados else None
            fecha_objeto = datetime.strptime(factura['DocDate'], "%Y%m%d")
            fecha_formateada = fecha_objeto.strftime("%Y-%m-%d")
            fecha_ven = datetime.strptime(factura['DocDueDate'], "%Y%m%d")
            fecha_formateada_ven = fecha_ven.strftime("%Y-%m-%d")
            factura_det={
                'referencia':factura['NumAtCard'],
                'fecha_formateada':fecha_formateada,
                'fecha_formateada_ven':fecha_formateada_ven,
                'fecha_vencimiento':factura['DocDueDate'],
                'factura':factura['DocNum'],
                'total':format(int(factura['DocTotal']), '0,.0f'),
                'fecha_factura': factura['DocDate'],
                'estado': estado,
            }
            facturas.append(factura_det)

        return render(request, "config_servicio_credilisto_registro.html", {'user': current_user,
                                                           'permiso_usuario': usuario_datos,
                                                            'empresa': empresa,
                                                           'facturas': facturas,
                                                           })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        hoy = date.today()
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        correo=empresa.email
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.pk
        numero_pedido = request.POST['numero_pedido']
        intereses = request.POST['intereses']
        factura_cliente = request.POST['factura_cliente']
        total_desembolso = request.POST['total_desembolso']
        fecha_pedido = request.POST['fecha_pedido']
        fecha_vencimiento = request.POST['fecha_vencimiento']
        valor_orden = request.POST['valor_orden']
        fecha_pedido = datetime.strptime(fecha_pedido, "%Y%m%d").date()
        fecha_vencimiento = datetime.strptime(fecha_vencimiento, "%Y%m%d").date()

        if CrediListoPreaprobado.objects.filter(NumeroFactura=numero_pedido).exists():
            messages.add_message(request, messages.ERROR,
                                 'La solicitud de crédito para la factura ' + str(numero_pedido) + ' ya fue realizada previamente. ')

            return HttpResponseRedirect('/configuracion/servicio_credilisto_registro/')

        preaprobado = CrediListoPreaprobado(
            Empresa_id=empresa,
            UsuarioSolicitado_id=current_user.id,
            TipoIdentificacion='NIT',
            ValorAprobado=total_desembolso,
            NumeroFactura=numero_pedido,
            NumeroFacturaCliente=factura_cliente,
            Interes=intereses,
            estado='Preaprobado',
            FechaEmision=fecha_pedido,
            FechaVencimiento=fecha_vencimiento,
            FechaSolicitud=hoy,
            ValorFactura=valor_orden,
            Correo=correo,
        )
        preaprobado.save()
        messages.add_message(request, messages.INFO,
                             'El credito para la factura ' + str(numero_pedido) + ' se encuentra preaprobado.')

        return HttpResponseRedirect('/configuracion/servicio_credilisto_registro/')





def config_servicio_codigoregistro_registro(request):
    # Render  administracion.html

    if request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        hoy = date.today()
        empresa = request.POST['empresa_usuario']
        codigoregistro = request.POST['codigoregistro']
        partes = empresa.split(u'-')
        if len(partes) >= 2:
            card_code = partes[0]
            card_name = partes[1]
        else:
            card_code = partes[0]
            card_name = u''



        preaprobado = CodigosRegistros(
            empresa=card_name,
            codigo_cliente=card_code,
            codigo=codigoregistro,
            creado=hoy,
        )
        preaprobado.save()
        messages.add_message(request, messages.INFO,
                             'Se ha creado el código número ' + str(codigoregistro) + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/usuarios/')




def config_servicio_registro_empresas_aut(request):
    # Render  administracion.html

    if request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        hoy = date.today()
        codigo = request.POST['codigo']
        empresa_exist=Empresas.objects.filter(codigo=codigo)



        url2 = IP_SAP + "SQLQueries('EmpresariosListaConsultas')/List?empresario= '" + codigo + "'"

        response = sap_request(url2)
        response = response.text
        response = response.replace('null', ' " " ')
        response = ast.literal_eval(response)
        response = response['value']
        if response == []:
            messages.add_message(request, messages.ERROR,
                                 'No se encontro el codigo  ' + str(codigo) + ' en SAP.')

            return HttpResponseRedirect('/configuracion/empresas/')
        for datos in response:
            id = datos['CardCode']
            nombre = datos['CardName']
            nit = datos['LicTradNum']
            telefono_emp = datos['Phone1']
            email_empre = datos['E_Mail']
            direccion_empresa = datos['Address']
            ean = datos['U_EAN']
            tipo = datos['CardType']
        if tipo== 'S':
            tipo='Empresario'
        else:
            tipo='Cadena'
        if empresa_exist:  # Si existe la PK, es una actualización
            empresa_exist = Empresas.objects.filter(codigo=codigo).first()
            empresasr = Empresas.objects.get(pk=empresa_exist.pk)
            empresasr.nombre = nombre
            empresasr.nit = nit
            empresasr.telefono = telefono_emp
            empresasr.movil = telefono_emp
            empresasr.responsable = nombre
            empresasr.tipo = tipo
            empresasr.email = email_empre
            empresasr.edi = ean
            empresasr.codigo = id
        else:  # Si no existe, es una creación
            empresasr = Empresas(
                nombre=nombre,
                nit=nit,
                telefono=telefono_emp,
                movil=telefono_emp,
                responsable=nombre,
                tipo=tipo,
                email=email_empre,
                edi=ean,
                codigo=id,
            )

        empresasr.save()
        messages.add_message(request, messages.INFO,
                             'Se ha actualizado la empresa ' + str(nombre) + ' satisfactoriamente.')


        return HttpResponseRedirect('/configuracion/empresas/')



def config_solicitudes_generales(request):

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_grupo = AreasAtencion.objects.all()
        lista_peticiones = Peticiones.objects.all()


        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_solicitudes_generales.html", {'user': current_user,
                                                        'lista_grupo': lista_grupo,
                                                        'lista_peticiones': lista_peticiones,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    else:
        pass



def informacion_complementaria_consulta_solicitud(request, ):
    if request.method == 'GET':

        lista_infoc = RespuestaPedido.objects.all()

        if request.GET.get('peticiones'):
            fecha_inicio = request.GET.get('fecha_inicio')
            fecha_fin = request.GET.get('fecha_fin')
            peticiones = request.GET.get('peticiones')
            estado = request.GET.get('estado')
            grupo = request.GET.get('grupo')
            if not fecha_inicio == '' and not fecha_fin == '':
                lista_infoc = lista_infoc.filter(fecha__range=[fecha_inicio, fecha_fin])

            if not peticiones == 'T':
                lista_infoc = lista_infoc.filter(peticion_id=peticiones)

            if not estado == '4':
                lista_infoc = lista_infoc.filter(estado=estado)

            if not grupo == 'T':
                lista_infoc = lista_infoc.filter(peticion__area=grupo)



        cuenta = lista_infoc.count()
        paginador = Paginator(lista_infoc, PAGINADOR)
        pagina = request.GET.get('page')

        try:
            infos = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            infos = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            infos = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        infos_dict = []

        for info in infos:
            info_dict = {

                'id': info.id,
                'num_pedido': info.num_pedido,
                'entry_pedido': info.entry_pedido,
                'empresa': info.empresa,
                'fecha': info.fecha,
                'estado': info.estado,
                'peticion': info.peticion.descripcion,

            }
            infos_dict.append(info_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': infos_dict
        }

        return JsonResponse(response_dict)


def informacion_pedidos_otros_canales_solicitud(request, ):
    if request.method == 'GET':

        lista_infoc = PedidosOtrosCanales.objects.all()

        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        empresa = request.GET.get('empresa')
        pedido = request.GET.get('pedido')
        estado = request.GET.get('estado')
        if estado=='enproceso':
            estado='en proceso'
        if not fecha_inicio == '' and not fecha_fin == '':
            lista_infoc = lista_infoc.filter(fecha__range=[fecha_inicio, fecha_fin])

        if not empresa == '':
            lista_infoc = lista_infoc.filter(empresa=empresa)

        if not estado == '':
            lista_infoc = lista_infoc.filter(estado=estado)

        if not pedido == '':
            lista_infoc = lista_infoc.filter(num_pedido=int(pedido))


        cuenta = lista_infoc.count()
        paginador = Paginator(lista_infoc, PAGINADOR)
        pagina = request.GET.get('page')

        try:
            infos = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            infos = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            infos = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        infos_dict = []

        for info in infos:
            fecha_entrega=info.fecha_entrega
            if info.fecha_entrega == None:
                fecha_entrega=''

            info_dict = {

                'num_pedido': info.num_pedido,
                'num_pedido_CLI': info.numero_pedido_cliente,
                'fecha': info.fecha,
                'fecha_entrega': fecha_entrega,
                'hora': info.hora,
                'estado': info.estado,

            }
            infos_dict.append(info_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': infos_dict
        }

        return JsonResponse(response_dict)


def informacion_pedidos_otros_canales_solicitud_cliente(request, ):
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_infoc = PedidosOtrosCanales.objects.all()

        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        pedido = request.GET.get('pedido')
        estado = request.GET.get('estado')
        lista_infoc = lista_infoc.filter(empresa_id=usuario_datos.empresa)
        if not fecha_inicio == '' and not fecha_fin == '':
            lista_infoc = lista_infoc.filter(fecha__range=[fecha_inicio, fecha_fin])



        if not pedido == '':
            lista_infoc = lista_infoc.filter(num_pedido=int(pedido))

        if not estado == '':
            if estado=='enproceso':
                estado='en proceso'
            lista_infoc = lista_infoc.filter(estado=str(estado))


        cuenta = lista_infoc.count()
        paginador = Paginator(lista_infoc, PAGINADOR)
        pagina = request.GET.get('page')

        try:
            infos = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            infos = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            infos = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        infos_dict = []

        for info in infos:
            fecha_entrega = info.fecha_entrega
            if info.fecha_entrega == None:
                fecha_entrega = ''
            info_dict = {

                'num_pedido': info.num_pedido,
                'fecha': info.fecha,
                'fecha_entrega': fecha_entrega,
                'hora': info.hora,
                'estado': info.estado,

            }
            infos_dict.append(info_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': infos_dict
        }

        return JsonResponse(response_dict)


def informacion_pedidos_otros_canales_empresario_solicitud(request, ):
    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.id

        lista_infoc = AsignacionPedidosOtrosCanales.objects.all()

        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        empresa_input = request.GET.get('empresa_input')
        pedido = request.GET.get('pedido')
        estado = request.GET.get('estado')
        codigo = request.GET.get('codigo')
        if estado=='enproceso':
            estado='en proceso'
        if not fecha_inicio == '' and not fecha_fin == '':
            lista_infoc = lista_infoc.filter(num_detalle__num_pedido__fecha__range=[fecha_inicio, fecha_fin])
        if not empresa_input == '':
            lista_infoc = lista_infoc.filter(empresa_id=empresa_input)
        if not pedido == '':
            lista_infoc= lista_infoc.filter(num_detalle__num_pedido__num_pedido=pedido)
        if not estado == '':
            lista_infoc= lista_infoc.filter(num_detalle__num_pedido__estado=estado)
        if not codigo == '':
            lista_infoc = lista_infoc.filter(empresa__codigo=codigo)
        cuenta = lista_infoc.count()
        paginador = Paginator(lista_infoc, PAGINADOR)
        pagina = request.GET.get('page')

        try:
            infos = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            infos = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            infos = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        infos_dict = []

        for info in infos:
            fecha=info.fecha
            if fecha == None:
                fecha =''

            info_dict = {
                'identificador': info.pk,
                'num_pedido': info.num_detalle.num_pedido.num_pedido,
                'cantidad': info.cantidad,
                'codigo': info.empresa.codigo,
                'fecha': fecha,
                'referencia': info.num_detalle.referencia,
                'nombre': info.num_detalle.nombre,
                'observaciones': info.num_detalle.observaciones,
                'empresa':info.empresa.nombre,
                'cliente':info.num_detalle.num_pedido.empresa.nombre

            }
            infos_dict.append(info_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': infos_dict
        }

        return JsonResponse(response_dict)

def informacion_pedidos_otros_canales_empresario_facturar(request, ):
    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.id

        lista_infoc = AsignacionPedidosOtrosCanales.objects.all()

        fecha_inicio = request.GET.get('fecha_inicio') or ''
        fecha_fin = request.GET.get('fecha_fin') or ''
        empresa_input = request.GET.get('empresa_input') or ''
        pedido = request.GET.get('pedido') or ''
        estado = request.GET.get('estado') or ''
        referencia = request.GET.get('referencia') or ''
        u_plu = request.GET.get('u_plu') or ''

        if estado=='enproceso':
            estado='en proceso'
        if not fecha_inicio == '' and not fecha_fin == '':
            lista_infoc = lista_infoc.filter(num_detalle__num_pedido__fecha__range=[fecha_inicio, fecha_fin])
        if not empresa_input == '':
            lista_infoc = lista_infoc.filter(num_detalle__num_pedido__empresa_id=empresa_input)
        if not pedido == '':
            lista_infoc= lista_infoc.filter(num_detalle__num_pedido__num_pedido=pedido)
        if not estado == '':
            lista_infoc= lista_infoc.filter(num_detalle__num_pedido__estado=estado)
        if not referencia == '':
            lista_infoc = lista_infoc.filter(num_detalle__referencia=referencia)
        if not u_plu == '':
            lista_infoc = lista_infoc.filter(num_detalle__u_plu=u_plu)
        cuenta = lista_infoc.count()
        paginador = Paginator(lista_infoc, PAGINADOR)
        pagina = request.GET.get('page')

        try:
            infos = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            infos = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            infos = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        infos_dict = []

        for info in infos:
            fecha=info.fecha
            if fecha == None:
                fecha =''
            cantidad_real=info.cantidad-info.cantidadfacturada
            info_dict = {

                'num_pedido': info.num_detalle.num_pedido.num_pedido,
                'cantidad': cantidad_real,
                'cantidadfactura': info.cantidadfacturada,
                'pk': info.pk,
                'codigo': info.empresa.codigo,
                'fecha': fecha,
                'referencia': info.num_detalle.referencia,
                'u_plu': info.num_detalle.u_plu,
                'nombre': info.num_detalle.nombre,
                'observaciones': info.num_detalle.observaciones,
                'empresa':info.empresa.nombre,
                'cliente':info.num_detalle.num_pedido.empresa.nombre

            }
            infos_dict.append(info_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': infos_dict
        }

        return JsonResponse(response_dict)





def informacion_pedidos_otros_canales_empresario_recibo(request, ):
    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.id

        lista_infoc = AsignacionPedidosOtrosCanales.objects.exclude(cantidad=F('cantidadrecibo'))

        fecha_inicio = request.GET.get('fecha_inicio') or ''
        fecha_fin = request.GET.get('fecha_fin') or ''
        empresa_input = request.GET.get('empresa_input') or ''
        pedido = request.GET.get('pedido') or ''
        estado = request.GET.get('estado') or ''
        referencia = request.GET.get('referencia') or ''
        u_plu = request.GET.get('u_plu') or ''
        if estado=='enproceso':
            estado='en proceso'
        if not fecha_inicio == '' and not fecha_fin == '':
            lista_infoc = lista_infoc.filter(num_detalle__num_pedido__fecha__range=[fecha_inicio, fecha_fin])
        if not empresa_input == '':
            lista_infoc = lista_infoc.filter(empresa_id=empresa_input)
        if not pedido == '':
            lista_infoc= lista_infoc.filter(num_detalle__num_pedido__num_pedido=pedido)
        if not estado == '':
            lista_infoc= lista_infoc.filter(num_detalle__num_pedido__estado=estado)
        if not referencia == '':
            lista_infoc = lista_infoc.filter(num_detalle__referencia=referencia)
        if not u_plu == '':
            lista_infoc = lista_infoc.filter(num_detalle__u_plu=u_plu)
        cuenta = lista_infoc.count()
        paginador = Paginator(lista_infoc, PAGINADOR)
        pagina = request.GET.get('page')

        try:
            infos = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            infos = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            infos = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        infos_dict = []

        for info in infos:
            try:
                imagenes= ImagenesOtrosCanales.objects.filter(referencia=info.num_detalle.referencia).first()
                imagenes= imagenes.imagen
            except:
                imagenes=''
            fecha=info.fecha
            if fecha == None:
                fecha =''
            cantidad_real=info.cantidad-info.cantidadrecibo
            info_dict = {

                'num_pedido': info.num_detalle.num_pedido.num_pedido,
                'cantidad': cantidad_real,
                'imagen': imagenes,
                'cantidadped': info.cantidad,
                'cantidadrecibo': info.cantidadrecibo,
                'pk': info.pk,
                'codigo': info.empresa.codigo,
                'fecha': fecha,
                'u_plu': info.num_detalle.u_plu,
                'referencia': info.num_detalle.referencia,
                'nombre': info.num_detalle.nombre,
                'observaciones': info.num_detalle.observaciones,
                'empresa':info.empresa.nombre,
                'cliente':info.num_detalle.num_pedido.empresa.nombre

            }
            infos_dict.append(info_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': infos_dict
        }

        return JsonResponse(response_dict)





def admin_graficas_powerbi(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_graficas = Graficas.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "admin_graficas_powerbi.html", {'user': current_user,
                                                          'lista_graficas': lista_graficas,
                                                          'permiso_usuario': usuario_datos})
    elif request.method == 'POST':
        codigo = request.POST['codigo']
        nombre = request.POST['nombre']
        grafica = request.POST['grafica']
        nivel = request.POST['nivel']
        tipo_usuario = request.POST['tipo_usuario']
        tabla = request.POST.get('tabla', '')
        campo = request.POST.get('campo', '')
        area = request.POST.get('area', '')
        if codigo == '':
            graficas = Graficas(nombre=nombre, grafico=grafica, tipo_usuario=tipo_usuario, tabla=tabla, campo=campo, area=area,nivel=int(nivel))
        else:
            graficas = Graficas(pk=codigo, nombre=nombre, grafico=grafica, tipo_usuario=tipo_usuario, tabla=tabla, campo=campo, area=area,nivel=int(nivel))
        graficas.save()

        return HttpResponseRedirect('/configuracion/definiciones/graficas/')


def facturacion_otros_canales(request):
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_graficas = Graficas.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "admin_graficas_powerbi.html", {
            'user': current_user,
            'lista_graficas': lista_graficas,
            'permiso_usuario': usuario_datos
        })

    elif request.method == 'POST':
        items = request.POST.dict()
        indices = sorted(set([k.split('[')[1].split(']')[0] for k in items.keys() if k.startswith('items[')]))

        pedido_unico = None

        for i in indices:
            pk = items.get('items[%s][pk]' % i)
            cantidad = int(items.get('items[%s][cantidad]' % i, 0))

            asignacion = AsignacionPedidosOtrosCanales.objects.filter(pk=pk).first()
            if asignacion:
                if pedido_unico is None:
                    pedido_unico = asignacion.num_detalle.num_pedido_id
                elif pedido_unico != asignacion.num_detalle.num_pedido_id:
                    # Protección backend si alguien manipula HTML
                    return HttpResponseRedirect('/configuracion/orden_empresiario_otroscanales_facturar/')

                asignacion.cantidadfacturada += cantidad
                asignacion.save()

                pedido_id = asignacion.num_detalle.num_pedido_id
                detalles = DetallesPedidosOtrosCanales.objects.filter(num_pedido_id=pedido_id)

                completado = True
                for detalle in detalles:
                    total_facturado = AsignacionPedidosOtrosCanales.objects.filter(
                        num_detalle=detalle
                    ).aggregate(total=models.Sum('cantidadfacturada'))['total'] or 0

                    if total_facturado < detalle.cantidad:
                        completado = False
                        break

                estado = 'completado' if completado else 'pendiente'
                PedidosOtrosCanales.objects.filter(pk=pedido_id).update(estado=estado)

        return HttpResponseRedirect('/configuracion/orden_empresiario_otroscanales_facturar/')




def recibo_otros_canales(request):
    if request.method != 'POST':
        return HttpResponse(json.dumps({'ok': False, 'error': 'Método no permitido'}), content_type='application/json', status=405)

    try:
        try:
            data = json.loads(request.body)
        except ValueError:
            return HttpResponse(json.dumps({'ok': False, 'error': 'JSON inválido'}), content_type='application/json', status=400)

        pedidos = data.get('pedidos', [])
        if not pedidos:
            return HttpResponse(json.dumps({'ok': False, 'error': 'No se recibieron pedidos'}), content_type='application/json', status=400)

        # Agrupar pedidos por empresa_id (si en el JS mandas nombre en vez de id, aquí se agrupará por ese nombre
        # pero luego buscaremos la empresa real desde la asignacion)
        agrupados = defaultdict(list)
        for p in pedidos:
            # si en el JS mandas el nombre de la empresa, 'empresa' contendrá ese nombre; si mandas id, contendrá el id.
            agrupados[p.get('empresa', 'Desconocida')].append(p)

        resultado_log = []
        for empresa_key, lista in agrupados.items():
            cuerpo = u"Estimado/a,\n\nSe han recibido los siguientes pedidos:\n\n"
            destinatarios = set()  # emails de usuarios asociados a la empresa (evita duplicados)
            empresa_obj = None

            for p in lista:
                pedido_id = p.get('id')
                try:
                    cantidad = int(p.get('cantidad', 0))
                except (TypeError, ValueError):
                    cantidad = 0
                novedad = (p.get('novedad', '') or '').strip()

                if not pedido_id or cantidad <= 0:
                    # ignorar entradas inválidas
                    resultado_log.append({'id': pedido_id, 'status': 'invalid_cantidad_o_id'})
                    continue

                asignacion = AsignacionPedidosOtrosCanales.objects.filter(pk=pedido_id).first()
                if not asignacion:
                    resultado_log.append({'id': pedido_id, 'status': 'asignacion_no_encontrada'})
                    continue

                # Guardamos la empresa real para luego buscar usuarios
                if not empresa_obj:
                    empresa_obj = asignacion.empresa  # es un FK a Empresas

                # Actualizar registro tal como antes (sumando cantidad recibida)
                try:
                    asignacion.cantidadrecibo = (asignacion.cantidadrecibo or 0) + cantidad
                    asignacion.save()
                except Exception as e:
                    resultado_log.append({'id': pedido_id, 'status': 'error_guardado', 'error': str(e)})
                    # seguir con el resto sin detener todo

                # Datos para el cuerpo del correo
                detalles = getattr(asignacion, 'num_detalle', None)
                if detalles is not None:
                    pedido_rel = getattr(detalles, 'num_pedido', None)
                    referencia = getattr(detalles, 'referencia', u'')
                    nombre_detalle = getattr(detalles, 'nombre', u'')
                    numero_pedido = pedido_rel.pk if pedido_rel else u'--'
                else:
                    referencia = u''
                    nombre_detalle = u''
                    numero_pedido = u'--'

                cuerpo += (
                    u"Pedido Nº: {0}\n"
                    u"Referencia: {1}\n"
                    u"Descripción: {2}\n"
                    u"Cantidad recibida: {3}\n"
                    u"Novedades: {4}\n\n"
                ).format(
                    numero_pedido,
                    referencia,
                    nombre_detalle,
                    cantidad,
                    novedad or u"Sin novedades registradas."
                )

                resultado_log.append({'id': pedido_id, 'status': 'procesado', 'cantidad': cantidad})

            cuerpo += u"Saludos cordiales,\n\nSistema de Recibo de Pedidos"

            # Buscar usuarios de la empresa para obtener emails
            if empresa_obj:
                try:
                    usuarios = Usuarios_datos.objects.filter(empresa=empresa_obj)
                    for u in usuarios:
                        try:
                            if u.usuario and getattr(u.usuario, 'email', None):
                                destinatarios.add(u.usuario.email)
                        except Exception:
                            # ignorar usuarios con estructura rara
                            pass
                except Exception as e:
                    # problema al consultar usuarios: registrar pero no detener todo
                    resultado_log.append({'empresa': empresa_key, 'status': 'error_consulta_usuarios', 'error': str(e)})

                # Si no hay destinatarios tomamos como respaldo el campo email de la empresa (si existe)
                if not destinatarios:
                    # Intentamos leer un campo 'email' en el modelo Empresas (si existe)
                    try:
                        if empresa_obj and getattr(empresa_obj, 'email', None):
                            destinatarios.add(empresa_obj.email)
                    except Exception:
                        pass

            # Enviar correo si hay destinatarios
            if destinatarios:
                try:
                    recipients = list(destinatarios)  # lista de emails
                    # Para pruebas forzar a un único correo (descomenta si lo necesitas):
                    #recipients = ['juansebastianduartes@gmail.com']

                    email = EmailMessage(
                        subject=u"Recibo de pedidos - {0}".format(getattr(empresa_obj, 'nombre', empresa_key)),
                        body=cuerpo,
                        to=recipients,
                        # from_email="tu_email@dominio.com"  # opcional: especifica from si lo deseas
                    )
                    email.send(fail_silently=False)  # que lance excepción si falla
                    resultado_log.append(
                        {'empresa': getattr(empresa_obj, 'nombre', empresa_key), 'status': 'correo_enviado',
                         'to': recipients})
                except Exception as e:
                    resultado_log.append(
                        {'empresa': getattr(empresa_obj, 'nombre', empresa_key), 'status': 'error_envio',
                         'error': str(e)})
            else:
                resultado_log.append({'empresa': getattr(empresa_obj, 'nombre', empresa_key), 'status': 'sin_destinatarios'})

        return HttpResponse(json.dumps({'ok': True, 'log': resultado_log}), content_type='application/json')

    except Exception as e:
        # Error inesperado
        return HttpResponse(json.dumps({'ok': False, 'error': str(e)}), content_type='application/json', status=500)




def detalle_grafica(request, grafica_id):

    try:
        grafica = Graficas.objects.get(pk=grafica_id)
        data_grafica =  {"codigo":grafica.pk ,
                      "nombre":grafica.nombre,
                      "tipo_usuario":grafica.tipo_usuario,
                       "grafica": grafica.grafico,
                       "tabla": grafica.tabla,
                       "campo": grafica.campo,
                       "area": grafica.area,
                       "nivel": grafica.nivel,
                     }
    except:
        data_grafica = {"codigo":"",
                     "nombre": "",
                     "tipo_usuario": "",
                     "grafica": "",
                     "tabla": "",
                     "campo": "",
                     "area": "",
                     "nivel": "",
                    }

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_graficas = Graficas.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "admin_graficas_powerbi.html",
                      {'user_name': current_user.first_name,
                       'permiso_usuario': usuario_datos,
                       'lista_graficas': lista_graficas,
                       'data_grafica':data_grafica})

    elif request.method == 'DELETE':
        pais = Pais.objects.get(pk=pais_id)
        pais.delete()
        return JsonResponse({'status': '200', 'mensaje': 'El pais : ' + str(pais) +
                                                         ' has sido  borrado exitosamente'})



def config_graficas_borrar(request, grafica_id):
    if request.method == 'GET':

        grafica = Graficas.objects.get(pk=grafica_id)
        usuario_actual = request.user
        current_user = request.user


        grafica.delete()
        messages.add_message(request, messages.WARNING,
                                 'Se ha borrado la grafica ' + str(grafica_id) + ' satisfactoriamente')

        return HttpResponseRedirect('/configuracion/definiciones/graficas/')




def admin_graficas_actuales_pcs_powerbi(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "admin_graficas_actuales_pcs_powerbi.html", {'user': current_user,
                                                          'permiso_usuario': usuario_datos})
    else:
        pass



def admin_graficas_empresario_cliente_powerbi(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_graficas = Graficas.objects.filter(tipo_usuario='CLIENTE').order_by('nivel')
        url = 'https://login.microsoftonline.com/common/oauth2/token'

        data = {
            'grant_type': 'password',
            'resource': 'https://analysis.windows.net/powerbi/api',
            'client_id': 'a77e1c67-79c8-4dfb-bc49-27f416a217fb',
            'client_secret': 'a2a8Q~zOcyCcEAgUjc9d38WxNUPoGs6r44pZrcH_',
            'username': 'analistati@pcsocial.org',
            'password': 'csAti2017*'
        }

        response = requests.post(url, data=data)
        token = response.json()['access_token']
        lista_de_graficas=[]
        for graficas in lista_graficas:
            diccionario={
                'nombre':graficas.nombre,
                'grafico':str(graficas.grafico)+"&filter="+str(graficas.tabla)+'/'+str(graficas.campo)+" eq '"+str(usuario_datos.empresa.nombre)+ "'&filterPaneEnabled=false&token="+str(token),
            }
            lista_de_graficas.append(diccionario)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "admin_graficas_empresarios_cliente_powerbi.html", {'user': current_user,
                                                          'lista_graficas': lista_de_graficas,
                                                          'permiso_usuario': usuario_datos})
    else:
        pass



@xframe_options_exempt
def admin_graficas_actuales_empresarios_powerbi(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_graficas = Graficas.objects.filter(tipo_usuario='EMPRESARIO').order_by('nivel')
        url = 'https://login.microsoftonline.com/common/oauth2/token'

        data = {
            'grant_type': 'password',
            'resource': 'https://analysis.windows.net/powerbi/api',
            'client_id': 'a77e1c67-79c8-4dfb-bc49-27f416a217fb',
            'client_secret': 'a2a8Q~zOcyCcEAgUjc9d38WxNUPoGs6r44pZrcH_',
            'username': 'analistati@pcsocial.org',
            'password': 'csAti2017*'
        }

        response = requests.post(url, data=data)
        token = response.json()['access_token']
        lista_de_graficas=[]
        for graficas in lista_graficas:
            diccionario={
                'nombre':graficas.nombre,
                'grafico':str(graficas.grafico)+"&filter="+str(graficas.tabla)+'/'+str(graficas.campo)+" eq '"+str(usuario_datos.empresa.nombre)+ "'&filterPaneEnabled=false&token="+str(token),
            }
            lista_de_graficas.append(diccionario)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "admin_graficas_actuales_empresarios_powerbi.html", {'user': current_user,
                                                          'lista_graficas': lista_de_graficas,
                                                          'permiso_usuario': usuario_datos})
    else:
        pass



def admin_graficas_comerciales_powerbi(request,grafica_id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_graficas = Graficas.objects.filter(tipo_usuario='PCS',area=str(grafica_id)).order_by('nivel')
        url = 'https://login.microsoftonline.com/common/oauth2/token'

        data = {
            'grant_type': 'password',
            'resource': 'https://analysis.windows.net/powerbi/api',
            'client_id': 'a77e1c67-79c8-4dfb-bc49-27f416a217fb',
            'client_secret': 'a2a8Q~zOcyCcEAgUjc9d38WxNUPoGs6r44pZrcH_',
            'username': 'analistati@pcsocial.org',
            'password': 'csAti2017*'
        }

        response = requests.post(url, data=data)
        token = response.json()['access_token']
        lista_de_graficas=[]
        for graficas in lista_graficas:
            diccionario={
                'nombre':graficas.nombre,
                'grafico':str(graficas.grafico)+"&filter="+str(graficas.tabla)+'/'+str(graficas.campo)+" eq '"+str(usuario_datos.empresa.nombre)+ "'&filterPaneEnabled=false&token="+str(token),
            }
            lista_de_graficas.append(diccionario)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "admin_graficas_comerciales_powerbi.html", {'user': current_user,
                                                          'lista_graficas': lista_de_graficas,
                                                          'grafica_id': grafica_id,
                                                          'permiso_usuario': usuario_datos})
    else:
        pass


def admin_graficas_exportaciones_powerbi(request,grafica_id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_graficas = Graficas.objects.filter(tipo_usuario='PCS',area=str(grafica_id)).order_by('nivel')
        url = 'https://login.microsoftonline.com/common/oauth2/token'

        data = {
            'grant_type': 'password',
            'resource': 'https://analysis.windows.net/powerbi/api',
            'client_id': 'a77e1c67-79c8-4dfb-bc49-27f416a217fb',
            'client_secret': 'a2a8Q~zOcyCcEAgUjc9d38WxNUPoGs6r44pZrcH_',
            'username': 'analistati@pcsocial.org',
            'password': 'csAti2017*'
        }

        response = requests.post(url, data=data)
        token = response.json()['access_token']
        lista_de_graficas=[]
        for graficas in lista_graficas:
            diccionario={
                'nombre':graficas.nombre,
                'grafico':str(graficas.grafico)+"&filter="+str(graficas.tabla)+'/'+str(graficas.campo)+" eq '"+str(usuario_datos.empresa.nombre)+ "'&filterPaneEnabled=false&token="+str(token),
            }
            lista_de_graficas.append(diccionario)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "admin_graficas_exportaciones_powerbi.html", {'user': current_user,
                                                          'lista_graficas': lista_de_graficas,
                                                          'grafica_id': grafica_id,
                                                          'permiso_usuario': usuario_datos})
    else:
        pass



def admin_graficas_desarrollo_empresarial_powerbi(request,grafica_id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_graficas = Graficas.objects.filter(tipo_usuario='PCS',area=str(grafica_id)).order_by('nivel')
        url = 'https://login.microsoftonline.com/common/oauth2/token'

        data = {
            'grant_type': 'password',
            'resource': 'https://analysis.windows.net/powerbi/api',
            'client_id': 'a77e1c67-79c8-4dfb-bc49-27f416a217fb',
            'client_secret': 'a2a8Q~zOcyCcEAgUjc9d38WxNUPoGs6r44pZrcH_',
            'username': 'analistati@pcsocial.org',
            'password': 'csAti2017*'
        }

        response = requests.post(url, data=data)
        token = response.json()['access_token']
        lista_de_graficas=[]
        for graficas in lista_graficas:
            diccionario={
                'nombre':graficas.nombre,
                'grafico':str(graficas.grafico)+"&filter="+str(graficas.tabla)+'/'+str(graficas.campo)+" eq '"+str(usuario_datos.empresa.nombre)+ "'&filterPaneEnabled=false&token="+str(token),
            }
            lista_de_graficas.append(diccionario)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "admin_graficas_desarrollo_empresarial_powerbi.html", {'user': current_user,
                                                          'lista_graficas': lista_de_graficas,
                                                          'grafica_id': grafica_id,
                                                          'permiso_usuario': usuario_datos})
    else:
        pass


def admin_graficas_administrativo_financiero_powerbi(request,grafica_id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_graficas = Graficas.objects.filter(tipo_usuario='PCS',area=str(grafica_id)).order_by('nivel')
        url = 'https://login.microsoftonline.com/common/oauth2/token'

        data = {
            'grant_type': 'password',
            'resource': 'https://analysis.windows.net/powerbi/api',
            'client_id': 'a77e1c67-79c8-4dfb-bc49-27f416a217fb',
            'client_secret': 'a2a8Q~zOcyCcEAgUjc9d38WxNUPoGs6r44pZrcH_',
            'username': 'analistati@pcsocial.org',
            'password': 'csAti2017*'
        }

        response = requests.post(url, data=data)
        token = response.json()['access_token']
        lista_de_graficas=[]
        for graficas in lista_graficas:
            diccionario={
                'nombre':graficas.nombre,
                'grafico':str(graficas.grafico)+"&filter="+str(graficas.tabla)+'/'+str(graficas.campo)+" eq '"+str(usuario_datos.empresa.nombre)+ "'&filterPaneEnabled=false&token="+str(token),
            }
            lista_de_graficas.append(diccionario)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "admin_graficas_administrativo_financiero_powerbi.html", {'user': current_user,
                                                          'lista_graficas': lista_de_graficas,
                                                          'grafica_id': grafica_id,
                                                          'permiso_usuario': usuario_datos})
    else:
        pass


def admin_graficas_mercadeo_innovacion_powerbi(request,grafica_id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_graficas = Graficas.objects.filter(tipo_usuario='PCS',area=str(grafica_id)).order_by('nivel')
        url = 'https://login.microsoftonline.com/common/oauth2/token'

        data = {
            'grant_type': 'password',
            'resource': 'https://analysis.windows.net/powerbi/api',
            'client_id': 'a77e1c67-79c8-4dfb-bc49-27f416a217fb',
            'client_secret': 'a2a8Q~zOcyCcEAgUjc9d38WxNUPoGs6r44pZrcH_',
            'username': 'analistati@pcsocial.org',
            'password': 'csAti2017*'
        }

        response = requests.post(url, data=data)
        token = response.json()['access_token']
        lista_de_graficas=[]
        for graficas in lista_graficas:
            diccionario={
                'nombre':graficas.nombre,
                'grafico':str(graficas.grafico)+"&filter="+str(graficas.tabla)+'/'+str(graficas.campo)+" eq '"+str(usuario_datos.empresa.nombre)+ "'&filterPaneEnabled=false&token="+str(token),
            }
            lista_de_graficas.append(diccionario)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "admin_graficas_mercadeo_innovacion_powerbi.html", {'user': current_user,
                                                          'lista_graficas': lista_de_graficas,
                                                          'grafica_id': grafica_id,
                                                          'permiso_usuario': usuario_datos})
    else:
        pass



def admin_graficas_operaciones_logistica_powerbi(request,grafica_id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_graficas = Graficas.objects.filter(tipo_usuario='PCS',area=str(grafica_id)).order_by('nivel')
        url = 'https://login.microsoftonline.com/common/oauth2/token'

        data = {
            'grant_type': 'password',
            'resource': 'https://analysis.windows.net/powerbi/api',
            'client_id': 'a77e1c67-79c8-4dfb-bc49-27f416a217fb',
            'client_secret': 'a2a8Q~zOcyCcEAgUjc9d38WxNUPoGs6r44pZrcH_',
            'username': 'analistati@pcsocial.org',
            'password': 'csAti2017*'
        }

        response = requests.post(url, data=data)
        token = response.json()['access_token']
        lista_de_graficas=[]
        for graficas in lista_graficas:
            diccionario={
                'nombre':graficas.nombre,
                'grafico':str(graficas.grafico)+"&filter="+str(graficas.tabla)+'/'+str(graficas.campo)+" eq '"+str(usuario_datos.empresa.nombre)+ "'&filterPaneEnabled=false&token="+str(token),
            }
            lista_de_graficas.append(diccionario)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "admin_graficas_operaciones_logistica_powerbi.html", {'user': current_user,
                                                          'lista_graficas': lista_de_graficas,
                                                          'grafica_id': grafica_id,
                                                          'permiso_usuario': usuario_datos})
    else:
        pass





def admin_graficas_adicional_powerbi(request,grafica_id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_graficas = Graficas.objects.filter(tipo_usuario='PCS',area=str(grafica_id)).order_by('nivel')
        url = 'https://login.microsoftonline.com/common/oauth2/token'

        data = {
            'grant_type': 'password',
            'resource': 'https://analysis.windows.net/powerbi/api',
            'client_id': 'a77e1c67-79c8-4dfb-bc49-27f416a217fb',
            'client_secret': 'a2a8Q~zOcyCcEAgUjc9d38WxNUPoGs6r44pZrcH_',
            'username': 'analistati@pcsocial.org',
            'password': 'csAti2017*'
        }

        response = requests.post(url, data=data)
        token = response.json()['access_token']
        lista_de_graficas=[]
        for graficas in lista_graficas:
            diccionario={
                'nombre':graficas.nombre,
                'grafico':str(graficas.grafico)+"&filter="+str(graficas.tabla)+'/'+str(graficas.campo)+" eq '"+str(usuario_datos.empresa.nombre)+ "'&filterPaneEnabled=false&token="+str(token),
            }
            lista_de_graficas.append(diccionario)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "admin_graficas_adicional_powerbi.html", {'user': current_user,
                                                          'lista_graficas': lista_de_graficas,
                                                          'grafica_id': grafica_id,
                                                          'permiso_usuario': usuario_datos})
    else:
        pass

def admin_graficas_formacion_powerbi(request):
    # Render  administracion.html
    if request.method == 'GET':
        grafica_id='analista_senior_2'
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_graficas = Graficas.objects.filter(tipo_usuario='PCS',area=str(grafica_id)).order_by('nivel')
        url = 'https://login.microsoftonline.com/common/oauth2/token'

        data = {
            'grant_type': 'password',
            'resource': 'https://analysis.windows.net/powerbi/api',
            'client_id': 'a77e1c67-79c8-4dfb-bc49-27f416a217fb',
            'client_secret': 'a2a8Q~zOcyCcEAgUjc9d38WxNUPoGs6r44pZrcH_',
            'username': 'analistati@pcsocial.org',
            'password': 'csAti2017*'
        }

        response = requests.post(url, data=data)
        token = response.json()['access_token']
        lista_de_graficas=[]
        for graficas in lista_graficas:
            diccionario={
                'nombre':graficas.nombre,
                'grafico':str(graficas.grafico)+"&filter="+str(graficas.tabla)+'/'+str(graficas.campo)+" eq '"+str(usuario_datos.empresa.nombre)+ "'&filterPaneEnabled=false&token="+str(token),
            }
            lista_de_graficas.append(diccionario)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "admin_graficas_formacion_powerbi.html", {'user': current_user,
                                                          'lista_graficas': lista_de_graficas,
                                                          'grafica_id': grafica_id,
                                                          'permiso_usuario': usuario_datos})
    else:
        pass

def admin_graficas_formacion_empresario_powerbi(request):
    # Render  administracion.html
    if request.method == 'GET':
        grafica_id='analista_senior_3'
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_graficas = Graficas.objects.filter(tipo_usuario='PCS',area=str(grafica_id)).order_by('nivel')
        url = 'https://login.microsoftonline.com/common/oauth2/token'

        data = {
            'grant_type': 'password',
            'resource': 'https://analysis.windows.net/powerbi/api',
            'client_id': 'a77e1c67-79c8-4dfb-bc49-27f416a217fb',
            'client_secret': 'a2a8Q~zOcyCcEAgUjc9d38WxNUPoGs6r44pZrcH_',
            'username': 'analistati@pcsocial.org',
            'password': 'csAti2017*'
        }

        response = requests.post(url, data=data)
        token = response.json()['access_token']
        lista_de_graficas=[]
        for graficas in lista_graficas:
            diccionario={
                'nombre':graficas.nombre,
                'grafico':str(graficas.grafico)+"&filter="+str(graficas.tabla)+'/'+str(graficas.campo)+" eq '"+str(usuario_datos.empresa.nombre)+ "'&filterPaneEnabled=false&token="+str(token),
            }
            lista_de_graficas.append(diccionario)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "admin_graficas_formacion_powerbi.html", {'user': current_user,
                                                          'lista_graficas': lista_de_graficas,
                                                          'grafica_id': grafica_id,
                                                          'permiso_usuario': usuario_datos})
    else:
        pass



def admin_graficas_formacion_cliente_powerbi(request):
    # Render  administracion.html
    if request.method == 'GET':
        grafica_id='analista_senior_4'
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_graficas = Graficas.objects.filter(tipo_usuario='PCS',area=str(grafica_id)).order_by('nivel')
        url = 'https://login.microsoftonline.com/common/oauth2/token'

        data = {
            'grant_type': 'password',
            'resource': 'https://analysis.windows.net/powerbi/api',
            'client_id': 'a77e1c67-79c8-4dfb-bc49-27f416a217fb',
            'client_secret': 'a2a8Q~zOcyCcEAgUjc9d38WxNUPoGs6r44pZrcH_',
            'username': 'analistati@pcsocial.org',
            'password': 'csAti2017*'
        }

        response = requests.post(url, data=data)
        token = response.json()['access_token']
        lista_de_graficas=[]
        for graficas in lista_graficas:
            diccionario={
                'nombre':graficas.nombre,
                'grafico':str(graficas.grafico)+"&filter="+str(graficas.tabla)+'/'+str(graficas.campo)+" eq '"+str(usuario_datos.empresa.nombre)+ "'&filterPaneEnabled=false&token="+str(token),
            }
            lista_de_graficas.append(diccionario)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "admin_graficas_formacion_powerbi.html", {'user': current_user,
                                                          'lista_graficas': lista_de_graficas,
                                                          'grafica_id': grafica_id,
                                                          'permiso_usuario': usuario_datos})
    else:
        pass




def config_respuesta_pedido(request):
    # Render  administracion.html
    if request.method == 'POST':
        hoy = date.today()
        hora=datetime.now().time()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        numero_pedido = request.POST['numero_pedido']
        entry_pedido = request.POST['entry_pedido']
        peticion = request.POST['peticion']
        fecha_problema = request.POST['fecha_problema']
        fecha_compro_prob = datetime.strptime(fecha_problema, '%Y-%m-%d')
        justificacion = request.POST['justificacion']
        fecha_vencimiento = str(request.POST['fecha_vencimiento'])
        fecha_ped_ven = datetime.strptime(fecha_vencimiento, '%Y-%m-%d')
        justificacion_text=Justificacion.objects.filter(id=int(justificacion)).first()
        y = 0
        while (y != 3):
            y += 1
            fecha_ped_ven = fecha_ped_ven + timedelta(days=1)
            x = 1
            while (x >= 1):
                if fecha_ped_ven.weekday() == 6:
                    fecha_ped_ven = fecha_ped_ven + timedelta(days=1)
                elif holidays_co.is_holiday_date(fecha_ped_ven):
                    fecha_ped_ven = fecha_ped_ven + timedelta(days=1)
                else:
                    break
        if fecha_compro_prob > fecha_ped_ven:
            messages.add_message(request, messages.ERROR,
                                 'la fecha de prorroga no debe superar los 3 dias habiles ')

            return HttpResponseRedirect('/configuracion/solicitud_pedido_orden/detalle/'+ entry_pedido +'/')
        adicionales='Solicitud de prorroga para el  '+str(fecha_problema) +' con la justificacion '+str(justificacion_text.descripcion)
        dato_area = Peticiones.objects.filter(descripcion=peticion).first()
        respuesta = RespuestaPedido(
            peticion_id=dato_area.id,
            num_pedido=numero_pedido,
            entry_pedido=entry_pedido,
            email=usuario_datos.usuario.email,
            fecha=fecha_problema,
            estado='pendiente',
            justificacion_id=int(justificacion),
            empresa=str(usuario_datos.empresa.nombre)
        )
        respuesta.save()
        emails=PersonasAtencion.objects.filter(area_id=dato_area.area.id)
        for correos in emails:
            try:
                email = EmailMessage('SOLICITUD DE PRORROGA PARA EL PEDIDO'+str(numero_pedido) ,
                                     'la empresa '+ usuario_datos.empresa.nombre + ' solicita la prórroga '+
                                    ' para el pedido '+str(numero_pedido) +' Información : '+ adicionales + ' Para ver más ingrese en '+
                                     IP_SERVIDOR +'/configuracion/solicitud_pedido_orden/problema/' +
                                    entry_pedido + '/',
                                     to=[correos.email])
                email.send()
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(usuario_datos.empresa.nombre),
                    accion='Envio de respuesta al pedido'
                )
                log.save()
            except:
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(usuario_datos.empresa.nombre),
                    accion='fallo al enviar la respuesta al pedido'
                )
                log.save()
        return HttpResponseRedirect('/configuracion/solicitud_pedido_orden/detalle/'+ entry_pedido +'/')
    else:
        pass


def config_respuesta_seg_pedido(request):
    # Render  administracion.html
    if request.method == 'POST':
        hoy = date.today()
        hora = datetime.now().time()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        numero_pedido = request.POST['numero_pedido']
        entry_pedido = request.POST['entry_pedido']
        peticion = request.POST['peticion']
        justificacion = request.POST['justificacion']
        cliente = request.POST['cliente']
        empresa = request.POST['empresa']
        adicionales='Se solicito la factura para el pedido '+str(numero_pedido)
        dato_area = Peticiones.objects.filter(descripcion=peticion).first()
        justificacion_id=1
        texto_adicional=''
        if RespuestaPedido.objects.filter(peticion_id=dato_area.id,num_pedido=numero_pedido).exists():
            justificacion_id=6
            texto_adicional='(ANTERIORMENTE SOLICITADO)'
        respuesta = RespuestaPedido(
            peticion_id=dato_area.id,
            num_pedido=numero_pedido,
            entry_pedido=entry_pedido,
            fecha=hoy,
            estado='pendiente',
            justificacion_id=justificacion_id,
            email=usuario_datos.usuario.email,
            empresa=str(empresa)
        )
        respuesta.save()
        emails=PersonasAtencion.objects.filter(area_id=dato_area.area.id)
        for correos in emails:
            try:
                email = EmailMessage('SOLICITUD DE FACTURA PARA EL PEDIDO '+str(numero_pedido)+texto_adicional ,
                                     'La empresa '+ str(empresa) + ' solicito la factura '
                                    + ' para el pedido '+str(numero_pedido) +'\nInformación  : '+ adicionales + '\nCliente: '+str(cliente) +'\nPara ver más ingrese en ' +
                                     IP_SERVIDOR +'/configuracion/solicitud_pedido_orden/problema/' +
                                    entry_pedido + '/',
                                     to=[correos.email])
                email.send()
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(empresa),
                    accion='Envio de respuesta al pedido'
                )
                log.save()
            except:
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(empresa),
                    accion='fallo al enviar la respuesta al pedido'
                )
                log.save()
        return HttpResponseRedirect('/configuracion/solicitud_pedido_orden/detalle/'+ entry_pedido +'/')
    else:
        pass


def config_respuesta_ter_pedido(request):
    # Render  administracion.html
    if request.method == 'POST':
        hoy = date.today()
        hora = datetime.now().time()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        numero_pedido = request.POST['numero_pedido']
        entry_pedido = request.POST['entry_pedido']
        peticion = request.POST['peticion']
        justificacion = request.POST['justificacion']
        adicionales='No se despachara  el pedido '+str(numero_pedido)+' por '+str(justificacion)
        dato_area = Peticiones.objects.filter(descripcion=peticion).first()
        respuesta = RespuestaPedido(
            peticion_id=dato_area.id,
            num_pedido=numero_pedido,
            entry_pedido=entry_pedido,
            fecha=hoy,
            estado='pendiente',
            justificacion_id=1,
            justificacion_adicional=justificacion,
            email=usuario_datos.usuario.email,
            empresa=str(usuario_datos.empresa.nombre)
        )
        respuesta.save()
        emails=PersonasAtencion.objects.filter(area_id=dato_area.area.id)
        for correos in emails:
            try:
                email = EmailMessage('SOLICITUD DE NO DESPACHO PARA EL PEDIDO'+str(numero_pedido) ,
                                     'La empresa '+ usuario_datos.empresa.nombre + ' solicita no despachar  '
                                    + '  el pedido '+str(numero_pedido) +' Información : '+ adicionales + ' Para ver más ingrese en ' +
                                     IP_SERVIDOR + '/configuracion/solicitud_pedido_orden/problema/' +
                                    entry_pedido + '/',
                                     to=[correos.email])
                email.send()
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(usuario_datos.empresa.nombre),
                    accion='Envio de respuesta al pedido'
                )
                log.save()
            except:
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(usuario_datos.empresa.nombre),
                    accion='fallo al enviar la respuesta al pedido'
                )
                log.save()
        return HttpResponseRedirect('/configuracion/solicitud_pedido_orden/detalle/'+ entry_pedido +'/')
    else:
        pass



def config_respuesta_quin_pedido(request):
    # Render  administracion.html
    if request.method == 'POST':
        hoy = date.today()
        hora = datetime.now().time()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        numero_pedido = request.POST['numero_pedido']
        entry_pedido = request.POST['entry_pedido']
        peticion = request.POST['peticion']
        justificacion = request.POST['justificacion']
        adicionales='Se suspendera el producto del pedido'+str(numero_pedido)+' por '+str(justificacion)
        dato_area = Peticiones.objects.filter(descripcion=peticion).first()
        respuesta = RespuestaPedido(
            peticion_id=dato_area.id,
            num_pedido=numero_pedido,
            entry_pedido=entry_pedido,
            fecha=hoy,
            estado='pendiente',
            justificacion_adicional=justificacion,
            justificacion_id=1,
            email=usuario_datos.usuario.email,
            empresa=str(usuario_datos.empresa.nombre)
        )
        respuesta.save()
        emails=PersonasAtencion.objects.filter(area_id=dato_area.area.id)
        for correos in emails:
            try:
                email = EmailMessage('SOLICITUD DE SUSPENCION DE PRODUCTO PARA EL PEDIDO'+str(numero_pedido) ,
                                     'La empresa '+ usuario_datos.empresa.nombre + ' solicita la suspencion del producto para  '
                                    + '  el pedido '+str(numero_pedido) +' Informacion : '+ adicionales + ' Para ver mas ingrese en ' +
                                     + IP_SERVIDOR +'/configuracion/solicitud_pedido_orden/problema/' +
                                    entry_pedido + '/',
                                     to=[correos.email])
                email.send()
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(usuario_datos.empresa.nombre),
                    accion='Envio de respuesta al pedido'
                )
                log.save()
            except:
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(usuario_datos.empresa.nombre),
                    accion='fallo al enviar la respuesta al pedido'
                )
                log.save()
        return HttpResponseRedirect('/configuracion/solicitud_pedido_orden/detalle/'+ entry_pedido +'/')
    else:
        pass


def config_respuesta_sex_pedido(request):
    # Render  administracion.html
    if request.method == 'POST':
        hoy = date.today()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        numero_pedido = request.POST['numero_pedido']
        entry_pedido = request.POST['entry_pedido']
        try:
            peticion = request.POST['peticion']
        except:
            peticion='Solicitud de cita entrega'
        try:
            codigo_pedido= request.POST['codigo_pedido']
        except:
            codigo_pedido=numero_pedido
        unidades = request.POST['unidades']
        cajas = request.POST['cajas']
        vehiculo = request.POST['vehiculo']
        fecha_problema = request.POST['fecha_problema']
        hora = request.POST['hora']
        justificacion = request.POST['justificacion']
        adicionales='Se solicito una cita de entrega para el pedido '+str(numero_pedido)
        dato_area = Peticiones.objects.filter(descripcion=peticion).first()
        respuesta = RespuestaPedido(
            peticion_id=dato_area.id,
            num_pedido=numero_pedido,
            entry_pedido=entry_pedido,
            fecha=fecha_problema,
            hora=hora,
            estado='pendiente',
            justificacion_id=1,
            email=usuario_datos.usuario.email,
            empresa=str(usuario_datos.empresa.nombre)
        )
        respuesta.save()
        cita = RespuestaCita(
            unidades=unidades,
            orden_compra=codigo_pedido,
            cajas=cajas,
            vehiculo=vehiculo,
            peticion_id=respuesta.pk,
            fecha=fecha_problema,
            hora=hora,
        )
        cita.save()
        emails=PersonasAtencion.objects.filter(area_id=dato_area.area.id)
        for correos in emails:
            try:
                subject, from_email, to = 'SOLICITUD DE CITA DE ENTREGA PARA EL PEDIDO ' + str(
                    numero_pedido), 'conectaportalweb@gmail.com', correos.email
                text_content = 'La empresa ' + usuario_datos.empresa.nombre + ' dio como respuesta ' + str(
                    dato_area.descripcion) + ' a la orden ' + str(numero_pedido)

                html_content = '<br><p>Buenos dias</p><br> <p>La empresa ' + usuario_datos.empresa.nombre + ' solicita agendar una cita para la orden ' + str(numero_pedido) + '</p>' \
                        '<br><strong>DATOS PARA SOLICITAR CITA: </strong><br>' \
                        '<table > ' \
                        '<thead>' \
                        '<tr>' \
                        '<th>Orden de Compra</th><th>Numero de unidades</th><th>Numero de cajas</th><th>Tipo de vehículo</th><th>Fecha a entregar</th><th>Hora a entregar</th>' \
                        '</tr></thead>' \
                        '<tbody>' \
                        '<tr><td>'+ codigo_pedido +'</td><td>'+ unidades +'</td><td>'+ cajas +'</td><td>'+ vehiculo +'</td><td>'+ fecha_problema +'</td><td>'+ hora +'</td></tr></tbody></table>' \
                        '<br><p>Para ver mas ingrese en http://160.153.178.159/configuracion/solicitud_pedido_orden/problema/'+entry_pedido + '/</p>'

                msg = EmailMultiAlternatives(subject, text_content, from_email, [to])
                msg.attach_alternative(html_content, "text/html")
                msg.send()
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(usuario_datos.empresa.nombre),
                    accion='Envio de respuesta al pedido'
                )
                log.save()
            except:
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(usuario_datos.empresa.nombre),
                    accion = 'fallo al enviar la respuesta al pedido'
                )
                log.save()
        return HttpResponseRedirect('/configuracion/solicitud_pedido_orden/detalle/'+ entry_pedido +'/')
    else:
        pass


def config_respuesta_sept_pedido(request):
    # Render  administracion.html
    if request.method == 'POST':
        hoy = date.today()
        hora = datetime.now().time()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        numero_pedido = request.POST['numero_pedido']
        entry_pedido = request.POST['entry_pedido']
        peticion = request.POST['peticion']
        justificacion = request.POST['justificacion']
        adicionales='Se solicito el despacho para el pedido '+str(numero_pedido)
        dato_area = Peticiones.objects.filter(descripcion=peticion).first()
        justificacion_id = 1
        texto_adicional = ''
        if RespuestaPedido.objects.filter(peticion_id=dato_area.id, num_pedido=numero_pedido).exists():
            justificacion_id = 6
            texto_adicional = '(ANTERIORMENTE SOLICITADO)'
        respuesta = RespuestaPedido(
            peticion_id=dato_area.id,
            num_pedido=numero_pedido,
            entry_pedido=entry_pedido,
            fecha=hoy,
            estado='pendiente',
            justificacion_id=1,
            email=usuario_datos.usuario.email,
            empresa=str(usuario_datos.empresa.nombre)
        )
        respuesta.save()
        emails=PersonasAtencion.objects.filter(area_id=dato_area.area.id)
        for correos in emails:
            try:
                email = EmailMessage('SOLICITUD DE DESPACHO PARA EL PEDIDO '+str(numero_pedido)+texto_adicional ,
                                     'La empresa '+ usuario_datos.empresa.nombre + ' solicito el despacho '
                                    + ' para el pedido '+str(numero_pedido) +' Información  : '+ adicionales + ' Para ver más ingrese en ' +
                                     IP_SERVIDOR + '/configuracion/solicitud_pedido_orden/problema/' +
                                    entry_pedido + '/',
                                     to=[correos.email])
                email.send()
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(usuario_datos.empresa.nombre),
                    accion='Envio de respuesta al pedido'
                )
                log.save()
            except:
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(usuario_datos.empresa.nombre),
                    accion='fallo al enviar la respuesta al pedido'
                )
                log.save()
                
        return HttpResponseRedirect('/configuracion/solicitud_pedido_orden/detalle/'+ entry_pedido +'/')
    else:
        pass


def config_respuesta_oct_pedido(request):
    # Render  administracion.html
    if request.method == 'POST':
        hoy = date.today()
        hora = datetime.now().time()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        cantidad = request.POST.getlist('cantidad[]')
        comprobante_cantidad = request.POST.getlist('comprobante_cantidad[]')
        nombre_prod = request.POST.getlist('nombre_prod[]')
        numero_prod = request.POST.getlist('numero_prod[]')
        lista_tamano = len(cantidad)
        numero_pedido = request.POST['numero_pedido']
        entry_pedido = request.POST['entry_pedido']
        peticion = request.POST['peticion']
        justificacion = request.POST['justificacion']
        adicionales = 'Despachar el pedido ' + str(numero_pedido) + ' con las siguientes novedades '
        dato_area = Peticiones.objects.filter(descripcion=peticion).first()
        justificacion_id = 1
        texto_adicional = ''
        if RespuestaPedido.objects.filter(peticion_id=dato_area.id, num_pedido=numero_pedido).exists():
            justificacion_id = 6
            texto_adicional = '(ANTERIORMENTE SOLICITADO)'
        respuesta = RespuestaPedido(
            peticion_id=dato_area.id,
            num_pedido=numero_pedido,
            entry_pedido=entry_pedido,
            fecha=hoy,
            estado='pendiente',
            justificacion_id=1,
            email=usuario_datos.usuario.email,
            empresa=str(usuario_datos.empresa.nombre)
        )
        respuesta.save()
        for i in range(lista_tamano):
            modificado = 'NO'
            if cantidad[i] != comprobante_cantidad[i]:
                modificado = 'SI'
            pedidonovedad = PedidosNovedades(
                numero=numero_prod[i],
                nombre=nombre_prod[i],
                cantidad=cantidad[i],
                modificado=modificado,
                peticion_id=respuesta.pk,
            )
            pedidonovedad.save()
        emails = PersonasAtencion.objects.filter(area_id=dato_area.area.id)
        for correos in emails:
            try:
                subject, from_email, to = 'SOLICITUD DE DESPACHO CON NOVEDADES PARA EL PEDIDO' + str(
                    numero_pedido)+texto_adicional, 'conectaportalweb@gmail.com', correos.email
                text_content = 'La empresa ' + usuario_datos.empresa.nombre + ' dio como respuesta ' + str(
                    dato_area.descripcion) + ' a la orden ' + str(numero_pedido)

                html_content = '<br> <p>La empresa ' + usuario_datos.empresa.nombre + ' solicita un despacho con las siguientes novedades para la orden ' + str(
                    numero_pedido) + '</p>' \
                                     '<br><strong>Despachar el pedido con las siguientes novedades: </strong><br>' \
                                     '<table > ' \
                                     '<thead>' \
                                     '<tr>' \
                                     '<th>Codigo</th><th>Nombre de articulo</th><th>Cantidad</th><th>Modificado</th>' \
                                     '</tr></thead>' \
                                     '<tbody>'
                content_x = ''
                for i in range(lista_tamano):
                    html_content2 = '<tr>' \
                                    '<td>' + numero_prod[i] + '</td>' \
                                                              '<td>' + nombre_prod[i] + '</td>' \
                                                                                        '<td>' + cantidad[i] + '</td>'
                    if cantidad[i] != comprobante_cantidad[i]:
                        html_content2 = html_content2 + '<td> SI </td>'
                    content_x = content_x + html_content2
                html_content3 = '</tr></tbody></table>' \
                                '<br><p>Para ver mas ingrese en http://160.153.178.159/configuracion/solicitud_pedido_orden/problema/' + entry_pedido + '/</p>'
                html_content = html_content + content_x + html_content3
                msg = EmailMultiAlternatives(subject, text_content, from_email, [to])
                msg.attach_alternative(html_content, "text/html")
                msg.send()
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(usuario_datos.empresa.nombre),
                    accion='Envio de respuesta al pedido'
                )
                log.save()
            except:
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(usuario_datos.empresa.nombre),
                    accion='fallo al enviar la respuesta al pedido'
                )
                log.save()

        return HttpResponseRedirect('/configuracion/solicitud_pedido_orden/detalle/' + entry_pedido + '/')
    else:
        pass


def config_respuesta_cuar_pedido(request):
    # Render  administracion.html
    if request.method == 'POST':
        hoy = date.today()
        hora = datetime.now().time()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        cantidad = request.POST.getlist('cantidad[]')
        comprobante_cantidad = request.POST.getlist('comprobante_cantidad[]')
        nombre_prod = request.POST.getlist('nombre_prod[]')
        numero_prod = request.POST.getlist('numero_prod[]')
        lista_tamano=len(cantidad)
        numero_pedido = request.POST['numero_pedido']
        entry_pedido = request.POST['entry_pedido']
        cliente = request.POST['cliente']
        peticion = request.POST['peticion']
        justificacion = request.POST['justificacion']
        empresa = request.POST['empresa']
        adicionales='Facturar el pedido '+str(numero_pedido)+' con las siguientes novedades '
        dato_area = Peticiones.objects.filter(descripcion=peticion).first()
        justificacion_id = 1
        texto_adicional = ''
        if RespuestaPedido.objects.filter(peticion_id=dato_area.id, num_pedido=numero_pedido).exists():
            justificacion_id = 6
            texto_adicional = '(ANTERIORMENTE SOLICITADO)'
        respuesta = RespuestaPedido(
            peticion_id=dato_area.id,
            num_pedido=numero_pedido,
            entry_pedido=entry_pedido,
            fecha=hoy,
            estado='pendiente',
            justificacion_id=1,
            email=usuario_datos.usuario.email,
            empresa=str(empresa)
        )
        respuesta.save()
        for i in range(lista_tamano):
            modificado = 'NO'
            if cantidad[i] != comprobante_cantidad[i]:
                modificado = 'SI'
            pedidonovedad = PedidosNovedades(
                numero=numero_prod[i],
                nombre=nombre_prod[i],
                cantidad=cantidad[i],
                modificado=modificado,
                peticion_id=respuesta.pk,
            )
            pedidonovedad.save()
        emails=PersonasAtencion.objects.filter(area_id=dato_area.area.id)
        for correos in emails:
            try:
                subject, from_email, to = 'SOLICITUD DE FACTURA CON NOVEDADES PARA EL PEDIDO '+str(numero_pedido)+texto_adicional, 'conectaportalweb@gmail.com', correos.email
                text_content = 'La empresa '+ empresa + ' dio como respuesta '+str(dato_area.descripcion) +' a la orden '+str(numero_pedido)

                html_content = '<br> <p>La empresa '+ empresa + ' solicita una Factura con las siguientes novedades para la orden '+str(numero_pedido)+'</p>' \
                               '<br><strong>Facturar el pedido con las siguientes novedades: </strong><br>' \
                                               '<table > ' \
                                               '<thead>' \
                                               '<tr>' \
                                               '<th>Codigo</th><th>Nombre de articulo</th><th>Cantidad</th><th>Modificado</th>' \
                                               '</tr></thead>' \
                                               '<tbody>' \

                content_x=''
                for i in range(lista_tamano):
                    html_content2='<tr>' \
                                  '<td>'+ numero_prod[i] +'</td>' \
                                    '<td>'+ nombre_prod[i] +'</td>' \
                                    '<td>'+ cantidad[i] +'</td>'
                    if cantidad[i] != comprobante_cantidad[i]:
                        html_content2=html_content2+'<td> SI </td>'
                    content_x=content_x+html_content2
                html_content3='</tr></tbody></table>' \
                              '<br><p>Cliente: '+cliente+'</p><br><p>Para ver mas ingrese en http://160.153.178.159/configuracion/solicitud_pedido_orden/problema/'+entry_pedido + '/</p>'
                html_content=html_content+content_x+html_content3
                msg = EmailMultiAlternatives(subject, text_content, from_email, [to])
                msg.attach_alternative(html_content, "text/html")
                msg.send()
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(empresa),
                    accion='Envio de respuesta al pedido'
                )
                log.save()
            except:
                log = LogRespuestaPedido(
                    peticion_id=dato_area.id,
                    num_pedido=numero_pedido,
                    fecha=hoy,
                    hora=hora,
                    email=correos.email,
                    empresa=str(empresa),
                    accion='fallo al enviar la respuesta al pedido'
                )
                log.save()

        return HttpResponseRedirect('/configuracion/solicitud_pedido_orden/detalle/'+ entry_pedido +'/')
    else:
        pass

def config_respuesta_peticion(request):
    # Render  administracion.html
    if request.method == 'POST':
        hoy = date.today()
        hora = datetime.now().time()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        problema = int(request.POST['problema'])
        entry_pedido = request.POST['entry_pedido']
        respuesta = request.POST['respuesta']
        direccion = request.POST['direccion']
        doc = request.POST['doc']
        fecha_prorroga = request.POST.get('fecha_prorroga', '')
        datospersona=RespuestaPedido.objects.filter(id=problema).first()
        if fecha_prorroga!="":
            respuesta=str(respuesta)+" con fecha del "+str(fecha_prorroga)
        try:
            cita = request.POST['cita']
            tipo_cita = request.POST['tipo_cita']
            numero_cita=request.POST['numero_cita']
            fecha_cita=request.POST['fecha_cita']
            hora_cita=request.POST['hora_cita']
            respuesta = 'Buenos Dias. Su cita quedo agendada para el dia '+ fecha_cita +' a las '+ hora_cita +' Con el No. de Cita'+ numero_cita

        except:
            pass
        if doc=='1':
            try:
                adjunto = request.POST['adjunto']
            except:
                adjunto = request.FILES['adjunto']

            if adjunto != '':
                myfile = request.FILES['adjunto']
                fs = FileSystemStorage()
                filename = fs.save(myfile.name, myfile)
                uploaded_file_url = fs.url(filename)
            else:
                uploaded_file_url = None

            RespuestaPedido.objects.filter(id=problema).update(respuesta=respuesta,estado='respondido',doc_respuesta=uploaded_file_url)
            email = EmailMessage('RESPUESTA PETICION ' + str(datospersona.num_pedido),
                                 str(respuesta)+', Documento adjunto: '+ IP_SERVIDOR +str(uploaded_file_url),
                                 to=[datospersona.email])
            email.send()
            log = LogRespuestaPedido(
                peticion_id=datospersona.peticion.id,
                num_pedido=datospersona.num_pedido,
                fecha=hoy,
                hora=hora,
                email=datospersona.email,
                empresa=str(datospersona.empresa),
                accion='Respuesta a la solicitud'
            )
            log.save()

        else:
            RespuestaPedido.objects.filter(id=problema).update(respuesta=respuesta, estado='respondido')
            email = EmailMessage('RESPUESTA PETICION ' + str(datospersona.num_pedido),
                                 str(respuesta),
                                 to=[datospersona.email])
            email.send()
            log = LogRespuestaPedido(
                peticion_id=datospersona.peticion.id,
                num_pedido=datospersona.num_pedido,
                fecha=hoy,
                hora=hora,
                email=datospersona.email,
                empresa=str(datospersona.empresa),
                accion='Respuesta a la solicitud'
            )
            log.save()
        if direccion == '1':
            return HttpResponseRedirect('/configuracion/solicitud_pedido_orden/problema/'+ entry_pedido +'/')
        else:
            return HttpResponseRedirect('/configuracion/solicitudes/')
    else:
        pass



# Tipos de Radicados
def config_tipos_radicados(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_tipos_radicados = Tipos_radicados.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_tipo_radicado == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_tipos_radicados.html", {'user': current_user,
                                                          'lista_tipos_radicados': lista_tipos_radicados,
                                                          'permisos': permisos,
                                                        'permiso_usuario': usuario_datos,})
    else:
        pass

def config_tipos_radicados_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_tipo_radicado == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_radicados/')

        return render(request, "config_tipos_radicados_registrar.html", {'user': current_user,
                                                                   'permisos': permisos,
                                                                    'permiso_usuario': usuario_datos, })
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.registrar_tipo_radicado == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_radicados/')

        id = request.POST['id']
        nombre_tr = request.POST['nombre_tr']
        genera_radicado_salida = request.POST['genera_radicado_salida']

        tipos_radicados = Tipos_radicados(
            id=None if not id else id,
            nombre_tr=nombre_tr,
            genera_radicado_salida=genera_radicado_salida,
        )
        tipos_radicados.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el Tipo de Radicado ' + (nombre_tr).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/tipos_radicados/')

def config_tipos_radicados_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        tipos_radicados = Tipos_radicados.objects.get(pk=id)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_tipo_radicado == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_radicados/')

        return render(request, "config_tipos_radicados_editar.html", {'user': current_user,
                                                                         'permisos': permisos,
                                                                         'tipos_radicados': tipos_radicados,
                                                                      'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.editar_tipo_radicado == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_radicados/')

        id = request.POST['id']
        nombre_tr = request.POST['nombre_tr']
        genera_radicado_salida = request.POST['genera_radicado_salida']

        tipos_radicados = Tipos_radicados(
            id=None if not id else id,
            nombre_tr=nombre_tr,
            genera_radicado_salida=genera_radicado_salida,
        )
        tipos_radicados.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado el Tipo de Radicado ' + (nombre_tr).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/tipos_radicados/')

def config_tipos_radicados_borrar(request, id):

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        tipos_radicados = Tipos_radicados.objects.get(pk=id)
        usuario_actual = request.user

        if usuario_datos.perfiles.borrar_tipo_radicado == False:
            messages.add_message(request, messages.ERROR,
                                 'No tienes permisos para eliminar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_radicados/')

        tipos_radicados.delete()
        messages.add_message(request, messages.WARNING,
                             'Se ha borrado el tipo de radicado ' + str(id) + ' satisfactoriamente')

        return HttpResponseRedirect('/configuracion/tipos_radicados/')


# Tipos de Envios
def config_tipos_envios(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_tipos_envios = Tipos_envios.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_tipo_envio == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_tipos_envios.html", {'user': current_user,
                                                          'lista_tipos_envios': lista_tipos_envios,
                                                          'permisos': permisos,
                                                        'permiso_usuario': usuario_datos, })
    else:
        pass

def config_tipos_envios_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_tipo_envio == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_envios/')

        return render(request, "config_tipos_envios_registrar.html", {'user': current_user,
                                                                   'permisos': permisos,
                                                                    'permiso_usuario': usuario_datos, })
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.registrar_tipo_envio == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_envios/')

        id = request.POST['id']
        descripcion = request.POST['descripcion']


        tipos_envios = Tipos_envios(
            id=None if not id else id,
            descripcion=descripcion,

        )
        tipos_envios.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el Tipo de Envio ' + (descripcion).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/tipos_envios/')

def config_tipos_envios_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        tipos_envios = Tipos_envios.objects.get(pk=id)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_tipo_envio == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_envios/')

        return render(request, "config_tipos_envios_editar.html", {'user': current_user,
                                                                         'permisos': permisos,
                                                                         'tipos_envios': tipos_envios,
                                                                   'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.editar_tipo_envio == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_envios/')

        id = request.POST['id']
        descripcion = request.POST['descripcion']


        tipos_envios = Tipos_envios(
            id=None if not id else id,
            descripcion=descripcion,

        )
        tipos_envios.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado el Tipo de Envio ' + (descripcion).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/tipos_envios/')

def config_tipos_envios_borrar(request, id):

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        tipos_envios = Tipos_envios.objects.get(pk=id)
        usuario_actual = request.user

        if usuario_datos.perfiles.borrar_tipo_envio == False:
            messages.add_message(request, messages.ERROR,
                                 'No tienes permisos para eliminar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_envios/')

        tipos_envios.delete()
        messages.add_message(request, messages.WARNING,
                             'Se ha borrado el tipo de envio ' + str(id) + ' satisfactoriamente')

        return HttpResponseRedirect('/configuracion/tipos_envios/')


# Formas de Envios
def config_formas_envios(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_formas_envios = Formas_envios.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_formas_envio == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_formas_envios.html", {'user': current_user,
                                                          'lista_formas_envios': lista_formas_envios,
                                                          'permisos': permisos,
                                                        'permiso_usuario': usuario_datos, })
    else:
        pass

def config_formas_envios_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()

        lista_tipos_envios = Tipos_envios.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_formas_envio == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/formas_envios/')

        return render(request, "config_formas_envios_registrar.html", {'user': current_user,
                                                                   'permisos': permisos,
                                                                       'lista_tipos_envios': lista_tipos_envios,
                                                                       'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.registrar_formas_envio == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/formas_envios/')

        id = request.POST['id']
        tipos_envios = request.POST['tipos_envios']
        nombre_fe = request.POST['nombre_fe']
        estado = request.POST['estado']
        genera_planilla = request.POST['genera_planilla']

        formas_envios = Formas_envios(
            id=None if not id else id,
            tipos_envios_id=int(tipos_envios),
            nombre_fe=nombre_fe,
            estado=estado,
            genera_planilla=genera_planilla,
        )
        formas_envios.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado la Formas de Envios ' + (nombre_fe).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/formas_envios/')

def config_formas_envios_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        formas_envios = Formas_envios.objects.get(pk=id)

        lista_tipos_envios = Tipos_envios.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_formas_envio == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/formas_envios/')

        return render(request, "config_formas_envios_editar.html", {'user': current_user,
                                                                         'permisos': permisos,
                                                                         'lista_tipos_envios': lista_tipos_envios,
                                                                         'formas_envios': formas_envios,
                                                                    'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.editar_formas_envio == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/formas_envios/')

        id = request.POST['id']
        tipos_envios = request.POST['tipos_envios']
        nombre_fe = request.POST['nombre_fe']
        estado = request.POST['estado']
        genera_planilla = request.POST['genera_planilla']

        formas_envios = Formas_envios(
            id=None if not id else id,
            tipos_envios_id=int(tipos_envios),
            nombre_fe=nombre_fe,
            estado=estado,
            genera_planilla=genera_planilla,
        )
        formas_envios.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado la Forma de Envio ' + (nombre_fe).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/formas_envios/')

def config_formas_envios_borrar(request, id):

    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        formas_envio = Formas_envios.objects.get(pk=id)
        usuario_actual = request.user

        if usuario_datos.perfiles.borrar_formas_envio == False:
            messages.add_message(request, messages.ERROR,
                                 'No tienes permisos para eliminar en este modulo')
            return HttpResponseRedirect('/configuracion/formas_envios/')

        formas_envio.delete()
        messages.add_message(request, messages.WARNING,
                             'Se ha borrado la forma de envio ' + str(id) + ' satisfactoriamente')

        return HttpResponseRedirect('/configuracion/formas_envios/')

# Soportes
def config_soportes(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_soportes = Soportes.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_soportes == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_soportes.html", {'user': current_user,
                                                          'lista_soportes': lista_soportes,
                                                          'permisos': permisos,
                                                        'permiso_usuario': usuario_datos,})
    else:
        pass

def config_soportes_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_soportes == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/soportes/')

        return render(request, "config_soportes_registrar.html", {'user': current_user,
                                                                      'permisos': permisos,
                                                                  'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.registrar_soportes == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/soportes/')

        id = request.POST['id']
        descripcion = request.POST['descripcion']

        lista_soportes = Soportes.objects.all()
        lista_soportes = lista_soportes.filter(id=id)
        if lista_soportes.exists():
            messages.add_message(request, messages.ERROR,
                                 'Registro ya existe con el codigo ' + (id).encode(
                                     'utf-8').strip())

            return HttpResponseRedirect('/configuracion/soportes/')

        soportes = Soportes(
            id=id,
            descripcion=descripcion,

        )
        soportes.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el Soporte ' + (descripcion).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/soportes/')

def config_soportes_editar(request, id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        soportes = Soportes.objects.get(pk=id)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_soportes == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/soportes/')

        return render(request, "config_soportes_editar.html", {'user': current_user,
                                                                   'permisos': permisos,
                                                                   'soportes': soportes,
                                                               'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.editar_soportes == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/soportes/')

        id = request.POST['id']
        descripcion = request.POST['descripcion']

        soportes = Soportes(
            id=id,
            descripcion=descripcion,

        )
        soportes.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado el Soporte ' + (descripcion).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/soportes/')

def config_soportes_borrar(request, id):

    if request.method == 'GET':

        soporte = Soportes.objects.get(pk=id)
        usuario_actual = request.user
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.borrar_soportes == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para eliminar en este modulo')
            return HttpResponseRedirect('/configuracion/soportes/')


        soporte.delete()
        messages.add_message(request, messages.WARNING,
                             'Se ha borrado el Soporte ' + str(id) + ' satisfactoriamente')

        return HttpResponseRedirect('/configuracion/soportes/')

# Medios de Recepcion
def config_medios_recepcion(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_medios_recepcion = Medios_recepcion.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_medios == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_medios_recepcion.html", {'user': current_user,
                                                          'lista_medios_recepcion': lista_medios_recepcion,
                                                          'permisos': permisos,
                                                            'permiso_usuario': usuario_datos,})
    else:
        pass

def config_medios_recepcion_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_medios == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/medios_recepcion/')

        return render(request, "config_medios_recepcion_registrar.html", {'user': current_user,
                                                                  'permisos': permisos,
                                                                'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.registrar_medios == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/medios_recepcion/')

        id = request.POST['id']
        descripcion = request.POST['descripcion']

        medios_recepcion = Medios_recepcion(
            id=None if not id else id,
            descripcion=descripcion,

        )
        medios_recepcion.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el Medio de Recepcion ' + (descripcion).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/medios_recepcion/')

def config_medios_recepcion_editar(request, id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        medios_recepcion = Medios_recepcion.objects.get(pk=id)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_medios == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/medios_recepcion/')

        return render(request, "config_medios_recepcion_editar.html", {'user': current_user,
                                                               'permisos': permisos,
                                                               'medios_recepcion': medios_recepcion,
                                                                'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.editar_medios == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/medios_recepcion/')

        id = request.POST['id']
        descripcion = request.POST['descripcion']

        medios_recepcion = Medios_recepcion(
            id=id,
            descripcion=descripcion,

        )
        medios_recepcion.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado el Medios_recepcion ' + (descripcion).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/medios_recepcion/')

def config_medios_recepcion_borrar(request, id):

    if request.method == 'GET':
        medio_recepcion = Medios_recepcion.objects.get(pk=id)
        usuario_actual = request.user
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.borrar_medios == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para eliminar en este modulo')
            return HttpResponseRedirect('/configuracion/medios_recepcion/')

        medio_recepcion.delete()
        messages.add_message(request, messages.WARNING,
                             'Se ha borrado el Medio de Recepcion ' + str(id) + ' satisfactoriamente')

        return HttpResponseRedirect('/configuracion/medios_recepcion/')

#Tipos Devoluciones
def config_tipos_devoluciones(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_tipos_devoluciones = Tipos_devoluciones.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_motivos_dev == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_tipos_devoluciones.html", {'user': current_user,
                                                          'lista_tipos_devoluciones': lista_tipos_devoluciones,
                                                          'permisos': permisos,
                                                        'permiso_usuario': usuario_datos,})
    else:
        pass

def config_tipos_devoluciones_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_motivos_dev == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_devoluciones/')

        return render(request, "config_tipos_devoluciones_registrar.html", {'user': current_user,
                                                                   'permisos': permisos,
                                                                'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.registrar_motivos_dev == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_devoluciones/')

        id = request.POST['id']
        descripcion = request.POST['descripcion']

        tipos_devoluciones = Tipos_devoluciones(
            id=None if not id else id,
            descripcion=descripcion,

        )
        tipos_devoluciones.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el Concepto Devolucion ' + (descripcion).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/tipos_devoluciones/')

def config_tipos_devoluciones_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        tipos_devoluciones = Tipos_devoluciones.objects.get(pk=id)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_motivos_dev == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_devoluciones/')

        return render(request, "config_tipos_devoluciones_editar.html", {'user': current_user,
                                                                         'permisos': permisos,
                                                                         'tipos_devoluciones': tipos_devoluciones,
                                                                         'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.editar_motivos_dev == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_devoluciones/')

        id = request.POST['id']
        descripcion = request.POST['descripcion']

        tipos_devoluciones = Tipos_devoluciones(
            id=id,
            descripcion=descripcion,

        )
        tipos_devoluciones.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado Motivo Devolucion ' + (descripcion).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/tipos_devoluciones/')

def config_tipos_devoluciones_borrar(request, id):

        if request.method == 'GET':
            current_user = request.user
            usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
            tipo_devolucion = Tipos_devoluciones.objects.get(pk=id)
            usuario_actual = request.user

            if usuario_datos.perfiles.borrar_motivos_dev == False:
                messages.add_message(request, messages.ERROR, 'No tienes permisos para eliminar en este modulo')
                return HttpResponseRedirect('/configuracion/tipos_devoluciones/')

            tipo_devolucion.delete()
            messages.add_message(request, messages.WARNING,
                             'Se ha borrado Motivo Devolucion ' + str(id) + ' satisfactoriamente')

            return HttpResponseRedirect('/configuracion/tipos_devoluciones/')

#Tipos Anulaciones
def config_tipos_anulaciones(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_tipos_anulaciones = Tipos_anulaciones.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_anulaciones == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_tipos_anulaciones.html", {'user': current_user,
                                                          'lista_tipos_anulaciones': lista_tipos_anulaciones,
                                                          'permisos': permisos,
                                                        'permiso_usuario': usuario_datos,})
    else:
        pass

def config_tipos_anulaciones_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_anulaciones == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_anulaciones/')

        return render(request, "config_tipos_anulaciones_registrar.html", {'user': current_user,
                                                                   'permisos': permisos,
                                                                'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.registrar_anulaciones == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_anulaciones/')

        id = request.POST['id']
        descripcion = request.POST['descripcion']

        tipos_anulaciones = Tipos_anulaciones(
            id=None if not id else id,
            descripcion=descripcion,

        )
        tipos_anulaciones.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el Concepto Anulacion ' + (descripcion).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/tipos_anulaciones/')

def config_tipos_anulaciones_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        tipos_anulaciones = Tipos_anulaciones.objects.get(pk=id)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_anulaciones == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_anulaciones/')

        return render(request, "config_tipos_anulaciones_editar.html", {'user': current_user,
                                                                         'permisos': permisos,
                                                                         'tipos_anulaciones': tipos_anulaciones,
                                                                        'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.editar_anulaciones == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_anulaciones/')

        id = request.POST['id']
        descripcion = request.POST['descripcion']

        tipos_anulaciones = Tipos_anulaciones(
            id=id,
            descripcion=descripcion,

        )
        tipos_anulaciones.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado Solicitud Anulacion ' + (descripcion).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/tipos_anulaciones/')

def config_tipos_anulaciones_borrar(request, id):

        if request.method == 'GET':

            current_user = request.user
            usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

            if usuario_datos.perfiles.borrar_anulaciones == False:
                messages.add_message(request, messages.ERROR, 'No tienes permisos para eliminar en este modulo')
                return HttpResponseRedirect('/configuracion/tipos_anulaciones/')

            tipo_anulacion = Tipos_anulaciones.objects.get(pk=id)
            usuario_actual = request.user

            tipo_anulacion.delete()
            messages.add_message(request, messages.WARNING,
                             'Se ha borrado Concepto Anulacion ' + str(id) + ' satisfactoriamente')

            return HttpResponseRedirect('/configuracion/tipos_anulaciones/')



# Fechas Festivos
def config_fechas_festivos(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_fechas_festivos = Fechas_festivos.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_fechas_festivos == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_fechas_festivos.html", {'user': current_user,
                                                          'lista_fechas_festivos': lista_fechas_festivos,
                                                          'permisos': permisos,
                                                         'permiso_usuario': usuario_datos})
    else:
        pass

def config_fechas_festivos_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_fechas_festivos == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/fechas_festivos/')

        return render(request, "config_fechas_festivos_registrar.html", {'user': current_user,
                                                                   'permisos': permisos,
                                                                'permiso_usuario': usuario_datos})
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        if usuario_datos.perfiles.registrar_fechas_festivos == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/fechas_festivos/')

#        id = request.POST['id']
        fecha = request.POST['fecha']


        fechas_festivos = Fechas_festivos(
#            id=id,
            fecha=fecha,

        )
        fechas_festivos.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado la Fecha Festivo ' + (fecha).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/fechas_festivos/')


def config_fechas_festivos_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        fechas_festivos = Fechas_festivos.objects.get(pk=id)
        fecha=fechas_festivos.fecha.isoformat()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_fechas_festivos == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/fechas_festivos/')

        return render(request, "config_fechas_festivos_editar.html", {'user': current_user,
                                                                         'permisos': permisos,
                                                                         'fechas_festivos': fechas_festivos,
                                                                        'fecha':fecha,
                                                                      'permiso_usuario': usuario_datos})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.editar_fechas_festivos == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/fechas_festivos/')

        id = request.POST['id']
        fecha = request.POST['fecha']


        fechas_festivos = Fechas_festivos(
            id=id,
            fecha=fecha,

        )
        fechas_festivos.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado la Fecha Festivo ' + (fecha).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/fechas_festivos/')

def config_fechas_festivos_borrar(request,id):
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        fechas_festivos = Fechas_festivos.objects.get(pk=id)
        fecha = fechas_festivos.fecha.isoformat()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.borrar_fechas_festivos == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para eliminar en este modulo')
            return HttpResponseRedirect('/configuracion/fechas_festivos/')
        else:
            fechas_festivos.delete()
            messages.add_message(request, messages.WARNING,
                                    'Se ha borrado la fecha ' + str(id) + ' satisfactoriamente')

            return HttpResponseRedirect('/configuracion/fechas_festivos/')

# Tipos de Documentos
def config_tipos_documentos(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_tipos_documentos = Tipos_documentos.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_tipos_documentos == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_tipos_documentos.html", {'user': current_user,
                                                          'lista_tipos_documentos': lista_tipos_documentos,
                                                          'permisos': permisos,
                                                        'permiso_usuario': usuario_datos,})
    else:
        pass

def config_tipos_documentos_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()

        lista_dependencias = Dependencias.objects.all()
        lista_tipos_radicados = Tipos_radicados.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_tipos_documentos == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_documentos/')

        return render(request, "config_tipos_documentos_registrar.html", {'user': current_user,
                                                                   'lista_dependencias': lista_dependencias,
                                                                   'lista_tipos_radicados': lista_tipos_radicados,
                                                                'permisos': permisos,
                                                            'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.registrar_tipos_documentos == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_documentos/')

        dependencias = request.POST['dependencias']
        tipos_radicados = request.POST['tipos_radicados']
        descripcion = request.POST['descripcion']
        dias_tramite = request.POST['dias_tramite']
        publicar = request.POST['publicar']

        tipos_documentos = Tipos_documentos(
            dependencias_id=dependencias,
            tipos_radicados_id=int(tipos_radicados),
            descripcion=descripcion,
            dias_tramite=dias_tramite,
            publicar=publicar,
            creado=timezone.now(),

        )
        tipos_documentos.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el Tipo Documento ' + (descripcion).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/tipos_documentos/')

def config_tipos_documentos_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        tipos_documentos = Tipos_documentos.objects.get(pk=id)
        lista_tipos_radicados = Tipos_radicados.objects.all()
        lista_dependencias=Dependencias.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_tipos_documentos == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_documentos/')

        return render(request, "config_tipos_documentos_editar.html", {'user': current_user,
                                                                         'permisos': permisos,
                                                                         'lista_tipos_radicados': lista_tipos_radicados,
                                                                         'tipos_documentos': tipos_documentos,
                                                                       'permiso_usuario': usuario_datos,
                                                                       'lista_dependencias':lista_dependencias})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.editar_tipos_documentos == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_documentos/')

        id = request.POST['id']
        dependencias = request.POST['dependencias']
        tipos_radicados = request.POST['tipo_radicado']
        descripcion = request.POST['descripcion']
        dias_tramite = request.POST['dias_tramite']
        publicar = request.POST['publicar']

        tipos_documentos = Tipos_documentos(
            id=id,
            dependencias_id=dependencias,
            tipos_radicados_id=int(tipos_radicados),
            descripcion=descripcion,
            dias_tramite=dias_tramite,
            publicar=publicar,
            modificado=timezone.now(),

        )
        tipos_documentos.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado el Tipo de Documento ' + (descripcion).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/tipos_documentos/')

def config_tipos_documentos_borrar(request, id):
    if request.method == 'GET':

        tipo_documento = Tipos_documentos.objects.get(pk=id)
        usuario_actual = request.user
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.editar_tipos_documentos == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para eliminar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_documentos/')

        tipo_documento.delete()
        messages.add_message(request, messages.WARNING,
                             'Se ha borrado el Tipo Documento ' + str(id) + ' satisfactoriamente')

        return HttpResponseRedirect('/configuracion/tipos_documentos/')

# Terceros
def config_terceros(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_terceros = Terceros.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_terceros == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_terceros.html", {'user': current_user,
                                                          'lista_terceros': lista_terceros,
                                                          'permisos': permisos,
                                                        'permiso_usuario': usuario_datos,})
    else:
        pass

def config_terceros_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()

        lista_continentes = Continentes.objects.all()
        lista_paises = Paises.objects.all()
        lista_departamentos = Departamentos.objects.all()
        lista_municipios = Municipios.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_terceros == False:
            messages.add_message(request, messages.ERROR,
                                    'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/terceros/')

        return render(request, "config_terceros_registrar.html", {'user': current_user,
                                                                  'lista_continentes': lista_continentes,
                                                                  'lista_paises': lista_paises,
                                                                  'lista_departamentos': lista_departamentos,
                                                                  'lista_municipios': lista_municipios,
                                                                    'permisos': permisos,
                                                                  'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.registrar_terceros == False:
            messages.add_message(request, messages.ERROR,
                                    'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/terceros/')

        nombre = request.POST['nombre']
        sigla = request.POST['sigla']
        direccion = request.POST['direccion']
        telefono = request.POST['telefono']
        telefono2 = request.POST['telefono2']
        celular = request.POST['celular']
        nit = request.POST['nit']
        dv = request.POST['dv']
        representante_legal = request.POST['representante_legal']
        continentes = request.POST['continentes']
        paises = request.POST['paises']
        departamentos = request.POST['departamentos']
        municipios = request.POST['municipios']
        tipo_tercero = request.POST['tipo_tercero']
        email = request.POST['email']

        terceros = Terceros(


            nombre=nombre,
            sigla=sigla,
            direccion = direccion,
            telefono = telefono,
            telefono2 = telefono2,
            celular = celular,
            nit = nit,
            dv = dv,
            representante_legal = representante_legal,
            continentes_id = int(continentes),
            paises_id = int(paises),
            departamentos_id = int(departamentos),
            municipios_id = int(municipios),
            tipo_tercero = tipo_tercero,
            email = email,
            creado=timezone.now(),

        )
        terceros.save()

        messages.add_message(request, messages.INFO,
                                 'Se ha registrado el Tercero ' + (nombre).encode(
                                     'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/terceros/')

def config_terceros_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        terceros = Terceros.objects.get(pk=id)
        lista_continentes = Continentes.objects.all()
        lista_paises = Paises.objects.all()
        lista_departamentos = Departamentos.objects.all()
        lista_municipios = Municipios.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_terceros == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/tipos_documentos/')

        return render(request, "config_terceros_editar.html", {'user': current_user,
                                                                         'permisos': permisos,
                                                                         'terceros': terceros,
                                                                       'lista_continentes': lista_continentes,
                                                                       'lista_paises': lista_paises,
                                                                       'lista_departamentos': lista_departamentos,
                                                                       'lista_municipios': lista_municipios,
                                                                       'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.registrar_terceros == False:
            messages.add_message(request, messages.ERROR,
                                 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/terceros/')

        id=request.POST['id']
        nombre = request.POST['nombre']
        sigla = request.POST['sigla']
        direccion = request.POST['direccion']
        telefono = request.POST['telefono']
        telefono2 = request.POST['telefono2']
        celular = request.POST['celular']
        nit = request.POST['nit']
        dv = request.POST['dv']
        representante_legal = request.POST['representante_legal']
        continentes = request.POST['continentes']
        paises = request.POST['paises']
        departamentos = request.POST['departamentos']
        municipios = request.POST['municipios']
        tipo_tercero = request.POST['tipo_tercero']
        email = request.POST['email']

        terceros = Terceros(
            id=id,
            nombre=nombre,
            sigla=sigla,
            direccion=direccion,
            telefono=telefono,
            telefono2=telefono2,
            celular=celular,
            nit=nit,
            dv=dv,
            representante_legal=representante_legal,
            continentes_id=int(continentes),
            paises_id=int(paises),
            departamentos_id=int(departamentos),
            municipios_id=int(municipios),
            tipo_tercero=tipo_tercero,
            email=email,
            creado=timezone.now(),

        )
        terceros.save()

        messages.add_message(request, messages.INFO,
                             'Se ha editado el Tercero ' + (nombre).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')
        return HttpResponseRedirect('/configuracion/terceros/')


# Perfiles
def config_empresas(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        #permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_empresas = Empresas.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.consultar_usuarios == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            #return HttpResponseRedirect('/administracion/')

        return render(request, "config_empresas.html", {'user': current_user,
                                                      'lista_empresas': lista_empresas,
                                                      #'permisos': permisos,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    else:
        pass

def config_empresas_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        lista_continentes = Continentes.objects.all()
        lista_municipios = Municipios.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')



        return render(request, "config_empresas_registrar.html", {'user': current_user,
                                                                  'lista_continentes': lista_continentes,
                                                                  'lista_municipios': lista_municipios,
                                                                #'permisos': permisos,
                                                                  #'permiso_usuario': usuario_datos,
                                                                  })
    elif request.method == 'POST':

        current_user = request.user

        nombre = request.POST['nombre']
        nit = request.POST['nit']
        telefono = request.POST['telefono']
        movil = request.POST['movil']
        responsable = request.POST['responsable']
        tipo = request.POST['tipo']
        continentes = request.POST['continentes']
        paises = request.POST['paises']
        departamentos = request.POST['departamentos']
        municipios = request.POST['municipios']
        email = request.POST['email']
        direccion = request.POST['direccion']



        empresas = Empresas(
            nombre=nombre,
            nit=nit,
            telefono=telefono,
            movil=movil,
            responsable=responsable,
            tipo=tipo,
            email=email,
            direccion=direccion,
            continentes_id=continentes,
            paises_id=paises,
            departamentos_id=departamentos,
            municipios_id=municipios,

        )
        empresas.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado la empresa ' + nombre + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/empresas/')

def config_empresas_editar(request, id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        empresas = Empresas.objects.get(pk=id)
        lista_continentes = Continentes.objects.all()
        lista_paises = Paises.objects.all()
        lista_departamentos = Departamentos.objects.all()
        lista_municipios = Municipios.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')



        return render(request, "config_empresas_editar.html", {'user': current_user,
                                                             'permisos': permisos,
                                                             'empresas': empresas,
                                                            'lista_continentes': lista_continentes,
                                                            'lista_paises': lista_paises,
                                                            'lista_departamentos': lista_departamentos,
                                                            'lista_municipios': lista_municipios,
                                                            'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':



        id = request.POST['id']
        nombre = request.POST['nombre']
        nit = request.POST['nit']
        telefono = request.POST['telefono']
        movil = request.POST['movil']
        responsable = request.POST['responsable']
        tipo = request.POST['tipo']
        continentes = request.POST['continentes']
        paises = request.POST['paises']
        departamentos = request.POST['departamentos']
        municipios = request.POST['municipios']
        email = request.POST['email']
        direccion = request.POST['direccion']
        temporada = request.POST['temporada']
        cliente = request.POST['cliente']
        codigo = request.POST['codigo']

        empresas = Empresas(
            id=id,
            nombre=nombre,
            nit=nit,
            telefono=telefono,
            movil=movil,
            responsable=responsable,
            tipo=tipo,
            email=email,
            direccion=direccion,
            continentes_id=continentes,
            paises_id=paises,
            departamentos_id=departamentos,
            municipios_id=municipios,
            temporada=int(temporada),
            cliente=int(cliente),
            codigo=codigo,

        )
        empresas.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado la empresa ' + nombre + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/empresas/')

def config_empresas_borrar(request, id):

    if request.method == 'GET':

        empresa = Empresas.objects.get(pk=id)
        usuario_actual = request.user
        current_user = request.user

        if empresa.usuarios_datos_set.all().exists():
            messages.add_message(request, messages.ERROR,
                                 'No se puede borrar la empresa ' + str(
                                     id) + ' porque tiene un usuario asociado')

            return HttpResponseRedirect('/configuracion/empresas/')

        else:

            empresa.delete()
            messages.add_message(request, messages.WARNING,
                                 'Se ha borrado la empresa ' + str(id) + ' satisfactoriamente')

            return HttpResponseRedirect('/configuracion/empresas/')



def config_historial_empresas(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_empresas.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        fecha_fin_a = datetime.strptime(fecha_fin, "%Y-%m-%d").date()

        # Combinar fecha_fin con la hora máxima para incluir todo el día
        fecha_fin_ajustada = datetime.combine(fecha_fin_a, time(23, 59, 59))
        historial = HistoriaUsuario.objects.filter(fecha__range=[fecha_inicio, fecha_fin_ajustada])
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_empresas.html", {'user': current_user,
                                                               'permiso_usuario': usuario_datos,
                                                               'lista_historial': historial,
                                                               'fecha_inicio': fecha_inicio,
                                                               'fecha_fin': fecha_fin,
                                                               })


def config_historial_detalle_pedidos(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_detalle_pedidos.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pedido = request.POST['pedido']
        fecha_fin_a = datetime.strptime(fecha_fin, "%Y-%m-%d").date()

        # Combinar fecha_fin con la hora máxima para incluir todo el día
        fecha_fin_ajustada = datetime.combine(fecha_fin_a, time(23, 59, 59))

        if pedido:
            historial = HistoriaDetallePedidos.objects.filter(fecha__range=[fecha_inicio, fecha_fin_ajustada],pedido=pedido)
        else:
            historial = HistoriaDetallePedidos.objects.filter(fecha__range=[fecha_inicio, fecha_fin_ajustada])
            pedido=''
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_detalle_pedidos.html", {'user': current_user,
                                                               'permiso_usuario': usuario_datos,
                                                               'lista_historial': historial,
                                                               'fecha_inicio': fecha_inicio,
                                                               'fecha_fin': fecha_fin,
                                                               'pedido': pedido,
                                                               })




def config_servicio_crediya(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_servicio_crediya.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        pass





def config_informe_recibo(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_informe_recibo.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':

            current_user = request.user
            nombre = current_user.username
            fecha_inicio= request.POST['fecha_inicio']
            fecha_fin= request.POST['fecha_fin']
            historial =  AsignacionPedidosOtrosCanales.objects.filter(
            fecha__range=[fecha_inicio, fecha_fin],
            cantidadrecibo__gt=0
            )

            subtitulo = "InformeRecibo"

            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="INFORMERECIBO.xls"'

            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet(subtitulo)

            # Sheet header, first row
            row_num = 0

            font_style = xlwt.XFStyle()
            font_style.font.bold = True

            columns = ['Nombre artesano',
                       'Código artesano',
                       'Código  artículo',
                       'Descripción artículo',
                       'Número pedido',
                       'Fecha pedido',
                       'Nro unidades pedidas',
                       'Nro unidades pendientes',
                       'Nro unidades entregadas',
                       'Valor costo unidad',
                       'Valor venta unidad',
                       'Valor costo pedido',
                       'Valor costo pendiente',
                       'Valor venta pedido',
                       'Partida arancelaria',
                       ]

            for col_num in range(len(columns)):
                cwidth = ws.col(col_num).width
                if (len(columns[col_num]) * 367) > cwidth:
                    ws.col(col_num).width = (len(columns) * 367)
                ws.write(row_num, col_num, columns[col_num], font_style)

            # Sheet body, remaining rows
            font_style = xlwt.XFStyle()

            rows = []

            for d in historial:
                url3 = IP_SAP + "SQLQueries('consultadatospreciosfacturas1')/List?codigo='" + str(
                    d.num_detalle.referencia) + "'"

                response1 = sap_request(url3)
                response1 = response1.text.replace('null', ' " " ')
                response1 = ast.literal_eval(response1)['value']

                precioventa = 0
                preciocompra = 0
                codigo_arancelaria = ''

                for cuenta in response1:
                    codigo_arancelaria = cuenta['U_HBT1_ARANCEL']
                    if cuenta['PriceList'] == 2:
                        precioventa = cuenta['Price']
                    if cuenta['PriceList'] == 11:
                        preciocompra = cuenta['Price']
                cantidad_pendiente=d.cantidad-d.cantidadrecibo
                nombre_artesano=d.empresa.nombre
                codigo_artesano=d.empresa.codigo
                codigo_articulo=str(d.num_detalle.referencia)
                descripcion_articulo=d.num_detalle.nombre
                numero_pedido = str(d.pk),
                fecha_pedido = (d.fecha).strftime('%Y-%m-%d'),
                numero_u_pedidas = str(d.cantidad)
                numero_u_pendientes = str(cantidad_pendiente)
                numero_u_entregadas = str(d.cantidadrecibo)
                valor_c_unidad = str(preciocompra)
                valor_v_unidad = str(precioventa)
                valor_c_pedido = str(preciocompra*d.cantidad)
                valor_c_pendiente = str(preciocompra*cantidad_pendiente)
                valor_v_pedido = str(precioventa*d.cantidad)
                partida_arancelaria = codigo_arancelaria
                datos = [(
                    nombre_artesano,
                    codigo_artesano,
                    codigo_articulo,
                    descripcion_articulo,
                    numero_pedido,
                    fecha_pedido,
                    numero_u_pedidas,
                    numero_u_pendientes,
                    numero_u_entregadas,
                    valor_c_unidad,
                    valor_v_unidad,
                    valor_c_pedido,
                    valor_c_pendiente,
                    valor_v_pedido,
                    partida_arancelaria,
                )]
                rows.extend(datos)

            for row in rows:
                row_num += 1
                for col_num in range(len(row)):
                    ws.write(row_num, col_num, row[col_num], font_style)

            wb.save(response)
            return response


def config_informe_pedido_otros_can(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresas = Empresas.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_informe_pedido_otros_can.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'empresas': empresas,
                                                        })
    elif request.method == 'POST':

            current_user = request.user
            nombre = current_user.username
            fecha_inicio= request.POST['fecha_inicio']
            fecha_fin= request.POST['fecha_fin']
            empresa= request.POST['empresa']
            historial =  AsignacionPedidosOtrosCanales.objects.filter(
            num_detalle__num_pedido__fecha__range=[fecha_inicio, fecha_fin],
            num_detalle__num_pedido__empresa__pk=int(empresa)
            )

            subtitulo = "PedidoDetalladoCliente"

            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="PedidoDetalladoCliente.xls"'

            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet(subtitulo)

            # Sheet header, first row
            row_num = 0

            font_style = xlwt.XFStyle()
            font_style.font.bold = True

            columns = ['PEDIDO SAP',
                       'PEDIDO CLIENTE',
                       'CLIENTE',
                       'CANTIDADA',
                       'REFERENCIA SAP',
                       'REFERENCIA CLIENTE',
                       'DESCRIPCION',
                       'EMPRESAA',
                       'OBSERVACIONES',
                       ]

            for col_num in range(len(columns)):
                cwidth = ws.col(col_num).width
                if (len(columns[col_num]) * 367) > cwidth:
                    ws.col(col_num).width = (len(columns) * 367)
                ws.write(row_num, col_num, columns[col_num], font_style)

            # Sheet body, remaining rows
            font_style = xlwt.XFStyle()

            rows = []

            for d in historial:

                pedido_sap=d.num_detalle.num_pedido.num_pedido
                pedido_cliente=d.num_detalle.num_pedido.numero_pedido_cliente
                cliente=d.num_detalle.num_pedido.empresa.nombre
                cantidad=d.cantidad
                referencia_sap=d.num_detalle.referencia
                descripcion=d.num_detalle.nombre
                referencia_cliente=''
                try:
                    empresaa = d.empresa.nombre or ""
                except AttributeError:
                    empresaa = ""
                observaciones = d.num_detalle.observaciones
                datos = [(
                    pedido_sap,
                    pedido_cliente,
                    cliente,
                    cantidad,
                    referencia_sap,
                    referencia_cliente,
                    descripcion,

                    empresaa,
                    observaciones,
                )]
                rows.extend(datos)

            for row in rows:
                row_num += 1
                for col_num in range(len(row)):
                    ws.write(row_num, col_num, row[col_num], font_style)

            wb.save(response)
            return response



def config_informe_pedido_otros_can_cliente(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresas = Empresas.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_informe_pedido_otros_can_cliente.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'empresas': empresas,
                                                        })
    elif request.method == 'POST':

            current_user = request.user
            nombre = current_user.username
            fecha_inicio= request.POST['fecha_inicio']
            fecha_fin= request.POST['fecha_fin']
            empresa= Usuarios_datos.objects.filter(usuario=current_user.pk).first()
            empresa= empresa.empresa.pk
            historial =  AsignacionPedidosOtrosCanales.objects.filter(
            num_detalle__num_pedido__fecha__range=[fecha_inicio, fecha_fin],
            num_detalle__num_pedido__empresa__pk=int(empresa)
            )

            subtitulo = "PedidoDetalladoCliente"

            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="PedidoDetalladoCliente.xls"'

            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet(subtitulo)

            # Sheet header, first row
            row_num = 0

            font_style = xlwt.XFStyle()
            font_style.font.bold = True

            columns = ['PEDIDO SAP',
                       'PEDIDO CLIENTE',
                       'CLIENTE',
                       'CANTIDADA',
                       'REFERENCIA SAP',
                       'REFERENCIA CLIENTE',
                       'DESCRIPCION',
                       'EMPRESAA',
                       'OBSERVACIONES',
                       ]

            for col_num in range(len(columns)):
                cwidth = ws.col(col_num).width
                if (len(columns[col_num]) * 367) > cwidth:
                    ws.col(col_num).width = (len(columns) * 367)
                ws.write(row_num, col_num, columns[col_num], font_style)

            # Sheet body, remaining rows
            font_style = xlwt.XFStyle()

            rows = []

            for d in historial:

                pedido_sap=d.num_detalle.num_pedido.num_pedido
                pedido_cliente=d.num_detalle.num_pedido.numero_pedido_cliente
                cliente=d.num_detalle.num_pedido.empresa.nombre
                cantidad=d.cantidad
                referencia_sap=d.num_detalle.referencia
                descripcion=d.num_detalle.nombre
                referencia_cliente=''
                try:
                    empresaa = d.empresa.nombre or ""
                except AttributeError:
                    empresaa = ""
                observaciones = d.num_detalle.observaciones
                datos = [(
                    pedido_sap,
                    pedido_cliente,
                    cliente,
                    cantidad,
                    referencia_sap,
                    referencia_cliente,
                    descripcion,

                    empresaa,
                    observaciones,
                )]
                rows.extend(datos)

            for row in rows:
                row_num += 1
                for col_num in range(len(row)):
                    ws.write(row_num, col_num, row[col_num], font_style)

            wb.save(response)
            return response




def config_informe_cliente_fact(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresas = Empresas.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_informe_cliente_fact.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'empresas': empresas,
                                                        })
    elif request.method == 'POST':

            current_user = request.user
            nombre = current_user.username
            hoy = date.today().strftime("%Y-%m-%d")
            dolar = 4065.22

            url3 = IP_SAP + "SQLQueries('obtenertrm')/List?hoy='" + str(
                hoy) + "'"

            response1 = sap_request(url3)
            response1 = response1.text.replace('null', ' " " ')
            dolar = ast.literal_eval(response1)['value'][0]['Rate']
            fecha_inicio= request.POST['fecha_inicio']
            fecha_fin= request.POST['fecha_fin']
            empresa= request.POST['empresa']
            historial =  AsignacionPedidosOtrosCanales.objects.filter(
            num_detalle__num_pedido__fecha__range=[fecha_inicio, fecha_fin],
            num_detalle__num_pedido__empresa__pk=int(empresa)
            )

            subtitulo = "InformeCliente"

            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="InformeCliente.xls"'

            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet(subtitulo)

            # Sheet header, first row
            row_num = 0

            font_style = xlwt.XFStyle()
            font_style.font.bold = True

            columns = ['Empresario',
                       'Item',
                       'Número de Artículo',
                       'Descripción Español',
                       'Descripción inglés',
                       'Cantidad a facturar',
                       'Costo 2025 ($)',
                       'Venta 2025 ($) ',
                       'Base Unitaria Precio USD',
                       'Total Costo',
                       'Total Venta',
                       'Total Venta USD',
                       'Código Partida Arancelaria',
                       'Margen',
                       'Tipo de Cambio',
                       'No de Pedido Sugar',
                       'No. Envío',
                       'No. Factura',
                       ]

            for col_num in range(len(columns)):
                cwidth = ws.col(col_num).width
                if (len(columns[col_num]) * 367) > cwidth:
                    ws.col(col_num).width = (len(columns) * 367)
                ws.write(row_num, col_num, columns[col_num], font_style)

            # Sheet body, remaining rows
            font_style = xlwt.XFStyle()

            rows = []

            for d in historial:

                url3 = IP_SAP + "SQLQueries('consultadatospreciosfacturas1')/List?codigo='" + str(
                    d.num_detalle.referencia) + "'"

                response1 = sap_request(url3)
                response1 = response1.text.replace('null', ' " " ')
                response1 = ast.literal_eval(response1)['value']

                precioventa = 0
                preciocompra = 0
                codigo_arancelaria = ''

                for cuenta in response1:
                    codigo_arancelaria = cuenta['U_HBT1_ARANCEL']
                    if cuenta['PriceList'] == 2:
                        precioventa = cuenta['Price']
                    if cuenta['PriceList'] == 11:
                        preciocompra = cuenta['Price']

                cantidad_pedido = d.cantidad
                cantidad_facturar = d.cantidadfacturada
                cantidad_pendiente = cantidad_pedido - cantidad_facturar
                base_unitaria = round(precioventa / dolar, 2)
                total_costo = cantidad_facturar * preciocompra
                total_venta = cantidad_facturar * precioventa
                total_venta_usd = base_unitaria * cantidad_facturar
                resultado = 1.0 - (preciocompra / precioventa)


                try:
                    empresaa = d.empresa.nombre or ""
                except AttributeError:
                    empresaa = ""

                item=''
                numero_articulo=d.num_detalle.referencia
                descripcion_esp=d.num_detalle.nombre
                descripcion_ing=''
                cantidad_fact=d.cantidadfacturada
                costo=preciocompra
                venta=precioventa
                partida_arance=codigo_arancelaria
                margen=resultado
                tipo_cambio='usd'
                no_envio=''
                no_factura=''
                no_pedido=d.num_detalle.num_pedido.num_pedido
                datos = [(
                    empresaa,
                    numero_articulo,
                    item,
                    descripcion_esp,
                    descripcion_ing,
                    cantidad_fact,
                    costo,
                    venta,
                    base_unitaria,
                    total_costo,
                    total_venta,
                    total_venta_usd,
                    partida_arance,
                    margen,
                    tipo_cambio,
                    no_pedido,
                    no_envio,
                    no_factura,

                )]
                rows.extend(datos)

            for row in rows:
                row_num += 1
                for col_num in range(len(row)):
                    ws.write(row_num, col_num, row[col_num], font_style)

            wb.save(response)
            return response


def config_servicio_crediya_informe(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_servicio_crediya_informe.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        pass


def config_servicio_credilisto_informe(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_servicio_credilisto_informe.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        pass



def config_servicio_credilisto(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_servicio_credilisto.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        pass



def reporte_historial_empresa(request):

    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        historial = HistoriaUsuario.objects.filter(fecha__range=[fecha_inicio, fecha_fin])

        subtitulo ="Lista_HistorialEmail"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="LISTA_HISTORIALEMAIL.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['USUARIO',
                   'EMPRESA',
                   'ACCION',
                   'FECHA',
                   'HORA',
                   ]

        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []

        for d in historial:
            usuario =str(d.usuario.usuario),
            empresa = d.empresa.nombre,
            fecha =(d.fecha).strftime('%Y-%m-%d'),
            hora =(d.fecha).strftime('%H:%M:%S'),
            accion = d.accion,
            datos = [(
                usuario,
                empresa,
                accion,
                fecha,
                hora,
            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response







def reporte_generacion_facturas(request):
    if request.method == 'POST':
        hoy = date.today().strftime("%Y-%m-%d")
        dolar = 4065.22



        url3 = IP_SAP + "SQLQueries('obtenertrm')/List?hoy='" + str(
            hoy) + "'"

        response1 = sap_request(url3)
        response1 = response1.text.replace('null', ' " " ')
        dolar = ast.literal_eval(response1)['value'][0]['Rate']
        current_user = request.user
        nombre = current_user.username

        # 🚩 Obtener y procesar lista de pedidos seleccionados
        numero_pedidoss = request.POST.get('numero_pedidoss')
        lista_pedidos = numero_pedidoss.split(",")

        # 🚩 Obtener todos los registros relacionados
        historial = AsignacionPedidosOtrosCanales.objects.filter(num_detalle__num_pedido__num_pedido__in=lista_pedidos)

        # 🚩 Obtener uno solo para los datos generales
        historial_busqueda = historial.first()

        # Obtener campos del formulario
        puerto_embarque = request.POST.get('puerto_embarque')
        pais = request.POST.get('pais')
        puerto_destino = request.POST.get('puerto_destino')
        pais_destino = request.POST.get('pais_destino')
        via = request.POST.get('via')
        forma_pago = request.POST.get('forma_pago')
        termino_negociacion = request.POST.get('termino_negociacion')
        numero_piezas = request.POST.get('numero_piezas')
        peso_neto = request.POST.get('peso_neto')
        peso_bruto = request.POST.get('peso_bruto')

        # Excel
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="FACTURAS.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet("Lista_Facturas")

        enunciado_style = xlwt.XFStyle()
        enunciado_font = xlwt.Font()
        enunciado_font.bold = True
        enunciado_style.font = enunciado_font

        enunciados = [
            ("ORDEN DE VENTA", ''),
            ("Codigo Cliente", historial_busqueda.num_detalle.num_pedido.empresa.codigo),
            ("Nombre", historial_busqueda.num_detalle.num_pedido.empresa.nombre),
            ("Moneda USD:", dolar),
            ("Fecha Contabilizacion", hoy),
            ("Fecha Entrega", hoy),
            ("Fecha del Documento", hoy),
            ("Puerto de Embarque", puerto_embarque),
            ("Pais de Origen ", pais),
            ("Puerto Destino", puerto_destino),
            ("Pais Destino", pais_destino),
            ("Via/Via", via),
            ("Forma de Pago", forma_pago),
            ("Termino de Negociacion", termino_negociacion),
            ("Número piezas", numero_piezas),
            ("Peso neto", peso_neto),
            ("Peso bruto", peso_bruto),
        ]

        row_num = 0
        for clave, valor in enunciados:
            ws.write(row_num, 0, clave, enunciado_style)
            ws.write(row_num, 1, valor)
            row_num += 1

        row_num += 1
        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['Nro pedido cliente', 'Nro pedido SAP', 'Cliente', 'Codigo cliente', 'Número de Artículo',
                   'Descripción', 'Cantidad Pedido', 'Cantidad Facturar', 'Cantidad Pendiente', 'Costo', 'Venta',
                   'Base Unitaria Precio USD', 'Total Costo', 'Total Venta', 'Total Venta USD',
                   'Código Partida Arancelaria', 'Margen']

        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        font_style = xlwt.XFStyle()
        rows = []

        for d in historial:
            url3 = IP_SAP + "SQLQueries('consultadatospreciosfacturas1')/List?codigo='" + str(
                d.num_detalle.referencia) + "'"

            response1 = sap_request(url3)
            response1 = response1.text.replace('null', ' " " ')
            response1 = ast.literal_eval(response1)['value']

            precioventa = 0
            preciocompra = 0
            codigo_arancelaria = ''

            for cuenta in response1:
                codigo_arancelaria = cuenta['U_HBT1_ARANCEL']
                if cuenta['PriceList'] == 2:
                    precioventa = cuenta['Price']
                if cuenta['PriceList'] == 11:
                    preciocompra = cuenta['Price']

            cantidad_pedido = d.cantidad
            cantidad_facturar = d.cantidadfacturada
            cantidad_pendiente = cantidad_pedido - cantidad_facturar
            base_unitaria = round(precioventa / dolar, 2)
            total_costo = cantidad_facturar * preciocompra
            total_venta = cantidad_facturar * precioventa
            total_venta_usd = base_unitaria * cantidad_facturar
            resultado = 1.0 - (preciocompra / precioventa)

            nro_pedido_cliente = d.num_detalle.num_pedido.numero_pedido_cliente
            nro_pedido = d.num_detalle.num_pedido.num_pedido
            cliente = d.empresa.nombre
            codigo_cliente = d.empresa.codigo
            numero_articulo = d.num_detalle.referencia
            descripcion = d.num_detalle.nombre

            datos = [(nro_pedido_cliente, nro_pedido, cliente, codigo_cliente, numero_articulo, descripcion,
                      cantidad_pedido, cantidad_facturar, cantidad_pendiente, preciocompra, precioventa,
                      base_unitaria, total_costo, total_venta, total_venta_usd, codigo_arancelaria, resultado)]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response




def reporte_servicio_crediya(request):

    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        historial = RespuestaOrdenCompraApi.objects.filter(FechaEmision__range=[fecha_inicio, fecha_fin])

        subtitulo ="Lista_Servicio_CrediYa"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="LISTA_SERVICIO_CREDIYA.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['IDENTIFICACION',
                   'TIPO_IDENTIFICACION',
                   'VALOR_APROBADO',
                   'FECHA_DEL_PEDIDO',
                   'NUMERO_ORDEN_COMPRA',
                   'VALOR_ORDEN_DE_COMPRA',
                   'INTERES',
                   'VALOR 50%',
                   'EMPRESARIO',
                   'USUARIO',
                   ]

        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []

        for d in historial:
            try:
                cliente = CrediyaPreaprobado.objects.filter(NumeroOrdenCompra=d.NumeroOrdenCompra).first()
                valor_orden_compra = str(cliente.ValorOrden)
                empresario = cliente.Empresa.nombre
                if cliente.UsuarioSolicitado.pk==1:
                    usuario=''
                else:
                    usuario = cliente.UsuarioSolicitado.username
            except:
                cliente = ClientesApi.objects.filter(NumeroPedido=d.NumeroOrdenCompra).first()
                valor_orden_compra = str(cliente.ValorOrden),
                empresario = cliente.NombreEmpresa,
                usuario = ''

            identificacion =str(d.Identificacion),
            tipo_identificacion = d.TipoIdentificacion,
            valor_aprobado = d.ValorAprobado,
            fecha =(d.FechaEmision).strftime('%Y-%m-%d'),
            numero_orden_compra = d.NumeroOrdenCompra,
            interes = d.Interes,
            total=str(int(d.ValorAprobado)+int(d.Interes))
            datos = [(
                identificacion,
                tipo_identificacion,
                valor_aprobado,
                fecha,
                numero_orden_compra,
                valor_orden_compra,
                interes,
                total,
                empresario,
                usuario,
            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response



def reporte_facturas_activas(request):
    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        ahora = datetime.now(pytz.timezone('America/Bogota'))
        fecha_futura = ahora + timedelta(days=10)
        fecha_futura_str = fecha_futura.strftime('%Y-%m-%d')
        subtitulo = u"Facturas_Activas"

        # Estilo para fecha en Excel
        date_style = xlwt.XFStyle()
        date_style.num_format_str = 'DD/MM/YYYY'

        # Construir URL sin f-string
        url2 = IP_SAP + "SQLQueries('ConsultasFacturasActivasV88')/List?fecha='" + fecha_futura_str + "'"

        response = sap_request(url2)
        response = response.text.replace('null', ' " " ')
        data = ast.literal_eval(response)
        facturas = data['value']

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="FACTURAS_ACTIVAS.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)

        # Encabezados
        row_num = 0
        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['Numero Factura Empresa', 'Valor', 'Pagado', 'Saldo', 'Fecha Vencimiento', 'Nombre Empresario']

        for col_num in range(len(columns)):
            ws.col(col_num).width = (len(columns[col_num]) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Cuerpo
        font_style = xlwt.XFStyle()

        for d in facturas:
            numero_fact_emp = unicode(d['NumAtCard'])
            valor = float(d['DocTotal'])
            pagado = float(d['PaidToDate'])
            fecha_ven_raw = str(d['DocDueDate'])
            nombre_emp = unicode(d['CardName'])

            # Convertir fecha
            fecha_ven = datetime.strptime(fecha_ven_raw, "%Y%m%d")

            # Calcular saldo
            saldo_calculado = valor - pagado

            # Escribir fila
            row_num += 1
            ws.write(row_num, 0, numero_fact_emp)
            ws.write(row_num, 1, valor)
            ws.write(row_num, 2, pagado)
            ws.write(row_num, 3, saldo_calculado)
            ws.write(row_num, 4, fecha_ven, date_style)  # Formato fecha
            ws.write(row_num, 5, nombre_emp)

        wb.save(response)
        return response




def reporte_servicio_crediya_informe(request):

    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        estado = request.GET.get('estado')
        historial = CrediyaPreaprobado.objects.filter(FechaSolicitud__range=[fecha_inicio, fecha_fin],estado=estado)

        subtitulo ="Lista_Servicio_CrediYa"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="INFORMECREDIYA.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['EMPRESA',
                   'PEDIDO',
                   'FECHA APROBACION',
                   'VALOR APROBADO',
                   'INTERES',
                   'EGRESO CREADO',
                   ]

        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []

        for d in historial:
            empresa = str(d.Empresa.nombre) if d.Empresa and d.Empresa.nombre is not None else ''
            pedido = d.NumeroOrdenCompra if d.NumeroOrdenCompra is not None else ''
            fecha_aprobacion = d.FechaRespuesta.strftime('%Y-%m-%d') if d.FechaRespuesta else ''
            valor_aprobado = d.ValorAprobado if d.ValorAprobado is not None else ''
            interes = d.Interes if d.Interes is not None else ''
            egreso = d.EgresoCreado if d.EgresoCreado is not None else ''
            datos = [(
                empresa,
                pedido,
                fecha_aprobacion,
                valor_aprobado,
                interes,
                egreso,
            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response




def reporte_servicio_credilisto_informe(request):

    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        estado = request.GET.get('estado')
        historial = CrediListoPreaprobado.objects.filter(FechaSolicitud__range=[fecha_inicio, fecha_fin],estado=estado)

        subtitulo ="Lista_Servicio_CrediListo"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="INFORMECREDILISTO.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['EMPRESA',
                   'FACTURA',
                   'REFERENCIA',
                   'FECHA APROBACION',
                   'VALOR APROBADO',
                   'VALOR_APLICADO',
                   'INTERES',
                   'INTERES APLICADO',
                   'EGRESO CREADO',
                   ]

        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []

        for d in historial:
            empresa = str(d.Empresa.nombre) if d.Empresa and d.Empresa.nombre is not None else ''
            pedido = d.NumeroFactura if d.NumeroFactura is not None else ''
            referencia = d.NumeroFacturaCliente if d.NumeroFacturaCliente is not None else ''
            fecha_aprobacion = d.FechaRespuesta.strftime('%Y-%m-%d') if d.FechaRespuesta else ''
            valor_aprobado = d.ValorAprobado if d.ValorAprobado is not None else ''
            valor_aplicado = d.ValorAprobado_aplicado if d.ValorAprobado_aplicado is not None else ''
            interes = d.Interes if d.Interes is not None else ''
            interes_aplicado = d.Interes_aplicado if d.Interes_aplicado is not None else ''
            egreso = d.EgresoCreado if d.EgresoCreado is not None else ''
            datos = [(
                empresa,
                pedido,
                referencia,
                fecha_aprobacion,
                valor_aprobado,
                valor_aplicado,
                interes,
                interes_aplicado,
                egreso,
            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response




def reporte_servicio_credilisto(request):

    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')

        historial = RespuestaFacturaApi.objects.filter(FechaEmision__range=[fecha_inicio, fecha_fin])

        subtitulo ="Lista_Servicio_CrediListo"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="LISTA_SERVICIO_CREDILISTO.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['IDENTIFICACION',
                   'TIPO_IDENTIFICACION',
                   'VALOR_APROBADO',
                   'VALOR_APLICADO',
                   'FECHA_DE_FACTURA',
                   'NUMERO FACTURA PCS',
                   'NO FACTURA EMPRESARIO',
                   'VALOR FACTURA',
                   'INTERES',
                   'INTERES APLICADO',
                   'EMPRESARIO',
                   'USUARIO',
                   'CRUCES NOTAS/DESCUENTOS',
                   ]

        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []

        for d in historial:
            try:
                cliente = CrediListoPreaprobado.objects.filter(NumeroFactura=d.NumeroFactura).first()
                numero_factura_empresario = str(cliente.NumeroFacturaCliente)
                valor_factura = str(cliente.ValorFactura)
                empresario = cliente.Empresa.nombre
                if cliente.UsuarioSolicitado.pk==1:
                    usuario=''
                else:
                    usuario = cliente.UsuarioSolicitado.username
                valor_aplicado = cliente.ValorAprobado_aplicado
                interes_aplicado = cliente.Interes_aplicado
                cruces = cliente.CrucesAplicados
            except:
                cliente = FacturasApi.objects.filter(NumeroFactura=d.NumeroFactura).first()
                numero_factura_empresario = str(cliente.Referencia2),
                valor_factura = str(cliente.ValorFacturaEmitida),
                empresario = cliente.NombreEmpresa,
                usuario=''

            identificacion =str(d.Identificacion),
            tipo_identificacion = d.TipoIdentificacion,
            valor_aprobado = d.ValorAprobado,

            fecha =(d.FechaEmision).strftime('%Y-%m-%d'),
            numero_factura = d.NumeroFactura,
            interes = d.Interes,

            datos = [(
                identificacion,
                tipo_identificacion,
                valor_aprobado,
                valor_aplicado,
                fecha,
                numero_factura,
                numero_factura_empresario,
                valor_factura,
                interes,
                interes_aplicado,
                empresario,
                usuario,
                cruces,
            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response


def config_historial_email(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_email.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        historial=HistorialErrorTarea.objects.filter(fecha__range=[fecha_inicio,fecha_fin])
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_email.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'lista_historial': historial,
                                                        'fecha_inicio': fecha_inicio,
                                                        'fecha_fin': fecha_fin,
                                                        })




def config_reenviar_pedido(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_reenviar_email.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        historial=HistorialEmailReEnviados.objects.filter(fecha__range=[fecha_inicio,fecha_fin])
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_reenviar_email.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'lista_historial': historial,
                                                        'fecha_inicio': fecha_inicio,
                                                        'fecha_fin': fecha_fin,
                                                        })


def config_historial_estado_sistema(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_estado_sistema.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        historial=HistoriaEstadoSistema.objects.filter(fecha__range=[fecha_inicio,fecha_fin])
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_estado_sistema.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'lista_historial': historial,
                                                        'fecha_inicio': fecha_inicio,
                                                        'fecha_fin': fecha_fin,
                                                        })


def config_estado_sistema(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        estado=EstadoSistema.objects.all().first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_estado_sistema.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'estado': estado,
                                                        })
    elif request.method == 'POST':
        usuarios=User.objects.all()
        now = datetime.now(pytz.timezone('America/Bogota'))
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        estado = request.POST['estado']
        try:
            razon = request.POST['razon']
        except:
            razon=''
        if estado=='True':
            errores = HistoriaEstadoSistema(
                accion='Desactivacion Sistema',
                fecha=now,
                usuario_id=current_user.id,
                razon=str(razon),
            )
            errores.save()
            EstadoSistema.objects.filter(id=1).update(estado=0)
            for usuario in usuarios:
                try:
                    subject = 'PORTAL WEB TEMPORALMENTE FUERA DE SERVICIO'
                    from_email = 'conectaportalweb@gmail.com'
                    to = usuario.email

                    text_content = ''

                    html_content = '<html>' \
                                   '<head></head>' \
                                   '<body>' \
                                   '<h1>¡Hola!</h1>' \
                                   '<p>Te informamos que nuestro portal web <strong>Conecta</strong> estará temporalmente fuera de servicio debido a ' + str(
                        razon) + '. Agradecemos tu comprensión y te mantendremos informado(a) sobre cualquier actualización. Disculpa las molestias.</p>' \
                                 '<br>' \
                                 '<p>Saludos,</p>' \
                                 '<p><strong>CONECTA</strong></p>' \
                                 '</body>' \
                                 '</html>'

                    msg = EmailMultiAlternatives(subject, text_content, from_email, [to])
                    msg.attach_alternative(html_content, "text/html")
                    msg.send()
                except:
                    pass

        else:
            errores = HistoriaEstadoSistema(
                accion='Activacion Sistema',
                fecha=now,
                usuario_id=current_user.id,
                razon=str(razon),
            )
            errores.save()
            EstadoSistema.objects.filter(id=1).update(estado=1)
            for usuario in usuarios:
                try:
                    subject = 'RESTAURACIÓN DEL SERVICIO EN EL PORTAL WEB'
                    from_email = 'conectaportalweb@gmail.com'
                    to = usuario.email

                    text_content = ''

                    html_content = '<html>' \
                                   '<head></head>' \
                                   '<body>' \
                                   '<h1>¡Hola!</h1>' \
                                   '<p>Es un placer informarte que el servicio en nuestro portal web <strong>Conecta</strong> ha sido restaurado satisfactoriamente.</p>' \
                                   '<p>A partir de ahora, tendrás pleno acceso nuevamente a todas las funcionalidades y características de nuestro portal. Nos disculpamos sinceramente por los inconvenientes que pudiste haber experimentado durante el período de interrupción y agradecemos tu comprensión y paciencia mientras trabajábamos en la restauración.</p>' \
                                   '<br>' \
                                   '<p>Queremos asegurarte que hemos tomado las medidas necesarias para evitar futuras interrupciones y estamos comprometidos en brindarte un servicio confiable y de calidad.</p>' \
                                   '<br>' \
                                   '<p>¡Te damos la bienvenida de nuevo y esperamos que disfrutes de una experiencia excepcional en nuestro portal web!</p>' \
                                   '<br>' \
                                   '<p>Atentamente,</p>' \
                                   '<p><strong>CONECTA</strong></p>' \
                                   '</body>' \
                                   '</html>'

                    msg = EmailMultiAlternatives(subject, text_content, from_email, [to])
                    msg.attach_alternative(html_content, "text/html")
                    msg.send()
                except:
                    pass

        return HttpResponseRedirect('/configuracion/estado_sistema/')




def config_excel_pedidos_externos(request):
    # Render de la vista de configuración (GET)
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_excel_pedidos_externos.html", {
            'user': current_user,
            'permiso_usuario': usuario_datos,
        })

    # Procesamiento del Excel (POST)
    elif request.method == 'POST':
        hoy = date.today()
        hora = datetime.now().time()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        # Validar archivo
        if 'excel' not in request.FILES:
            messages.add_message(request, messages.ERROR, 'Debe adjuntar un archivo de Excel.')
            return HttpResponseRedirect('/configuracion/excel_pedidos_externos/')

        excel_file = request.FILES['excel']

        try:
            wb = openpyxl.load_workbook(excel_file)
        except Exception:
            messages.add_message(request, messages.ERROR, 'No se pudo leer el archivo de Excel.')
            return HttpResponseRedirect('/configuracion/excel_pedidos_externos/')

        sheet = wb.active

        # Último registro para consecutivo
        ultimo_registro = PedidosOtrosCanales.objects.all().last()
        consecutivo = 1 if ultimo_registro is None else (ultimo_registro.num_pedido + 1)

        # Campos opcionales desde el Excel
        nro_pedido = sheet['B3'].value if sheet['B3'].value else None
        fecha_pediw = sheet['B4'].value if sheet['B4'].value else None

        # Guardar encabezado del pedido
        pedidootros = PedidosOtrosCanales(
            num_pedido=consecutivo,
            empresa=usuario_datos.empresa,
            fecha=hoy,
            hora=hora,
            fecha_maxima=fecha_pediw,           # Puede ser None
            fecha_entrega=fecha_pediw,          # Puede ser None
            numero_pedido_cliente=nro_pedido,   # Puede ser None
        )
        pedidootros.save()

        # Recorrer filas desde la 7 (A7, B7, C7, ...)
        for row in sheet.iter_rows(min_row=7, values_only=True):
            u_plu = row[0]
            necesidad = row[1]
            observacion = row[2]

            # Validar cantidad numérica (Py2.7)
            if isinstance(necesidad, (int, float, long)):
                # Validar que haya PLU diligenciado
                if u_plu is not None and (not isinstance(u_plu, basestring) or u_plu.strip() != u''):
                    # Normalizar u_plu si es cadena
                    if isinstance(u_plu, basestring):
                        u_plu = u_plu.strip()

                    # Buscar en Portal por PLU
                    articulos_plu_qs = MaestroArticulo.objects.filter(u_plu=u_plu)

                    if not articulos_plu_qs.exists():
                        # No está en Portal → consultar SAP
                        url3 = (
                            IP_SAP + "SQLQueries('ProductosOtrosCanalesv1')/List?producto='"
                            + unicode(u_plu) + "'"
                        )

                        raw = sap_request(url3)
                        # Parseo robusto del cuerpo (preferir JSON)
                        try:
                            data_sap = json.loads(raw.text)  # {'value': [...]}
                        except ValueError:
                            # Fallback si viene como texto tipo dict
                            try:
                                data_sap = ast.literal_eval(raw.text)
                            except Exception:
                                data_sap = {}

                        response_list = data_sap.get('value') or []
                        count = len(response_list)

                        if count == 0:
                            # No hay datos en SAP para ese PLU
                            pedidootros.delete()
                            messages.add_message(
                                request, messages.ERROR,
                                'No se encuentra el PLU ' + unicode(u_plu) + ' en el sistema SAP'
                            )
                            return HttpResponseRedirect('/configuracion/excel_pedidos_externos/')

                        elif count == 1:
                            # Un único resultado
                            item = response_list[0]

                            # Ver si ya existe en Portal por itemCode
                            articulo = MaestroArticulo.objects.filter(itemCode=item.get('ItemCode')).first()
                            if not articulo:
                                # Crear en Portal con datos de SAP
                                maestroarticulo = MaestroArticulo(
                                    itemCode=item.get('ItemCode'),
                                    codeBars=item.get('CodeBars'),
                                    itemName=item.get('ItemName'),
                                    proveedorCodigo=item.get('ProveedorCodigo'),
                                    proveedorNombre=item.get('ProveedorNombre'),
                                    u_plu=item.get('U_PLU'),
                                )
                                maestroarticulo.save()
                                articulo = maestroarticulo

                            # Validar proveedor en Portal
                            empresario = Empresas.objects.filter(codigo=item.get('ProveedorCodigo')).first()
                            if not empresario:
                                pedidootros.delete()
                                messages.add_message(
                                    request, messages.ERROR,
                                    'En Portal Conecta no se encuentra registrado el proveedor: '+item.get('ProveedorNombre')+' - '+item.get('ProveedorCodigo')
                                )
                                return HttpResponseRedirect('/configuracion/excel_pedidos_externos/')

                            # Crear detalle completo
                            detallepedidootros = DetallesPedidosOtrosCanales(
                                num_pedido_id=int(consecutivo),
                                cantidad=necesidad,
                                nombre=articulo.itemName,
                                referencia=articulo.itemCode,
                                u_plu=articulo.u_plu,
                                observaciones=observacion,
                                empresa=empresario,
                                #articulo=articulo,
                            )
                            detallepedidootros.save()

                            asignacion = AsignacionPedidosOtrosCanales(
                                num_detalle_id=detallepedidootros.id,
                                cantidad=necesidad,
                                empresa_id=detallepedidootros.empresa.id,
                                fecha=hoy
                            )
                            asignacion.save()

                        else:
                            items = response_list
                            for item in items:
                                # Ver si ya existe en Portal por itemCode
                                articulo = MaestroArticulo.objects.filter(itemCode=item.get('ItemCode'))
                                if not articulo:
                                    # Crear en Portal con datos de SAP
                                    maestroarticulo = MaestroArticulo(
                                        itemCode=item.get('ItemCode'),
                                        codeBars=item.get('CodeBars'),
                                        itemName=item.get('ItemName'),
                                        proveedorCodigo=item.get('ProveedorCodigo'),
                                        proveedorNombre=item.get('ProveedorNombre'),
                                        u_plu=item.get('U_PLU'),
                                    )
                                    maestroarticulo.save()

                                # Validar proveedor en Portal
                                empresario = Empresas.objects.filter(codigo=item.get('ProveedorCodigo')).first()
                                if not empresario:
                                    pedidootros.delete()
                                    messages.add_message(
                                        request, messages.ERROR,
                                        'En Portal Conecta no se encuentra registrado el proveedor: '+item.get('ProveedorNombre')+' - '+item.get('ProveedorCodigo')
                                    )
                                    return HttpResponseRedirect('/configuracion/excel_pedidos_externos/')

                                # detallepedidootros = DetallesPedidosOtrosCanales(
                                #     num_pedido_id=int(consecutivo),
                                #     cantidad=necesidad,
                                #     nombre=item.get('ItemName'),
                                #     referencia=item.get('ItemCode'),
                                #     u_plu=item.get('U_PLU'),
                                #     observaciones=observacion,
                                #     empresa=empresario,
                                #     # articulo=articulo,
                                # )
                                # detallepedidootros.save()

                            detallepedidootros_plu = DetallesPedidosOtrosCanales_plus(
                                num_pedido_id=int(consecutivo),
                                cantidad=necesidad,
                                u_plu=u_plu,
                                observaciones=observacion,
                            )
                            detallepedidootros_plu.save()

                            # # Más de un resultado → ambigüedad → guardar detalle parcial con PLU
                            # detallepedidootros = DetallesPedidosOtrosCanales(
                            #     num_pedido_id=int(consecutivo),
                            #     cantidad=necesidad,
                            #     u_plu=u_plu,
                            #     observaciones=observacion,
                            # )
                            # detallepedidootros.save()

                    else:
                        # Ya existe al menos un artículo con ese PLU en Portal
                        count_plu = articulos_plu_qs.count()

                        if count_plu == 1:
                            articulo = articulos_plu_qs.first()
                            # Validar proveedor
                            empresario = Empresas.objects.filter(codigo=articulo.proveedorCodigo).first()
                            if not empresario:
                                pedidootros.delete()
                                messages.add_message(
                                    request, messages.ERROR,
                                    'No se encuentra registrado el proveedor del producto en Portal Conecta'
                                )
                                return HttpResponseRedirect('/configuracion/excel_pedidos_externos/')

                            # Crear detalle completo
                            detallepedidootros = DetallesPedidosOtrosCanales(
                                num_pedido_id=int(consecutivo),
                                cantidad=necesidad,
                                nombre=articulo.itemName,
                                referencia=articulo.itemCode,
                                u_plu=articulo.u_plu,
                                observaciones=observacion,
                                empresa=empresario,
                                #articulo=articulo,
                            )
                            detallepedidootros.save()

                            asignacion = AsignacionPedidosOtrosCanales(
                                num_detalle_id=detallepedidootros.id,
                                cantidad=necesidad,
                                empresa_id=detallepedidootros.empresa.id,
                                fecha=hoy
                            )
                            asignacion.save()

                        else:
                            for articulo_plu in articulos_plu_qs:
                                # Validar proveedor
                                empresario = Empresas.objects.filter(codigo=articulo_plu.proveedorCodigo).first()
                                if not empresario:
                                    pedidootros.delete()
                                    messages.add_message(
                                        request, messages.ERROR,
                                        'No se encuentra registrado el proveedor del producto en Portal Conecta'
                                    )
                                    return HttpResponseRedirect('/configuracion/excel_pedidos_externos/')

                                # Crear detalle completo
                                # detallepedidootros = DetallesPedidosOtrosCanales(
                                #     num_pedido_id=int(consecutivo),
                                #     cantidad=necesidad,
                                #     nombre=articulo_plu.itemName,
                                #     referencia=articulo_plu.itemCode,
                                #     u_plu=articulo_plu.u_plu,
                                #     observaciones=observacion,
                                #     empresa=empresario,
                                #     # articulo=articulo,
                                # )
                                # detallepedidootros.save()

                            detallepedidootros_plu = DetallesPedidosOtrosCanales_plus(
                                num_pedido_id=int(consecutivo),
                                cantidad=necesidad,
                                u_plu=u_plu,
                                observaciones=observacion,
                            )
                            detallepedidootros_plu.save()

                            # Hay múltiples artículos con el mismo PLU → ambigüedad
                            # detallepedidootros = DetallesPedidosOtrosCanales(
                            #     num_pedido_id=int(consecutivo),
                            #     cantidad=necesidad,
                            #     u_plu=u_plu,
                            #     observaciones=observacion,
                            # )
                            # detallepedidootros.save()

                else:
                    # PLU vacío o None
                    pedidootros.delete()
                    messages.add_message(
                        request, messages.ERROR,
                        'Debe diligenciar el código PLU'
                    )
                    return HttpResponseRedirect('/configuracion/excel_pedidos_externos/')

            else:
                # Cantidad no numérica
                if necesidad is None:
                    pedidootros.delete()
                    messages.add_message(
                        request, messages.ERROR,
                        'Hay un valor en el campo de cantidad que se encuentra sin diligenciar'
                    )
                    return HttpResponseRedirect('/configuracion/excel_pedidos_externos/')
                else:
                    pedidootros.delete()
                    messages.add_message(
                        request, messages.ERROR,
                        'El campo de cantidad solo puede contener valores numéricos'
                    )
                    return HttpResponseRedirect('/configuracion/excel_pedidos_externos/')

        # Enviar correo a usuarios con perfiles específicos
        usuarios_correo = Usuarios_datos.objects.filter(perfil_pcs_id__in=[8, 21, 23])
        for correos in usuarios_correo:
            email = EmailMessage(
                'Nuevo Pedido Recibido de Otros Canales – Pedido ' + str(nro_pedido),
                'Te informamos que has recibido un nuevo pedido a través de otros canales de la empresa ' +
                str(usuario_datos.empresa.nombre),
                to=[correos.usuario.email]
            )
            # email.send()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el pedido #'+str(consecutivo)+' exitosamente.')

        # Cerrar sesión en SAP

        return HttpResponseRedirect('/configuracion/excel_pedidos_externos/')


def config_excel_pedidos_externos_distribuido(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_excel_pedidos_externos_distribuido.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        hoy = date.today()
        hora = datetime.now().time()
        usuarios=User.objects.all()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        excel_file = request.FILES['excel']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        ultimo_registro = PedidosOtrosCanales.objects.all().last()


        if ultimo_registro == None:
            consecutivo = 1
        else:
            consecutivo = ultimo_registro.num_pedido + 1
        nro_pedido = sheet['A2'].value
        pedidootros = PedidosOtrosCanales(
            num_pedido=consecutivo,
            empresa=usuario_datos.empresa,
            fecha=hoy,
            hora=hora,
            estado='predistribuido',
            numero_pedido_cliente=nro_pedido,
        )
        pedidootros.save()
        detalle_pedido=0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            ean = row[1]
            necesidad = row[2]
            localizacion = row[4]
            unidades = row[5]
            if ean:
                cantidad_unidades_totales=0
                if isinstance(necesidad, (int,long, float)):

                    if ean != None:
                        url3 = IP_SAP + "SQLQueries('ConsultaPorEanOtrosC')/List?Ean='" + str(
                            ean) + "'"

                        response1 = sap_request(url3)
                        response1 = response1.text
                        response1 = response1.replace('null', ' " " ')
                        response1 = ast.literal_eval(response1)
                        response1 = response1['value']
                        if response1 != []:
                            response1 = response1[0]
                            detallepedidootros = DetallesPedidosOtrosCanales(
                                num_pedido_id=int(consecutivo),
                                cantidad=necesidad,
                                nombre=response1['ItemName'],
                                referencia=response1['ItemCode'],
                                observaciones='',
                            )
                        else:
                            pedidootros.delete()
                            messages.add_message(request, messages.ERROR,
                                                    'No se encuentra el ean '+str(ean) +' en el sistema'
                                                    )
                            return HttpResponseRedirect('/configuracion/excel_pedidos_externos_distribuido/')
                    else:
                        pedidootros.delete()
                        messages.add_message(request, messages.ERROR,
                                                'Hay un valor en el campo de ean que se encuentra sin diligenciar'
                                                )
                        return HttpResponseRedirect('/configuracion/excel_pedidos_externos_distribuido/')


                else:
                    if necesidad==None:
                        pedidootros.delete()
                        messages.add_message(request, messages.ERROR,
                                             'Hay un valor en el campo de unidades totales que se encuentra sin diligenciar'
                                            )
                        return HttpResponseRedirect('/configuracion/excel_pedidos_externos_distribuido/')
                    else:
                        pedidootros.delete()
                        messages.add_message(request, messages.ERROR,
                                             'El campo de unidades totales solo puede contener valores numéricos')
                        return HttpResponseRedirect('/configuracion/excel_pedidos_externos_distribuido/')
                detallepedidootros.save()
                detalle_pedido=detallepedidootros.pk

            if isinstance(unidades, (int, long, float)):
                url3 = IP_SAP + "SQLQueries('ConsultaEmpresarioPoreEan')/List?localizacion='" + str(
                    localizacion) + "'&codigo='"+str(usuario_datos.empresa.codigo)+"'"

                response1 = sap_request(url3)
                response1 = response1.text
                response1 = response1.replace('null', ' " " ')
                response1 = ast.literal_eval(response1)
                response1 = response1['value']
                if response1 != []:
                    response1 = response1[0]
                    empresa_asignada=Empresas.objects.filter(nombre=response1['Address']).first()
                    detallepedidootros = AsignacionPedidosOtrosCanales(
                        num_detalle_id=int(detalle_pedido),
                        cantidad=unidades,
                        empresa_id=empresa_asignada.pk,
                        fecha=hoy,
                    )
                    detallepedidootros.save()
                    cantidad_unidades_totales=cantidad_unidades_totales+unidades
                    DetallesPedidosOtrosCanales.objects.filter(pk=detalle_pedido).update(cantidad=cantidad_unidades_totales)

                else:
                    pedidootros.delete()
                    messages.add_message(request, messages.ERROR,
                                         'No se encuentra la localizacion edi ' + str(localizacion) + ' en el sistema'
                                         )
                    return HttpResponseRedirect('/configuracion/excel_pedidos_externos_distribuido/')
            else:
                if unidades == None:
                    pedidootros.delete()
                    messages.add_message(request, messages.ERROR,
                                         'Hay un valor en el campo de unidades que se encuentra sin diligenciar'
                                         )
                    return HttpResponseRedirect('/configuracion/excel_pedidos_externos_distribuido/')
                else:
                    pedidootros.delete()
                    messages.add_message(request, messages.ERROR,
                                         'El campo de unidades solo puede contener valores numéricos')
                    return HttpResponseRedirect('/configuracion/excel_pedidos_externos_distribuido/')

        usuarios_correo = Usuarios_datos.objects.filter(perfil_pcs_id__in=[8, 21, 23])

        for correos in usuarios_correo:
            email = EmailMessage(' Nuevo Pedido Recibido de Otros Canales – Pedido  ' + str(nro_pedido),
                                 'Te informamos que has recibido un nuevo pedido a través de otros canales de la empresa . ' + str(
                                     usuario_datos.empresa.nombre),
                                 to=[correos.usuario.email])
            email.send()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el pedido satisfactoriamente')
        return HttpResponseRedirect('/configuracion/excel_pedidos_externos_distribuido/')



def config_plantilla_excel_pedidos_externos(request):
    # Render  administracion.html
    if request.method == 'GET':
        wb = openpyxl.Workbook()

        # Crear una hoja en el libro de trabajo
        sheet = wb.active
        sheet.title = 'PEDIDO'
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        # Obtener el valor de la variable empresario (puedes ajustar esto según tus necesidades)
        empresario = str(usuario_datos.empresa.edi)

        # Escribir el valor de empresario en el archivo Excel
        sheet['A2'] = 'Cliente:'
        sheet['B2'] = empresario
        sheet['A3'] = 'Nro pedido:'
        sheet['A4'] = 'Fecha entrega:'
        sheet['A6'] = 'Referencia / EAN'
        sheet['B6'] = 'Cantidad'
        sheet['C6'] = 'Observaciones'
        sheet['B4'] = 'AAAA/MM/DD'

        # Guardar el libro de trabajo en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Crear una respuesta HTTP con el archivo adjunto
        response = HttpResponse(output.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="PEDIDO.xlsx"'

        return response

def config_plantilla_excel_pedidos_externos_distribuido(request):
    # Render  administracion.html
    if request.method == 'GET':
        wb = openpyxl.Workbook()

        # Crear una hoja en el libro de trabajo
        sheet = wb.active
        sheet.title = 'ORDEN DISTRIBUCIÓN'
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        # Obtener el valor de la variable empresario (puedes ajustar esto según tus necesidades)
        empresario = str(usuario_datos.empresa.edi)

        # Escribir el valor de empresario en el archivo Excel
        sheet['A1'] = 'NRO_OD'
        sheet['B1'] = 'EAN'
        sheet['C1'] = 'UNIDADES TOTALES'
        sheet['D1'] = 'NOMBRE DEPENDENCIA'
        sheet['E1'] = 'LOCALIZACION EDI'
        sheet['F1'] = 'UNIDADES'

        # Guardar el libro de trabajo en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Crear una respuesta HTTP con el archivo adjunto
        response = HttpResponse(output.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="ORDEN DISTRIBUCIÓN.xlsx"'

        return response

def config_historial_respuesta_pedido(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_respuesta_pedido.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        historial=LogRespuestaPedido.objects.filter(fecha__range=[fecha_inicio,fecha_fin])
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_respuesta_pedido.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'lista_historial': historial,
                                                        'fecha_inicio': fecha_inicio,
                                                        'fecha_fin': fecha_fin,
                                                        })



def config_historial_correos_enviados(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresas=Empresas.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_correos_enviados.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'empresas': empresas,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        empresa = request.POST.get('empresa', None)
        email = request.POST.get('email', None)
        pedido = request.POST.get('pedido', None)
        empresas = Empresas.objects.all()
        filtros = {}
        filtros['fecha__range'] = (fecha_inicio, fecha_fin)
        if empresa:
            filtros['empresa'] = empresa
        else:
            empresa=''
        if pedido:
            filtros['pedido'] = pedido
        else:
            pedido=''
        if email:
            filtros['email'] = email
        else:
            email=''

        historial=HistorialEmailEnviados.objects.filter(**filtros)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_correos_enviados.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'lista_historial': historial,
                                                        'fecha_inicio': fecha_inicio,
                                                        'fecha_fin': fecha_fin,
                                                        'empresas': empresas,
                                                        'empresa': empresa,
                                                        'pedido': pedido,
                                                        'email': email,
                                                        })


def config_historial_codigos_registro(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresas=Empresas.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_codigos_registro.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'empresas': empresas,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        fecha_inicio = request.POST.get('fecha_inicio', None)
        fecha_fin = request.POST.get('fecha_fin', None)
        empresa = request.POST.get('empresa', None)
        Codigo = request.POST.get('Codigo', None)
        usuario = request.POST.get('usuario', None)
        empresas = Empresas.objects.all()
        filtros = {}
        filtro= {}
        if fecha_inicio and fecha_fin:
            filtros['codigoregistro__creado__range'] = (fecha_inicio, fecha_fin)
            filtro['creado__range'] = (fecha_inicio, fecha_fin)
        else:
            fecha_inicio = ''
            fecha_fin = ''
        if empresa:
            filtros['empresa__pk'] = empresa
        else:
            empresa=''
        if Codigo:
            filtros['codigoregistro__codigo'] = Codigo
            filtro['codigo'] = Codigo
        else:
            Codigo=''
        if usuario:
            filtros['usuario__username'] = usuario
        else:
            usuario=''
        filtro['activo'] = 1
        historial=Usuarios_datos.objects.filter(**filtros).exclude(codigoregistro=None)
        historial_es=CodigosRegistros.objects.filter(**filtro)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_codigos_registro.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'lista_historial': historial,
                                                        'lista_historiales': historial_es,
                                                        'fecha_inicio': fecha_inicio,
                                                        'fecha_fin': fecha_fin,
                                                        'empresas': empresas,
                                                        'empresa': empresa,
                                                        'Codigo': Codigo,
                                                        'usuario': usuario,
                                                        })






def config_consulta_empresario_tipo_codigos(request):
    if request.method == 'GET':

        if request.GET.get('tipo_empresa'):
            tipo_empresa = request.GET.get('tipo_empresa')




        url2 = IP_SAP + "BusinessPartners?$select=CardCode,CardName&$filter=CardType eq '" + tipo_empresa + "'"

        response = sap_request(url2)
        response = response.text


        response = response.replace('null', ' " " ')
        response = ast.literal_eval(response)
        response = response['value']

        cuenta = len(response)
        paginador = Paginator(response,cuenta)
        pagina = request.GET.get('page')

        try:
            matriz = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            matriz = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            matriz = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        matrices_dict = []

        for dato in matriz:
            card_code = unicode(dato['CardCode'], 'utf-8')
            card_name = unicode(dato['CardName'], 'utf-8')
            codigo_enviar = u"{}-{}".format(card_code, card_name)
            matriz_dict = {
                'codigo':codigo_enviar,
                'nombre':codigo_enviar,
                'tipo_empresa': tipo_empresa
            }

            matrices_dict.append(matriz_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': matrices_dict
        }
        url = IP_SAP + "Logout"
        responselogout = requests.request("POST", url, verify=False)
        return JsonResponse(response_dict)





def config_historial_correos_creditos(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresas=Empresas.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_correos_creditos.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'empresas': empresas,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pedido = request.POST.get('pedido', None)
        tipo_credito = request.POST.get('tipo_credito', None)
        empresas = Empresas.objects.all()
        filtros = {}
        filtros['fecha__range'] = (fecha_inicio, fecha_fin)

        if pedido:
            filtros['pedido'] = pedido
        else:
            pedido=''
        if tipo_credito !='':
            filtros['tipo'] = tipo_credito
        else:
            tipo_credito=''


        historial=HistorialErrorApi.objects.filter(**filtros)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_correos_creditos.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'lista_historial': historial,
                                                        'fecha_inicio': fecha_inicio,
                                                        'fecha_fin': fecha_fin,
                                                        'empresas': empresas,
                                                        'pedido': pedido,
                                                        'tipo_credito': tipo_credito,
                                                        })


def config_historial_correos_creditos_error(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresas=Empresas.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_correos_creditos_error.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'empresas': empresas,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pedido = request.POST.get('pedido', None)
        tipo_credito = request.POST.get('tipo_credito', None)
        empresas = Empresas.objects.all()
        filtros = {}
        filtros['fecha__range'] = (fecha_inicio, fecha_fin)

        if pedido:
            filtros['pedido'] = pedido
        else:
            pedido=''

        if tipo_credito !='':
            filtros['tipo'] = tipo_credito
        else:
            tipo_credito=''


        historial=HistorialErrorEnvioSap.objects.filter(**filtros)

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "config_historial_correos_creditos_error.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'lista_historial': historial,
                                                        'fecha_inicio': fecha_inicio,
                                                        'fecha_fin': fecha_fin,
                                                        'empresas': empresas,
                                                        'tipo_credito': tipo_credito,
                                                        'pedido': pedido,
                                                        })

def config_historial_correos_no_enviados(request):
    # Render  administracion.html
    if request.method == 'GET':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')
        fecha_inicio=request.GET.get('fecha_inicio', '')
        fecha_fin=request.GET.get('fecha_fin', '')
        now = datetime.now(pytz.timezone('America/Bogota'))
        hoy = now.date()
        hora = now.time()
        url2 = IP_SAP + "PurchaseOrders?$select=DocNum,CardName,DocDate,DocEntry,CardCode&$filter=DocDate le '" + fecha_fin + "' and DocDate ge '" + fecha_inicio +"'"
        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']

        datos_no_en_queryset=[]
        for campos in  response:
            if HistorialEmailEnviados.objects.filter(pedido=campos['DocNum'],tipo='enviado').exists() | HistorialEmailEnviados.objects.filter(pedido=campos['DocNum'],tipo='noregistrado').exists() :
                pass
            else:
                datos_query = {
                    "empresa": campos['CardName'],
                    "pedido": campos['DocNum'],
                    "num_pedido": campos['DocEntry'],
                    "fecha": campos['DocDate'],
                    "codigo": campos['CardCode'],
                }
                datos_no_en_queryset.append(datos_query)

        return render(request, "config_historial_correos_no_enviados.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'datos_no_en_queryset': datos_no_en_queryset,
                                                        'fecha_inicio':fecha_inicio,
                                                        'fecha_fin':fecha_fin,
                                                        })
    elif request.method == 'POST':
        pass




def config_enviar_correos_no_enviados(request):
    if request.method == 'POST':
        fecha_fin = request.POST['fecha_fin']
        fecha_inicio = request.POST['fecha_inicio']
        now = datetime.now(pytz.timezone('America/Bogota'))
        hoy = now.date()
        hora = now.time()
        datos_seleccionados = request.POST.getlist('datos_seleccionados[]')


        for dato in datos_seleccionados:
            variable=dato
            variable = str(variable).split(",")

            dependencias = 'LOGISTICA Y DESPACHOS'
            url3 = IP_SAP + "SQLQueries('ConsultaEmailEmpresa')/List?empresa='" + \
                       variable[3] + "'&dependencia='" + dependencias + "'"
            response2 = sap_request(url3)
            response2 = ast.literal_eval(response2.text)
            response2 = response2['value']
            if response2 == []:
                errores = HistorialErrorTarea(
                    accion='No se tiene correo asignado al titulo LOGISTICA Y DESPACHOS',
                    fecha=hoy,
                    hora=hora,
                    empresa=str(variable[0]),
                    pedido=str(variable[1]),
                )
                errores.save()
            else:
                response2 = response2[0]
                response2 = response2['E_MailL']
                response2 = str(response2).split(";")
                for correos in response2:
                    correos = correos.lstrip()
                    expresion_regular = r"(?i)(?:[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*|\"(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21\x23-\x5b\x5d-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])*\")@(?:(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?|\[(?:(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9]))\.){3}(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9])|[a-z0-9-]*[a-z0-9]:(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21-\x5a\x53-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])+)\])"
                    valido = re.match(expresion_regular, correos) is not None
                    if valido == True:
                        try:
                            email = EmailMessage(str(variable[0])+' TIENES UN NUEVO PEDIDO ' + str(variable[1]),
                                                     'Ha recibido un pedido nuevo.Para conocer el detalle del pedido ingresa al siguiente link '
                                                     + IP_SERVIDOR + '/configuracion/solicitud_pedido_orden/detalle/' + str(
                                                         variable[2]) + '/',
                                                     to=[correos])
                            email.send()
                            enviados = HistorialEmailReEnviados(
                                fecha=hoy,
                                hora=hora,
                                empresa=str(variable[0]),
                                pedido=str(variable[1]),
                                tipo='enviado',
                                email=correos
                            )
                            enviados.save()
                            enviados = HistorialEmailEnviados(
                                fecha=hoy,
                                hora=hora,
                                empresa=str(variable[0]),
                                pedido=str(variable[1]),
                                tipo='enviado',
                                email=correos
                            )
                            enviados.save() 
                            registros_enviados = HistorialErrorTarea.objects.filter(pedido=variable[1])
                            registros_enviados.delete()
                        except:
                            now = datetime.now(pytz.timezone('America/Bogota'))
                            hoy = now.date()
                            hora = now.time()
                            errores = HistorialErrorTarea(
                                accion='Fallo al enviar al correo' + str(correos),
                                fecha=hoy,
                                hora=hora,
                                empresa=str(variable[0]),
                                pedido=str(variable[1]),
                            )
                            errores.save()
                    elif valido == False:
                        now = datetime.now(pytz.timezone('America/Bogota'))
                        hoy = now.date()
                        hora = now.time()
                        errores = HistorialErrorTarea(
                            accion='No se reconoce el correo ' + str(correos),
                            fecha=hoy,
                            hora=hora,
                            empresa=str(variable[0]),
                            pedido=str(variable[1]),
                        )
                        errores.save()

        return HttpResponseRedirect('/configuracion/historial_correos_no_enviados/?fecha_inicio=' + fecha_inicio+'&fecha_fin='+fecha_fin)

def config_historial_correos_no_registrados(request):
    # Render  administracion.html
    if request.method == 'GET':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')
        fecha_inicio=request.GET.get('fecha_inicio', '')
        fecha_fin=request.GET.get('fecha_fin', '')
        now = datetime.now(pytz.timezone('America/Bogota'))
        hoy = now.date()
        hora = now.time()
        url2 = IP_SAP + "PurchaseOrders?$select=DocNum,CardName,DocDate&$filter=DocDate le '" + fecha_fin + "' and DocDate ge '" + fecha_inicio +"'"
        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']

        datos_no_en_queryset=[]
        for campos in  response:
            if HistorialEmailEnviados.objects.filter(pedido=campos['DocNum'],tipo='noregistrado').exists():
                datos_query = {
                    "empresa": campos['CardName'],
                    "pedido": campos['DocNum'],
                    "fecha": campos['DocDate'],
                }
                datos_no_en_queryset.append(datos_query)
            else:
                pass

        return render(request, "config_historial_correos_no_registrados.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        'datos_no_en_queryset': datos_no_en_queryset,
                                                        })
    elif request.method == 'POST':
        pass


def config_indicadores_envio_emails(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        return render(request, "indicador_envio_email.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        pass



def indienvio_mail (request):
    if request.method == 'GET':
        
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')

        url2 = IP_SAP + "PurchaseOrders/$count?$filter=DocDate le '" + fecha_fin + "' and DocDate ge '" + fecha_inicio +"'"

        response = sap_request(url2)
        total = ast.literal_eval(response.text)
        n_casos_pendientes = HistorialEmailEnviados.objects.filter(fecha__range=(fecha_inicio, fecha_fin),tipo='enviado').values('pedido').annotate(total=Count('pedido')).count()
        n_casos_no_pertenece = HistorialEmailEnviados.objects.filter(fecha__range=(fecha_inicio, fecha_fin),tipo='noregistrado').values('pedido').annotate(total=Count('pedido')).count()
        n_casos_finalizado = total-n_casos_pendientes-n_casos_no_pertenece
        p_casos_pendientes = (n_casos_pendientes / float(total)) * 100
        p_casos_no_pertenece = (n_casos_no_pertenece / float(total)) * 100
        p_casos_finalizado = (n_casos_finalizado / float(total)) * 100

        p_casos_pendientes = round(p_casos_pendientes, 2)
        p_casos_no_pertenece = round(p_casos_no_pertenece, 2)
        p_casos_finalizado = round(p_casos_finalizado, 2)

        response_dict = {
            'total': total,
            'n_casos_pendientes': n_casos_pendientes,
            'n_casos_no_pertenece': n_casos_no_pertenece,
            'n_casos_finalizado': n_casos_finalizado,
            'p_casos_pendientes': p_casos_pendientes,
            'p_casos_no_pertenece': p_casos_no_pertenece,
            'p_casos_finalizado': p_casos_finalizado,
        }

        return JsonResponse(response_dict)



def reporte_historial(request):

    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        historial = HistorialErrorTarea.objects.filter(fecha__range=[fecha_inicio, fecha_fin])

        subtitulo ="Lista_HistorialEmail"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="LISTA_HISTORIALEMAIL.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['ACCION',
                   'FECHA',
                   'HORA',
                   'EMPRESA',
                   'PEDIDO',
                   ]

        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []

        for d in historial:
            accion =d.accion,
            pedido = d.pedido,
            empresa = d.empresa,
            fecha = str(d.fecha),
            hora = str(d.hora),
            datos = [(
                accion,
                fecha,
                hora,
                empresa,
                pedido,
            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response




def reporte_historial_reenviado(request):

    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        historial = HistorialEmailReEnviados.objects.filter(fecha__range=[fecha_inicio, fecha_fin])

        subtitulo ="Lista_HistorialEmail"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="LISTA_HISTORIALEMAIL.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['PEDIDO',
                   'EMPRESA',
                   'FECHA',
                   'HORA',
                   'EMAIL',
                   ]

        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []

        for d in historial:
            email =d.email,
            pedido = d.pedido,
            empresa = d.empresa,
            fecha = str(d.fecha),
            hora = str(d.hora),
            datos = [(
                pedido,
                empresa,
                fecha,
                hora,
                email,
            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response



def reporte_historial_respuesta(request):

    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        historial = LogRespuestaPedido.objects.filter(fecha__range=[fecha_inicio, fecha_fin])

        subtitulo ="Lista_HistorialRespuestaPedido"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="LISTA_HISTORIAL_RESPUESTA_PEDIDO.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['ACCION',
                   'FECHA',
                   'HORA',
                   'EMPRESA',
                   'PEDIDO',
                   'EMAIL',
                   'PETICION',
                   ]

        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []

        for d in historial:
            accion =d.accion,
            pedido = d.num_pedido,
            empresa = d.empresa,
            email = d.email,
            peticion = d.peticion.descripcion,
            fecha = str(d.fecha),
            hora = str(d.hora),
            datos = [(
                accion,
                fecha,
                hora,
                empresa,
                pedido,
                email,
                peticion
            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response


def reporte_correos_enviados(request):

    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        empresas = request.GET.get('empresa')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        empresa = request.GET.get('empresa', None)
        email = request.GET.get('email', None)
        pedido = request.GET.get('pedido', None)
        filtros = {}
        filtros['fecha__range'] = (fecha_inicio, fecha_fin)
        if empresa:
            filtros['empresa'] = empresa
        else:
            empresa = ''
        if pedido:
            filtros['pedido'] = pedido
        else:
            pedido = ''
        if email:
            filtros['email'] = email
        else:
            email = ''

        historial = HistorialEmailEnviados.objects.filter(**filtros)

        subtitulo ="ListaHistorialCorreosEnviados"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="LISTA_HISTORIAL_CORREOS_ENVIADOS.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['EMPRESA',
                   'PEDIDO',
                   'FECHA',
                   'HORA',
                   'EMAIL',
                   ]

        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []

        for d in historial:
            empresa =d.empresa,
            pedido = d.pedido,
            fecha = str(d.fecha),
            hora = str(d.hora),
            email = d.email or 'sin correo',
            datos = [(
                empresa,
                pedido,
                fecha,
                hora,
                email,
            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response


# Usuarios
def config_usuarios(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_usuarios = User.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "config_opciones.html", {'user': current_user,
                                                          'lista_usuarios': lista_usuarios,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    else:
        pass





def procesar_pdfs(request):
    carpeta = os.path.join(settings.MEDIA_ROOT, 'pdfs')

    if not os.path.exists(carpeta):
        os.makedirs(carpeta)

    for archivo in os.listdir(carpeta):
        if archivo.lower().endswith('.pdf'):

            # Evitar reprocesar el mismo archivo
            if Documento.objects.filter(archivo=archivo).exists():
                continue

            partes = archivo.split('_')
            if len(partes) >= 4:
                nit = partes[1]
                nombre = partes[2]
                fecha_str = partes[3]

                # Parsear fecha DDMMYYYY
                try:
                    fecha_obj = datetime.strptime(fecha_str, "%d%m%Y").date()
                except ValueError:
                    continue

                # Construir la URL accesible (ej: /bodega/pdfs/archivo.pdf)
                ruta_url = os.path.join(settings.MEDIA_URL, 'pdfs', archivo)

                # Guardar en BD
                Documento.objects.create(
                    nit=nit,
                    nombre=nombre,
                    fecha=fecha_obj,
                    ruta=ruta_url,
                    archivo=archivo
                )

    return HttpResponseRedirect('/configuracion/usuarios/')






def config_usuarios_perfil(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        lista_usuarios = User.objects.all()
        lista_empresas = Empresas.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "config_perfil_usuario.html", {'user': current_user,
                                                          'lista_usuarios': lista_usuarios,
                                                        'permiso_usuario': usuario_datos,
                                                        'lista_empresas': lista_empresas,
                                                        })

    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        clave = request.POST['password']
        clave2 = request.POST['password2']
        if clave == clave2:
            clavef = clave
        else:
            messages.add_message(request, messages.ERROR,
                                 'No coincide la contraseña con la confirmación de contraseña')
            usuario_actual = request.user
            current_user = request.user
            usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


            if not current_user.is_staff:
                return HttpResponseRedirect('/login/')


            return render(request, "config_perfil_usuario.html",
                          {
                           'user': current_user,
                            'permiso_usuario': usuario_datos,
                           })
        usuario=User.objects.filter(id=usuario_datos.usuario_id).first()
        usuario.set_password(clavef)
        usuario.save()

        messages.add_message(request, messages.INFO,

                             'Se ha actualizado su contraseña '  + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/usuarios/perfil/')



def config_usuarios_perfil_empresa(request):
    if request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresa = request.POST['empresas']

        usuario=Usuarios_datos.objects.filter(id=usuario_datos.usuario_id).update(empresa_id=empresa)

        messages.add_message(request, messages.INFO,

                             'Se ha actualizado la empresa '  + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/usuarios/perfil/')



def config_usuarios_perfil_empresa_despacho(request):
    if request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresa = request.POST['empresas']

        usuario=Usuarios_datos.objects.filter(id=usuario_datos.usuario_id).update(empresa_id=empresa)

        messages.add_message(request, messages.INFO,

                             'Se ha actualizado la empresa '  + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/solicitud_pedido_orden/')


def config_usuarios_registrar(request):
    if request.method == 'GET':
        current_user = request.user
        #usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        #permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_empresas = Empresas.objects.all()
        perfiles_pcs = Perfiles_PCS.objects.all
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.registrar_usuarios == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            #return HttpResponseRedirect('/configuracion/usuarios/')

        return render(request, "config_usuarios_registrar.html", {'user_name': current_user.first_name,
                                                                  #'permisos': permisos,
                                                                  'lista_empresas': lista_empresas,
                                                                  'perfiles_pcs': perfiles_pcs,
                                                                  #'permiso_usuario': usuario_datos,
                                                                  })
    elif request.method == 'POST':

        current_user = request.user
        nombres = request.POST['nombres']
        apellidos = request.POST['apellidos']
        nombre_usuario = request.POST['nombre_usuario']
        clave = request.POST['password']
        clave2 = request.POST['password2']
        correo = request.POST['email']

        empresas = request.POST['empresas']
        telefono = request.POST['telefono']
        cargo = request.POST['cargo']
        publico = request.POST['publico']
        atencion = request.POST['atencion']
        pcs_user = request.POST['pcs_user']
        perfil_pcs = request.POST['perfil_pcs']

        if User.objects.filter(email=correo).exists():
            messages.add_message(request, messages.INFO,
                                 'Este email ya esta registrado, Porfavor seleccione uno diferente')
            usuario_actual = request.user
            current_user = request.user




            return render(request, "config_usuarios_registrar.html",
                          {'nombres': nombres,
                           'apellidos': apellidos,
                           'usuario': nombre_usuario,
                           'email': correo,
                           'usuario_actual': usuario_actual,
                           })
        
        if clave == clave2:
            clavef = clave
        else:
            messages.add_message(request, messages.INFO,
                                 'No coincide la contraseña con la confirmación de contraseña')
            usuario_actual = request.user
            current_user = request.user


            if not current_user.is_staff:
                return HttpResponseRedirect('/login/')


            return render(request, "config_usuarios_registrar.html",
                          {'nombres': nombres,
                           'apellidos': apellidos,
                           'usuario': nombre_usuario,
                           'email': correo,
                           'usuario_actual': usuario_actual,
                           })

        user = User.objects.all()
        for us in user:
            if us.username == nombre_usuario:
                usuario_actual = request.user
                current_user = request.user
                permisos = Permisos.objects.filter(usuario_id=current_user.id).first()

                if not current_user.is_staff:
                    return HttpResponseRedirect('/login/')

                messages.add_message(request, messages.INFO,

                                     'Ya existe el nombre de usuario : ' + (nombre_usuario).encode('utf-8').strip())

                usuario_actual = request.user
                return render(request, "config_usuarios_registrar.html",
                              {'nombres': nombres,
                               'apellidos': apellidos,
                               'usuario': nombre_usuario,
                               'email': correo,
                               'usuario_actual': usuario_actual,
                               'permisos': permisos,
                               })


        usuario = User(
            first_name=nombres,
            last_name=apellidos,
            username=nombre_usuario,
            email=correo,
            is_superuser=1,
            is_staff=1,
        )

        usuario.set_password(clavef)
        usuario.save()
        if telefono:
            telefono=telefono
        else:
            telefono=None
        ultimo_usuario = User.objects.all().last()
        id_usuario = int(ultimo_usuario.id)
        usuarios_datos = Usuarios_datos(

            id=id_usuario,
            usuario_id=id_usuario,
            empresa_id=int(empresas),
            perfil_pcs_id=int(perfil_pcs),
            cargo=cargo,
            telefono=telefono,
            pcs=pcs_user,
            admin=publico,
            atencion=atencion,
            creado=timezone.now(),
        )

        usuarios_datos.save()
        messages.add_message(request, messages.INFO,

                             'Se ha registrado el administrador ' + (nombres).encode(

                                 'utf-8').strip() + ' ' + (apellidos).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/usuarios/')


def config_usuarios_registrar_externos(request):
    # Render  administracion.html
    if request.method == 'GET':
        codigo = request.GET.get("codigo")
        codigoregistro = request.GET.get("codigoregistro")

        url2 = IP_SAP + "BusinessPartners?$select=CardCode,CardName,CardType&$filter=CardCode eq '" + codigo + "'"

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        for datos in response:
            id= datos['CardCode']
            nombre= datos['CardName']
            tipo= datos['CardType']
        return render(request, "registro_usuarios_externos.html", {
                                                        'codigo':codigo,
                                                        'id': id,
                                                        'codigoregistro': codigoregistro,
                                                        'nombre': nombre,
                                                        'tipo': tipo,
                                                        })
    else:
        empresa = request.POST['empresa']
        id = request.POST['id']
        codigoregistro = request.POST['codigoregistro']
        codigo = request.POST['codigo']
        tipo = request.POST['tipo']
        nombres = request.POST['nombres']
        apellidos = request.POST['apellidos']
        nombre_usuario = request.POST['nombre_usuario']
        email = request.POST['email']
        telefono = request.POST['telefono']
        cargo = request.POST['cargo']
        clave = request.POST['password']
        clave2 = request.POST['password2']
        if tipo == 'cSupplier':
            tipo = 'Empresario'
        else:
            tipo = 'Cadena'
        if User.objects.filter(email=email).exists():
            messages.add_message(request, messages.ERROR,
                                 'Este email ya esta registrado, Porfavor seleccione uno diferente')
            return render(request, "registro_usuarios_externos.html",
                          {'nombres': nombres,
                           'apellidos': apellidos,
                           'usuario': nombre_usuario,
                           'email': email,
                           'telefono': telefono,
                           'cargo': cargo,
                           'nombre': empresa,
                           'tipo': tipo,
                           'id': id,
                           'codigoregistro': codigoregistro,
                           'codigo': codigo,
                           })
        if clave == clave2:
            clavef = clave
        else:
            messages.add_message(request, messages.ERROR,
                                 'No coincide la contraseña con la confirmación de contraseña')

            return HttpResponseRedirect('/registrar/usuarios_externos/?codigo='+id+'&codigoregistro='+codigoregistro)

        user = User.objects.all()
        for us in user:
            if us.username == nombre_usuario:
                usuario_actual = request.user


                messages.add_message(request, messages.ERROR,

                                     'Ya existe el nombre de usuario : ' + (nombre_usuario).encode('utf-8').strip())

                return HttpResponseRedirect('/registrar/usuarios_externos/?codigo='+id+'&codigoregistro='+codigoregistro)
        try:
            codigoderegistro=CodigosRegistros.objects.filter(codigo=codigoregistro).first()
            empresa=Empresas.objects.get(nombre=empresa , tipo=tipo)
            usuario = User(
                first_name=nombres,
                last_name=apellidos,
                username=nombre_usuario,
                email=email,
                is_superuser=1,
                is_staff=1,
            )

            usuario.set_password(clavef)
            usuario.save()
            if telefono:
                telefono = telefono
            else:
                telefono = None
            ultimo_usuario = User.objects.all().last()
            id_usuario = int(ultimo_usuario.id)
            usuarios_datos = Usuarios_datos(

                id=id_usuario,
                usuario_id=id_usuario,
                empresa_id=int(empresa.id),
                cargo=cargo,
                telefono=telefono,
                admin=0,
                creado=timezone.now(),
                codigoregistro_id=int(codigoderegistro.pk)
            )

            usuarios_datos.save()
            CodigosRegistros.objects.filter(codigo=codigoregistro).update(activo=0,asignado=timezone.now())
            messages.add_message(request, messages.INFO,

                                 'Se ha registrado el usuario ' + nombres + ' ' + apellidos + ' satisfactoriamente.')

            return HttpResponseRedirect('/login/')
        except Empresas.DoesNotExist :
            codigoderegistro = CodigosRegistros.objects.filter(codigo=codigoregistro).first()


            url2 = IP_SAP + "SQLQueries('EmpresariosListaConsulta')/List?empresario= '" + id + "'"

            response = sap_request(url2)
            response = response.text



            response = response.replace('null', ' " " ')
            response = ast.literal_eval(response)
            response = response['value']
            for datos in response:
                    id = datos['CardCode']
                    nombre = datos['CardName']
                    nit = datos['LicTradNum']
                    telefono_emp = datos['Phone1']
                    email_empre = datos['E_Mail']
                    direccion_empresa = datos['Address']
                    ean = datos['U_EAN']
            empresasr = Empresas(

                    nombre=nombre,
                    nit=nit,
                    telefono=telefono_emp,
                    movil=telefono_emp,
                    responsable='',
                    tipo=tipo,
                    email=email_empre,
                    edi=ean,
                    codigo=codigo,
                )

            empresasr.save()


            empresa = Empresas.objects.get(nombre=empresa, tipo=tipo)
            usuario = User(
                first_name=nombres,
                last_name=apellidos,
                username=nombre_usuario,
                email=email,
                is_superuser=1,
                is_staff=1,
            )

            usuario.set_password(clavef)
            usuario.save()
            if telefono:
                telefono = telefono
            else:
                telefono = None
            ultimo_usuario = User.objects.all().last()
            id_usuario = int(ultimo_usuario.id)
            usuarios_datos = Usuarios_datos(

                id=id_usuario,
                usuario_id=id_usuario,
                empresa_id=int(empresa.id),
                cargo=cargo,
                telefono=telefono,
                admin=0,
                creado=timezone.now(),
                codigoregistro_id=int(codigoderegistro.pk)
            )

            usuarios_datos.save()
            CodigosRegistros.objects.filter(codigo=codigoregistro).update(activo=0, asignado=timezone.now())
            messages.add_message(request, messages.INFO,

                                 'Se ha registrado el usuario ' + nombres + ' ' + apellidos + ' satisfactoriamente.')


            return HttpResponseRedirect('/login/')


def config_usuarios_aten_registrar_externos(request):
    # Render  administracion.html
    if request.method == 'GET':

        return render(request, "registro_usuarios_externos_atencion.html", {
                                                        })
    else:
        empresa = request.POST['empresa']
        tipo = request.POST['tipo']
        nombres = request.POST['nombres']
        apellidos = request.POST['apellidos']
        nombre_usuario = request.POST['nombre_usuario']
        email = request.POST['email']
        telefono = request.POST['telefono']
        cargo = request.POST['cargo']
        clave = request.POST['password']
        clave2 = request.POST['password2']
        if User.objects.filter(email=email).exists():
            messages.add_message(request, messages.ERROR,
                                 'Este email ya esta registrado, Porfavor seleccione uno diferente')
            return render(request, "registro_usuarios_externos_atencion.html",
                          {'nombres': nombres,
                           'apellidos': apellidos,
                           'usuario': nombre_usuario,
                           'email': email,
                           'telefono': telefono,
                           'cargo': cargo,
                           })
        if clave == clave2:
            clavef = clave
        else:
            messages.add_message(request, messages.ERROR,
                                 'No coincide la contraseña con la confirmación de contraseña')

            return HttpResponseRedirect('/registrar/usuarios_externos_atencion/')
        user = User.objects.all()
        for us in user:
            if us.username == nombre_usuario:
                usuario_actual = request.user


                messages.add_message(request, messages.ERROR,

                                     'Ya existe el nombre de usuario : ' + (nombre_usuario).encode('utf-8').strip())

                return HttpResponseRedirect('/registrar/usuarios_externos_atencion/')

        usuario = User(
            first_name=nombres,
            last_name=apellidos,
            username=nombre_usuario,
            email=email,
            is_superuser=1,
            is_staff=1,
        )

        usuario.set_password(clavef)
        usuario.save()
        if telefono:
            telefono = telefono
        else:
            telefono = None
        ultimo_usuario = User.objects.all().last()
        id_usuario = int(ultimo_usuario.id)
        usuarios_datos = Usuarios_datos(

            id=id_usuario,
            usuario_id=id_usuario,
            empresa_id=6,
            cargo=cargo,
            telefono=telefono,
            admin=0,
            creado=timezone.now(),
            atencion=2,
        )

        usuarios_datos.save()
        messages.add_message(request, messages.INFO,

                             'Se ha registrado el usuario ' + nombres + ' ' + apellidos + ' satisfactoriamente.')

        return HttpResponseRedirect('/login/')



def config_usuarios_creditos_registrar_externos(request):
    # Render  administracion.html
    if request.method == 'GET':

        return render(request, "registro_usuarios_externos_creditos.html", {
                                                        })
    else:
        empresa = request.POST['empresa']
        tipo = request.POST['tipo']
        nombres = request.POST['nombres']
        apellidos = request.POST['apellidos']
        nombre_usuario = request.POST['nombre_usuario']
        email = request.POST['email']
        telefono = request.POST['telefono']
        cargo = request.POST['cargo']
        clave = request.POST['password']
        clave2 = request.POST['password2']
        if User.objects.filter(email=email).exists():
            messages.add_message(request, messages.ERROR,
                                 'Este email ya esta registrado, Porfavor seleccione uno diferente')
            return render(request, "registro_usuarios_externos_creditos.html",
                          {'nombres': nombres,
                           'apellidos': apellidos,
                           'usuario': nombre_usuario,
                           'email': email,
                           'telefono': telefono,
                           'cargo': cargo,
                           })
        if clave == clave2:
            clavef = clave
        else:
            messages.add_message(request, messages.ERROR,
                                 'No coincide la contraseña con la confirmación de contraseña')

            return HttpResponseRedirect('/registrar/usuarios_externos_creditos/')
        user = User.objects.all()
        for us in user:
            if us.username == nombre_usuario:
                usuario_actual = request.user


                messages.add_message(request, messages.ERROR,

                                     'Ya existe el nombre de usuario : ' + (nombre_usuario).encode('utf-8').strip())

                return HttpResponseRedirect('/registrar/usuarios_externos_creditos/')

        usuario = User(
            first_name=nombres,
            last_name=apellidos,
            username=nombre_usuario,
            email=email,
            is_superuser=1,
            is_staff=1,
        )

        usuario.set_password(clavef)
        usuario.save()
        if telefono:
            telefono = telefono
        else:
            telefono = None
        ultimo_usuario = User.objects.all().last()
        id_usuario = int(ultimo_usuario.id)
        usuarios_datos = Usuarios_datos(

            id=id_usuario,
            usuario_id=id_usuario,
            empresa_id=6,
            cargo=cargo,
            telefono=telefono,
            admin=0,
            creado=timezone.now(),
            atencion=4,
        )

        usuarios_datos.save()
        messages.add_message(request, messages.INFO,

                             'Se ha registrado el usuario ' + nombres + ' ' + apellidos + ' satisfactoriamente.')

        return HttpResponseRedirect('/login/')



def config_usuarios_creditos_registrar_despacho_nueve(request):
    # Render  administracion.html
    if request.method == 'GET':

        return render(request, "registro_usuarios_externos_despacho_nueve.html", {
                                                        })
    else:
        empresa = request.POST['empresa']
        tipo = request.POST['tipo']
        nombres = request.POST['nombres']
        apellidos = request.POST['apellidos']
        nombre_usuario = request.POST['nombre_usuario']
        email = request.POST['email']
        telefono = request.POST['telefono']
        cargo = request.POST['cargo']
        clave = request.POST['password']
        clave2 = request.POST['password2']
        if User.objects.filter(email=email).exists():
            messages.add_message(request, messages.ERROR,
                                 'Este email ya esta registrado, Porfavor seleccione uno diferente')
            return render(request, "registro_usuarios_externos_despacho_nueve.html",
                          {'nombres': nombres,
                           'apellidos': apellidos,
                           'usuario': nombre_usuario,
                           'email': email,
                           'telefono': telefono,
                           'cargo': cargo,
                           })
        if clave == clave2:
            clavef = clave
        else:
            messages.add_message(request, messages.ERROR,
                                 'No coincide la contraseña con la confirmación de contraseña')

            return HttpResponseRedirect('/registrar/usuarios_externos_despacho_diesnueve/')
        user = User.objects.all()
        for us in user:
            if us.username == nombre_usuario:
                usuario_actual = request.user


                messages.add_message(request, messages.ERROR,

                                     'Ya existe el nombre de usuario : ' + (nombre_usuario).encode('utf-8').strip())

                return HttpResponseRedirect('/registrar/usuarios_externos_despacho_diesnueve/')

        usuario = User(
            first_name=nombres,
            last_name=apellidos,
            username=nombre_usuario,
            email=email,
            is_superuser=1,
            is_staff=1,
        )

        usuario.set_password(clavef)
        usuario.save()
        if telefono:
            telefono = telefono
        else:
            telefono = None
        ultimo_usuario = User.objects.all().last()
        id_usuario = int(ultimo_usuario.id)
        usuarios_datos = Usuarios_datos(

            id=id_usuario,
            usuario_id=id_usuario,
            empresa_id=6,
            cargo=cargo,
            telefono=telefono,
            admin=0,
            creado=timezone.now(),
            atencion=5,
        )

        usuarios_datos.save()
        messages.add_message(request, messages.INFO,

                             'Se ha registrado el usuario ' + nombres + ' ' + apellidos + ' satisfactoriamente.')

        return HttpResponseRedirect('/login/')





def config_usuarios_creditos_registrar_despacho_dos(request):
    # Render  administracion.html
    if request.method == 'GET':

        return render(request, "registro_usuarios_externos_despacho_dos.html", {
                                                        })
    else:
        empresa = request.POST['empresa']
        tipo = request.POST['tipo']
        nombres = request.POST['nombres']
        apellidos = request.POST['apellidos']
        nombre_usuario = request.POST['nombre_usuario']
        email = request.POST['email']
        telefono = request.POST['telefono']
        cargo = request.POST['cargo']
        clave = request.POST['password']
        clave2 = request.POST['password2']
        if User.objects.filter(email=email).exists():
            messages.add_message(request, messages.ERROR,
                                 'Este email ya esta registrado, Porfavor seleccione uno diferente')
            return render(request, "registro_usuarios_externos_despacho_nueve.html",
                          {'nombres': nombres,
                           'apellidos': apellidos,
                           'usuario': nombre_usuario,
                           'email': email,
                           'telefono': telefono,
                           'cargo': cargo,
                           })
        if clave == clave2:
            clavef = clave
        else:
            messages.add_message(request, messages.ERROR,
                                 'No coincide la contraseña con la confirmación de contraseña')

            return HttpResponseRedirect('/registrar/usuarios_externos_despacho_diesnueve/')
        user = User.objects.all()
        for us in user:
            if us.username == nombre_usuario:
                usuario_actual = request.user


                messages.add_message(request, messages.ERROR,

                                     'Ya existe el nombre de usuario : ' + (nombre_usuario).encode('utf-8').strip())

                return HttpResponseRedirect('/registrar/usuarios_externos_despacho_diesnueve/')

        usuario = User(
            first_name=nombres,
            last_name=apellidos,
            username=nombre_usuario,
            email=email,
            is_superuser=1,
            is_staff=1,
        )

        usuario.set_password(clavef)
        usuario.save()
        if telefono:
            telefono = telefono
        else:
            telefono = None
        ultimo_usuario = User.objects.all().last()
        id_usuario = int(ultimo_usuario.id)
        usuarios_datos = Usuarios_datos(

            id=id_usuario,
            usuario_id=id_usuario,
            empresa_id=6,
            cargo=cargo,
            telefono=telefono,
            admin=0,
            creado=timezone.now(),
            atencion=6,
        )

        usuarios_datos.save()
        messages.add_message(request, messages.INFO,

                             'Se ha registrado el usuario ' + nombres + ' ' + apellidos + ' satisfactoriamente.')

        return HttpResponseRedirect('/login/')




def config_usuarios_otros_canales_registrar_externos(request):
    # Render  administracion.html
    if request.method == 'GET':
        codigo = request.GET.get("codigo")
        return render(request, "registro_usuarios_otros_canales_atencion.html", {
                                    'codigo':codigo
                                                        })
    else:
        empresa = request.POST['empresa']
        tipo = request.POST['tipo']
        nombres = request.POST['nombres']
        apellidos = request.POST['apellidos']
        nombre_usuario = request.POST['nombre_usuario']
        email = request.POST['email']
        telefono = request.POST['telefono']
        cargo = request.POST['cargo']
        clave = request.POST['password']
        clave2 = request.POST['password2']
        codigo = request.POST['codigo']
        if User.objects.filter(email=email).exists():
            messages.add_message(request, messages.ERROR,
                                 'Este email ya esta registrado, Porfavor seleccione uno diferente')
            return render(request, "registro_usuarios_otros_canales_atencion.html",
                          {'nombres': nombres,
                           'apellidos': apellidos,
                           'usuario': nombre_usuario,
                           'email': email,
                           'telefono': telefono,
                           'cargo': cargo,
                           })
        if clave == clave2:
            clavef = clave
        else:
            messages.add_message(request, messages.ERROR,
                                 'No coincide la contraseña con la confirmación de contraseña')

            return HttpResponseRedirect('/registrar/usuarios_otros_canales_atencion/?codigo='+codigo)
        user = User.objects.all()
        for us in user:
            if us.username == nombre_usuario:
                usuario_actual = request.user


                messages.add_message(request, messages.ERROR,

                                     'Ya existe el nombre de usuario : ' + (nombre_usuario).encode('utf-8').strip())

                return HttpResponseRedirect('/registrar/usuarios_otros_canales_atencion/?codigo='+codigo)

        usuario = User(
            first_name=nombres,
            last_name=apellidos,
            username=nombre_usuario,
            email=email,
            is_superuser=1,
            is_staff=1,
        )

        usuario.set_password(clavef)
        usuario.save()
        if telefono:
            telefono = telefono
        else:
            telefono = None
        ultimo_usuario = User.objects.all().last()
        id_usuario = int(ultimo_usuario.id)
        usuarios_datos = Usuarios_datos(

            id=id_usuario,
            usuario_id=id_usuario,
            empresa_id=307,
            cargo=cargo,
            telefono=telefono,
            admin=0,
            creado=timezone.now(),
            atencion=1,
            edi=codigo
        )

        usuarios_datos.save()
        messages.add_message(request, messages.INFO,

                             'Se ha registrado el usuario ' + nombres + ' ' + apellidos + ' satisfactoriamente.')

        return HttpResponseRedirect('/login/')


def config_usuarios_pcs_registrar_externos(request,id):
    # Render  administracion.html
    if request.method == 'GET':

        return render(request, "registro_usuarios_externos_pcs.html", {
                                                        'id':id
                                                        })
    else:
        perfil = request.POST['perfil']
        empresa = request.POST['empresa']
        tipo = request.POST['tipo']
        nombres = request.POST['nombres']
        apellidos = request.POST['apellidos']
        nombre_usuario = request.POST['nombre_usuario']
        email = request.POST['email']
        telefono = request.POST['telefono']
        cargo = request.POST['cargo']
        clave = request.POST['password']
        clave2 = request.POST['password2']
        if User.objects.filter(email=email).exists():
            messages.add_message(request, messages.ERROR,
                                 'Este email ya esta registrado, Porfavor seleccione uno diferente')
            return render(request, "registro_usuarios_externos_pcs.html",
                          {'nombres': nombres,
                           'apellidos': apellidos,
                           'usuario': nombre_usuario,
                           'email': email,
                           'telefono': telefono,
                           'cargo': cargo,
                           'id': perfil,
                           })
        if clave == clave2:
            clavef = clave
        else:
            messages.add_message(request, messages.ERROR,
                                 'No coincide la contraseña con la confirmación de contraseña')

            return HttpResponseRedirect('/registrar/usuarios_externos_pcs/'+str(perfil)+'/')
        user = User.objects.all()
        for us in user:
            if us.username == nombre_usuario:
                usuario_actual = request.user


                messages.add_message(request, messages.ERROR,

                                     'Ya existe el nombre de usuario : ' + (nombre_usuario).encode('utf-8').strip())

                return HttpResponseRedirect('/registrar/usuarios_externos_pcs/'+str(perfil)+'/')

        usuario = User(
            first_name=nombres,
            last_name=apellidos,
            username=nombre_usuario,
            email=email,
            is_superuser=1,
            is_staff=1,
        )

        usuario.set_password(clavef)
        usuario.save()
        if telefono:
            telefono = telefono
        else:
            telefono = None
        ultimo_usuario = User.objects.all().last()
        id_usuario = int(ultimo_usuario.id)
        usuarios_datos = Usuarios_datos(

            id=id_usuario,
            usuario_id=id_usuario,
            empresa_id=6,
            cargo=cargo,
            telefono=telefono,
            admin=0,
            creado=timezone.now(),
            atencion=1,
            pcs=1,
            perfil_pcs_id=perfil,
        )

        usuarios_datos.save()
        messages.add_message(request, messages.INFO,

                             'Se ha registrado el usuario ' + nombres + ' ' + apellidos + ' satisfactoriamente.')

        return HttpResponseRedirect('/login/')




def config_usuarios_registrar_externos_comp(request):
    if request.method == 'GET':
        return render(request, "registro_usuarios_externos_comp.html", {})

    codigo = request.POST['codigo']
    codigo_interno = request.POST['codigo_interno']

    # Obtener o crear el objeto de LoginAttempt para el usuario
    login_attempt, created = LoginAttempt.objects.get_or_create(codigo=codigo)

    # Verificar si el usuario está bloqueado
    if login_attempt.is_blocked():
        error_user = 'Usuario bloqueado. Intenta de nuevo después de {}'.format(login_attempt.bloqueado_hasta)
        return render(request, "registro_usuarios_externos_comp.html", {
            'error_user': error_user,
        })

    # Lógica de redirección existente
    if codigo == 'pcsSAPConecta2022.':
        login_attempt.reset_attempts()  # Resetear intentos si es exitoso
        return redirect('/registrar/usuarios_externos_atencion/')
    if codigo == 'pcsSAPConecta2024*':
        login_attempt.reset_attempts()
        return redirect('/registrar/usuarios_externos_creditos/')
    if codigo == 'PcsCOnecta2024BV19':
        login_attempt.reset_attempts()
        return redirect('/registrar/usuarios_externos_despacho_diesnueve/')
    if codigo == 'PcsCOnecta2024BB22':
        login_attempt.reset_attempts()
        return redirect('/registrar/usuarios_externos_despacho_veintedos/')

    # Validar código en base de datos
    if Perfiles_PCS.objects.filter(codigo_conecta__contains=codigo).exists():
        id_perfil = Perfiles_PCS.objects.filter(codigo_conecta=codigo).first().pk
        login_attempt.reset_attempts()
        return redirect('/registrar/usuarios_externos_pcs/{}/'.format(id_perfil))

    if CodigosRegistros.objects.filter(codigo=codigo_interno, activo=True, codigo_cliente=codigo).exists():
        login_attempt.reset_attempts()
    else:
        # Incrementar intentos fallidos y verificar bloqueo
        login_attempt.intentos_fallidos += 1
        login_attempt.ultimo_intento = timezone.now()

        if login_attempt.intentos_fallidos >= 5:
            # Establecer bloqueado_hasta usando timezone.now() para mantener la conciencia de zona horaria
            login_attempt.bloqueado_hasta = timezone.now() + timedelta(minutes=5)
            login_attempt.save()
            error_user = 'Demasiados intentos fallidos. Usuario bloqueado por 5 minutos.'
        else:
            login_attempt.save()
            error_user = 'El código de registro no es válido en el sistema.'

        return render(request, "registro_usuarios_externos_comp.html", {
            'error_user': error_user,
        })

    # Continuar con la lógica de autenticación

    url2 = IP_SAP + "BusinessPartners?$select=CardCode&$filter=CardCode eq '{}'".format(codigo)

    response = sap_request(url2)
    response = ast.literal_eval(response.text)
    response = response['value']

    if not response:
        url2 = IP_SAP + "SQLQueries('validacionEDIS')/List?EAN= '{}'".format(codigo)
        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']

        if not response:
            error_user = 'El Código no pertenece a un usuario asociado a PCS'
            return render(request, "registro_usuarios_externos_comp.html", {
                'error_user': error_user,
            })
        else:
            login_attempt.reset_attempts()
            return redirect('/registrar/usuarios_otros_canales_atencion/?codigo=' + codigo)

    login_attempt.reset_attempts()
    return redirect('/registrar/usuarios_externos/?codigo=' + codigo + '&codigoregistro=' + codigo_interno)



def config_usuarios_olvidar_contrasena(request):
    # Render  administracion.html
    if request.method == 'GET':

        return render(request, "olvidar_contrasena_usuario.html", {
                                                        })
    else:
        email = request.POST['email']
        if User.objects.filter(email=email).exists():
            letters = string.ascii_uppercase + string.digits + string.ascii_lowercase
            palabra = ''
            for i in range(10):
                letra = random.choice(letters)
                palabra = palabra + str(letra)
            usuario = User.objects.filter(email=email).first()
            usuario.set_password(palabra)
            usuario.save()
            subject, from_email, to = 'CAMBIO DE CONTRASEÑA', 'conectaportalweb@gmail.com', email
            text_content = ''

            html_content = '<h1>¡hola!</h1><br>' \
                           '<p>Hola, se solicito un restablecimiento de contraseña para tu cuenta ' \
                           ''+str(email)+' ingresa con la siguiente clave temporal al sistema </p>' \
                           '<br><center><h2><mark>'+str(palabra)+'</mark></h2></center><br>' \
                           '<p>si tu no realizaste la solicitud de cambio de contraseña, solo ignora este mensaje</p><br>' \
                            '<br><p>Saludos,</p><br><p>CONECTA</p>'


            msg = EmailMultiAlternatives(subject, text_content, from_email, [to])
            msg.attach_alternative(html_content, "text/html")
            msg.send()
            messages.add_message(request, messages.INFO,
                                 'Se ha enviado el correo electronico')
            return HttpResponseRedirect('/login/')
        else:
            error_user = 'Este correo no se encuentra registrado en el sistema'
            return render(request, "olvidar_contrasena_usuario.html", {
                'error_user': error_user,
            })




def config_usuarios_borrar(request, id):
    try:
        if request.method == 'GET':

            usuario = User.objects.get(pk=id)
            usuarios_datos=Usuarios_datos.objects.get(usuario_id=id)
            usuario_actual = request.user
            current_user = request.user
            usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


            if usuario:
                usuario.delete()
                usuarios_datos.delete()

            messages.add_message(request, messages.WARNING,
                                    'Se ha borrado el usuario ' + str(id) + ' satisfactoriamente')

            return HttpResponseRedirect('/configuracion/usuarios/')
    except :
        messages.add_message(request, messages.WARNING,
                             'No se puede borrar,Faltan datos en la tabla' )

        return HttpResponseRedirect('/configuracion/usuarios/')

def config_usuarios_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        usuarios = User.objects.get(pk=id)
        lista_empresas = Empresas.objects.all()
        usuarios_datos=Usuarios_datos.objects.get(usuario_id=id)
        perfiles_pcs=Perfiles_PCS.objects.all


        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')


        return render(request, "config_usuarios_editar.html", {'user': current_user,
                                                                         'permisos': permisos,
                                                                         'usuarios': usuarios,
                                                                         'perfiles_pcs': perfiles_pcs,
                                                                        'lista_empresas': lista_empresas,
                                                                        'usuarios_datos':usuarios_datos,
                                                               'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


        id = request.POST['id']
        clave = request.POST['password']
        nombres = request.POST['nombres']
        apellidos = request.POST['apellidos']
        correo = request.POST['email']
        nombre_usuario = request.POST['nombre_usuario']
        empresas = request.POST['empresas']
        telefono = request.POST['telefono']
        cargo = request.POST['cargo']
        publico = request.POST['publico']
        atencion = request.POST['atencion']
        pcs_user = request.POST['pcs_user']
        perfil_pcs = request.POST['perfil_pcs']

        usuario = User(
            id=id,
            password=clave,
            first_name=nombres,
            last_name=apellidos,
            email=correo,
            username=nombre_usuario,
            is_superuser=1,
            is_staff=1,

        )

        usuario.save()

        usuarios_datos = Usuarios_datos(

            id=id,
            usuario_id=id,
            empresa_id=int(empresas),
            perfil_pcs_id=int(perfil_pcs),
            cargo=cargo,
            pcs=pcs_user,
            telefono=telefono,
            admin=publico,
            atencion=atencion,
            creado=timezone.now(),
        )

        usuarios_datos.save()
        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado el Usuarios ' + (nombres).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/usuarios/')

# Consecutivo

def config_consecutivos(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_consecutivos = Consecutivos.objects.all()
        lista_anos=Anos.objects.all()
        lista_dependencias=Dependencias.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_usuarios == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/usuarios/')

        return render(request, "config_consecutivos.html", {'user': current_user,
                                                          'lista_consecutivos': lista_consecutivos,
                                                          'permisos': permisos,
                                                          'permiso_usuario': usuario_datos,
                                                          'lista_anos':lista_anos,
                                                          'lista_dependencias':lista_dependencias,})
    else:
        pass


def config_consecutivos_historial_carpeta(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_carpetas=CarpetaAno.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_usuarios == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/usuarios/')

        return render(request, "config_consecutivos_historial_carpeta.html", {'user': current_user,
                                                          'permisos': permisos,
                                                          'permiso_usuario': usuario_datos,
                                                          'lista_carpetas':lista_carpetas,})
    else:
        pass

def config_consecutivos_historial(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_consecutivos=Consecutivos.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_usuarios == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/usuarios/')

        return render(request, "config_consecutivos_historial.html", {'user': current_user,
                                                          'permisos': permisos,
                                                          'permiso_usuario': usuario_datos,
                                                          'lista_consecutivos':lista_consecutivos,})
    else:
        pass

def config_consecutivos_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_dependencias = Dependencias.objects.all()
        lista_tipos_radicados = Tipos_radicados.objects.all()
        lista_anos = Anos.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_usuarios == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/usuarios/')

        return render(request, "config_consecutivos_registrar.html", {'user': current_user,
                                                                   'lista_dependencias': lista_dependencias,
                                                                   'lista_tipos_radicados': lista_tipos_radicados,
                                                                   'permisos': permisos,
                                                                    'permiso_usuario': usuario_datos,
                                                                    'lista_anos':lista_anos,})
    elif request.method == 'POST':
        lista_dependencias=Dependencias.objects.all()
        tipo_consecutivo=request.POST['consecutivo_tipo']
        ano = request.POST['ano']
        usuario_origen = request.POST['id']

        if tipo_consecutivo== 'general':
            codigo = ano
            lista_consecutivos = Consecutivos.objects.all()
            lista_consecutivos = lista_consecutivos.filter(codigo=codigo)
            if lista_consecutivos.exists():
                messages.add_message(request, messages.INFO,
                                      'Registro ya existe con el codigo ' + (codigo).encode(
                                     'utf-8').strip())
                return HttpResponseRedirect('/configuracion/consecutivos/')

            consecutivos = Consecutivos(

                codigo=codigo,
                ano=ano,
                secuencia='1'.zfill(7),
                tipo='general',
                usuario_origen_id=usuario_origen,
                accion='Creacion de Consecutivo'
            )
            consecutivos.save()

            messages.add_message(request, messages.INFO,
                                 'Se ha registrado el Consecutivo ' + (codigo).encode(
                                     'utf-8').strip() + ' satisfactoriamente.')

            return HttpResponseRedirect('/configuracion/consecutivos/')

        elif tipo_consecutivo=='dependencias':
            for dependencia in lista_dependencias:
                codigo = ano+dependencia.codigo
                lista_consecutivos = Consecutivos.objects.all()
                lista_consecutivos = lista_consecutivos.filter(codigo=codigo)
                if lista_consecutivos.exists():
                    messages.add_message(request, messages.INFO,
                                         'Registro ya existe con el codigo ' + (codigo).encode(
                                             'utf-8').strip())
                    return HttpResponseRedirect('/configuracion/consecutivos/')

                consecutivos = Consecutivos(

                    codigo=codigo,
                    ano=ano,
                    secuencia='1'.zfill(7),
                    dependencias_id=dependencia.codigo,
                    tipo='dependencias',
                    usuario_origen_id=usuario_origen,
                    accion='Creacion de Consecutivo'
                )
                consecutivos.save()

            messages.add_message(request, messages.INFO,
                                    'Se ha registrado el Consecutivo ' + (codigo).encode(
                                        'utf-8').strip() + ' satisfactoriamente.')

            return HttpResponseRedirect('/configuracion/consecutivos/')

        else:
            pass


def config_consecutivos_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        consecutivos=Consecutivos.objects.get(codigo=id)
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_dependencias = Dependencias.objects.all()
        lista_tipos_radicados = Tipos_radicados.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if permisos.personas == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_consecutivos_editar.html", {'user': current_user,
                                                                   'lista_dependencias': lista_dependencias,
                                                                   'lista_tipos_radicados': lista_tipos_radicados,
                                                                   'permisos': permisos,
                                                                    'consecutivos':consecutivos})
    elif request.method == 'POST':
        codigo=request.POST['id']
        ano = request.POST['ano']
        dependencias = request.POST['dependencias']
        tipos_radicados = request.POST['tipos_radicados']


        consecutivos = Consecutivos(
            codigo=codigo,
            ano=ano,
            dependencias_id= dependencias,
            tipos_radicados_id= int(tipos_radicados),
            secuencia = 1,

        )
        consecutivos.save()

        messages.add_message(request, messages.INFO,
                             'Se ha editado el Consecutivo ' + (codigo).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/consecutivos/')

def config_consecutivos_borrar(request,id):
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        consecutivos = Consecutivos.objects.get(codigo=id)


        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_usuarios == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')
        else:
            consecutivos.delete()
            messages.add_message(request, messages.WARNING,
                                    'Se ha borrado el codigo ' + str(id) + ' satisfactoriamente')

            return HttpResponseRedirect('/configuracion/consecutivos/')


def config_consecutivos_anos(request):
    # Render  administracion.html
    if request.method == 'POST':
        lista_dependencias = Dependencias.objects.all()

        ano = request.POST['ano']
        usuario_origen = request.POST['id']
        url = os.path.join(settings.MEDIA_ROOT, ano)
        os.mkdir(url)
        anos = Anos(
            ano=ano
        )
        anos.save()
        for campo in lista_dependencias:

            url_dependencias = os.path.join(url, str(campo.codigo))
            os.mkdir(url_dependencias)
            url_docs = os.path.join(url_dependencias, 'docs')
            os.mkdir(url_docs)
            url_expediente = os.path.join(url_dependencias, 'expedientes')
            os.mkdir(url_expediente)

            carpeta = CarpetaAno(
                ano_id=ano,
                dependencias_id=campo.codigo,
                accion='creacion de carpeta',
                usuario_origen_id=usuario_origen,
            )
            carpeta.save()

        messages.add_message(request, messages.INFO,
                             'Se ha creado el año ' + (ano).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/consecutivos/')


def config_consecutivos_anos_actualizar(request):
    # Render  administracion.html
    if request.method == 'POST':
        ano = request.POST['ano']
        usuario_origen = request.POST['id']
        dependencia = request.POST['dependencia']

        url = os.path.join(settings.MEDIA_ROOT, ano)
        lista_dependencias=CarpetaAno.objects.filter(ano_id=ano,dependencias_id=dependencia)
        if lista_dependencias.exists():
            messages.add_message(request, messages.ERROR,
                                 'Ya existe un carpeta para esta dependencia y año ')

            return HttpResponseRedirect('/configuracion/consecutivos/')

        else:
            url_dependencias = os.path.join(url, dependencia)
            os.mkdir(url_dependencias)
            url_docs = os.path.join(url_dependencias, 'docs')
            os.mkdir(url_docs)

            carpeta = CarpetaAno(
                ano_id=ano,
                dependencias_id=dependencia,
                accion='actualizacion de carpeta',
                usuario_origen_id=usuario_origen,
            )
            carpeta.save()

            messages.add_message(request, messages.INFO,
                                 'Se ha actualizado la carpeta del año ' + (ano).encode(
                                     'utf-8').strip() + ' satisfactoriamente.')

            return HttpResponseRedirect('/configuracion/consecutivos/')

def config_consecutivos_actualizar(request):
    # Render  administracion.html
    if request.method == 'POST':
        ano = request.POST['ano']
        usuario_origen = request.POST['id']
        dependencia = request.POST['dependencia']

        lista_dependencias= Consecutivos.objects.filter(ano=ano,dependencias_id=dependencia)
        if lista_dependencias.exists():
            messages.add_message(request, messages.ERROR,
                                 'Ya existe un consecutivo para esta dependencia y año ')

            return HttpResponseRedirect('/configuracion/consecutivos/')

        else:

            consecutivo = Consecutivos(
                secuencia='1'.zfill(7),
                dependencias_id=dependencia,
                ano=ano,
                codigo=ano+dependencia,
                tipo='dependencias',
                usuario_origen_id=usuario_origen,
                accion='Creacion de Consecutivo',
            )
            consecutivo.save()

            messages.add_message(request, messages.INFO,
                                 'Se ha actualizado el consecutivo del año ' + (ano).encode(
                                     'utf-8').strip() + ' satisfactoriamente.')

            return HttpResponseRedirect('/configuracion/consecutivos/')


# Dependencias
def config_dependencias(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        #usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        #permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_dependencias = Dependencias.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.consultar_dependencias == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            #return HttpResponseRedirect('/administracion/')

        return render(request, "config_dependencias.html", {'user': current_user,
                                                          'lista_dependencias': lista_dependencias,
                                                          #'permisos': permisos,
                                                        #'permiso_usuario': usuario_datos
                                                            })
    else:
        pass

def config_dependencias_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        #usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        #permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_continentes = Continentes.objects.all()
        lista_municipios = Municipios.objects.all()
        lista_dependencias = Dependencias.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        #if usuario_datos.perfiles.registrar_dependencias == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            #return HttpResponseRedirect('/configuracion/dependencias/')

        return render(request, "config_dependencias_registrar.html", {'user': current_user,
                                                                   'lista_continentes': lista_continentes,
                                                                   'lista_dependencias': lista_dependencias,
                                                                   'lista_municipios': lista_municipios,
                                                                   #'permisos': permisos,
                                                                #'permiso_usuario': usuario_datos
                                                                      })
    elif request.method == 'POST':

        current_user = request.user
        #usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        #if usuario_datos.perfiles.registrar_dependencias == False:
            #messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            #return HttpResponseRedirect('/configuracion/dependencias/')

        codigo = request.POST['codigo']
        codigopadre = request.POST['codigopadre']
        codigoterritorio = request.POST['codigoterritorio']
        sigla = request.POST['sigla']
        descripcion = request.POST['descripcion']
        estado = request.POST['estado']
        continentes = request.POST['continentes']
        paises = request.POST['paises']
        departamentos = request.POST['departamentos']
        municipios = request.POST['municipios']
        direccion = request.POST['direccion']

        dependencias = Dependencias(
            codigo=codigo,
            codigopadre=codigopadre,
            codigoterritorio=codigoterritorio,
            sigla=sigla,
            descripcion=descripcion,
            estado=estado,
            continentes_id=int(continentes),
            paises_id=int(paises),
            departamentos_id=int(departamentos),
            municipios_id=int(municipios),
            direccion=direccion,
            creado=timezone.now(),
        )
        dependencias.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado la Dependencia ' + (descripcion).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/dependencias/')

def config_dependencias_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        dependencias = Dependencias.objects.get(codigo=id)
        lista_continentes = Continentes.objects.all()
        lista_paises=Paises.objects.all()
        lista_departamentos=Departamentos.objects.all()
        lista_municipios = Municipios.objects.all()
        lista_dependencias = Dependencias.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_dependencias == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/dependencias/')

        return render(request, "config_dependencias_editar.html", {'user': current_user,
                                                                         'permisos': permisos,
                                                                         'dependencias': dependencias,
                                                                        'lista_continentes': lista_continentes,
                                                                        'lista_paises':lista_paises,
                                                                        'lista_departamentos':lista_departamentos,
                                                                        'lista_dependencias': lista_dependencias,
                                                                        'lista_municipios': lista_municipios,
                                                                        'permiso_usuario': usuario_datos})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.editar_dependencias == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/dependencias/')

        codigo = request.POST['codigo']
        codigopadre = request.POST['codigopadre']
        codigoterritorio = request.POST['codigoterritorio']
        sigla = request.POST['sigla']
        descripcion = request.POST['descripcion']
        estado = request.POST['estado']
        continentes = request.POST['continentes']
        paises = request.POST['paises']
        departamentos = request.POST['departamentos']
        municipios = request.POST['municipios']
        direccion = request.POST['direccion']

        dependencias = Dependencias(
            codigo=codigo,
            codigopadre=codigopadre,
            codigoterritorio=codigoterritorio,
            sigla=sigla,
            descripcion=descripcion,
            estado=estado,
            continentes_id=int(continentes),
            paises_id=int(paises),
            departamentos_id=int(departamentos),
            municipios_id=int(municipios),
            direccion=direccion,
            modificado=timezone.now(),
        )
        dependencias.save()

        messages.add_message(request, messages.SUCCESS,
                             'Se ha editado la Dependencia ' + (descripcion).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/dependencias/')


def config_dependencias_borrar(request, id):

        if request.method == 'GET':

            dependencia = Dependencias.objects.get(pk=id)
            current_user = request.user
            usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
            if usuario_datos.perfiles.borrar_dependencias == False:
                messages.add_message(request, messages.ERROR, 'No tienes permisos para eliminar en este modulo')
                return HttpResponseRedirect('/configuracion/dependencias/')

            dependencia.delete()
            messages.add_message(request, messages.WARNING,
                                'Se ha borrado la Dependencia ' + str(id) + ' satisfactoriamente')

            return HttpResponseRedirect('/configuracion/dependencias/')








# Sub Series
def config_subseries(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_subseries = Subseries.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_subseries == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_subseries.html", {'user': current_user,
                                                          'lista_subseries': lista_subseries,
                                                          'permisos': permisos,
                                                         'permiso_usuario': usuario_datos,})
    else:
        pass


def config_subseries_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        series = Series.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_subseries == False:
            messages.add_message(request, messages.ERROR,
                                    'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/subseries/')

        return render(request, "config_subseries_registrar.html", {'user': current_user,
                                                                  'series': series,
                                                                    'permisos': permisos,
                                                                  'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.registrar_subseries == False:
            messages.add_message(request, messages.ERROR,
                                    'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/subseries/')

        codigo = request.POST['codigo']
        series = request.POST['series']
        descripcion = request.POST['descripcion']
        proceso = request.POST['proceso']
        procedimiento = request.POST['procedimiento']
        soporte = request.POST['soporte']
        retencion_ag = request.POST['retencion_ag']
        retencion_ac = request.POST['retencion_ac']
        observacion = request.POST['observacion']
        dfinal= request.POST['dfinal']

        subseries = Subseries(
            codigo=codigo,
            series_id=series,
            descripcion=descripcion,
            proceso=proceso,
            procedimiento=procedimiento,
            soporte=soporte,
            retencionag=retencion_ag,
            retencionac=retencion_ac,
            dispocicion_final=dfinal,
            observacion=observacion,
            creado=timezone.now(),

        )
        subseries.save()

        messages.add_message(request, messages.INFO,
                                 'Se ha registrado la subserie ' + (descripcion).encode(
                                     'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/subseries/')


def config_subseries_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        subseries = Subseries.objects.get(pk=id)
        lista_series=Series.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.editar_subseries == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para editar en este modulo')
            return HttpResponseRedirect('/configuracion/subseries/')

        return render(request, "config_subseries_editar.html", {'user': current_user,
                                                                         'permisos': permisos,
                                                                         'subseries': subseries,
                                                                    'permiso_usuario': usuario_datos,
                                                                    'lista_series':lista_series})
    elif request.method == 'POST':

        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        if usuario_datos.perfiles.editar_subseries == False:
            messages.add_message(request, messages.ERROR,
                                 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/subseries/')

        id=request.POST['id']
        codigo = request.POST['codigo']
        series = request.POST['series']
        descripcion = request.POST['descripcion']
        proceso = request.POST['proceso']
        procedimiento = request.POST['procedimiento']
        soporte = request.POST['soporte']
        retencion_ag = request.POST['retencion_ag']
        retencion_ac = request.POST['retencion_ac']
        observacion = request.POST['observacion']
        dfinal= request.POST['dfinal']


        subseries = Subseries(
            id=id,
            codigo=codigo,
            series_id=series,
            descripcion=descripcion,
            proceso=proceso,
            procedimiento=procedimiento,
            soporte=soporte,
            retencionag=retencion_ag,
            retencionac=retencion_ac,
            dispocicion_final=dfinal,
            observacion=observacion,
            modificado=timezone.now(),

        )
        subseries.save()

        messages.add_message(request, messages.INFO,
                             'Se ha editado la subserie ' + (descripcion).encode(
                                 'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/subseries/')

def config_trd(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        tipos_documentos=Tipos_documentos.objects.all()
        lista_dependencias = Dependencias.objects.all()
        lista_series=Series.objects.all()
        lista_subseries=Subseries.objects.all()
        lista_trd=Matriz.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_subseries == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_trd.html", {'user': current_user,
                                                          'permisos': permisos,
                                                         'permiso_usuario': usuario_datos,
                                                        'tipos_documentos':tipos_documentos,
                                                        'lista_dependencias': lista_dependencias,
                                                        'lista_series':lista_series,
                                                        'lista_subseries':lista_subseries,})
    else:
        pass

def config_trd_modificacion(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        tipos_documentos=Tipos_documentos.objects.all()
        lista_dependencias = Dependencias.objects.all()
        lista_series=Series.objects.all()
        lista_subseries=Subseries.objects.all()
        lista_trd=Matriz.objects.all()
        tipos_documentos = Tipos_documentos.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_subseries == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_trd_modificacion.html", {'user': current_user,
                                                          'permisos': permisos,
                                                         'permiso_usuario': usuario_datos,
                                                        'tipos_documentos':tipos_documentos,
                                                        'lista_dependencias': lista_dependencias,
                                                        'lista_series':lista_series,
                                                        'lista_subseries':lista_subseries,})
    elif request.method == 'POST':

        dependencias = request.POST['dependencias']
        series = request.POST['series']
        subseries = request.POST['subseries']
        tipo_documento = request.POST['tipo_documento']
        boton=request.POST['boton']

        lista_trd = Matriz.objects.filter(dependencia_id=dependencias)
        lista_trd = lista_trd.filter(serie_id=series)
        lista_trd = lista_trd.filter(subserie_id=subseries)
        lista_trd = lista_trd.filter(tipos_documentos_id=tipo_documento).first()
        if boton=='activar':
            Matriz.objects.filter(id=lista_trd.id).update(estado='activo')
        elif boton=='desactivar':
            Matriz.objects.filter(id=lista_trd.id).update(estado='inactivo')
        messages.add_message(request, messages.INFO,
                                 'Se ha editado el estado  a la matriz TRD'.encode(
                                     'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/trd/modificacion/')


def config_trd_datos(request):
    if request.method == 'GET':

        subseries = Subseries.objects.all()

        cuenta = subseries.count()
        paginador = Paginator(subseries,9)
        pagina = request.GET.get('page')

        try:
            sub = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            sub = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            sub = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        subseries_dict = []

        for subserie in subseries:
            subserie_dict = {
                'id': subserie.id,
                'descripcion':str(subserie.descripcion),
                'series': str(subserie.series_id),
                    }
            subseries_dict.append(subserie_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': subseries_dict
        }

        return JsonResponse(response_dict)


def config_trd_dependencia_datos(request):
    if request.method == 'GET':

        trd = Matriz.objects.all()
        if request.GET.get('dependencias'):
            dependencias = request.GET.get('dependencias')
            trd = trd.filter(dependencia_id=dependencias)

        if request.GET.get('series'):
            series = request.GET.get('series')
            trd = trd.filter(serie_id=series)

        if request.GET.get('subseries'):
            subseries = request.GET.get('subseries')
            trd = trd.filter(subserie_id=subseries)

        cuenta = trd.count()
        paginador = Paginator(trd,9)
        pagina = request.GET.get('page')

        try:
            matriz = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            matriz = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            matriz = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        matrices_dict = []

        for dato in matriz:
            matriz_dict = {
                'id': dato.tipos_documentos_id,
                'descripcion':str(dato.tipos_documentos.descripcion),
                'subseries':str(dato.subserie_id),
                    }
            matrices_dict.append(matriz_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': matrices_dict
        }

        return JsonResponse(response_dict)

def config_trd_r_documentos(request):

    if request.method == 'GET':

        lista_matriz = Matriz.objects.all()

        if request.GET.get('subseries'):
            subseries = request.GET.get('subseries')
            lista_matriz = lista_matriz.filter(subserie_id=subseries)


        if request.GET.get('soporte'):
            soporte = request.GET.get('soporte')
            lista_matriz = lista_matriz.filter(soporte=soporte)

        if subseries:
            if request.GET.get('dependencias'):
                dependencias = request.GET.get('dependencias')
                lista_matriz = lista_matriz.filter(dependencia_id=dependencias)
        else:
            pass

        cuenta = lista_matriz.count()
        paginador = Paginator(lista_matriz, 9)
        pagina = request.GET.get('page')

        try:
            sub = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            sub = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            sub = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        matrices_dict = []

        for matriz in lista_matriz:
            matriz_dict = {
                'id': matriz.id,
                'codigo': str(matriz.tipos_documentos_id),
                'descripcion': str(matriz.tipos_documentos.descripcion),
            }
            matrices_dict.append(matriz_dict)
        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': matrices_dict
        }

        return JsonResponse(response_dict)


    return HttpResponseRedirect('/administracion/usuarios/casos/')


def config_trd_sin_asignar(request):

    if request.method == 'GET':

        lista_matriz = Matriz.objects.all()

        if request.GET.get('subseries'):
            subseries = request.GET.get('subseries')
            lista_matriz = lista_matriz.filter(subserie_id=subseries)

        if request.GET.get('soporte'):
            soporte = request.GET.get('soporte')
            lista_matriz = lista_matriz.filter(soporte=soporte)

        if subseries:
            if request.GET.get('dependencias'):
                dependencias = request.GET.get('dependencias')
                lista_matriz = lista_matriz.filter(dependencia_id=dependencias)
        else:
            pass



        cuenta = lista_matriz.count()
        paginador = Paginator(lista_matriz, 9)
        pagina = request.GET.get('page')

        try:
            sub = paginador.page(pagina)
        except PageNotAnInteger:
            # si la pagina no es un entero
            sub = paginador.page(1)
            pagina = 1
        except EmptyPage:
            # si la pagina excede la cantidad total
            sub = paginador.page(paginador.num_pages)
            pagina = paginador.num_pages

        # crear objeto json por  cada pareja
        matrices_dict = []
        documento=[]
        for matriz in lista_matriz:
            documentos=matriz.tipos_documentos_id
            documento.append(documentos)

        lista_documento =Tipos_documentos.objects.exclude(id__in=documento)


        for documento in lista_documento:
            matriz_dict = {
                'id':documento.id,
                'descripcion': documento.descripcion,
            }
            matrices_dict.append(matriz_dict)


        # ejemplo  usando  list comprehensions
        response_dict = {
            'pagina': pagina,
            'total_items': cuenta,
            'datos': matrices_dict
        }

        return JsonResponse(response_dict)



def config_trd_registrar(request):

    if request.method == 'POST':

        dependencias = request.POST['dependencias']
        series = request.POST['series']
        subseries = request.POST['subseries']
        soporte = request.POST['soporte']
        tipo_documento = request.POST.getlist('tipo_documento[]')

        for tipo in tipo_documento:
            matriz_trd = Matriz(
                dependencia_id=dependencias,
                serie_id=series,
                subserie_id=subseries,
                soporte=soporte,
                tipos_documentos_id=tipo,
            )
            matriz_trd.save()

        messages.add_message(request, messages.INFO,
                                 'Se ha agregado los tipos de documentos a la matriz ' .encode(
                                     'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/trd/')

def config_trd_informa(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_dependencias = Dependencias.objects.all()
        lista_series=Series.objects.all()
        lista_trd=Matriz.objects.filter(estado='activo').order_by('-dependencia')



        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_admin_anulacion == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_trd_informe.html", {'user': current_user,
                                                          'permisos': permisos,
                                                         'permiso_usuario': usuario_datos,
                                                         'lista_dependencias':lista_dependencias,
                                                        'lista_series':lista_series,
                                                        'lista_trd':lista_trd,
                                                           })
    elif request.method == 'POST':
            tipo_requerimiento=request.POST['tipo_requerimiento']
            current_user = request.user
            usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
            permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
            lista_dependencias = Dependencias.objects.all()
            lista_series = Series.objects.all()
            lista_trd = Matriz.objects.filter(dependencia_id=tipo_requerimiento,estado='activo').order_by('-dependencia')


            return render(request, "config_trd_informe.html", {'user': current_user,
                                                               'permisos': permisos,
                                                               'permiso_usuario': usuario_datos,
                                                               'lista_dependencias': lista_dependencias,
                                                               'lista_series': lista_series,
                                                               'lista_trd': lista_trd,
                                                               })









def config_ubicacion(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        lista_ubicacion = Ubicacion.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_localizacion == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_ubicacion.html", {'user': current_user,
                                                          'lista_ubicacion': lista_ubicacion,
                                                          'permisos': permisos,
                                                           'permiso_usuario': usuario_datos })
    else:
        pass

def config_ubicacion_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()

        lista_continentes = Continentes.objects.all()
        lista_paises = Paises.objects.all()
        lista_departamentos = Departamentos.objects.all()
        lista_municipios = Municipios.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_terceros == False:
            messages.add_message(request, messages.ERROR,
                                    'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/terceros/')

        return render(request, "config_ubicacion_registrar.html", {'user': current_user,
                                                                  'lista_continentes': lista_continentes,
                                                                  'lista_paises': lista_paises,
                                                                  'lista_departamentos': lista_departamentos,
                                                                  'lista_municipios': lista_municipios,
                                                                    'permisos': permisos,
                                                                  'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        codigo = request.POST['codigo']
        descripcion = request.POST['descripcion']
        continentes = request.POST['continentes']
        paises = request.POST['paises']
        departamentos = request.POST['departamentos']
        municipios = request.POST['municipios']


        ubicaciones = Ubicacion(
            codigo=codigo,
            descripcion=descripcion,
            continentes_id = continentes,
            paises_id = paises,
            departamentos_id = departamentos,
            municipios_id = municipios,


        )
        ubicaciones.save()

        messages.add_message(request, messages.INFO,
                                 'Se ha registrado la ubicacion ' + (descripcion).encode(
                                     'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/ubicacion/')



def config_ubicacion_editar(request,id):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        ubicacion=Ubicacion.objects.get(pk=id)
        lista_continentes = Continentes.objects.all()
        lista_paises = Paises.objects.all()
        lista_departamentos = Departamentos.objects.all()
        lista_municipios = Municipios.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_terceros == False:
            messages.add_message(request, messages.ERROR,
                                    'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/terceros/')

        return render(request, "config_ubicacion_editar.html", {'user': current_user,
                                                                  'lista_continentes': lista_continentes,
                                                                  'lista_paises': lista_paises,
                                                                  'lista_departamentos': lista_departamentos,
                                                                  'lista_municipios': lista_municipios,
                                                                    'permisos': permisos,
                                                                  'permiso_usuario': usuario_datos,
                                                                    'ubicacion':ubicacion,})
    elif request.method == 'POST':
        id = request.POST['id']
        codigo = request.POST['codigo']
        descripcion = request.POST['descripcion']
        continentes = request.POST['continentes']
        paises = request.POST['paises']
        departamentos = request.POST['departamentos']
        municipios = request.POST['municipios']


        ubicaciones = Ubicacion(
            id=id,
            codigo=codigo,
            descripcion=descripcion,
            continentes_id = continentes,
            paises_id = paises,
            departamentos_id = departamentos,
            municipios_id = municipios,


        )
        ubicaciones.save()

        messages.add_message(request, messages.INFO,
                                 'Se ha editado la ubicacion ' + (descripcion).encode(
                                     'utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/ubicacion/')


def config_piso(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        lista_piso = Piso.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_localizacion == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_piso.html", {'user': current_user,
                                                          'lista_piso': lista_piso,
                                                          'permisos': permisos,
                                                           'permiso_usuario': usuario_datos })
    else:
        pass

def config_piso_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()

        lista_ubicacion = Ubicacion.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_localizacion == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/paises/')

        return render(request, "config_piso_registrar.html", {'user': current_user,
                                                                   'lista_ubicacion': lista_ubicacion,
                                                                'permisos': permisos,
                                                                'permiso_usuario': usuario_datos,})
    elif request.method == 'POST':

        ubicacion = request.POST['ubicacion']
        numero = request.POST['numero']


        pisos = Piso(
            numero=int(numero),
            ubicacion_id=ubicacion,
        )
        pisos.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el Piso ' + (numero).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/piso/')


def config_salones(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        lista_salones = Salones.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_localizacion == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_salones.html", {'user': current_user,
                                                          'lista_salones': lista_salones,
                                                          'permisos': permisos,
                                                           'permiso_usuario': usuario_datos })
    else:
        pass


def config_salones_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_ubicacion = Ubicacion.objects.all()
        lista_piso=Piso.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_localizacion == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/paises/')

        return render(request, "config_salones_registrar.html", {'user': current_user,
                                                                   'lista_ubicacion': lista_ubicacion,
                                                                'permisos': permisos,
                                                                'permiso_usuario': usuario_datos,
                                                                 'lista_piso':lista_piso,})
    elif request.method == 'POST':

        ubicacion = request.POST['ubicacion']
        piso = request.POST['piso']
        salon = request.POST['salon']


        salones = Salones(
            piso_id=int(piso),
            ubicacion_id=ubicacion,
            salon=salon,
        )
        salones.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el salon ' + (salon).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/salones/')


def config_estante(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        lista_estantes = Estantes.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_localizacion == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_estante.html", {'user': current_user,
                                                          'lista_estantes': lista_estantes,
                                                          'permisos': permisos,
                                                           'permiso_usuario': usuario_datos })
    else:
        pass



def config_estante_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_ubicacion = Ubicacion.objects.all()
        lista_piso=Piso.objects.all()
        lista_salon=Salones.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_localizacion == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/paises/')

        return render(request, "config_estante_registrar.html", {'user': current_user,
                                                                   'lista_ubicacion': lista_ubicacion,
                                                                'permisos': permisos,
                                                                'permiso_usuario': usuario_datos,
                                                                 'lista_piso':lista_piso,
                                                                 'lista_salon':lista_salon})
    elif request.method == 'POST':

        ubicacion = request.POST['ubicacion']
        piso = request.POST['piso']
        salon = request.POST['salon']
        estantes = request.POST['estante']


        estante = Estantes(
            piso_id=int(piso),
            ubicacion_id=ubicacion,
            salon_id=salon,
            estante=estantes,
        )
        estante.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el estante ' + (estantes).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/estante/')


def config_nivel_estante(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        lista_niveles = NivelEstantes.objects.all()

        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.consultar_localizacion == False:
            messages.add_message(request, messages.ERROR, 'No tienes permitido el acceso a ese modulo')
            return HttpResponseRedirect('/administracion/')

        return render(request, "config_nivel_estante.html", {'user': current_user,
                                                          'lista_niveles': lista_niveles,
                                                          'permisos': permisos,
                                                           'permiso_usuario': usuario_datos })
    else:
        pass


def config_nivel_estante_registrar(request):
    # Render  administracion.html
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        permisos = Permisos.objects.filter(usuario_id=current_user.id).first()
        lista_ubicacion = Ubicacion.objects.all()
        lista_piso=Piso.objects.all()
        lista_salon=Salones.objects.all()
        lista_estantes=Estantes.objects.all()
        if not current_user.is_staff:
            return HttpResponseRedirect('/login/')

        if usuario_datos.perfiles.registrar_localizacion == False:
            messages.add_message(request, messages.ERROR, 'No tienes permisos para registrar en este modulo')
            return HttpResponseRedirect('/configuracion/paises/')

        return render(request, "config_nivel_estante_registrar.html", {'user': current_user,
                                                                   'lista_ubicacion': lista_ubicacion,
                                                                'permisos': permisos,
                                                                'permiso_usuario': usuario_datos,
                                                                 'lista_piso':lista_piso,
                                                                 'lista_salon':lista_salon,
                                                                'lista_estantes':lista_estantes,})
    elif request.method == 'POST':

        ubicacion = request.POST['ubicacion']
        piso = request.POST['piso']
        salon = request.POST['salon']
        estantes = request.POST['estante']
        nivel = request.POST['nivel']


        niveles = NivelEstantes(
            piso_id=int(piso),
            ubicacion_id=ubicacion,
            salon_id=salon,
            estante_id=estantes,
            nivel=nivel,
        )
        niveles.save()

        messages.add_message(request, messages.INFO,
                             'Se ha registrado el nivel ' + (nivel).encode('utf-8').strip() + ' satisfactoriamente.')

        return HttpResponseRedirect('/configuracion/nivel_estante/')


def config_solicitud_pedido(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


        return render(request, "config_solicitud_pedido.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pagina = request.POST['pagina']

        try:
            paginador = request.POST['paginador']
            if paginador == 'atras':
                pagina = str(int(pagina) - 20)
                if int(pagina) < 0:
                    pagina = str(0)
            elif paginador == 'adelante':
                pagina = str(int(pagina) + 20)
            elif paginador == 'primera':
                pagina = str(0)
        except:
            pass
        try:
            estado = request.POST['estado']
            if estado == 'None':
                estado = None
        except:
            estado = None



        if estado== None:
            url2 = IP_SAP + "Orders?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                + empresa + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina
        elif estado=='tYES':
            url2 = IP_SAP + "Orders?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and Cancelled eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina
        elif estado=='bost_Open' or estado=='bost_Close':
            url2 = IP_SAP + "Orders?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and DocumentStatus eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'n_pedido': datos['DocNum'],
                'fecha_contabilizacion': datos['DocDate'],
                'fecha_entrega': datos['DocDueDate'],
                'total': format(datos['DocTotal'], '0,.0f') ,
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_pedido.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                'estado':estado,
                                                                'permiso_usuario': usuario_datos,
                                                                'pagina': pagina,
                                                                'pagina_fin': str(int(pagina)+20),
                                                                })



def vista_formula(request, form_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


    url2 = IP_SAP + "Orders?$select=DocNum,DocDate,DocDueDate,TaxDate,CardName&$filter=DocNum eq " + form_id
    url3 = IP_SAP + "Orders?$select=DocumentLines&$filter=DocNum eq " + form_id

    response = sap_request(url2)
    response = ast.literal_eval(response.text)
    response = response['value']
    response2 = sap_request(url3)
    response2=json.loads(response2.text)
    response2=response2['value']
    response2=response2[0]
    response2=response2['DocumentLines']
    articulos=[]
    for d in response2:
        data_articulos={
            'numero': d['ItemCode'],
            'nombre': d['ItemDescription'],
            'cantidad': d['Quantity'],
            'precio_unidad': d['UnitPrice'],
            'indicador_impuestos': d['TaxCode'],
            'codigo_unidad': d['UoMCode'],
            'total': format(d['LineTotal'], '0,.0f'),
        }
        articulos.append(data_articulos)
    dato_lista = []
    for datos in response:
        data_prueba = {
            'n_pedido': datos['DocNum'],
            'fecha_contabilizacion': datos['DocDate'],
            'fecha_entrega':datos['DocDueDate'],
            'fecha_documento':datos['TaxDate'],
            'nombre':datos['CardName'],
            'articulos':articulos
        }
        dato_lista.append(data_prueba)
    return render(request, "solicitud_detalle.html", {"form_id": form_id,
                                                       'user':current_user,
                                                       'lista_solicitud':dato_lista,
                                                      'permiso_usuario': usuario_datos,
                                                       })



def config_solicitud_entrega(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        return render(request, "config_solicitud_entrega.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                        'permiso_usuario': usuario_datos,
                                                        })

    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pagina = request.POST['pagina']

        try:
            paginador = request.POST['paginador']
            if paginador == 'atras':
                pagina = str(int(pagina) - 20)
                if int(pagina) < 0:
                    pagina = str(0)
            elif paginador == 'adelante':
                pagina = str(int(pagina) + 20)
            elif paginador == 'primera':
                pagina = str(0)
        except:
            pass
        try:
            estado = request.POST['estado']
            if estado == 'None':
                estado = None
        except:
            estado = None



        if estado== None:
            url2 = IP_SAP + "DeliveryNotes?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                + empresa + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina
        elif estado=='tYES':
            url2 = IP_SAP + "DeliveryNotes?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and Cancelled eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina
        elif estado=='bost_Open' or estado=='bost_Close':
            url2 = IP_SAP + "DeliveryNotes?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and DocumentStatus eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'n_pedido': datos['DocNum'],
                'fecha_contabilizacion': datos['DocDate'],
                'fecha_entrega': datos['DocDueDate'],
                'total': format(datos['DocTotal'], '0,.0f') ,
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_entrega.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                'estado': estado,
                                                                'permiso_usuario': usuario_datos,
                                                                'pagina': pagina,
                                                                'pagina_fin': str(int(pagina) + 20),
                                                                })



def entrega_detalle(request, form_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


    url2 = IP_SAP + "DeliveryNotes?$select=DocNum,DocDate,DocDueDate,TaxDate,CardName&$filter=DocNum eq " + form_id
    url3 = IP_SAP + "DeliveryNotes?$select=DocumentLines&$filter=DocNum eq " + form_id

    response = sap_request(url2)
    response = ast.literal_eval(response.text)
    response = response['value']
    response2 = sap_request(url3)
    response2 = json.loads(response2.text)
    response2 = response2['value']
    response2 = response2[0]
    response2 = response2['DocumentLines']
    articulos = []
    for d in response2:
        data_articulos = {
            'numero': d['ItemCode'],
            'nombre': d['ItemDescription'],
            'cantidad': d['Quantity'],
            'precio_unidad': d['UnitPrice'],
            'indicador_impuestos': d['TaxCode'],
            'codigo_unidad': d['UoMCode'],
            'total': format(d['LineTotal'], '0,.0f'),
        }
        articulos.append(data_articulos)
    dato_lista = []
    for datos in response:
        data_prueba = {
            'n_pedido': datos['DocNum'],
            'fecha_contabilizacion': datos['DocDate'],
            'fecha_entrega': datos['DocDueDate'],
            'fecha_documento': datos['TaxDate'],
            'nombre': datos['CardName'],
            'articulos': articulos
        }
        dato_lista.append(data_prueba)
    return render(request, "entrega_detalle.html", {"form_id": form_id,
                                                      'user': current_user,
                                                      'lista_solicitud': dato_lista,
                                                    'permiso_usuario': usuario_datos,
                                                      })


def config_solicitud_devolucion(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        return render(request, "config_solicitud_devolucion.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                                    'permiso_usuario': usuario_datos,
                                                        })

    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pagina = request.POST['pagina']

        try:
            paginador = request.POST['paginador']
            if paginador == 'atras':
                pagina = str(int(pagina) - 20)
                if int(pagina) < 0:
                    pagina = str(0)
            elif paginador == 'adelante':
                pagina = str(int(pagina) + 20)
            elif paginador == 'primera':
                pagina = str(0)
        except:
            pass
        try:
            estado = request.POST['estado']
            if estado == 'None':
                estado = None
        except:
            estado = None



        if estado== None:
            url2 = IP_SAP + "Returns?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                + empresa + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina
        elif estado=='tYES':
            url2 = IP_SAP + "Returns?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and Cancelled eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina
        elif estado=='bost_Open' or estado=='bost_Close':
            url2 = IP_SAP + "Returns?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and DocumentStatus eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'n_pedido': datos['DocNum'],
                'fecha_contabilizacion': datos['DocDate'],
                'fecha_entrega': datos['DocDueDate'],
                'total': format(datos['DocTotal'], '0,.0f') ,
            }
            dato_lista.append(data_prueba)

        return render(request, "config_solicitud_devolucion.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                'estado': estado,
                                                                'permiso_usuario': usuario_datos,
                                                                'pagina': pagina,
                                                                'pagina_fin': str(int(pagina) + 20),
                                                                })


def devolucion_detalle(request, form_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


    url2 = IP_SAP + "Returns?$select=DocNum,DocDate,DocDueDate,TaxDate,CardName&$filter=DocNum eq " + form_id
    url3 = IP_SAP + "Returns?$select=DocumentLines&$filter=DocNum eq " + form_id

    response = sap_request(url2)
    response = ast.literal_eval(response.text)
    response = response['value']
    response2 = sap_request(url3)
    response2 = json.loads(response2.text)
    response2 = response2['value']
    response2 = response2[0]
    response2 = response2['DocumentLines']
    articulos = []
    for d in response2:
        data_articulos = {
            'numero': d['ItemCode'],
            'nombre': d['ItemDescription'],
            'cantidad': d['Quantity'],
            'precio_unidad': d['UnitPrice'],
            'indicador_impuestos': d['TaxCode'],
            'codigo_unidad': d['UoMCode'],
            'total': format(d['LineTotal'], '0,.0f'),
        }
        articulos.append(data_articulos)
    dato_lista = []
    for datos in response:
        data_prueba = {
            'n_pedido': datos['DocNum'],
            'fecha_contabilizacion': datos['DocDate'],
            'fecha_entrega': datos['DocDueDate'],
            'fecha_documento': datos['TaxDate'],
            'nombre': datos['CardName'],
            'articulos': articulos
        }
        dato_lista.append(data_prueba)
    return render(request, "devolucion_detalle.html", {"form_id": form_id,
                                                    'user': current_user,
                                                    'lista_solicitud': dato_lista,
                                                       'permiso_usuario': usuario_datos,
                                                    })


def config_solicitud_notas_debito(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


        return render(request, "config_solicitud_notas_debito.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                                      'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pagina = request.POST['pagina']

        try:
            paginador = request.POST['paginador']
            if paginador == 'atras':
                pagina = str(int(pagina) - 20)
                if int(pagina) < 0:
                    pagina = str(0)
            elif paginador == 'adelante':
                pagina = str(int(pagina) + 20)
            elif paginador == 'primera':
                pagina = str(0)
        except:
            pass
        try:
            estado = request.POST['estado']
            if estado == 'None':
                estado = None
        except:
            estado = None



        if estado== None:
            url2 = IP_SAP + "Invoices?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                + empresa + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'and DocumentSubType eq 'bod_DebitMemo'&$skip="+pagina
        elif estado=='tYES':
            url2 = IP_SAP + "Invoices?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and Cancelled eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'and DocumentSubType eq 'bod_DebitMemo'&$skip="+pagina
        elif estado=='bost_Open' or estado=='bost_Close':
            url2 = IP_SAP + "Invoices?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and DocumentStatus eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'and DocumentSubType eq 'bod_DebitMemo'&$skip="+pagina

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'n_pedido': datos['DocNum'],
                'fecha_contabilizacion': datos['DocDate'],
                'fecha_entrega': datos['DocDueDate'],
                'total': format(datos['DocTotal'], '0,.0f') ,
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_notas_debito.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                'estado': estado,
                                                                'permiso_usuario': usuario_datos,
                                                                'pagina': pagina,
                                                                'pagina_fin': str(int(pagina) + 20),
                                                                })




def solicitud_notas_debito_detalle(request, form_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


    url2 = IP_SAP + "Invoices?$select=DocNum,DocDate,DocDueDate,TaxDate,CardName&$filter=DocNum eq " + form_id + "and DocumentSubType eq 'bod_DebitMemo'"
    url3 = IP_SAP + "Invoices?$select=DocumentLines&$filter=DocNum eq " + form_id + "and DocumentSubType eq 'bod_DebitMemo'"

    response = sap_request(url2)
    response = ast.literal_eval(response.text)
    response = response['value']
    response2 = sap_request(url3)
    response2 = json.loads(response2.text)
    response2 = response2['value']
    response2 = response2[0]
    response2 = response2['DocumentLines']
    articulos = []
    for d in response2:
        data_articulos = {
            'numero': d['ItemCode'],
            'nombre': d['ItemDescription'],
            'cantidad': d['Quantity'],
            'precio_unidad': d['UnitPrice'],
            'indicador_impuestos': d['TaxCode'],
            'codigo_unidad': d['UoMCode'],
            'total': format(d['LineTotal'], '0,.0f'),
        }
        articulos.append(data_articulos)
    dato_lista = []
    for datos in response:
        data_prueba = {
            'n_pedido': datos['DocNum'],
            'fecha_contabilizacion': datos['DocDate'],
            'fecha_entrega': datos['DocDueDate'],
            'fecha_documento': datos['TaxDate'],
            'nombre': datos['CardName'],
            'articulos': articulos
        }
        dato_lista.append(data_prueba)
    url = IP_SAP + "Logout"
    responselogout = requests.request("POST", url, verify=False)
    return render(request, "notas_debito_detalle.html", {"form_id": form_id,
                                                       'user': current_user,
                                                       'lista_solicitud': dato_lista,
                                                         'permiso_usuario': usuario_datos,
                                                       })


def config_solicitud_notas_credito(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        return render(request, "config_solicitud_notas_credito.html", {'user': current_user,
                                                        'lista_formulas': formulas,

                                                                       'permiso_usuario': usuario_datos,
                                                        })

    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pagina = request.POST['pagina']

        try:
            paginador = request.POST['paginador']
            if paginador == 'atras':
                pagina = str(int(pagina) - 20)
                if int(pagina) < 0:
                    pagina = str(0)
            elif paginador == 'adelante':
                pagina = str(int(pagina) + 20)
            elif paginador == 'primera':
                pagina = str(0)
        except:
            pass
        try:
            estado = request.POST['estado']
            if estado == 'None':
                estado = None
        except:
            estado = None



        if estado== None:
            url2 = IP_SAP + "CreditNotes?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                + empresa + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina
        elif estado=='tYES':
            url2 = IP_SAP + "CreditNotes?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and Cancelled eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina
        elif estado=='bost_Open' or estado=='bost_Close':
            url2 = IP_SAP + "CreditNotes?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and DocumentStatus eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'n_pedido': datos['DocNum'],
                'fecha_contabilizacion': datos['DocDate'],
                'fecha_entrega': datos['DocDueDate'],
                'total': format(datos['DocTotal'], '0,.0f') ,
            }
            dato_lista.append(data_prueba)

        return render(request, "config_solicitud_notas_credito.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                'estado': estado,
                                                                'permiso_usuario': usuario_datos,
                                                                'pagina': pagina,
                                                                'pagina_fin': str(int(pagina) + 20),
                                                                })



def solicitud_notas_credito_detalle(request, form_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()



    url2 = IP_SAP + "CreditNotes?$select=DocNum,DocDate,DocDueDate,TaxDate,CardName&$filter=DocNum eq " + form_id
    url3 = IP_SAP + "CreditNotes?$select=DocumentLines&$filter=DocNum eq " + form_id

    response = sap_request(url2)
    response = ast.literal_eval(response.text)
    response = response['value']
    response2 = sap_request(url3)
    response2 = json.loads(response2.text)
    response2 = response2['value']
    response2 = response2[0]
    response2 = response2['DocumentLines']
    articulos = []
    for d in response2:
        data_articulos = {
            'numero': d['ItemCode'],
            'nombre': d['ItemDescription'],
            'cantidad': d['Quantity'],
            'precio_unidad': d['UnitPrice'],
            'indicador_impuestos': d['TaxCode'],
            'codigo_unidad': d['UoMCode'],
            'total': format(d['LineTotal'], '0,.0f'),
        }
        articulos.append(data_articulos)
    dato_lista = []
    for datos in response:
        data_prueba = {
            'n_pedido': datos['DocNum'],
            'fecha_contabilizacion': datos['DocDate'],
            'fecha_entrega': datos['DocDueDate'],
            'fecha_documento': datos['TaxDate'],
            'nombre': datos['CardName'],
            'articulos': articulos
        }
        dato_lista.append(data_prueba)
    return render(request, "notas_credito_detalle.html", {"form_id": form_id,
                                                       'user': current_user,
                                                       'lista_solicitud': dato_lista,
                                                          'permiso_usuario': usuario_datos,
                                                       })



def config_solicitud_factura_deudores(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        return render(request, "config_solicitud_factura_deudores.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                                          'permiso_usuario': usuario_datos,
                                                        })

    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pagina = request.POST['pagina']

        try:
            paginador = request.POST['paginador']
            if paginador == 'atras':
                pagina = str(int(pagina) - 20)
                if int(pagina) < 0:
                    pagina = str(0)
            elif paginador == 'adelante':
                pagina = str(int(pagina) + 20)
            elif paginador == 'primera':
                pagina = str(0)
        except:
            pass
        try:
            estado = request.POST['estado']
            if estado == 'None':
                estado = None
        except:
            estado = None



        if estado== None:
            url2 = IP_SAP + "Invoices?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                + empresa + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'and DocumentSubType eq 'bod_None'&$skip="+pagina
        elif estado=='tYES':
            url2 = IP_SAP + "Invoices?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and Cancelled eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'and DocumentSubType eq 'bod_None'&$skip="+pagina
        elif estado=='bost_Open' or estado=='bost_Close':
            url2 = IP_SAP + "Invoices?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and DocumentStatus eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'and DocumentSubType eq 'bod_None'&$skip="+pagina

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'n_pedido': datos['DocNum'],
                'fecha_contabilizacion': datos['DocDate'],
                'fecha_entrega': datos['DocDueDate'],
                'total': format(datos['DocTotal'], '0,.0f') ,
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_factura_deudores.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                'estado':estado,
                                                                'permiso_usuario': usuario_datos,
                                                                'pagina': pagina,
                                                                'pagina_fin': str(int(pagina)+20),
                                                                })




def factura_deudores_detalle(request, form_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


    url2 = IP_SAP + "Invoices?$select=DocNum,DocDate,DocDueDate,TaxDate,CardName&$filter=DocNum eq " + form_id
    url3 = IP_SAP + "Invoices?$select=DocumentLines&$filter=DocNum eq " + form_id

    response = sap_request(url2)
    response = ast.literal_eval(response.text)
    response = response['value']
    response2 = sap_request(url3)
    response2 = json.loads(response2.text)
    response2 = response2['value']
    response2 = response2[0]
    response2 = response2['DocumentLines']
    articulos = []
    for d in response2:
        data_articulos = {
            'numero': d['ItemCode'],
            'nombre': d['ItemDescription'],
            'cantidad': d['Quantity'],
            'precio_unidad': d['UnitPrice'],
            'indicador_impuestos': d['TaxCode'],
            'codigo_unidad': d['UoMCode'],
            'total': format(d['LineTotal'], '0,.0f'),
        }
        articulos.append(data_articulos)
    dato_lista = []
    for datos in response:
        data_prueba = {
            'n_pedido': datos['DocNum'],
            'fecha_contabilizacion': datos['DocDate'],
            'fecha_entrega': datos['DocDueDate'],
            'fecha_documento': datos['TaxDate'],
            'nombre': datos['CardName'],
            'articulos': articulos,
        }
        dato_lista.append(data_prueba)
    return render(request, "factura_deudores_detalle.html", {"form_id": form_id,
                                                          'user': current_user,
                                                          'lista_solicitud': dato_lista,
                                                             'permiso_usuario': usuario_datos,
                                                          })


def factura_deudores_proveedor(request, form_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
    datos_proveedor = Empresas.objects.filter(id=usuario_datos.empresa_id).first()


    url2 = IP_SAP + "PurchaseInvoices?$select=DocNum,DocTotal,VatSum,WTApplied,TotalDiscount,BaseAmount,NumAtCard,DocDate,DocDueDate,TaxDate,CardName&$filter=DocNum eq " + form_id
    url3 = IP_SAP + "PurchaseInvoices?$select=DocumentLines&$filter=DocNum eq " + form_id

    response = sap_request(url2)
    response = ast.literal_eval(response.text)
    response = response['value']
    response2 = sap_request(url3)
    response2 = json.loads(response2.text)
    response2 = response2['value']
    response2 = response2[0]
    response2 = response2['DocumentLines']
    articulos = []
    for d in response2:
        data_articulos = {
            'numero': d['ItemCode'],
            'nombre': d['ItemDescription'],
            'cantidad': d['Quantity'],
            'precio_unidad': d['UnitPrice'],
            'indicador_impuestos': d['TaxCode'],
            'codigo_unidad': d['UoMCode'],
            'total': format(d['LineTotal'], '0,.0f'),
        }
        articulos.append(data_articulos)
    dato_lista = []
    for datos in response:
        data_prueba = {
            'n_pedido': datos['DocNum'],
            'fecha_contabilizacion': datos['DocDate'],
            'fecha_entrega': datos['DocDueDate'],
            'fecha_documento': datos['TaxDate'],
            'nombre': datos['CardName'],
            'referencia': datos['NumAtCard'],
            'subtotal': format(datos['BaseAmount'], '0,.0f'),
            'descuento': format(datos['TotalDiscount'], '0,.0f'),
            'impuestos': format(datos['VatSum'], '0,.0f'),
            'retenciones': format(datos['WTApplied'], '0,.0f'),
            'total': format(datos['DocTotal'], '0,.0f'),
            'articulos': articulos,
        }
        dato_lista.append(data_prueba)
    return render(request, "factura_proveedor_detalle.html", {"form_id": form_id,
                                                          'user': current_user,
                                                          'lista_solicitud': dato_lista,
                                                              'datos_empresa': datos_proveedor,
                                                             'permiso_usuario': usuario_datos,
                                                          })

def config_solicitud_pedido_orden(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.nombre
        pagina=str(0)
        estado='bost_Open'
	usuario_actual = request.user
        if not usuario_actual.is_staff:
            return HttpResponseRedirect('/login/')

        url2 = IP_SAP + "PurchaseOrders?$orderby=DocDate desc&$select=DocNum,DocEntry,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
               + empresa + "'and DocumentStatus eq '" + estado + "'&$skip=" + pagina

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        condicion_facturar_sin_novedades = Q(peticion__descripcion='Facturar pedido sin novedades')
        condicion_facturar_con_novedades = Q(peticion__descripcion='Facturar pedido con novedades')
        condiciones = condicion_facturar_sin_novedades | condicion_facturar_con_novedades
        dato_lista = []
        for datos in response:
            facturas_respondidas=RespuestaPedido.objects.filter(num_pedido=datos['DocNum'],estado='respondido').filter(condiciones)
            facturas_respondidas = facturas_respondidas.exists()
            facturas_no_respondidas=RespuestaPedido.objects.filter(num_pedido=datos['DocNum'],estado='pendiente').filter(condiciones)
            facturas_no_respondidas = facturas_no_respondidas.exists()
            if facturas_respondidas==True:
                facturas_respondidas=RespuestaPedido.objects.filter(num_pedido=datos['DocNum'],estado='respondido').filter(condiciones).first()
                documento_facturas_respondidas=facturas_respondidas.doc_respuesta
            else:
                documento_facturas_respondidas=None
            if facturas_no_respondidas==True:
                facturas_no_respondidas='True'
            else:
                facturas_no_respondidas=None

            fecha_contabilizacion = datetime.strptime(datos['DocDate'], '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d')
            fecha_entrega = datetime.strptime(datos['DocDueDate'], '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d')
            data_prueba = {
                'dato_interno': datos['DocEntry'],
                'n_pedido': datos['DocNum'],
                'fecha_contabilizacion': fecha_contabilizacion,
                'fecha_entrega': fecha_entrega,
                'total': format(datos['DocTotal'], '0,.0f'),
                'documento_facturas_respondidas':documento_facturas_respondidas,
                'facturas_no_respondidas':facturas_no_respondidas,
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_pedido_orden.html", {'user': current_user,
                                                        'lista_prueba': dato_lista,
                                                        'lista_formulas': formulas,
                                                        'estado':estado,
                                                        'permiso_usuario': usuario_datos,
                                                        'pagina': pagina,
                                                        'pagina_fin': str(int(pagina)+20),
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pagina = request.POST['pagina']
        pedido = request.POST.get('pedido', '')

        try:
            paginador = request.POST['paginador']
            if paginador=='atras':
                pagina=str(int(pagina)-20)
                if int(pagina)<0:
                    pagina=str(0)
            elif paginador=='adelante':
                pagina=str(int(pagina)+20)
            elif paginador=='primera':
                pagina=str(0)
        except:
            pass
        try :
            estado = request.POST['estado']
            if estado=='None':
                estado = None
        except:
            estado = None
        try:
            secundario=request.POST['secundarios']
        except:
            secundario='primario'

        if pedido!="":
            url2 = IP_SAP + "PurchaseOrders?$orderby=DocDate desc&$select=DocNum,DocEntry,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and DocNum eq " + pedido + "&$skip=" + pagina
        elif secundario=='primario':
            url2 = IP_SAP + "PurchaseOrders?$orderby=DocDate desc&$select=DocNum,DocEntry,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and DocumentStatus eq '" + estado + "'&$skip=" + pagina
        elif estado== None:
            url2 = IP_SAP + "PurchaseOrders?$orderby=DocDate desc&$select=DocNum,DocEntry,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                + empresa + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina
        elif estado=='tYES':
            url2 = IP_SAP + "PurchaseOrders?$orderby=DocDate desc&$select=DocNum,DocEntry,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and Cancelled eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina
        elif estado=='bost_Open' or estado=='bost_Close':
            url2 = IP_SAP + "PurchaseOrders?$orderby=DocDate desc&$select=DocNum,DocEntry,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and DocumentStatus eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        condicion_facturar_sin_novedades = Q(peticion__descripcion='Facturar pedido sin novedades')
        condicion_facturar_con_novedades = Q(peticion__descripcion='Facturar pedido con novedades')
        condiciones = condicion_facturar_sin_novedades | condicion_facturar_con_novedades
        dato_lista = []
        for datos in response:
            facturas_respondidas = RespuestaPedido.objects.filter(num_pedido=datos['DocNum'],
                                                                  estado='respondido').filter(condiciones)
            facturas_respondidas = facturas_respondidas.exists()
            facturas_no_respondidas = RespuestaPedido.objects.filter(num_pedido=datos['DocNum'],
                                                                     estado='pendiente').filter(condiciones)
            facturas_no_respondidas = facturas_no_respondidas.exists()
            if facturas_respondidas==True:
                facturas_respondidas=RespuestaPedido.objects.filter(num_pedido=datos['DocNum'],estado='respondido').filter(condiciones).first()
                documento_facturas_respondidas=facturas_respondidas.doc_respuesta
            else:
                documento_facturas_respondidas=None
            if facturas_no_respondidas==True:
                facturas_no_respondidas='True'
            else:
                facturas_no_respondidas=None
            fecha_contabilizacion = datetime.strptime(datos['DocDate'], '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d')
            fecha_entrega = datetime.strptime(datos['DocDueDate'], '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d')
            data_prueba = {
                'dato_interno': datos['DocEntry'],
                'n_pedido': datos['DocNum'],
                'fecha_contabilizacion': fecha_contabilizacion,
                'fecha_entrega': fecha_entrega,
                'total': format(datos['DocTotal'], '0,.0f') ,
                'documento_facturas_respondidas': documento_facturas_respondidas,
                'facturas_no_respondidas': facturas_no_respondidas,
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_pedido_orden.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                'estado':estado,
                                                                'pedido':pedido,
                                                                'permiso_usuario': usuario_datos,
                                                                'pagina': pagina,
                                                                'pagina_fin': str(int(pagina)+20),
                                                                })







def config_solicitud_pedido_orden_bodegas(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.nombre
        empresas_clientes=Empresas.objects.filter(tipo='Cadena')
        pagina=str(0)
        estado='O'
        if usuario_datos.atencion == 5:
            bodega='19'
        else:
            bodega='22'
	usuario_actual = request.user
        if not usuario_actual.is_staff:
            return HttpResponseRedirect('/login/')

        url2 = IP_SAP + "SQLQueries('Consultasbodegaspedidos1')/List?estado='" \
               + estado + "'&bodega=" + bodega+ "&$skip=" + pagina

        response = sap_request(url2)
        response = response.text
        response = response.replace('null', ' " " ')
        response = ast.literal_eval(response)
        response = response['value']
        condicion_facturar_sin_novedades = Q(peticion__descripcion='Facturar pedido sin novedades')
        condicion_facturar_con_novedades = Q(peticion__descripcion='Facturar pedido con novedades')
        condiciones = condicion_facturar_sin_novedades | condicion_facturar_con_novedades
        dato_lista = []
        for datos in response:
            facturas_respondidas=RespuestaPedido.objects.filter(num_pedido=datos['DocNum'],estado='respondido').filter(condiciones)
            facturas_respondidas = facturas_respondidas.exists()
            facturas_no_respondidas=RespuestaPedido.objects.filter(num_pedido=datos['DocNum'],estado='pendiente').filter(condiciones)
            facturas_no_respondidas = facturas_no_respondidas.exists()
            if facturas_respondidas==True:
                facturas_respondidas=RespuestaPedido.objects.filter(num_pedido=datos['DocNum'],estado='respondido').filter(condiciones).first()
                documento_facturas_respondidas=facturas_respondidas.doc_respuesta
            else:
                documento_facturas_respondidas=None
            if facturas_no_respondidas==True:
                facturas_no_respondidas='True'
            else:
                facturas_no_respondidas=None

            fecha_contabilizacion = datetime.strptime(datos['DocDate'], '%Y%m%d').strftime('%Y-%m-%d')
            fecha_entrega = datetime.strptime(datos['DocDueDate'], '%Y%m%d').strftime('%Y-%m-%d')
            data_prueba = {
                'dato_interno': datos['DocEntry'],
                'n_pedido': datos['DocNum'],
                'cliente': datos['CardName'],
                'n_cliente': datos['NumAtCard'],
                'fecha_contabilizacion': fecha_contabilizacion,
                'fecha_entrega': fecha_entrega,
                'total': format(datos['DocTotal'], '0,.0f'),
                'documento_facturas_respondidas':documento_facturas_respondidas,
                'facturas_no_respondidas':facturas_no_respondidas,
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_pedido_orden_bodega.html", {'user': current_user,
                                                        'lista_prueba': dato_lista,
                                                        'lista_formulas': formulas,
                                                        'estado':estado,
                                                        'permiso_usuario': usuario_datos,
                                                        'pagina': pagina,
                                                        'empresas_clientes': empresas_clientes,
                                                        'pagina_fin': str(int(pagina)+20),
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresas_clientes = Empresas.objects.filter(tipo='Cadena')
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pagina = request.POST['pagina']
        pedido = request.POST.get('pedido', '')
        cliente = request.POST.get('cliente', '')
        estado = request.POST.get('estado', '')
        if usuario_datos.atencion == 5:
            bodega='19'
        else:
            bodega='22'

        try:
            paginador = request.POST['paginador']
            if paginador=='atras':
                pagina=str(int(pagina)-20)
                if int(pagina)<0:
                    pagina=str(0)
            elif paginador=='adelante':
                pagina=str(int(pagina)+20)
            elif paginador=='primera':
                pagina=str(0)
        except:
            pass



        if pedido!="":
            url2 = IP_SAP + "SQLQueries('Consultasbodegaspedidos2s')/List?estado='" \
                   + estado + "'&num_pedido='" + pedido + "'&$skip=" + pagina
        elif cliente!="":
            url2 = IP_SAP + "SQLQueries('Consultasbodegaspedidos4')/List?estado='" \
                   + estado + "'&bodega=" + bodega+ "&cliente='" + cliente+ "'&fecha_minima='" + fecha_inicio+ "'&fecha_maxima='" + fecha_fin + "'&$skip=" + pagina
        else:
            url2 = IP_SAP + "SQLQueries('Consultasbodegaspedidos5')/List?estado='" \
                   + estado + "'&bodega=" + bodega  + "&fecha_minima='" + fecha_inicio + "'&fecha_maxima='" + fecha_fin + "'&$skip=" + pagina

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        condicion_facturar_sin_novedades = Q(peticion__descripcion='Facturar pedido sin novedades')
        condicion_facturar_con_novedades = Q(peticion__descripcion='Facturar pedido con novedades')
        condiciones = condicion_facturar_sin_novedades | condicion_facturar_con_novedades
        dato_lista = []
        for datos in response:
            facturas_respondidas = RespuestaPedido.objects.filter(num_pedido=datos['DocNum'],
                                                                  estado='respondido').filter(condiciones)
            facturas_respondidas = facturas_respondidas.exists()
            facturas_no_respondidas = RespuestaPedido.objects.filter(num_pedido=datos['DocNum'],
                                                                     estado='pendiente').filter(condiciones)
            facturas_no_respondidas = facturas_no_respondidas.exists()
            if facturas_respondidas==True:
                facturas_respondidas=RespuestaPedido.objects.filter(num_pedido=datos['DocNum'],estado='respondido').filter(condiciones).first()
                documento_facturas_respondidas=facturas_respondidas.doc_respuesta
            else:
                documento_facturas_respondidas=None
            if facturas_no_respondidas==True:
                facturas_no_respondidas='True'
            else:
                facturas_no_respondidas=None
            fecha_contabilizacion = datetime.strptime(datos['DocDate'], '%Y%m%d').strftime('%Y-%m-%d')
            fecha_entrega = datetime.strptime(datos['DocDueDate'], '%Y%m%d').strftime('%Y-%m-%d')
            data_prueba = {
                'dato_interno': datos['DocEntry'],
                'n_pedido': datos['DocNum'],
                'cliente': datos['CardName'],
                'n_cliente': datos['NumAtCard'],
                'fecha_contabilizacion': fecha_contabilizacion,
                'fecha_entrega': fecha_entrega,
                'total': format(datos['DocTotal'], '0,.0f') ,
                'documento_facturas_respondidas': documento_facturas_respondidas,
                'facturas_no_respondidas': facturas_no_respondidas,
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_pedido_orden_bodega.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                'estado':estado,
                                                                'cliente':cliente,
                                                                'empresas_clientes':empresas_clientes,
                                                                'pedido':pedido,
                                                                'permiso_usuario': usuario_datos,
                                                                'pagina': pagina,
                                                                'pagina_fin': str(int(pagina)+20),
                                                                })






def config_solicitud_pedido_orden_bodega(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.nombre
        empresas_clientes=Empresas.objects.filter(tipo='Cadena')
        pagina=str(0)
        estado='O'
        if usuario_datos.atencion == 5:
            bodega='19'
        else:
            bodega='22'
	usuario_actual = request.user
        if not usuario_actual.is_staff:
            return HttpResponseRedirect('/login/')

        url2 = IP_SAP + "SQLQueries('Consultasbodegaspedidosven1')/List?estado='" \
               + estado + "'&bodega=" + bodega+ "&$skip=" + pagina
        response = sap_request(url2)
        response = response.text
        response = response.replace('null', ' " " ')
        response = ast.literal_eval(response)
        response = response['value']
        dato_lista = []
        for datos in response:
            facturas_respondidas=OrdenVenta.objects.filter(pedido=datos['DocNum'],estado='respondido')
            facturas_respondidas = facturas_respondidas.exists()
            facturas_no_respondidas=OrdenVenta.objects.filter(pedido=datos['DocNum'],estado='pendiente')
            facturas_no_respondidas = facturas_no_respondidas.exists()
            if facturas_respondidas==True:
                facturas_respondidas=OrdenVenta.objects.filter(pedido=datos['DocNum'],estado='respondido')
                documento_facturas_respondidas=[]
                documento_edi_respondidas=[]
                for datos_lista in facturas_respondidas:
                    documento_facturas_respondidas.append(str(datos_lista.doc_respuesta))
                    documento_edi_respondidas.append(str(datos_lista.doc_edi))
            else:
                documento_facturas_respondidas=None
                documento_edi_respondidas=None
            if facturas_no_respondidas==True:
                facturas_no_respondidas='True'
            else:
                facturas_no_respondidas=None

            fecha_contabilizacion = datetime.strptime(datos['DocDate'], '%Y%m%d').strftime('%Y-%m-%d')
            fecha_entrega = datetime.strptime(datos['DocDueDate'], '%Y%m%d').strftime('%Y-%m-%d')
            data_prueba = {
                'dato_interno': datos['DocEntry'],
                'n_pedido': datos['DocNum'],
                'cliente': datos['CardName'],
                'n_cliente': datos['NumAtCard'],
                'fecha_contabilizacion': fecha_contabilizacion,
                'fecha_entrega': fecha_entrega,
                'total': format(datos['DocTotal'], '0,.0f'),
                'documento_facturas_respondidas':documento_facturas_respondidas,
                'documento_edi_respondidas':documento_edi_respondidas,
                'facturas_no_respondidas':facturas_no_respondidas,
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_pedido_orden_bodega.html", {'user': current_user,
                                                        'lista_prueba': dato_lista,
                                                        'lista_formulas': formulas,
                                                        'estado':estado,
                                                        'permiso_usuario': usuario_datos,
                                                        'pagina': pagina,
                                                        'empresas_clientes': empresas_clientes,
                                                        'pagina_fin': str(int(pagina)+20),
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresas_clientes = Empresas.objects.filter(tipo='Cadena')
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pagina = request.POST['pagina']
        pedido = request.POST.get('pedido', '')
        cliente = request.POST.get('cliente', '')
        estado = request.POST.get('estado', '')
        pedido_cliente = request.POST.get('pedido_cliente', '')
        if usuario_datos.atencion == 5:
            bodega='19'
        else:
            bodega='22'

        try:
            paginador = request.POST['paginador']
            if paginador=='atras':
                pagina=str(int(pagina)-20)
                if int(pagina)<0:
                    pagina=str(0)
            elif paginador=='adelante':
                pagina=str(int(pagina)+20)
            elif paginador=='primera':
                pagina=str(0)
        except:
            pass



        if pedido!="":
            url2 = IP_SAP + "SQLQueries('Consultasbodegaspedidos2a')/List?estado='" \
                   + estado + "'&num_pedido='" + pedido + "'&$skip=" + pagina
        elif pedido_cliente!="":
            url2 = IP_SAP + "SQLQueries('Consultasbodegaspedidosclientes')/List?num_pedido='" + pedido_cliente + "'&$skip=" + pagina
        elif cliente!="":
            url2 = IP_SAP + "SQLQueries('Consultasbodegaspedidos4a')/List?estado='" \
                   + estado + "'&bodega=" + bodega+ "&cliente='" + cliente+ "'&fecha_minima='" + fecha_inicio+ "'&fecha_maxima='" + fecha_fin + "'&$skip=" + pagina
        else:
            url2 = IP_SAP + "SQLQueries('Consultasbodegaspedidos5a')/List?estado='" \
                   + estado + "'&bodega=" + bodega  + "&fecha_minima='" + fecha_inicio + "'&fecha_maxima='" + fecha_fin + "'&$skip=" + pagina

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        condicion_facturar_sin_novedades = Q(peticion__descripcion='Facturar pedido sin novedades')
        condicion_facturar_con_novedades = Q(peticion__descripcion='Facturar pedido con novedades')
        condiciones = condicion_facturar_sin_novedades | condicion_facturar_con_novedades
        dato_lista = []
        for datos in response:
            facturas_respondidas = OrdenVenta.objects.filter(pedido=datos['DocNum'], estado='respondido')
            facturas_respondidas = facturas_respondidas.exists()
            facturas_no_respondidas = OrdenVenta.objects.filter(pedido=datos['DocNum'], estado='pendiente')
            facturas_no_respondidas = facturas_no_respondidas.exists()
            if facturas_respondidas == True:
                facturas_respondidas = OrdenVenta.objects.filter(pedido=datos['DocNum'], estado='respondido')
                documento_facturas_respondidas = []
                documento_edi_respondidas = []
                for datos_lista in facturas_respondidas:
                    documento_facturas_respondidas.append(str(datos_lista.doc_respuesta))
                    documento_edi_respondidas.append(str(datos_lista.doc_edi))
            else:
                documento_facturas_respondidas = None
                documento_edi_respondidas = None
            if facturas_no_respondidas == True:
                facturas_no_respondidas = 'True'
            else:
                facturas_no_respondidas = None
            fecha_contabilizacion = datetime.strptime(datos['DocDate'], '%Y%m%d').strftime('%Y-%m-%d')
            fecha_entrega = datetime.strptime(datos['DocDueDate'], '%Y%m%d').strftime('%Y-%m-%d')
            data_prueba = {
                'dato_interno': datos['DocEntry'],
                'n_pedido': datos['DocNum'],
                'cliente': datos['CardName'],
                'n_cliente': datos['NumAtCard'],
                'fecha_contabilizacion': fecha_contabilizacion,
                'fecha_entrega': fecha_entrega,
                'total': format(datos['DocTotal'], '0,.0f') ,
                'documento_facturas_respondidas': documento_facturas_respondidas,
                'facturas_no_respondidas': facturas_no_respondidas,
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_pedido_orden_bodega.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                'estado':estado,
                                                                'cliente':cliente,
                                                                'empresas_clientes':empresas_clientes,
                                                                'pedido':pedido,
                                                                'pedido_cliente':pedido_cliente,
                                                                'permiso_usuario': usuario_datos,
                                                                'pagina': pagina,
                                                                'pagina_fin': str(int(pagina)+20),
                                                                })



def reporte_pedido(request):

    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.nombre
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        estado = request.GET.get('estado')

        if estado == '':
            url2 = IP_SAP + "PurchaseOrders?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and DocDate lt '" + fecha_fin + "' and DocDate gt '" + fecha_inicio + "'"
        elif estado == 'tYES':
            url2 = IP_SAP + "PurchaseOrders?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and Cancelled eq '" + estado + "'and DocDate lt '" + fecha_fin + "' and DocDate gt '" + fecha_inicio + "'"
        elif estado == 'bost_Open' or estado == 'bost_Close':
            url2 = IP_SAP + "PurchaseOrders?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and DocumentStatus eq '" + estado + "'and DocDate lt '" + fecha_fin + "' and DocDate gt '" + fecha_inicio + "'"

        response = sap_request(url2)
        response2 = ast.literal_eval(response2.text)
        response2 = response2['value']

        subtitulo ="Lista_Pedidos"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="LISTA_PEDIDOS.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['NUMERO PEDIDO',
                   'FECHA DE CONTABILIZACION',
                   'FECHA DE ENTREGA',
                   'TOTAL',
                   ]

        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []

        for d in response2:
            numero = d['DocNum'],
            fecha_contabilizacion = d['DocDate'],
            fecha_entrega = d['DocDueDate'],
            total = format(d['DocTotal'], '0,.0f'),
            datos = [(
                int(numero[0]),
                str(fecha_contabilizacion[0]),
                str(fecha_entrega[0]),
                str(total[0]),

            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response



def reporte_otroscanales(request):

    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.id

        lista_infoc = AsignacionPedidosOtrosCanales.objects.all()

        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        empresa_input = request.GET.get('empresa_input')
        pedido = request.GET.get('pedido')
        estado = request.GET.get('estado')
        if estado == 'enproceso':
            estado = 'en proceso'
        if not fecha_inicio == '' and not fecha_fin == '':
            lista_infoc = lista_infoc.filter(num_detalle__num_pedido__fecha__range=[fecha_inicio, fecha_fin])
        if not empresa_input == '':
            lista_infoc = lista_infoc.filter(empresa_id=empresa_input)
        if not pedido == '':
            lista_infoc = lista_infoc.filter(num_detalle__num_pedido__num_pedido=pedido)
        if not estado == '':
            lista_infoc = lista_infoc.filter(num_detalle__num_pedido__estado=estado)

        subtitulo ="Lista_Asignacion_Otros_Canales"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="LISTA_ASIGNACION_OTROS_CANALES.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['PEDIDO',
                   'CLIENTE',
                   'CANTIDADA',
                   'REFERENCIA',
                   'DESCRIPCION',
                   'EMPRESAA',
                   'OBSERVACIONES',
                   ]

        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []

        for d in lista_infoc:
            num_pedido= str(d.num_detalle.num_pedido.num_pedido),
            cantidad= str(d.cantidad),
            referencia= str(d.num_detalle.referencia),
            nombre= str(d.num_detalle.nombre),
            observaciones= str(d.num_detalle.observaciones),
            empresa=str(d.empresa.nombre),
            cliente=d.num_detalle.num_pedido.empresa.nombre
            datos = [(
                num_pedido,
                str(cliente),
                cantidad,
                referencia,
                nombre,
                empresa,
                observaciones,



            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response


def reporte_otroscanales_cliente(request):

    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.id

        lista_infoc = PedidosOtrosCanales.objects.all()

        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        pedido = request.GET.get('pedido')
        estado = request.GET.get('estado')
        if estado == 'enproceso':
            estado = 'en proceso'
        if not fecha_inicio == '' and not fecha_fin == '':
            lista_infoc = lista_infoc.filter(fecha__range=[fecha_inicio, fecha_fin])
        if not pedido == '':
            lista_infoc = lista_infoc.filter(num_pedido=pedido)
        if not estado == '':
            lista_infoc = lista_infoc.filter(estado=estado)

        subtitulo ="Lista_Mis_Pedidos_Otros_Canales"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="LISTA_PEDIDOS_OTROS_CANALES.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['PEDIDO',
                   'FECHA',
                   'HORA',
                   'ESTADO',
                   ]

        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []

        for d in lista_infoc:
            num_pedido= str(d.num_pedido),
            fecha= str(d.fecha),
            hora= str(d.hora),
            estado= str(d.estado),
            datos = [(
                num_pedido,
                fecha,
                hora,
                estado,



            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response




def pedido_detalle(request, form_id):
    hoy = datetime.today().strftime('%Y-%m-%d')
    hoy_log = datetime.now()
    peticiones=Peticiones.objects.all()
    lista_justificaciones=Justificacion.objects.all()
    respuesta=RespuestaPedido.objects.filter(entry_pedido=form_id).exists()
    if respuesta:
        peticion_activa='si'
    else:
        peticion_activa = 'no'
    usuario_actual = request.user
    if not usuario_actual.is_staff:
        return HttpResponseRedirect('/login/')
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
    datos_proveedor= Empresas.objects.filter(id=usuario_datos.empresa_id).first()

    try:
        url6 = IP_SAP + "PurchaseOrders?$select=OpeningRemarks&$filter=DocEntry eq " + form_id

        response6 = sap_request(url6)
        response6 = ast.literal_eval(response6.text)
        response6 = response6['value']
        response6=response6[0]
        comentarios_iniciales=response6['OpeningRemarks']
        WhsCode=''

    except:
        try:
            url6 = IP_SAP + "SQLQueries('ConsultadetalleAlmacenes')/List?pedido=" + form_id

            response6 = sap_request(url6)
            response6 = response6.text
            response6 = response6.replace('null', ' " " ')
            response6 = ast.literal_eval(response6)
            response6 = response6['value']
            for d in response6:
                WhsCode=d['WhsCode']
                if d['WhsCode'] == '19' :
                    fecha_ped_ven = datetime.strptime(d['DocDate'], '%Y%m%d')
                    fecha_ped_ven = fecha_ped_ven.date()
                    y=0
                    while (y != 2):
                        y+=1
                        fecha_ped_ven = fecha_ped_ven + timedelta(days=1)
                        x = 1
                        while (x >= 1):
                            if fecha_ped_ven.weekday() == 5:
                                fecha_ped_ven = fecha_ped_ven + timedelta(days=2)
                            elif fecha_ped_ven.weekday() == 6:
                                fecha_ped_ven = fecha_ped_ven + timedelta(days=1)
                            elif holidays_co.is_holiday_date(fecha_ped_ven):
                                fecha_ped_ven = fecha_ped_ven + timedelta(days=1)
                            else:
                                break
                    comentarios_iniciales = 'Por favor entregar mercancía en <b style="color: #0b93ff"> <strong> bodega PCS Bogota, Cra. Avenida 68 # 9 - 77 a más tardar el día '+str(fecha_ped_ven)+ ' </strong></b> Para ingreso al CEDI deben presentar Cedula, Parafiscales, Documento para ingreso, Botas punteras Empresario si usted entrega directamente a la cadena haga caso omiso de esta información.'
                elif d['WhsCode'] == '15':
                    fecha_ped_ven = datetime.strptime(d['DocDate'], '%Y%m%d')
                    fecha_ped_ven = fecha_ped_ven.date()
                    y = 0
                    while (y != 3):
                        y += 1
                        fecha_ped_ven = fecha_ped_ven + timedelta(days=1)
                        x = 1
                        while (x >= 1):
                            if fecha_ped_ven.weekday() == 5:
                                fecha_ped_ven = fecha_ped_ven + timedelta(days=2)
                            elif fecha_ped_ven.weekday() == 6:
                                fecha_ped_ven = fecha_ped_ven + timedelta(days=1)
                            elif holidays_co.is_holiday_date(fecha_ped_ven):
                                fecha_ped_ven = fecha_ped_ven + timedelta(days=1)
                            else:
                                break
                    comentarios_iniciales = 'Por favor entregar mercancía en <b style="color: #0b93ff"><strong> bodega SPE Medellin, CLL 10 # 56-06 a más tardar el día ' +str(fecha_ped_ven)+ ' antes de las 2pm </strong></b>Solicitar cita para entrega en Cedi SPE a los números 3104737173, 604 4448481 ext.201-202, 604 4440050 ext.201, ingresospromotora@spe.com.co Empresario si usted entrega directamente a la cadena haga caso omiso de esta información.'
                else:
                    comentarios_iniciales = 'No'
        except:
            comentarios_iniciales = 'No'
    try:
        url2 = IP_SAP + "PurchaseOrders?$select=DocNum,Address,VatSum,DocTotal,DocDate,DocDueDate,TaxDate,CardName,DocumentStatus,Cancelled&$filter=DocEntry eq " + form_id
        url3 = IP_SAP + "sml.svc/DETALLE_PEDIDO?$filter=DocEntry eq " + form_id

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        response2 = sap_request(url3)
        response2 = response2.text
        response2 = response2.replace('null', ' " " ')
        response2 = ast.literal_eval(response2)
        response2 = response2['value']
        articulos=[]
        cantidad_pedidos=0
        cantidad_cajas=0
        for d in response2:
            cantidad_pedidos=cantidad_pedidos+int(d['Quantity'])
            try:
                division_cajas=int(d['Quantity'])/int(d['U_HBT1_CANTMINOR'])
            except:
                division_cajas=0
            cantidad_cajas=cantidad_cajas+division_cajas
            try:
                url4 = IP_SAP + "sml.svc/DET_DIR_PEDIDOS?$filter=U_ItemCode eq '"+ d['ItemCode']+"' and DocEntry eq '"+form_id+"'"
                response4 = sap_request(url4)
                response4 = ast.literal_eval(response4.text)
                response4 = response4['value']
                detallesdir=[]
                dependencias=[]
                for l in response4:
                    if l['Dependencia'] not in dependencias:
                        data_dir = {
                            'direccion':l['DirDestino'],
                            'cantidad':l['CantDestino'],
                            'Dependencia':l['Dependencia']
                        }
                        detallesdir.append(data_dir)
                        dependencias.append(l['Dependencia'])
                    else:
                        pass
            except:
                detallesdir=''
            data_articulos = {
                'numero': d['ItemCode'],
                'nombre': d['Dscription'],
                'cantidad': d['Quantity'],
                'precio_unidad':format(d['PriceBefDi'], '0,.0f'),
                'indicador_impuestos': d['TaxCode'],
                'total': format(d['LineTotal'], '0,.0f'),
                'detalledir':detallesdir,
            }
            articulos.append(data_articulos)

        dato_lista = []
        for datos in response:
            pedido_log = datos['DocNum']
            if datos['Cancelled']=='tYES':
                estado='cancelado'
            elif datos['DocumentStatus']=='bost_Open':
                estado='abierto'
            else:
                estado='cerrado'
            fecha_contabilizacion = datetime.strptime(datos['DocDate'], '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d')
            fecha_entrega = datetime.strptime(datos['DocDueDate'], '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d')
            fecha_documento = datetime.strptime(datos['TaxDate'], '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d')
            data_prueba = {
                'n_pedido': datos['DocNum'],
                'direccion': datos['Address'],
                'impuestos':format(datos['VatSum'], '0,.0f') ,
                'fecha_contabilizacion': fecha_contabilizacion,
                'fecha_entrega': fecha_entrega,
                'fecha_documento': fecha_documento,
                'nombre': datos['CardName'],
                'estado': estado,
                'articulos': articulos,
                'total': format(datos['DocTotal'], '0,.0f'),
                'subtotal':format(int(datos['DocTotal']) - int(datos['VatSum']), '0,.0f'),
            }
            dato_lista.append(data_prueba)

        articulos = []
        for d in response2:
            data_articulos = {
                'referencia_cliente': d['BaseRef'],
                'cliente': d['CardName'],
                'direccion': d['Address2'],
                'numero_referencia':d['NumAtCard'],
                'fecha_minima':d['TaxDate'],
                'fecha_maxima':d['DocDueDate'],
                'destino':d['ShipToCode'],
                'codigo_cliente':d['CardCode'],
            }
        articulos.append(data_articulos)
	try:
            variable_probar=WhsCode
        except:
            WhsCode=''
        consulta_log_detalle = HistoriaDetallePedidos.objects.filter(pedido=pedido_log).first()
        if not consulta_log_detalle:
            historial = HistoriaDetallePedidos(
                usuario_id=int(current_user.id),
                empresa_id=int(usuario_datos.empresa_id),
                accion='Ingreso a Detalle de Pedido',
                fecha=hoy_log,
                pedido=pedido_log,
                estado=1,
            )
            historial.save()
        elif consulta_log_detalle.estado == 0:
            historial = HistoriaDetallePedidos(
                usuario_id=int(current_user.id),
                empresa_id=int(usuario_datos.empresa_id),
                accion='Ingreso a Detalle de Pedido',
                fecha=hoy_log,
                pedido=pedido_log,
                estado=1,
            )
            historial.save()
        return render(request, "pedido_detalle.html", {"form_id": form_id,
                                                          'user': current_user,
                                                          'lista_solicitud': dato_lista,
                                                          'lista_justificaciones': lista_justificaciones,
                                                          'peticiones': peticiones,
                                                          'hoy':hoy,
                                                          'cantidad_pedidos':cantidad_pedidos,
                                                          'cantidad_cajas':cantidad_cajas,
                                                          'lista_cliente': articulos,
                                                          'peticion_activa': peticion_activa,
                                                       'permiso_usuario': usuario_datos,
                                                       'datos_empresa': datos_proveedor,
                                                       'WhsCode': WhsCode,
                                                       'comentarios_iniciales': comentarios_iniciales,
                                                          })
    except:
        url2 = IP_SAP + "PurchaseOrders?$select=DocNum,Address,VatSum,DocTotal,DocDate,DocDueDate,TaxDate,CardName,DocumentStatus,Cancelled&$filter=DocEntry eq " + form_id
        url3 = IP_SAP + "SQLQueries('consultadetallepedidosvaci2')/List?pedido=" + form_id

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        response2 = sap_request(url3)
        response2 = response2.text
        response2 = response2.replace('null', ' " " ')
        response2 = ast.literal_eval(response2)
        response2 = response2['value']
        articulos = []
        cantidad_pedidos=0
        cantidad_cajas=0
        for d in response2:
            cantidad_pedidos = cantidad_pedidos + int(d['Quantity'])
            try:
                division_cajas = int(d['Quantity']) / int(d['U_HBT1_CANTMINOR'])
            except:
                division_cajas=0
            cantidad_cajas = cantidad_cajas + division_cajas
            try:
                url4 = IP_SAP + "sml.svc/DET_DIR_PEDIDOS?$filter=U_ItemCode eq '"+ d['ItemCode']+"' and DocEntry eq '"+form_id+"'"
                response4 = sap_request(url4)
                response4 = ast.literal_eval(response4.text)
                response4 = response4['value']
                detallesdir = []
                dependencias = []
                for l in response4:
                    if l['Dependencia'] not in dependencias:
                        data_dir = {
                            'direccion': l['DirDestino'],
                            'cantidad': l['CantDestino'],
                            'Dependencia': l['Dependencia']
                        }
                        detallesdir.append(data_dir)
                        dependencias.append(l['Dependencia'])
                    else:
                        pass
            except:
                detallesdir = ''
            data_articulos = {
                'numero': d['ItemCode'],
                'nombre': d['Dscription'],
                'cantidad': d['Quantity'],
                'precio_unidad': format(d['PriceBefDi'], '0,.0f'),
                'indicador_impuestos': d['TaxCode'],
                'total': format(d['LineTotal'], '0,.0f'),
                'detalledir': detallesdir,
            }
            articulos.append(data_articulos)

        dato_lista = []
        for datos in response:
            pedido_log=datos['DocNum']
            if datos['Cancelled'] == 'tYES':
                estado = 'cancelado'
            elif datos['DocumentStatus'] == 'bost_Open':
                estado = 'abierto'
            else:
                estado = 'cerrado'
            fecha_contabilizacion = datetime.strptime(datos['DocDate'], '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d')
            fecha_entrega = datetime.strptime(datos['DocDueDate'], '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d')
            fecha_documento = datetime.strptime(datos['TaxDate'], '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d')
            data_prueba = {
                'n_pedido': datos['DocNum'],
                'direccion': datos['Address'],
                'impuestos': format(datos['VatSum'], '0,.0f'),
                'fecha_contabilizacion': fecha_contabilizacion,
                'fecha_entrega': fecha_entrega,
                'fecha_documento': fecha_documento,
                'nombre': datos['CardName'],
                'estado': estado,
                'articulos': articulos,
                'total': format(datos['DocTotal'], '0,.0f'),
                'subtotal': format(int(datos['DocTotal']) - int(datos['VatSum']), '0,.0f'),
            }
            dato_lista.append(data_prueba)


	try:
            variable_probar=WhsCode
        except:
            WhsCode=''
        consulta_log_detalle = HistoriaDetallePedidos.objects.filter(pedido=pedido_log).first()
        if not consulta_log_detalle:
            historial = HistoriaDetallePedidos(
                usuario_id=int(current_user.id),
                empresa_id=int(usuario_datos.empresa_id),
                accion='Ingreso a Detalle de Pedido',
                fecha=hoy_log,
                pedido=pedido_log,
                estado=1,
            )
            historial.save()
        elif consulta_log_detalle.estado == 0:
            historial = HistoriaDetallePedidos(
                usuario_id=int(current_user.id),
                empresa_id=int(usuario_datos.empresa_id),
                accion='Ingreso a Detalle de Pedido',
                fecha=hoy_log,
                pedido=pedido_log,
                estado=1,
            )
            historial.save()
        return render(request, "pedido_detalle.html", {"form_id": form_id,
                                                       'user': current_user,
                                                       'lista_solicitud': dato_lista,
                                                       'lista_justificaciones': lista_justificaciones,
                                                       'permiso_usuario': usuario_datos,
                                                       'peticiones': peticiones,
                                                       'cantidad_pedidos': cantidad_pedidos,
                                                       'cantidad_cajas': cantidad_cajas,
                                                       'hoy': hoy,
                                                       'peticion_activa': peticion_activa,
                                                       'WhsCode': WhsCode,
                                                       'datos_empresa': datos_proveedor,
                                                       'comentarios_iniciales': comentarios_iniciales,

                                                       })



def pedido_asignacion_pdf(request, form_id):
    hoy = datetime.today().strftime('%Y-%m-%d')
    hoy_log = datetime.now()
    asignacion=AsignacionPedidosOtrosCanales.objects.filter(pk=form_id)
    asignacion_titulos=AsignacionPedidosOtrosCanales.objects.filter(pk=form_id).first()

    url3 = IP_SAP + "SQLQueries('consultadatospreciosfacturas1')/List?codigo='" + str(
        asignacion_titulos.num_detalle.referencia) + "'"

    response1 = sap_request(url3)
    response1 = response1.text
    response1 = response1.replace('null', ' " " ')
    response1 = ast.literal_eval(response1)
    response1 = response1['value']
    preciocompra=0
    for cuenta in response1:
        if cuenta['PriceList'] == 11:
            preciocompra = cuenta['Price']

    total=preciocompra*asignacion_titulos.cantidad
    usuario_actual = request.user
    if not usuario_actual.is_staff:
        return HttpResponseRedirect('/login/')
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
    datos_proveedor= Empresas.objects.filter(id=usuario_datos.empresa_id).first()


    return render(request, "pedido_asignacion_pdf.html", {"form_id": form_id,
                                                       'user': current_user,
                                                       'asignacion': asignacion,
                                                       'asignacion_titulos': asignacion_titulos,
                                                       'preciocompra': preciocompra,
                                                       'total': total,

                                                       })



def pedido_detalle_bodega(request, form_id):
    if request.method == 'GET':
        hoy = datetime.today().strftime('%Y-%m-%d')
        peticiones=Peticiones.objects.all()
        lista_justificaciones=Justificacion.objects.all()
        respuesta=OrdenVenta.objects.filter(entry=form_id).exists()
        if respuesta:
            peticion_activa='si'
        else:
            peticion_activa = 'no'
        usuario_actual = request.user
        if not usuario_actual.is_staff:
            return HttpResponseRedirect('/login/')
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        datos_proveedor= Empresas.objects.filter(id=usuario_datos.empresa_id).first()


        url6 = IP_SAP + "SQLQueries('consultapedidosordendeventa')/List?pedido=" + form_id

        response6 = sap_request(url6)
        response6 = response6.text
        response6 = response6.replace('null', ' " " ')
        response6 = ast.literal_eval(response6)
        response6 = response6['value']

        url3 = IP_SAP + "SQLQueries('consultadetalleordenesdeventas2')/List?pedido=" + form_id

        response2 = sap_request(url3)
        response2 = response2.text
        response2 = response2.replace('null', ' " " ')
        response2 = ast.literal_eval(response2)
        response2 = response2['value']
        articulos=[]
        for d in response2:

            url4 = IP_SAP + "SQLQueries('consultadetalleordenesdeventas1')/List?pedido=" + form_id+"&CodigoProd='"+d['ItemCode']+"'"

            response3 = sap_request(url4)
            response3 = response3.text
            response3 = response3.replace('null', ' " " ')
            response3 = ast.literal_eval(response3)
            response3 = response3['value']
            direcciones=[]
            for n in response3:
                empaques_detalles=int(n['CantDestino']) / int(d['SalPackUn'])
                data_direcciones = {
                    'cantidad': n['CantDestino'],
                    'dependencia': n['U_EAN'],
                    'direccion': n['DirDestino'],
                    'empaques': empaques_detalles,
                }
                direcciones.append(data_direcciones)

            empaques_despachados = int(d['Quantity']) / int(d['SalPackUn'])

            data_articulos = {
                'ean': d['CodeBars'],
                'descripc': d['Dscription'],
                'unidades': d['Quantity'],
                'canti_empa': d['SalPackUn'],
                'plu': d['U_HBT1_PLU'],
                'tota_bruto': d['GTotal'],
                'total_neto': d['LineTotal'],
                'emp_desp': empaques_despachados,
                'direcciones': direcciones,
            }
            articulos.append(data_articulos)

        return render(request, "pedido_detalle_bodega.html", {"form_id": form_id,
                                                       'user': current_user,
                                                       'lista_justificaciones': lista_justificaciones,
                                                       'lista_solicitud': response6,
                                                       'lista_detalle': articulos,
                                                       'peticiones': peticiones,
                                                       'hoy': hoy,
                                                       'permiso_usuario': usuario_datos,
                                                       'datos_empresa': datos_proveedor,
                                                       'peticion_activa': peticion_activa,
                                                       'direcciones': direcciones,
                                                       })

    else:
        hoy = date.today()
        hora = datetime.now().time()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        ean = request.POST.getlist('ean[]')
        descripcion = request.POST.getlist('descripcion[]')
        u_pedidas = request.POST.getlist('u_pedidas[]')
        c_uni_emp = request.POST.getlist('c_uni_emp[]')
        u_pedidas_desp = request.POST.getlist('u_pedidas_desp[]')
        plu = request.POST.getlist('plu[]')
        tota_bruto = request.POST.getlist('tota_bruto[]')
        total_neto = request.POST.getlist('total_neto[]')
        e_despachados = request.POST.getlist('e_despachados[]')
        dependencia = request.POST.getlist('dependencia[]')
        direccion = request.POST.getlist('direccion[]')
        check = request.POST.getlist('check[]')

        lista_tamano = len(ean)
        tipo_empaque = request.POST['tipo_empaque']
        tipo_embalaje = request.POST['tipo_embalaje']
        pedido_num = request.POST['pedido_num']

        orden_venta= OrdenVenta(
            tipo_empaque=tipo_empaque,
            tipo_embalaje=tipo_embalaje,
            entry=form_id,
            fecha=hoy,
            hora=hora,
            pedido=pedido_num,
        )
        orden_venta.save()
        if direccion:
            for i in range(lista_tamano):
                if dependencia[i] in check:
                    pedidonovedad = DetalleOrdenVenta(

                        ean=ean[i],
                        descripcion=descripcion[i],
                        u_pedidas=u_pedidas[i],
                        cant_un_ped=c_uni_emp[i],
                        unidades_despachar=u_pedidas_desp[i],
                        empaques_despachado=e_despachados[i],
                        direccion=direccion[i],
                        dependencia=dependencia[i],
                        plu=plu[i],
                        tota_bruto=tota_bruto[i],
                        total_neto=total_neto[i],
                        orden_id=orden_venta.pk,
                    )
                    pedidonovedad.save()
                else:
                    pass
        else:
            for i in range(lista_tamano):
                if ean[i] in check:
                    pedidonovedad = DetalleOrdenVenta(

                        ean=ean[i],
                        descripcion=descripcion[i],
                        u_pedidas=u_pedidas[i],
                        cant_un_ped=c_uni_emp[i],
                        unidades_despachar=u_pedidas_desp[i],
                        empaques_despachado=e_despachados[i],
                        plu=plu[i],
                        tota_bruto=tota_bruto[i],
                        total_neto=total_neto[i],
                        orden_id=orden_venta.pk,
                    )
                    pedidonovedad.save()
                else:
                    pass
        area=int(13)
        emails = PersonasAtencion.objects.filter(area__pk=area)
        for correos in emails:
            try:
                email = EmailMessage('SOLICITUD DE FACTURA PARA LA ORDEN DE VENTA ' + str(pedido_num),
                                     'La bodega 19 solicito la factura '
                                     + ' para la orden de venta ' + str(
                                         pedido_num) +   '\nPara ver más ingrese en ' +
                                     IP_SERVIDOR + '/configuracion/solicitud_pedido_orden/bodegas/problema/' +
                                     form_id + '/',
                                     to=[correos.email])
                email.send()
            except:
                pass

        return HttpResponseRedirect('/configuracion/solicitud_pedido_orden/bodegas/detalle/' + form_id + '/')


def pedido_problema_detalle_bodegas(request, form_id):
    if request.method == 'GET':
        usuario_actual = request.user
        if not usuario_actual.is_staff:
            return HttpResponseRedirect('/login/')
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        orden_venta= OrdenVenta.objects.filter(entry=form_id)
        titulos= OrdenVenta.objects.filter(entry=form_id).first()



        url6 = IP_SAP + "SQLQueries('consultadefacturasabiertas')/List"

        response6 = sap_request(url6)
        response6 = response6.text
        response6 = response6.replace('null', ' " " ')
        response6 = ast.literal_eval(response6)
        response6 = response6['value']
        return render(request, "pedido_problemas_detalle_bodega.html", {"form_id": form_id,
                                                                 "orden_venta": orden_venta,
                                                                 'permiso_usuario': usuario_datos,
                                                                 "titulos": titulos,
                                                                 "facturas": response6,
                                                                 })

    else:
        hoy = date.today()
        hora = datetime.now().time()
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        try:
            adjunto = request.POST['adjunto']
        except:
            adjunto = request.FILES['adjunto']

        if adjunto != '':
            myfile = request.FILES['adjunto']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            ruta_documento = fs.url(filename)
        else:
            ruta_documento = None

        pedido_pk = request.POST['num_pedido']
        factura = request.POST['factura']
        problema = request.POST['problema']
        fecha = datetime.now()
        fechas = fecha.strftime('%Y-%m-%d')
        fecha = fechas[2:4] + fechas[5:7] + fechas[8:10]
        fecha_actual = fechas[0:4] + fechas[5:7] + fechas[8:10]
        hora = timezone.now()
        hora = hora.strftime('%H%M')
        consecutivo_edi = Consecutivo.objects.all().first()
        consuc_edi = "000" + str(consecutivo_edi.valor)
        datospersona = RespuestaPedido.objects.filter(id=problema).first()


        url3 = IP_SAP + "SQLQueries('consultaavisodespachobodegas')/List?pedido='" + str(
            pedido_pk) + "'"

        response1 = sap_request(url3)
        response1 = response1.text
        response1 = response1.replace('null', ' " " ')
        response1 = ast.literal_eval(response1)
        response1 = response1['value']
        dato_cambiante_cantidad_cajas = "esteesundatocambiantedecajas"
        validador=DetalleOrdenVenta.objects.filter(orden__pedido=pedido_pk).first()
        validador=validador.dependencia
        for datos in response1:
            fecha_minima = datos['DocDate']
            fechas_maxima = datos['DocDueDate']
            numero_pedido = datos['DocNum']
            GLNCliente = datos['U_EAN']
            NumeroOrden = datos['NumAtCard']
            if validador:
                edi_data = "UNB+UNOA:2+7701081000011+" + datos['U_EAN'] + "+" + fecha + ":" + hora + '+PCS' + str(
                    consuc_edi) + "+PASSWORD+DESADV" \
                                  "\nUNH+1+DESADV:D:96A:UN:EAN005\nBGM+YB1+" + str(consuc_edi) + "+9\nDTM+137:" \
                           + fecha_minima + "0000:203\nDTM+11:" + fecha_actual + "0000:203\nDTM+17:" + fechas_maxima + "0000:203\n" \
                                                                                                                       "RFF+ON:" + \
                           datos['NumAtCard'] + "\nRFF+IV:" + str(factura) + "\nNAD+DP+" + datos[
                               'EanTienda'] + "::9\nRFF+VA:" + datos['LicTradNum'][:-2] + "\nNAD+BY+" + datos[
                               'U_EAN'] + "::9\n" \
                                          "RFF+VA:" + datos['LicTradNum'][
                                                      :-2] + "\nNAD+SU+7701081000011::9\nRFF+VA:890985438\nNAD+CA+7701081000011::9\n"
            else:
                edi_data = "UNB+UNOA:2+7701081000011+" + datos['U_EAN'] + "+" + fecha + ":" + hora + '+PCS' + str(
                    consuc_edi) + "+PASSWORD+DESADV" \
                                  "\nUNH+1+DESADV:D:96A:UN:EAN005\nBGM+351+" + str(consuc_edi) + "+9\nDTM+137:" \
                           + fecha_minima + "0000:203\nDTM+11:" + fecha_actual + "0000:203\nDTM+17:" + fechas_maxima + "0000:203\n" \
                                                                                                                       "RFF+ON:" + \
                           datos['NumAtCard'] + "\nRFF+IV:" + str(factura) + "\nNAD+DP+" + datos[
                               'EanTienda'] + "::9\nRFF+VA:" + datos['LicTradNum'][:-2] + "\nNAD+BY+" + datos[
                               'U_EAN'] + "::9\n" \
                                          "RFF+VA:" + datos['LicTradNum'][
                                                      :-2] + "\nNAD+SU+7701081000011::9\nRFF+VA:890985438\nNAD+CA+7701081000011::9\n" \
                                                             "TDT+20++30+31++++:::123456\nEQD+BX\nCPS+1\nPAC+" + dato_cambiante_cantidad_cajas + "++BX\nCPS+2+1\nPAC+" + dato_cambiante_cantidad_cajas + "++BX"
        contador = 0

        pedido_novedades = DetalleOrdenVenta.objects.filter(orden_id=problema)

        cantidad_embalada = 0
        validador_repetidor=" "
        for detallepedido in pedido_novedades:
            if validador:
                if validador_repetidor != detallepedido.ean:
                    validador_repetidor=detallepedido.ean
                    edi_pre = "PAT+1++9:3:D:30\nLIN+1++" + detallepedido.ean + ":EN\nPIA+1+" + detallepedido.plu + ":AT\n" \
                    "QTY+21:" + detallepedido.u_pedidas + ":NAR\nPRI+AAB:" + detallepedido.tota_bruto + "\nPRI+AAA:" + detallepedido.total_neto + "" \
                    "\nPAC+" + dato_cambiante_cantidad_cajas + "++BX"
                    edi_data = edi_data + edi_pre

                cantidad_embalada = cantidad_embalada + int(detallepedido.empaques_despachado)

                edi_complement = "\nLOC+7+"+detallepedido.dependencia+"::9\nQTY+12:"+str(detallepedido.unidades_despachar)+":NAR"
                edi_data = edi_data + edi_complement
            else:
                cantidad_embalada = cantidad_embalada + int(detallepedido.empaques_despachado)
                contador = contador + 1
                edi_complement = "\nLIN+" + str(contador) + "++" + detallepedido.ean + ":EN" \
                                                                                               "\nQTY+12:" + str(
                    detallepedido.unidades_despachar) + ":NAR"
                edi_data = edi_data + edi_complement

        edi_final = "\nCNT+2:2"

        edi_data = edi_data + edi_final
        edi_data = edi_data.replace(dato_cambiante_cantidad_cajas, str(cantidad_embalada))
        edi_count = edi_data.count("\n")
        edi_finales = "\nUNT+" + str(edi_count) + "+1\nUNZ+1+PCS" + str(consuc_edi)
        edi_data = edi_data + edi_finales

        name_file = str(GLNCliente) + '_' + str(NumeroOrden) + '_' + str(factura) + '_' + str(consuc_edi)+".edi"

        valor_edi = 1 + consecutivo_edi.valor
        Consecutivo.objects.filter(pk=consecutivo_edi.pk).update(valor=valor_edi)


        # Crear una instancia de FileSystemStorage
        fs = FileSystemStorage()

        # Guardar el archivo en el sistema de archivos
        with fs.open(name_file, 'w') as edi_file:
            edi_file.write(edi_data)

        # Obtener la URL del archivo guardado (opcional)
        uploaded_file_url = fs.url(name_file)



        OrdenVenta.objects.filter(id=problema).update(factura=factura, estado='respondido',
                                                               doc_respuesta=ruta_documento, doc_edi=uploaded_file_url)

        messages.add_message(request, messages.INFO,
                             'Se ha realizado el documento edi para el pedido ' + str(pedido_pk) + ' satisfactoriamente.')


        return HttpResponseRedirect('/configuracion/solicitud_pedido_orden/bodegas/problema/' + form_id + '/')






def pedido_problema_novedades_bodegas(request, form_id,novedad_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
    titulo=OrdenVenta.objects.filter(entry=form_id).first()
    pedido_novedades=DetalleOrdenVenta.objects.filter(orden_id=novedad_id)
    detalle_validacion=DetalleOrdenVenta.objects.filter(orden_id=novedad_id).first()

    return render(request, "pedido_problemas_novedades_bodega.html", {"form_id": form_id,
                                                    'permiso_usuario': usuario_datos,
                                                   "problemas_titulo": titulo,
                                                   "pedido_novedades": pedido_novedades,
                                                   "detalle_validacion": detalle_validacion,
                                                        })




def pedido_problema_detalle(request, form_id):
    usuario_actual = request.user
    if not usuario_actual.is_staff:
        return HttpResponseRedirect('/login/')
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
    personas = PersonasAtencion.objects.filter(usuario=usuario_datos.usuario.username).exists()
    problemas=RespuestaPedido.objects.filter(entry_pedido=form_id)
    problemas_titulo=RespuestaPedido.objects.filter(entry_pedido=form_id).first()
    lista_citas=RespuestaCita.objects.filter(peticion__estado='pendiente')
    if personas == True:
        personas = PersonasAtencion.objects.filter(usuario=usuario_datos.usuario.username)
        listaprueba=[]
        for persona in personas:
            problemas = RespuestaPedido.objects.filter(entry_pedido=form_id,peticion__area_id=persona.area_id)
            listaprueba=list(chain(listaprueba,problemas))
        problemas=listaprueba
        
    dato_lista = []
    for datos in problemas:
        if datos.peticion.descripcion=='No despachare pedido' or datos.peticion.descripcion=='Suspender producto':
            justificacion=datos.justificacion_adicional
        else:
            justificacion=datos.justificacion.descripcion
        data_prueba = {
            'id': datos.id,
            'num_pedido': datos.num_pedido,
            'entry_pedido': datos.entry_pedido,
            'empresa': datos.empresa,
            'fecha': datos.fecha.strftime('%Y-%m-%d'),
            'hora': datos.hora.strftime('%H:%M:%S'),
            'estado': datos.estado,
            'respuesta': datos.respuesta,
            'email': datos.email,
            'peticion': datos.peticion,
            'justificacion': justificacion,
            'doc_respuesta': datos.doc_respuesta,
        }
        dato_lista.append(data_prueba)
    
    return render(request, "pedido_problemas_detalle.html", {"form_id": form_id,
                                                   "problemas": dato_lista,
                                                    'permiso_usuario': usuario_datos,
                                                    'lista_citas': lista_citas,
                                                   "problemas_titulo": problemas_titulo,
                                                        })


def pedido_problema_novedades(request, form_id,novedad_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
    problemas_titulo=RespuestaPedido.objects.filter(entry_pedido=form_id).first()
    pedido_novedades=PedidosNovedades.objects.filter(peticion_id=novedad_id)

    return render(request, "pedido_problemas_novedades.html", {"form_id": form_id,
                                                    'permiso_usuario': usuario_datos,
                                                   "problemas_titulo": problemas_titulo,
                                                   "pedido_novedades": pedido_novedades,
                                                        })



def pedido_problema_cita_entrega(request, form_id,novedad_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
    problemas_titulo=RespuestaPedido.objects.filter(entry_pedido=form_id).first()
    pedido_novedades=RespuestaCita.objects.filter(peticion_id=novedad_id)

    return render(request, "pedido_problemas_cita_entrega.html", {"form_id": form_id,
                                                    'permiso_usuario': usuario_datos,
                                                   "problemas_titulo": problemas_titulo,
                                                   "pedido_novedades": pedido_novedades,
                                                        })


def reporte_pedido_detalle(request):

    if request.method == 'GET':

        solicitud = request.GET.get('solicitud')


        url2 = IP_SAP + "PurchaseOrders?$select=DocNum,DocDate,DocDueDate,TaxDate,CardName,DocTotal&$filter=DocNum eq " + solicitud
        url3 = IP_SAP + "PurchaseOrders?$select=DocumentLines&$filter=DocNum eq " + solicitud

        response3 = sap_request(url2)
        response3 = ast.literal_eval(response3.text)
        response3 = response3['value']
        response2 = sap_request(url3)
        response2 = json.loads(response2.text)
        response2 = response2['value']
        response2 = response2[0]
        response2 = response2['DocumentLines']

        subtitulo ="Factura #" + str(solicitud)

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="PEDIDO.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columnas = ['CLIENTE',
                   'NO.SOLICITUD',
                   'FECHA DE CONTABILIZACION',
                   'FECHA DE ENTREGA',
                   'FECHA DEL DOCUMENTO',
                   ]

        for col_num in range(len(columnas)):
            cwidth = ws.col(col_num).width
            if (len(columnas[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columnas) * 367)
            ws.write(row_num, col_num, columnas[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        filas = []
        for d in response3:
            nombre = d['CardName'],
            n_pedido = d['DocNum'],
            fecha_contabilizacion = d['DocDate'],
            fecha_entrega = d['DocDueDate'],
            fecha_documento = d['TaxDate'],
            datos = [(
                str(nombre[0]),
                int(n_pedido[0]),
                str(fecha_contabilizacion[0]),
                str(fecha_entrega[0]),
                str(fecha_documento[0]),

            )]
            filas.extend(datos)

        for row in filas:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)



        columns = ['NUMERO DEL ARTICULO',
                   'NOMBRE DEL ARTICULO',
                   'CANTIDAD',
                   'PRECIO POR UNIDAD',
                   'INDICADOR DE IMPUESTOS',
                   'CODIGO DE UNIDAD DE MEDIDA',
                   'TOTAL',
                   ]

        for col_num in range(len(columns)):
            row_num = 5
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []

        for d in response2:
            numero = d['ItemCode'],
            nombre = d['ItemDescription'],
            cantidad = d['Quantity'],
            precio_unidad = d['UnitPrice'],
            indicador_impuestos = d['TaxCode'],
            codigo_unidad = d['UoMCode'],
            total = format(d['LineTotal'], '0,.0f'),
            datos = [(
                int(numero[0]),
                str(nombre[0]),
                int(cantidad[0]),
                int(precio_unidad[0]),
                str(indicador_impuestos[0]),
                str(codigo_unidad[0]),
                str(total[0]),

            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)
        filas_resultante = row_num

        ws.write(filas_resultante+2, 5, 'TOTAL:', font_style)
        for d in response3:
            total = format(d['DocTotal'], '0,.0f') ,
            ws.write(filas_resultante + 2, 6, str(total[0]), font_style)

        wb.save(response)
        return response


def config_solicitud_catalogo_productos(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


        return render(request, "config_solicitud_catalogo_productos.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                                            'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.nombre
        codigo_prod = request.POST['codigo']



        url2 = IP_SAP + "SQLQueries('ConsultasInventarios')/List?Empresa='" + empresa +"'&NombreProducto='%"+codigo_prod+"%'"

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'n_producto': datos['ItemCode'],
                'nombre': datos['ItemName'],
                'stock': datos['OnHand'],
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_catalogo_productos.html", {'user': current_user,
                                                                      'lista_formulas': formulas,
                                                                      'lista_prueba': dato_lista,
                                                                      'codigo_prod': codigo_prod,
                                                                            'permiso_usuario': usuario_datos,
                                                                      })



def catalogo_productos_detalle(request, form_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


    url2 = IP_SAP + "Items?$select=ItemCode,ItemName,CreateDate,Mainsupplier,SalesUnit,SalesItemsPerUnit,QuantityOnStock&$filter=ItemCode eq '" + form_id + " ' "
    url3 = IP_SAP + "Items?$select=ItemWarehouseInfoCollection&$filter=ItemCode eq '" + form_id + " ' "


    response = sap_request(url2)
    response = ast.literal_eval(response.text)
    response = response['value']
    response2 = sap_request(url3)
    response2 = json.loads(response2.text)
    response2 = response2['value']
    response2 = response2[0]
    response2 = response2['ItemWarehouseInfoCollection']
    almacenes = []
    for d in response2:
        if d['InStock']==0.0:
            pass
        else:
            url4 = IP_SAP + "Warehouses?$select=WarehouseName&$filter=WarehouseCode eq '" + d['WarehouseCode'] + " ' "
            response3 = sap_request(url4)
            response3 = ast.literal_eval(response3.text)
            response3 = response3['value']
            response3 = response3[0]
            response3 = response3['WarehouseName']
            data_articulos = {
                'codigo': d['WarehouseCode'],
                'nombre': response3,
                'stock': d['InStock'],
                'compremetido': d['Committed'],
                'pedido': d['Ordered'],
                'costo': d['StandardAveragePrice'],
            }
            almacenes.append(data_articulos)

    dato_lista = []
    for datos in response:
        data_prueba = {
            'n_producto': datos['ItemCode'],
            'nombre': datos['ItemName'],
            'fecha_creacion': datos['CreateDate'],
            'proveedor': datos['Mainsupplier'],
            'unidad_medida': datos['SalesUnit'],
            'art_por_und': datos['SalesItemsPerUnit'],
            'stock': datos['QuantityOnStock'],
            'almacenes':almacenes,

        }
        dato_lista.append(data_prueba)
    return render(request, "catalogo_productos_detalle.html", {"form_id": form_id,
                                                      'user': current_user,
                                                      'lista_solicitud': dato_lista,
                                                               'permiso_usuario': usuario_datos,
                                                      })


def config_solicitud_debito_proveedores(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        return render(request, "config_solicitud_debito_proveedores.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                                            'permiso_usuario': usuario_datos,
                                                        })

    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pagina = request.POST['pagina']

        try:
            paginador = request.POST['paginador']
            if paginador == 'atras':
                pagina = str(int(pagina) - 20)
                if int(pagina) < 0:
                    pagina = str(0)
            elif paginador == 'adelante':
                pagina = str(int(pagina) + 20)
            elif paginador == 'primera':
                pagina = str(0)
        except:
            pass
        try :
            estado = request.POST['estado']
            if estado=='None':
                estado = None
        except:
            estado = None


        if estado== None:
            url2 = IP_SAP + "PurchaseInvoices?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                + empresa + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina
        elif estado=='tYES':
            url2 = IP_SAP + "PurchaseInvoices?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and Cancelled eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina
        elif estado=='bost_Open' or estado=='bost_Close':
            url2 = IP_SAP + "PurchaseInvoices?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and DocumentStatus eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'n_pedido': datos['DocNum'],
                'fecha_contabilizacion': datos['DocDate'],
                'fecha_entrega': datos['DocDueDate'],
                'total': format(datos['DocTotal'], '0,.0f') ,
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_debito_proveedores.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                'estado':estado,
                                                                'permiso_usuario': usuario_datos,
                                                                'pagina': pagina,
                                                                'pagina_fin': str(int(pagina)+20),
                                                                })





def debito_proveedores_detalle(request, form_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

    url2 = IP_SAP + "PurchaseInvoices?$select=DocNum,DocDate,DocDueDate,TaxDate,CardName&$filter=DocNum eq " + form_id
    url3 = IP_SAP + "PurchaseInvoices?$select=DocumentLines&$filter=DocNum eq " + form_id

    response = sap_request(url2)
    response = ast.literal_eval(response.text)
    response = response['value']
    response2 = sap_request(url3)
    response2 = json.loads(response2.text)
    response2 = response2['value']
    response2 = response2[0]
    response2 = response2['DocumentLines']
    articulos = []
    for d in response2:
        data_articulos = {
            'numero': d['ItemCode'],
            'nombre': d['ItemDescription'],
            'cantidad': d['Quantity'],
            'precio_unidad': d['UnitPrice'],
            'indicador_impuestos': d['TaxCode'],
            'codigo_unidad': d['UoMCode'],
            'total': format(d['LineTotal'], '0,.0f'),
        }
        articulos.append(data_articulos)
    dato_lista = []
    for datos in response:
        data_prueba = {
            'n_pedido': datos['DocNum'],
            'fecha_contabilizacion': datos['DocDate'],
            'fecha_entrega': datos['DocDueDate'],
            'fecha_documento': datos['TaxDate'],
            'nombre': datos['CardName'],
            'articulos': articulos
        }
        dato_lista.append(data_prueba)
    return render(request, "debito_proveedores_detalle.html", {"form_id": form_id,
                                                   'user': current_user,
                                                   'lista_solicitud': dato_lista,
                                                               'permiso_usuario': usuario_datos,
                                                   })



def config_solicitud_credito_proveedores(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        return render(request, "config_solicitud_credito_proveedores.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                                             'permiso_usuario': usuario_datos,
                                                        })

    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pagina = request.POST['pagina']
        nota = request.POST.get('nota', '')

        try:
            paginador = request.POST['paginador']
            if paginador == 'atras':
                pagina = str(int(pagina) - 20)
                if int(pagina) < 0:
                    pagina = str(0)
            elif paginador == 'adelante':
                pagina = str(int(pagina) + 20)
            elif paginador == 'primera':
                pagina = str(0)
        except:
            pass
        try:
            estado = request.POST['estado']
            if estado == 'None':
                estado = None
        except:
            estado = None


        if nota!="":
            url2 = IP_SAP + "PurchaseCreditNotes?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and DocNum eq " + nota +  "&$skip=" + pagina
        elif estado== None:
            url2 = IP_SAP + "PurchaseCreditNotes?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                + empresa + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina
        elif estado=='tYES':
            url2 = IP_SAP + "PurchaseCreditNotes?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and Cancelled eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina
        elif estado=='bost_Open' or estado=='bost_Close':
            url2 = IP_SAP + "PurchaseCreditNotes?$select=DocNum,DocDate,DocDueDate,DocTotal&$filter=CardName eq '" \
                   + empresa + "'and DocumentStatus eq '" + estado + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'&$skip="+pagina

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'n_pedido': datos['DocNum'],
                'fecha_contabilizacion': datos['DocDate'],
                'fecha_entrega': datos['DocDueDate'],
                'total': format(datos['DocTotal'], '0,.0f') ,
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_credito_proveedores.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                'estado':estado,
                                                                'nota':nota,
                                                                'permiso_usuario': usuario_datos,
                                                                'pagina': pagina,
                                                                'pagina_fin': str(int(pagina)+20),
                                                                })




def credito_proveedores_detalle(request, form_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
    datos_proveedor = Empresas.objects.filter(id=usuario_datos.empresa_id).first()

    try:
        url2 = IP_SAP + "PurchaseCreditNotes?$select=DocNum,DocTotal,Comments,WTApplied,VatSum,TotalDiscount,NumAtCard,Address,DocDate,DocDueDate,TaxDate,CardName,OpeningRemarks&$filter=DocNum eq " + form_id
        url3 = IP_SAP + "PurchaseCreditNotes?$select=DocumentLines&$filter=DocNum eq " + form_id

        response = sap_request(url2)
        response = ast.literal_eval(response.text)

        response = response['value']
        response2 = sap_request(url3)
        response2 = json.loads(response2.text)
        response2 = response2['value']
        response2 = response2[0]
        response2 = response2['DocumentLines']
        articulos = []
        for d in response2:
            data_articulos = {
                'numero': d['AccountCode'],
                'nombre': d['ItemDescription'],
                'cantidad': d['Quantity'],
                'precio_unidad': d['UnitPrice'],
                'indicador_impuestos': d['TaxCode'],
                'codigo_unidad': d['UoMCode'],
                'total': format(d['LineTotal'], '0,.0f'),
            }
            articulos.append(data_articulos)
        dato_lista = []
        for datos in response:
            data_prueba = {
                'n_pedido': datos['DocNum'],
                'fecha_contabilizacion': datos['DocDate'],
                'fecha_entrega': datos['DocDueDate'],
                'fecha_documento': datos['TaxDate'],
                'nombre': datos['CardName'],
                'referencia': datos['NumAtCard'],
                'direccion': datos['Address'],
                'comentarios': datos['Comments'],
                'descuento': format(datos['TotalDiscount'], '0,.0f'),
                'impuestos': format(datos['VatSum'], '0,.0f'),
                'retencion': format(datos['WTApplied'], '0,.0f'),
                'total': format(datos['DocTotal'], '0,.0f'),
                'observaciones': datos['OpeningRemarks'],
                'articulos': articulos
            }
            dato_lista.append(data_prueba)
    except:
        url2 = IP_SAP + "PurchaseCreditNotes?$select=DocNum,WTApplied,DocTotal,VatSum,TotalDiscount,Comments,NumAtCard,Address,DocDate,DocDueDate,TaxDate,CardName&$filter=DocNum eq " + form_id
        url3 = IP_SAP + "PurchaseCreditNotes?$select=DocumentLines&$filter=DocNum eq " + form_id

        response = sap_request(url2)
        response = response.text
        response = response.replace('null', ' " " ')
        response = ast.literal_eval(response)
        response = response['value']
        response2 = sap_request(url3)
        response2 = json.loads(response2.text)
        response2 = response2['value']
        response2 = response2[0]
        response2 = response2['DocumentLines']
        articulos = []
        for d in response2:
            data_articulos = {
                'numero': d['AccountCode'],
                'nombre': d['ItemDescription'],
                'cantidad': d['Quantity'],
                'precio_unidad': d['UnitPrice'],
                'indicador_impuestos': d['TaxCode'],
                'codigo_unidad': d['UoMCode'],
                'total': format(d['LineTotal'], '0,.0f'),
            }
            articulos.append(data_articulos)
        dato_lista = []
        for datos in response:
            data_prueba = {
                'n_pedido': datos['DocNum'],
                'fecha_contabilizacion': datos['DocDate'],
                'fecha_entrega': datos['DocDueDate'],
                'fecha_documento': datos['TaxDate'],
                'direccion': datos['Address'],
                'nombre': datos['CardName'],
                'comentarios': datos['Comments'],
                'descuento': format(datos['TotalDiscount'], '0,.0f'),
                'impuestos': format(datos['VatSum'], '0,.0f'),
                'retencion': format(datos['WTApplied'], '0,.0f'),
                'total': format(datos['DocTotal'], '0,.0f'),
                'referencia': datos['NumAtCard'],
                'observaciones': '',
                'articulos': articulos
            }
            dato_lista.append(data_prueba)
    return render(request, "credito_proveedores_detalle.html", {"form_id": form_id,
                                                               'user': current_user,
                                                               'lista_solicitud': dato_lista,
                                                                'permiso_usuario': usuario_datos,
                                                                'datos_empresa': datos_proveedor,
                                                               })


def config_solicitud_aviso_recibo(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        return render(request, "config_solicitud_aviso_recibo.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pagina = request.POST['pagina']
        orden_compra = request.POST['orden_compra']
        factura = request.POST['factura']

        try:
            paginador = request.POST['paginador']
            if paginador=='atras':
                pagina=str(int(pagina)-20)
                if int(pagina)<0:
                    pagina=str(0)
            elif paginador=='adelante':
                pagina=str(int(pagina)+20)
            elif paginador=='primera':
                pagina=str(0)
        except:
            pass



        if orden_compra != '':
            url2 = IP_SAP + "SQLQueries('ConsultasAvisoRecibosorden3')/List?Ordenventa='" + orden_compra +"'&Empresa='"+ str(empresa) +"'&$skip="+pagina
        elif factura != '':
            url2 = IP_SAP + "SQLQueries('ConsultasAvisoRecibosfacturas1')/List?Factura='" + factura +"'&Empresa='"+ str(empresa) + "'&$skip=" + pagina
        else:
            url2 = IP_SAP + "SQLQueries('ConsultasAvisoRecibosnor1')/List?FechaInicial='"+ fecha_inicio +"'&FechaFinal='"+ fecha_fin +"'&Empresa='"+ str(empresa) +"'&$skip=" + pagina

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'codigo': datos['Code'],
                'formato': datos['U_Formato'],
                'fecha_cargue': datos['U_FechaCargaDoc'],
                'doc_referencia': datos['U_ConsecutivoDoc'],
                'cod_mensaje': datos['U_CodFuncionMensaje'],
                'fecha_doc': datos['U_FechaDocumento'],
                'cod_ean': datos['U_CodigoClienteEAN'],
                'cod_cliente': datos['U_CardCode'],
                'nombre_cliente': datos['U_CardName'],
                'ean_depend': datos['U_EANDireccion'],
                'dependencia': datos['Address'],
                'ean_articulo': datos['U_CodigoItemEAN'],
                'codigo_art': datos['U_ItemCode'],
                'descripcion': datos['U_ItemName'],
                'precio_articulo': datos['U_Price'],
                'cantidad': datos['U_Cantidad'],
                'nro_orden': datos['U_DocRefOrden'],
                'nro_factura': datos['U_DocRefFactura'],
                'nro_remision': datos['U_DocRefRemision'],
                'cantidad_enviada': datos['U_CantidadEnviada'],
                'cantidad_perdida': datos['U_CantidadPedida'],
                'cantidad_aceptada': datos['U_CantidadAceptada'],
                'proovedor':str(empresa),
                'cantidad_recibida_no_acep': datos['U_CantRechazoDev'],
                'cantidad_recibida_ser_dest': datos['U_CantRechazoDest'],
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_aviso_recibo.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                'permiso_usuario': usuario_datos,
                                                                'pagina': pagina,
                                                                'pagina_fin': str(int(pagina)+20),
                                                                'orden_compra':orden_compra,
                                                                'factura':factura,
                                                                })


def reporte_aviso_recibo(request):

    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.nombre
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        orden_compra = request.GET.get('orden_compra')
        factura = request.GET.get('factura')


        if orden_compra != '':
            url2 = IP_SAP + "SQLQueries('ConsultasAvisoRecibosorden3')/List?Ordenventa='" + orden_compra + "'&Empresa='"+ str(empresa) +"'"
        elif factura != '':
            url2 = IP_SAP + "SQLQueries('ConsultasAvisoRecibosfacturas1')/List?Factura='" + factura + "'&Empresa='"+ str(empresa) +"'"
        else:
            url2 = IP_SAP + "SQLQueries('ConsultasAvisoRecibosnor1')/List?FechaInicial='" + fecha_inicio + "'&FechaFinal='" + fecha_fin + "'&Empresa='"+ str(empresa) +"'"

        response2 = sap_request(url2)
        response2 = ast.literal_eval(response2.text)
        response2 = response2['value']

        subtitulo ="AvisoRecibo"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="AVISORECIBO.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = [
                    'NUMERO DE PEDIDO',
                    'NRO_FACTURA',
                    'CODIGO',
                   'FECHA_CARGUE',
                   'FECHA_DOC',
                   'NOMBRE_CLIENTE',
                   'DEPENDENCIA',
                   'EAN_ARTICULO',
                   'CODIGO_ART',
                   'DESCRIPCION',
                   'PRECIO_ARTICULO',
                   'CANTIDAD',
                   'NRO_REMISION',
                   'CANTIDAD_ENVIADA',
                   'CANTIDAD_PEDIDA',
                   'CANTIDAD_ACEPTADA',
                   'CANTIDAD NO ACEPTADA',
                   'CANTIDAD DESTRUIDA',
                   ]
        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []

        for d in response2:
            codigo= d['Code'],
            fecha_cargue=d['U_FechaCargaDoc'],
            fecha_doc= d['U_FechaDocumento'],
            nombre_cliente= d['U_CardName'],
            dependencia= d['Address'],
            ean_articulo= d['U_CodigoItemEAN'],
            codigo_art= d['U_ItemCode'],
            descripcion= d['U_ItemName'],
            precio_articulo= d['U_Price'],
            cantidad= d['U_Cantidad'],
            nro_orden= d['U_DocRefOrden'],
            nro_factura= d['U_DocRefFactura'],
            nro_remision= d['U_DocRefRemision'],
            cantidad_enviada= d['U_CantidadEnviada'],
            cantidad_perdida= d['U_CantidadPedida'],
            cantidad_aceptada= d['U_CantidadAceptada'],
            cantidad_recibida_no_acep= d['U_CantRechazoDev'],
            cantidad_recibida_ser_dest= d['U_CantRechazoDest'],
            datos = [(
                nro_orden[0],
                nro_factura[0],
                codigo[0],
                fecha_cargue[0],
                fecha_doc[0],
                nombre_cliente[0],
                dependencia[0],
                ean_articulo[0],
                codigo_art[0],
                descripcion[0],
                precio_articulo[0],
                cantidad[0],
                nro_remision[0],
                cantidad_enviada[0],
                cantidad_perdida[0],
                cantidad_aceptada[0],
                cantidad_recibida_no_acep[0],
                cantidad_recibida_ser_dest[0],
            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response


def config_solicitud_inventarios(request):
    formulas = []
    if request.method == 'GET':
        ahora=datetime.utcnow()
        semana = ahora - timedelta(days=15)
        ahora=ahora.strftime('%Y/%m/%d')
        semana=semana.strftime('%Y/%m/%d')
        cliente = request.GET['cliente']
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresa = usuario_datos.empresa.nombre
        current_user = request.user
        url2 = IP_SAP + "BusinessPartners?$select=CardName&$filter=CardCode eq '" + cliente + "'"
        url6 = IP_SAP + "SQLQueries('ConsultasFechaInventarios')/List?NombreProveedor='" + str(
            empresa) + "'&FechaInicial='"+ semana + "'&FechaFinal='"+ ahora +"'&ClienteBuscado='"+ cliente +"'"

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value'][0]
        nom_cliente = response['CardName']
        response5 = sap_request(url6)
        response5 = ast.literal_eval(response5.text)
        response5 = response5['value']
        dato_fechas = []
        for dias in response5:
            hora_prueba={
                'ano':str(dias['U_TaxDate'])[0:4],
                'mes':str(dias['U_TaxDate'])[4:6],
                'dia':str(dias['U_TaxDate'])[6:8],
            }
            dato_fechas.append(hora_prueba)

        return render(request, "config_solicitud_inventarios.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                        'permiso_usuario': usuario_datos,
                                                        'permiso_usuario': usuario_datos,
                                                        'dato_fechas': dato_fechas,
                                                        'cliente': cliente,
                                                        'nom_cliente': nom_cliente,
                                                        })
    elif request.method == 'POST':
        ahora = datetime.utcnow()
        semana = ahora - timedelta(days=15)
        ahora = ahora.strftime('%Y/%m/%d')
        semana = semana.strftime('%Y/%m/%d')
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        cliente = request.POST['cliente']
        hoy_dia=request.POST['selected_date']
        hoy=request.POST['selected_date']
        if hoy=='':
            messages.add_message(request, messages.ERROR,
                                 'Seleccione una Fecha de corte')
            return HttpResponseRedirect('/configuracion/solicitud_inventarios/?cliente='+cliente)
        hoy=datetime.strptime(hoy,'%Y-%m-%d')
        hoy=hoy.strftime('%Y-%m-%d')
        pagina = request.POST['pagina']

        try:
            paginador = request.POST['paginador']
            if paginador=='atras':
                pagina=str(int(pagina)-20)
                if int(pagina)<0:
                    pagina=str(0)
            elif paginador=='adelante':
                pagina=str(int(pagina)+20)
            elif paginador=='primera':
                pagina=str(0)
        except:
            pass


        url3 = IP_SAP + "BusinessPartners?$select=CardName&$filter=CardCode eq '" + cliente + "'"
        url6 = IP_SAP + "SQLQueries('ConsultasFechaInventarios')/List?NombreProveedor='" + str(
            empresa) + "'&FechaInicial='" + semana + "'&FechaFinal='" + ahora +"'&ClienteBuscado='"+ cliente +"'"

        response1 = sap_request(url3)
        response1 = ast.literal_eval(response1.text)
        response1 = response1['value'][0]
        nom_cliente = response1['CardName']
        response5 = sap_request(url6)
        response5 = ast.literal_eval(response5.text)
        response5 = response5['value']
        dato_fechas = []
        for dias in response5:
            hora_prueba = {
                'ano': str(dias['U_TaxDate'])[0:4],
                'mes': str(dias['U_TaxDate'])[4:6],
                'dia': str(dias['U_TaxDate'])[6:8],
            }
            dato_fechas.append(hora_prueba)
        url2 = IP_SAP + "SQLQueries('ConsultaInventariosDescripcionas')/List?ClienteBusqueda='"+cliente+"'&FechaActual='"+hoy+"'&NombreProveedor='"+empresa+"'&$skip="+pagina

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'Descripcion': datos['Descripcion'],
                'Linea': datos['Linea'],
                'EAN': datos['EAN'],
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_inventarios.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'hoy': hoy_dia,
                                                                'diahoy': hoy,
                                                                'permiso_usuario': usuario_datos,
                                                                'pagina': pagina,
                                                                'cliente': cliente,
                                                                'nom_cliente': nom_cliente,
                                                                'dato_fechas': dato_fechas,
                                                                'pagina_fin': str(int(pagina)+20),
                                                                })



def config_solicitud_inventarios_cliente(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresa = usuario_datos.empresa.nombre
        current_user = request.user
        url2 = IP_SAP + "SQLQueries('ConsultasClientesVentas')/List?NombreProveedor='" + str(
            empresa) + "'"
        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'id': datos['CodigoCliente'],
                'nombre': datos['Cliente'],
            }
            dato_lista.append(data_prueba)

        return render(request, "config_solicitud_inventarios_cliente.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                        'permiso_usuario': usuario_datos,
                                                        'dato_cliente':dato_lista,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        ahora = datetime.utcnow()
        semana = ahora - timedelta(days=14)
        ahora = ahora.strftime('%Y/%m/%d')
        semana = semana.strftime('%Y/%m/%d')
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        cliente = request.POST['cliente']
        hoy_dia=request.POST['fecha_fin']
        hoy=request.POST['fecha_fin']
        hoy=datetime.strptime(hoy,'%d/%m/%Y')
        hoy=hoy.strftime('%Y-%m-%d')
        pagina = request.POST['pagina']

        try:
            paginador = request.POST['paginador']
            if paginador=='atras':
                pagina=str(int(pagina)-20)
                if int(pagina)<0:
                    pagina=str(0)
            elif paginador=='adelante':
                pagina=str(int(pagina)+20)
            elif paginador=='primera':
                pagina=str(0)
        except:
            pass


        url3 = IP_SAP + "SQLQueries('ConsultasClientesVentas')/List?NombreProveedor='" + str(
            empresa) + "'"
        url6 = IP_SAP + "SQLQueries('ConsultasFechaInventario')/List?NombreProveedor='" + str(
            empresa) + "'&FechaInicial='" + semana + "'&FechaFinal='" + ahora + "'"

        response1 = sap_request(url3)
        response1 = ast.literal_eval(response1.text)
        response1 = response1['value']
        response5 = sap_request(url6)
        response5 = ast.literal_eval(response5.text)
        response5 = response5['value']
        dato_fechas = []
        for dias in response5:
            hora_prueba = {
                'ano': str(dias['U_TaxDate'])[0:4],
                'mes': str(dias['U_TaxDate'])[4:6],
                'dia': str(dias['U_TaxDate'])[6:8],
            }
            dato_fechas.append(hora_prueba)
        dato_listas = []
        for datos in response1:
            data_prueba = {
                'id': datos['CodigoCliente'],
                'nombre': datos['Cliente'],
            }
            dato_listas.append(data_prueba)
        url2 = IP_SAP + "SQLQueries('ConsultaInventariosDescripcionas')/List?ClienteBusqueda='"+cliente+"'&FechaActual='"+hoy+"'&NombreProveedor='"+empresa+"'&$skip="+pagina

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'Descripcion': datos['Descripcion'],
                'Linea': datos['Linea'],
                'EAN': datos['EAN'],
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_inventarios.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'hoy': hoy_dia,
                                                                'diahoy': hoy,
                                                                'permiso_usuario': usuario_datos,
                                                                'pagina': pagina,
                                                                'cliente': cliente,
                                                                'dato_cliente': dato_listas,
                                                                'dato_fechas': dato_fechas,
                                                                'pagina_fin': str(int(pagina)+20),
                                                                })

def config_solicitud_inventarios_detalle(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.nombre
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        current_user = request.user
        ean = request.GET.get("ean")
        fecha = request.GET.get("fecha")
        cliente = request.GET.get("cliente")

        url2 = IP_SAP + "SQLQueries('ConsultaInventariosIndicDetalle')/List?ClienteBusqueda='" + cliente + "'&FechaActual='" + fecha + "'&NombreProveedor='" + empresa + "'&EanCode='"+ ean + "'"

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        contador=1
        for datos in response:
            if contador == 1:
                nombre_prod = datos['Descripcion']
                linea = datos['Linea']

        dato_lista = []
        repeticion = []
        for datos in response:
            try:
                comparacion = repeticion.index([datos['EAN'], datos['EAN_Dependencia']])
                pass
            except:
                data_prueba = {
                    'EanDependencia': datos['EAN_Dependencia'],
                    'Dependencia': datos['Dependencia'],
                    'Cantidad': format(datos['Cantidad'], '0,.0f') ,
                }
                dato_lista.append(data_prueba)
                repeticion.append([datos['EAN'], datos['EAN_Dependencia']])
        return render(request, "indicadores_inventario_detalle.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                        'permiso_usuario': usuario_datos,
                                                        'nombre_prod':nombre_prod,
                                                        'linea':linea,
                                                        'ean':ean,
                                                        'fecha':fecha,
                                                        'cliente':cliente,
                                                        'lista_detalle':dato_lista,
                                                        })
    elif request.method == 'POST':
        pass


def reporte_inventario(request):
    if request.method == 'GET':
        try:
            current_user = request.user
            nombre = current_user.username
            empresa_obj = User.objects.filter(username=nombre).first()
            empresa_datos = Usuarios_datos.objects.filter(usuario_id=empresa_obj.id).first()
            empresa = empresa_datos.empresa.nombre

            cliente = request.GET.get('cliente', '').replace(" ", "")
            hoy = request.GET.get('fecha_fin', '')
            hoy = datetime.strptime(hoy, '%Y-%m-%d').strftime('%Y-%m-%d')



            # Construir URL de consulta
            query_url = (
                IP_SAP + "SQLQueries('ConsultaInventariosIndicadores1')/List"
                "?ClienteBusqueda='" + cliente + "'&FechaActual='" + hoy + "'&NombreProveedor='" + empresa + "'"
            )


            # Reintentos automáticos si excede 210 segundos
            max_retries = 10
            intento = 0
            data = []

            while intento < max_retries:
                try:
                    response2 = sap_request(query_url)
                    if response2.status_code == 200:
                        texto = response2.text.replace('null', '" "')
                        data = ast.literal_eval(texto)['value']
                        break
                    else:
                        raise Exception("Status code: " + str(response2.status_code))
                except Exception as e:
                    intento += 1
                    if intento >= max_retries:
                        raise e
                    time.sleep(5)

            # Crear Excel
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="INVENTARIO.xls"'
            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet("Inventario")

            font_style = xlwt.XFStyle()
            font_style.font.bold = True
            columns = ['EAN_Dependencia', 'Dependencia', 'Descripcion', 'Linea', 'EAN', 'Fecha Corte \n' + str(hoy)]

            for col_num in range(len(columns)):
                ws.write(0, col_num, columns[col_num], font_style)
                ws.col(col_num).width = 5000

            font_style = xlwt.XFStyle()
            repetidos = []
            row_num = 1
            for d in data:
                clave = [d.get('EAN'), d.get('EAN_Dependencia')]
                if clave in repetidos:
                    continue
                repetidos.append(clave)

                fila = [
                    d.get('EAN_Dependencia', ''),
                    d.get('Dependencia', ''),
                    d.get('Descripcion', ''),
                    d.get('Linea', ''),
                    d.get('EAN', ''),
                    d.get('Cantidad', '')
                ]
                for col_num, valor in enumerate(fila):
                    ws.write(row_num, col_num, '' if valor is None else valor, font_style)
                row_num += 1

            wb.save(response)

        except Exception as e:
            response = HttpResponse("Error al generar el reporte: " + str(e), status=500)

        finally:
            pass

        return response


def config_solicitud_ventas(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        empresa=usuario_datos.empresa.nombre
        current_user = request.user
        url2 = IP_SAP + "SQLQueries('consultadetallepedidos14')/List?ProveedorC='"+str(empresa)+"'"
        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'id': datos['CodigoCliente'],
                'nombre': datos['Cliente'],
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_ventas.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                        'permiso_usuario': usuario_datos,
                                                        'dato_cliente':dato_lista,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        cliente = request.POST['cliente']
        fecha_fin=datetime.strptime(request.POST['fecha_fin'], '%Y-%m-%d')
        fecha_inicio=fecha_fin - timedelta(days=7)
        if fecha_inicio.weekday() == 6:
            fecha_inicio = fecha_inicio + timedelta(days=1)
        elif fecha_inicio.weekday() == 5:
            fecha_inicio = fecha_inicio + timedelta(days=2)
        elif fecha_inicio.weekday() == 4:
            fecha_inicio = fecha_inicio + timedelta(days=3)
        elif fecha_inicio.weekday() == 3:
            fecha_inicio = fecha_inicio + timedelta(days=4)
        elif fecha_inicio.weekday() == 2:
            fecha_inicio = fecha_inicio + timedelta(days=5)
        elif fecha_inicio.weekday() == 1:
            fecha_inicio = fecha_inicio + timedelta(days=6)
        elif fecha_inicio.weekday() == 0:
            fecha_inicio = fecha_inicio + timedelta(days=7)
        fecha_fin=fecha_fin.strftime('%Y-%m-%d')
        fecha_inicio=fecha_inicio.strftime('%Y-%m-%d')
        pagina = request.POST['pagina']

        try:
            paginador = request.POST['paginador']
            if paginador=='atras':
                pagina=str(int(pagina)-100)
                if int(pagina)<0:
                    pagina=str(0)
            elif paginador=='adelante':
                pagina=str(int(pagina)+100)
            elif paginador=='primera':
                pagina=str(0)
        except:
            pass


        url3 = IP_SAP + "SQLQueries('consultadetallepedidos14')/List?ProveedorC='"+str(empresa)+"'"

        response1 = sap_request(url3)
        response1 = response1.text
        response1 = response1.replace('null', ' " " ')
        response1 = ast.literal_eval(response1)
        response1 = response1['value']
        dato_listas = []
        for datos in response1:
            data_prueba = {
                'id': datos['CodigoCliente'],
                'nombre': datos['Cliente'],
            }
            dato_listas.append(data_prueba)
        url2 = IP_SAP + "SQLQueries('ConsultaVentaIndicadoresa')/List?Cliente='"+cliente+"'&FechaInicial='"+fecha_inicio+"'&FechaFinal='"+fecha_fin+"'&ProveedorC='"+empresa+"'&$skip="+pagina

        response = sap_request(url2)
        response = response.text
        response = response.replace('null', ' " " ')
        response = ast.literal_eval(response)
        response = response['value']
        repeticion = []
        dato_lista = []
        for datos in response:
            try:
                comparacion = repeticion.index([datos['EAN'], datos['EAN_Dependencia'], datos['Fecha_Creacion']])
                pass
            except:
                data_prueba = {
                    'EAN_Dependencia': datos['EAN_Dependencia'],
                    'Nomb_Dependencia': datos['Nomb_Dependencia'],
                    'EAN': datos['EAN'],
                    'Descripcion': datos['Descripcion'],
                    'Linea': datos['Linea'],
                    'Fecha': datos['Fecha_Creacion'],
                    'Cantidad': datos['Cantidad'],
                }
                dato_lista.append(data_prueba)
                repeticion.append([datos['EAN'], datos['EAN_Dependencia'],datos['Fecha_Creacion']])

        return render(request, "config_solicitud_ventas.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio': fecha_inicio,
                                                                'fecha_fin': fecha_fin,
                                                                'permiso_usuario': usuario_datos,
                                                                'pagina': pagina,
                                                                'cliente': cliente,
                                                                'dato_cliente': dato_listas,
                                                                'pagina_fin': str(int(pagina)+20),
                                                                })


def reporte_ventas(request):

    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.nombre
        cliente = request.GET.get('cliente')
        fin=datetime.strptime(request.GET.get('fecha_fin'), '%Y-%m-%d')
        inicio = fin - timedelta(days=7)
        if inicio.weekday() == 6:
            inicio = inicio + timedelta(days=1)
        elif inicio.weekday() == 5:
            inicio = inicio + timedelta(days=2)
        elif inicio.weekday() == 4:
            inicio = inicio + timedelta(days=3)
        elif inicio.weekday() == 3:
            inicio = inicio + timedelta(days=4)
        elif inicio.weekday() == 2:
            inicio = inicio + timedelta(days=5)
        elif inicio.weekday() == 1:
            inicio = inicio + timedelta(days=6)
        elif inicio.weekday() == 0:
            inicio = inicio + timedelta(days=7)
        fecha_fin = fin.strftime('%Y-%m-%d')
        fecha_inicio = inicio.strftime('%Y-%m-%d')




        url2 = IP_SAP + "SQLQueries('ConsultaVentaIndicadoresa')/List?Cliente='"+cliente+"'&FechaInicial='"+fecha_inicio+"'&FechaFinal='"+fecha_fin+"'&ProveedorC='"+empresa+"'"

        response2 = sap_request(url2)
        response2 = response2.text
        response2 = response2.replace('null', ' " " ')
        response2 = ast.literal_eval(response2)
        response2 = response2['value']

        subtitulo ="Ventas"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="INFORMEVENTAS.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        lista_fechas = [inicio + timedelta(days=d) for d in range((fin - inicio).days + 1)]

        columns = ['EAN_Dependencia',
                   'Nomb_Dependencia',
                   'EAN',
                   'Descripcion',
                   'Linea',
                    ]
        for d in lista_fechas:
            columns.append("'"+d.strftime('%Y-%m-%d')+"'")
        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []
        repeticion=[]
        contador=0
        for d in response2:
            try:
                comparacion=repeticion.index([d['EAN'],d['EAN_Dependencia']])
                fecha=datetime.strptime(d['Fecha_Reporte'], '%Y%m%d')
                fecha=lista_fechas.index(fecha)
                fecha=fecha+5
                rows[comparacion][fecha]=d['Cantidad']
            except:
                fecha = datetime.strptime(d['Fecha_Reporte'], '%Y%m%d')
                fecha = lista_fechas.index(fecha)
                fecha = fecha + 5
                EAN_Dependencia= d['EAN_Dependencia'],
                Nomb_Dependencia= d['Nomb_Dependencia'],
                EAN= d['EAN'],
                Descripcion= d['Descripcion'],
                Linea= d['Linea'],


                datos = [[
                    EAN_Dependencia[0],
                    Nomb_Dependencia[0],
                    EAN[0],
                    Descripcion[0],
                    Linea[0],
                    '','','','','','','','','','',
                ]]
                rows.extend(datos)
                rows[contador][fecha] = d['Cantidad']
                repeticion.append([d['EAN'],d['EAN_Dependencia']])
                contador+=1
        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response


def config_solicitud_pagos_recibidos(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        return render(request, "config_solicitud_pagos_recibidos.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                                         'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']



        url2 = IP_SAP + "IncomingPayments?$select=DocNum,DocDate,DueDate,TransferSum&$filter=CardName eq '" \
               + empresa + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'"

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'n_pedido': datos['DocNum'],
                'fecha_contabilizacion': datos['DocDate'],
                'fecha_vencimiento': datos['DueDate'],
                'importe_total':format(datos['TransferSum'], '0,.0f') ,
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_pagos_recibidos.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                         'permiso_usuario': usuario_datos,
                                                                })


def pagos_recibidos_detalles(request, form_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


    url2 = IP_SAP + "IncomingPayments?$select=DocNum,DocDate,DueDate,CardCode,CardName,Address&$filter=DocNum eq " + form_id
    url3 = IP_SAP + "VendorPayments?$select=PaymentInvoices&$filter=DocNum eq " + form_id

    response = sap_request(url2)
    response = ast.literal_eval(response.text)
    response = response['value']
    response2 = sap_request(url3)
    response2 = json.loads(response2.text)
    response2 = response2['value']
    response2 = response2[0]
    response2 = response2['PaymentInvoices']
    articulos = []
    for d in response2:
        data_articulos = {
            'numero': d['DocEntry'],
            'plazo': '1 a 1',
            'total': format(d['SumApplied'], '0,.0f'),
        }
        articulos.append(data_articulos)

    dato_lista = []
    for datos in response:
        data_prueba = {
            'codigo': datos['CardCode'],
            'nombre': datos['CardName'],
            'destinatorio_factura': datos['Address'],
            'n_pagos': datos['DocNum'],
            'fecha_contabilizacion': datos['DocDate'],
            'fecha_vencimiento': datos['DueDate'],
            'detalle': articulos,
        }
        dato_lista.append(data_prueba)
    return render(request, "pagos_recibidos_detalle.html", {"form_id": form_id,
                                                      'user': current_user,
                                                      'lista_solicitud': dato_lista,
                                                            'permiso_usuario': usuario_datos,
                                                      })



def config_solicitud_comprobante_egreso(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        return render(request, "config_solicitud_comprobante_egreso.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                                            'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        codigo_sistema = empresa.empresa.codigo
        empresa=empresa.empresa.nombre

        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']



        url2 = IP_SAP + "VendorPayments?$select=DocNum,DocEntry,DocDate,DueDate,TransferSum&$filter=CardCode eq '" \
               + codigo_sistema + "'and DocDate lt '"+ fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'"

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            fecha_vencimiento = datos['DueDate']
            fecha_vencimiento = datetime.strptime(fecha_vencimiento, "%Y-%m-%dT%H:%M:%SZ")
            fecha_vencimiento = fecha_vencimiento.strftime("%Y-%m-%d")
            fecha_contabilizacion = datos['DocDate']
            fecha_contabilizacion = datetime.strptime(fecha_contabilizacion, "%Y-%m-%dT%H:%M:%SZ")
            fecha_contabilizacion = fecha_contabilizacion.strftime("%Y-%m-%d")
            data_prueba = {
                'n_pedido': datos['DocNum'],
                'n_interno': datos['DocEntry'],
                'fecha_contabilizacion': fecha_contabilizacion,
                'fecha_vencimiento': fecha_vencimiento,
                'importe_total':format(datos['TransferSum'], '0,.0f') ,
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_comprobante_egreso.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                            'permiso_usuario': usuario_datos,
                                                                })


def comprobante_egreso_detalles(request, form_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()



    url2 = IP_SAP + "VendorPayments?$select=DocNum,DocDate,DueDate,CardCode,CardName,Address&$filter=DocEntry eq " + form_id
    url3 = IP_SAP + "SQLQueries('ComprobanteEgresos7')/List?NumeroDoc="+form_id

    response = sap_request(url2)
    response = ast.literal_eval(response.text)
    response = response['value']
    response2 = sap_request(url3)
    response2 = ast.literal_eval(response2.text)
    response2 = response2['value']
    articulos = []
    suma_total=0
    for d in response2:
        trans_id=d['TransId']
        DatoTrans=d['DatoTrans']
        suma_total= suma_total+int(d['SumApplied'])
        if d['InvType']=='19':
            bloque_dato = d['BaseRef']
            url4 = IP_SAP + "SQLQueries('ComprobanteEgresosdato')/List?NumeroDoc="+bloque_dato
            response5 = sap_request(url4)
            response5 = json.loads(response5.text)
            nota = response5['value']
            nota=nota[0]
            nota=nota['NumAtCard']
            factura=''

        else:
            factura=d['Ref2']
            if factura=='':
                url4 = IP_SAP + "SQLQueries('consultatransaccion')/List?datotrans=" + str(DatoTrans)
                response5 = sap_request(url4)
                response5 = json.loads(response5.text)
                factura = response5['value']
                if factura==[]:
                    factura=''
                else:
                    factura = factura[0]
                    factura = factura['NumAtCard']
            nota=''
        data_articulos = {
            'numero': d['InvoiceId'],
            'tipo': d['InvType'],
            'doc_num': d['Ref1'],
            'fecha': d['RefDate'],
            'descuento':format(d['DcntSum'], '0,.0f'),
            'factura': factura,
            'nota': nota,
            'total': format(d['SumApplied'], '0,.0f'),
        }
        articulos.append(data_articulos)
    try:
	if response2 ==[]:
            url2 = IP_SAP + "SQLQueries('ConsultaTransPagos')/List?numerodoc="+form_id

            response10 = sap_request(url2)
            response10 = ast.literal_eval(response10.text)
            trans_id = response10['value'][0]['TransId']
        url4 = IP_SAP + "SQLQueries('ComprobanteEgresosContabilidad')/List?numero_pago=" + str(trans_id)
        response3 = sap_request(url4)
        response3 = ast.literal_eval(response3.text)
        response3 = response3['value']
        suma_total = format(int(suma_total), '0,.0f')
        dato_contabilizacion = []
        suma_credito = 0
        suma_debito = 0
        for datos_contabilizacion in response3:
            suma_credito = suma_credito + int(datos_contabilizacion['Credit'])
            suma_debito = suma_debito + int(datos_contabilizacion['Debit'])
            data_contabilizacion = {
                'cuenta': datos_contabilizacion['Account'],
                'descripcion': datos_contabilizacion['AcctName'],
                'debito': datos_contabilizacion['Debit'],
                'credito': datos_contabilizacion['Credit'],
                'comentarios': datos_contabilizacion['LineMemo'],
            }
            dato_contabilizacion.append(data_contabilizacion)
        suma_credito = format(int(suma_credito), '0,.0f')
        suma_debito = format(int(suma_debito), '0,.0f')
    except:
        suma_total=0
        dato_contabilizacion=''
        suma_credito = 0
        suma_debito = 0



    dato_lista = []

    for datos in response:
        fecha_contabilizacion=datos['DocDate']
        fecha_contabilizacion = datetime.strptime(fecha_contabilizacion, "%Y-%m-%dT%H:%M:%SZ")
        fecha_contabilizacion = fecha_contabilizacion.strftime("%Y-%m-%d")
        fecha_vencimiento = datos['DueDate']
        fecha_vencimiento = datetime.strptime(fecha_vencimiento, "%Y-%m-%dT%H:%M:%SZ")
        fecha_vencimiento = fecha_vencimiento.strftime("%Y-%m-%d")
        data_prueba = {
            'codigo': datos['CardCode'],
            'nombre': datos['CardName'],
            'destinatorio_factura': datos['Address'],
            'n_pagos': datos['DocNum'],
            'fecha_contabilizacion': fecha_contabilizacion,
            'fecha_vencimiento': fecha_vencimiento,
            'detalle':articulos,
        }
        dato_lista.append(data_prueba)

    return render(request, "comprobante_egreso_detalle.html", {"form_id": form_id,
                                                      'user': current_user,
                                                      'suma_total': suma_total,
                                                      'suma_credito': suma_credito,
                                                      'suma_debito': suma_debito,
                                                      'lista_solicitud': dato_lista,
                                                      'lista_contabilizacion': dato_contabilizacion,
                                                               'permiso_usuario': usuario_datos,
                                                      })

def comprobante_egreso_print(request, form_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
    comprobante_print='si'


    url2 = IP_SAP + "VendorPayments?$select=DocNum,DocDate,DueDate,CardCode,CardName,Address&$filter=DocEntry eq " + form_id
    url3 = IP_SAP + "SQLQueries('ComprobanteEgresos6')/List?NumeroDoc="+form_id

    response = sap_request(url2)
    response = ast.literal_eval(response.text)
    response = response['value']
    response2 = sap_request(url3)
    response2 = ast.literal_eval(response2.text)
    response2 = response2['value']
    articulos = []
    suma_total=0
    for d in response2:
        trans_id = d['TransId']
        suma_total = suma_total + int(d['SumApplied'])
        if d['InvType']=='19':
            bloque_dato = d['BaseRef']
            url4 = IP_SAP + "SQLQueries('ComprobanteEgresosdato')/List?NumeroDoc="+bloque_dato
            response5 = sap_request(url4)
            response5 = json.loads(response5.text)
            nota = response5['value']
            nota=nota[0]
            nota=nota['NumAtCard']
            factura=''

        else:
            factura=d['Ref2']
            nota=''
        data_articulos = {
            'numero': d['InvoiceId'],
            'tipo': d['InvType'],
            'doc_num': d['Ref1'],
            'fecha': d['RefDate'],
            'descuento':format(d['DcntSum'], '0,.0f'),
            'factura': factura,
            'nota': nota,
            'total': format(d['SumApplied'], '0,.0f'),
        }
        articulos.append(data_articulos)

    try:
	if response2 ==[]:
            url2 = IP_SAP + "SQLQueries('ConsultaTransPagos')/List?numerodoc="+form_id

            response10 = sap_request(url2)
            response10 = ast.literal_eval(response10.text)
            trans_id = response10['value'][0]['TransId']
        url4 = IP_SAP + "SQLQueries('ComprobanteEgresosContabilidad')/List?numero_pago=" + str(
            trans_id)
        response3 = sap_request(url4)
        response3 = ast.literal_eval(response3.text)
        response3 = response3['value']

        suma_total = format(int(suma_total), '0,.0f')

        dato_contabilizacion = []
        suma_credito = 0
        suma_debito = 0
        for datos_contabilizacion in response3:
            suma_credito = suma_credito + int(datos_contabilizacion['Credit'])
            suma_debito = suma_debito + int(datos_contabilizacion['Debit'])
            data_contabilizacion = {
                'cuenta': datos_contabilizacion['Account'],
                'descripcion': datos_contabilizacion['AcctName'],
                'debito': datos_contabilizacion['Debit'],
                'credito': datos_contabilizacion['Credit'],
                'comentarios': datos_contabilizacion['LineMemo'],
            }
            dato_contabilizacion.append(data_contabilizacion)
        suma_credito = format(int(suma_credito), '0,.0f')
        suma_debito = format(int(suma_debito), '0,.0f')
    except:
        suma_total = 0
        dato_contabilizacion = ''
        suma_credito = 0
        suma_debito = 0
    dato_lista = []
    for datos in response:
        data_prueba = {
            'codigo': datos['CardCode'],
            'nombre': datos['CardName'],
            'destinatorio_factura': datos['Address'],
            'n_pagos': datos['DocNum'],
            'fecha_contabilizacion': datos['DocDate'],
            'fecha_vencimiento': datos['DueDate'],
            'detalle':articulos,
        }
        dato_lista.append(data_prueba)
    return render(request, "comprobante_egreso_print.html", {"form_id": form_id,
                                                      'user': current_user,
                                                      'lista_solicitud': dato_lista,
                                                      'suma_total': suma_total,
                                                    'suma_credito': suma_credito,
                                                    'suma_debito': suma_debito,
                                                    'lista_contabilizacion': dato_contabilizacion,
                                                      'comprobante_print': comprobante_print,
                                                               'permiso_usuario': usuario_datos,
                                                      })

def config_solicitud_comprobante_egreso_pcs(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        return render(request, "config_solicitud_comprobante_egreso_pcs.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                                                'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nombre
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']



        url2 = IP_SAP + "VendorPayments?$select=DocNum,DocDate,DueDate,TransferSum&$filter=DocDate lt '" + fecha_fin + "' and DocDate gt '"+ fecha_inicio +"'"

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'n_pedido': datos['DocNum'],
                'fecha_contabilizacion': datos['DocDate'],
                'fecha_vencimiento': datos['DueDate'],
                'importe_total':format(datos['TransferSum'], '0,.0f') ,
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_comprobante_egreso_pcs.html", {'user': current_user,
                                                                'lista_formulas': formulas,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                                'permiso_usuario': usuario_datos,
                                                                })


def config_solicitud_estado_cuenta(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        return render(request, "config_solicitud_estado_cuenta.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.nombre
        codigo_prod = request.POST['codigo']



        url2 = IP_SAP + "BusinessPartners?$select=CardCode,FederalTaxID,CardName,CurrentAccountBalance&$filter=CardCode  eq '" \
               + codigo_prod + "'"

        response = sap_request(url2)
        response = ast.literal_eval(response.text)
        response = response['value']
        dato_lista = []
        for datos in response:
            data_prueba = {
                'codigo': datos['CardCode'],
                'nombre': datos['CardName'],
                'saldo_cuenta':format(datos['CurrentAccountBalance'], '0,.0f')
            }
            dato_lista.append(data_prueba)
        return render(request, "config_solicitud_estado_cuenta.html", {'user': current_user,
                                                                      'lista_formulas': formulas,
                                                                      'lista_prueba': dato_lista,
                                                                      'codigo_prod': codigo_prod,
                                                                       'permiso_usuario': usuario_datos,
                                                                      })



def config_solicitud_estado_cuenta_grupos(request, form_id):
    current_user = request.user
    usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
    nombre = current_user.username
    empresa = User.objects.filter(username=nombre).first()
    empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
    empresa = empresa.empresa.nombre



    url2 = IP_SAP + "BusinessPartners?$select=CardCode,FederalTaxID,CardName,CurrentAccountBalance&$filter=FederalTaxID  eq '" \
               + form_id + "' and CurrentAccountBalance ne 0"

    response = sap_request(url2)
    response = ast.literal_eval(response.text)
    response = response['value']
    dato_lista = []
    for datos in response:
        data_prueba = {
            'codigo': datos['CardCode'],
            'nombre': datos['CardName'],
            'saldo_cuenta':format(datos['CurrentAccountBalance'], '0,.0f')
        }
        dato_lista.append(data_prueba)
    return render(request, "config_solicitud_estado_cuenta_grupos.html", {'user': current_user,
                                                                    'lista_prueba': dato_lista,
                                                                    'permiso_usuario': usuario_datos,
                                                                      })


def estado_cuenta_detalle(request, form_id):
    if request.method == 'GET':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()


        try:
            url2 = IP_SAP + "BusinessPartners?$select=CardName,CardType,CardCode,EmailAddress,Phone1,FederalTaxID&$filter=CardCode eq '" + form_id + "'"
            response = sap_request(url2)
            response = ast.literal_eval(response.text)
            response = response['value']
            dato_lista = []
            for datos in response:
                codigo_em = datos['CardCode']
                data_prueba = {
                    'nombre': datos['CardName'],
                    'tipo': datos['CardType'],
                    'email': datos['EmailAddress'],
                    'contacto': datos['Phone1'],
                    'codigo_em': datos['CardCode'],
                    'nit': datos['FederalTaxID'],
                }
            dato_lista.append(data_prueba)
        except:
            url2 = IP_SAP + "BusinessPartners?$select=CardName,CardType,CardCode,FederalTaxID&$filter=CardCode eq '" + form_id + "'"
            response = sap_request(url2)
            response = ast.literal_eval(response.text)
            response = response['value']
            dato_lista = []
            for datos in response:
                codigo_em = datos['CardCode']
                data_prueba = {
                    'nombre': datos['CardName'],
                    'tipo': datos['CardType'],
                    'email': '',
                    'contacto': '',
                    'codigo_em': datos['CardCode'],
                    'nit': datos['FederalTaxID'],
                }
            dato_lista.append(data_prueba)
        return render(request, "estado_cuenta_detalle.html", {"form_id": form_id,
                                                              'user': current_user,
                                                              'codigo_em':codigo_em,
                                                              'dato_empresa': dato_lista,
                                                              'permiso_usuario': usuario_datos,
                                                              })
    elif request.method == 'POST':
        current_user = request.user
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        pagina = request.POST['pagina']
        check = request.POST['check']

        try:
            paginador = request.POST['paginador']
            if paginador == 'atras':
                pagina = str(int(pagina) - 20)
                if int(pagina) < 0:
                    pagina = str(0)
            elif paginador == 'adelante':
                pagina = str(int(pagina) + 20)
            elif paginador == 'primera':
                pagina = str(0)
        except:
            pass


        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        try:
            url2 = IP_SAP + "BusinessPartners?$select=CardName,CardType,CardCode,EmailAddress,Phone1,FederalTaxID&$filter=CardCode eq '" + form_id+"'"
            response = sap_request(url2)
            response = ast.literal_eval(response.text)
            response = response['value']
            dato_lista = []
            for datos in response:
                codigo_em = datos['CardCode']
                data_prueba = {
                    'nombre': datos['CardName'],
                    'tipo': datos['CardType'],
                    'email': datos['EmailAddress'],
                    'contacto': datos['Phone1'],
                    'codigo_em': datos['CardCode'],
                    'nit': datos['FederalTaxID'],
                }
            dato_lista.append(data_prueba)
        except:
            url2 = IP_SAP + "BusinessPartners?$select=CardName,CardType,CardCode,FederalTaxID&$filter=CardCode eq '" + form_id + "'"
            response = sap_request(url2)
            response = ast.literal_eval(response.text)
            response = response['value']
            dato_lista = []
            for datos in response:
                codigo_em = datos['CardCode']
                data_prueba = {
                    'nombre': datos['CardName'],
                    'tipo': datos['CardType'],
                    'email': '',
                    'contacto': '',
                    'codigo_em': datos['CardCode'],
                    'nit': datos['FederalTaxID'],
                }
            dato_lista.append(data_prueba)
        if check=='no':
            url3 = IP_SAP + "SQLQueries('Pruebas6')/List?ShortName='"+form_id+"'&FechaInicial='"+fecha_inicio+"'&FechaFinal='"+fecha_fin+"'&$skip="+pagina
        elif check=='si':
            url3 = IP_SAP + "SQLQueries('EstadosCuentasnoreconc3')/List?ShortName='" + form_id + "'&FechaInicial='" + fecha_inicio + "'&FechaFinal='" + fecha_fin + "'&$skip=" + pagina
        response2 = sap_request(url3)
        response2 = json.loads(response2.text)
        response2 = response2['value']
        estados = []
        for estado in response2:
            if estado['PayBlckRef']==None:
                causal_bloc=''
            else:
                try:
                    bloque_dato=str(int(estado['PayBlckRef']))
                    url4 = IP_SAP + "PaymentBlocks?$select=PaymentBlockCode&$filter=AbsEntry eq " + bloque_dato
                    response5 = sap_request(url4)
                    response5 = ast.literal_eval(response5.text)
                    causal_bloc = response5['value']
                    causal_bloc=causal_bloc[0]
                    causal_bloc = causal_bloc['PaymentBlockCode']
                except:
                    causal_bloc = ''
            data_articulos = {
                'info': estado['LineMemo'],
                'bloqueo': estado['PayBlock'],
                'causal_bloqueo': causal_bloc,
                'credito': format(estado['Credit'], '0,.0f'),
                'fecha_contabilizacion': estado['RefDate'],
                'fecha_vencimiento': estado['DueDate'],
                'numero_origen':format(estado['WTSum'], '0,.0f'),
                'factura': estado['Ref2'],
                'debito': format(estado['Debit'], '0,.0f'),
                'saldo_debito': format(int(estado['BalDueDeb'])+int(estado['BalDueCred']), '0,.0f'),
                'saldo_credito': format(estado['BalDueCred'], '0,.0f'),
            }
            estados.append(data_articulos)


        return render(request, "estado_cuenta_detalle.html", {"form_id": form_id,
                                                          'user': current_user,
                                                          'lista_solicitud': estados,
                                                        'dato_empresa':dato_lista,
                                                        'fecha_inicio':fecha_inicio,
                                                        'fecha_fin':fecha_fin,
                                                        'permiso_usuario': usuario_datos,
                                                        'pagina': pagina,
                                                        'check': check,
                                                        'pagina_fin': str(int(pagina)+20),
                                                        'codigo_em': codigo_em,
                                                          })



def reporte_estado_cuenta(request):

    if request.method == 'GET':
        current_user = request.user
        nombre = current_user.username
        empresa = User.objects.filter(username=nombre).first()
        empresa = Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa = empresa.empresa.nombre
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        check = request.GET.get('check')
        form_id = request.GET.get('form_id')

        if check == 'no':
            url3 = IP_SAP + "SQLQueries('Pruebas6')/List?ShortName='" + form_id + "'&FechaInicial='" + fecha_inicio + "'&FechaFinal='" + fecha_fin + "'"
        elif check == 'si':
            url3 = IP_SAP + "SQLQueries('EstadosCuentasnoreconc3')/List?ShortName='" + form_id + "'&FechaInicial='" + fecha_inicio + "'&FechaFinal='" + fecha_fin + "'"
        headers = {
            'Prefer': 'odata.maxpagesize=9999',
            'Cookie': 'B1SESSION=' + respuesta['SessionId']
        }
        response2 = sap_request(url3)
        response2 = json.loads(response2.text)
        response2 = response2['value']

        subtitulo ="Estado_cuenta"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ESTADOCUENTA.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['FECHA DE VENCIMIENTO',
                   'FACTURA',
                   'INFO DETALLADA',
                   'VALOR',
                   'SALDO VENCIDO(ML)',
                   'VALOR RETENCION',
                   'BLOQUEADO',
                   'CAUSAL',
                   ]

        for col_num in range(len(columns)):
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []
        for d in response2:
            if d['PayBlckRef']==None:
                causal_bloc=''
            else:
                try:
                    bloque_dato=str(int(d['PayBlckRef']))
                    url4 = IP_SAP + "PaymentBlocks?$select=PaymentBlockCode&$filter=AbsEntry eq " + bloque_dato
                    response5 = sap_request(url4)
                    response5 = ast.literal_eval(response5.text)
                    causal_bloc = response5['value']
                    causal_bloc=causal_bloc[0]
                    causal_bloc = causal_bloc['PaymentBlockCode']
                except:
                    causal_bloc = ''
            fecha_vencimiento = d['DueDate'],
            bloqueado = d['PayBlock'],
            factura = d['Ref2'],
            info = d['LineMemo'],
            retencion=format(d['WTSum'], '0,.0f'),
            valor = format(int(d['Credit']) + int(d['Debit']), '0,.0f'),
            if factura[0]=='':
                valor = '-' + valor[0]
            else:
                valor = valor[0]
            saldo_debito = format(int(d['BalDueDeb']) + int(d['BalDueCred']), '0,.0f'),
            datos = [(
                str(fecha_vencimiento[0]),
                factura[0],
                info[0],
                str(valor),
                str(saldo_debito[0]),
                retencion[0],
                bloqueado[0],
                causal_bloc,
            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)


        wb.save(response)
        return response







def reporte_pedidos_csv(request):
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        current_user = request.user

        return render(request, "config_solicitud_pedidos_csv.html", {'user': current_user,
                                                        'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']


        url2 = IP_SAP + "SQLQueries('ConsultaPedidosIndicadorJSON')/List?FechaInicial='" + fecha_inicio + "'&FechaFinal='" + fecha_fin + "'"

        response3 = sap_request(url2)
        response3=response3.text
        response3=response3.replace('null',' " " ')
        response3 = ast.literal_eval(response3)
        response3 = response3['value']
        response3 = [dict(t) for t in set(tuple(sorted(d.items())) for d in response3)]
        response3 = sorted(response3, key=lambda x: x['OrdenEDI'])

        response = HttpResponse(content_type='text/csv')

        response['Content-Disposition'] = 'attachment; filename=orden_de_compra.csv'

        # The csv.writer function takes file-like object as argument i.e.,HttpResponse object
        writer = csv.writer(response)

        # For each row in your CSV file, call writer.writerow function by passing
        # a list or tuple to it.
        writer.writerow(['OrdenEDI','OrdenSAP', 'Cliente', 'Empresario','NombreCliente','NombreEmpresario','GLNEntrega','SitioEntrega','FechaExpedicionEDI','FechaMinimaEdi','FechaMaximaEdi','GLNCod.Dep','NombreDep','EANproducto','DescripArticulo','Cantidad','ValorUnitario','ValorTotal','SKU / REFERENCIA CLIENTE','SKU / REFERENCIA EMPRESARIO','GLNCliente','Tipopedido'])

        for d in response3:
            if d['U_EAN']==' ':
                Eandependencia=d['EANEntrega']
                NombreDep=d['SitioEntrega']
            else:
                Eandependencia=d['U_EAN']
                NombreDep = d['NombreDep']
            if d['U_HBT1_TIPO_ORD_EDI']=='220':
                Tipopedido='Almacenamiento'
            elif d['U_HBT1_TIPO_ORD_EDI']=='YB1':
                Tipopedido = 'Predistribuido'
            elif d['U_HBT1_TIPO_ORD_EDI']=='YA9':
                Tipopedido = 'Consolidado'
            else:
                Tipopedido='Sin Asignar'
            writer.writerow([d['OrdenEDI'],d['PedidoVta'], d['Cliente'], d['Empresario'],d['NombreCliente'],d['NombreEmpresario'],d['EANEntrega'],d['SitioEntrega'],d['FechaDocto'],d['FechaMinEnt'],d['FechaMaxEnt'],Eandependencia,NombreDep,d['EANArticulo'],d['DescripArticulo'],d['CantidadPedida'],d['PrecioPed'],d['TotalPed'],d['PLU'],d['CodArticulo'],d['Cliente'],Tipopedido])

        return response






def reporte_certificados_retencion(request):
    formulas = []
    if request.method == 'GET':
        current_user = request.user
        nombre=current_user.username
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()

        return render(request, "config_solicitud_certificados_retencion.html", {'user': current_user,
                                                        'lista_formulas': formulas,
                                                                            'permiso_usuario': usuario_datos,
                                                        })
    elif request.method == 'POST':
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()
        nombre = current_user.username
        empresa=User.objects.filter(username=nombre).first()
        empresa=Usuarios_datos.objects.filter(usuario_id=empresa.id).first()
        empresa=empresa.empresa.nit
        fecha_inicio = request.POST['fecha_inicio']
        fecha_fin = request.POST['fecha_fin']
        dato_lista=Documento.objects.filter(nit=empresa)

        return render(request, "config_solicitud_certificados_retencion.html", {'user': current_user,
                                                                'lista_prueba': dato_lista,
                                                                'fecha_inicio':fecha_inicio,
                                                                'fecha_fin':fecha_fin,
                                                                            'permiso_usuario': usuario_datos,
                                                                })






def reporte_comprobante_detalle(request):

    if request.method == 'GET':

        form_id = request.GET.get('solicitud')
        current_user = request.user
        usuario_datos = Usuarios_datos.objects.filter(usuario_id=current_user.id).first()



        url2 = IP_SAP + "VendorPayments?$select=DocNum,DocDate,DueDate,CardCode,CardName,Address&$filter=DocEntry eq " + form_id
        url3 = IP_SAP + "SQLQueries('ComprobanteEgresos6')/List?NumeroDoc=" + form_id

        response3 = sap_request(url2)
        response3 = ast.literal_eval(response3.text)
        response3 = response3['value']
        response2 = sap_request(url3)
        response2 = ast.literal_eval(response2.text)
        response2 = response2['value']


        subtitulo ="Comprobante #" + str(form_id)

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="COMPROBANTE.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(subtitulo)


        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columnas = ['NO.PAGO',
                   'PROVEEDOR ',
                   'DESTINATARIO FACTURA',
                   'FECHA DE CONTABILIZACION',
                   'FECHA DEL DOCUMENTO',
                   ]

        for col_num in range(len(columnas)):
            cwidth = ws.col(col_num).width
            if (len(columnas[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columnas) * 367)
            ws.write(row_num, col_num, columnas[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        filas = []
        for d in response3:
            numero_pago = d['DocNum'],
            proveedor = d['CardName'],
            destinatorio_factura = d['Address'],
            fecha_contabilizacion = d['DocDate'],
            fecha_documento = d['DueDate'],
            datos = [(
                str(numero_pago[0]),
                str(proveedor[0]),
                str(destinatorio_factura[0]),
                str(fecha_contabilizacion[0]),
                str(fecha_documento[0]),

            )]
            filas.extend(datos)

        for row in filas:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        columns = ['NO.',
                   'TIPO',
                   'DOC.NO',
                   'FECHA',
                   'NO.FACTURA',
                   'DESCUENTO',
                   'VALOR A PAGAR',
                   ]

        for col_num in range(len(columns)):
            row_num = 5
            cwidth = ws.col(col_num).width
            if (len(columns[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columns) * 367)
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = []
        suma_total=0
        for d in response2:
            trans_id = d['TransId']
            suma_total = suma_total + int(d['SumApplied'])
            if d['InvType'] == '19':
                bloque_dato = d['BaseRef']
                url4 = IP_SAP + "SQLQueries('ComprobanteEgresosdato')/List?NumeroDoc=" + bloque_dato
                response5 = sap_request(url4)
                response5 = json.loads(response5.text)
                nota = response5['value']
                nota = nota[0]
                factura = nota['NumAtCard']
                tipo='Notas Credito Provedor'

            else:
                factura = d['Ref2']
                tipo='Factura Proveedor'

            numero = d['InvoiceId'],
            doc_num = d['Ref1'],
            fecha = d['RefDate'],
            descuento = format(d['DcntSum'], '0,.0f'),
            total = format(d['SumApplied'], '0,.0f'),
            datos = [(
                int(numero[0]),
                str(tipo),
                str(doc_num[0]),
                str(fecha[0]),
                str(factura),
                str(descuento[0]),
                str(total[0]),

            )]
            rows.extend(datos)

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)
        filas_resultante = row_num
        ws.write(filas_resultante + 2, 5, 'TOTAL:', font_style)
        for d in response3:
            total = format(int(suma_total), '0,.0f'),
            ws.write(filas_resultante + 2, 6, str(total[0]), font_style)

        url6 = IP_SAP + "SQLQueries('ComprobanteEgresosContabilidad')/List?numero_pago=" + str(
            trans_id)
        response7 = sap_request(url6)
        response7 = ast.literal_eval(response7.text)
        response7 = response7['value']
        columnas = ['CUENTA',
                    'DESCRIPCION ',
                    'DEBITO',
                    'CREDITO',
                    'COMENTARIOS',
                    ]

        for col_num in range(len(columnas)):
            cwidth = ws.col(col_num).width
            if (len(columnas[col_num]) * 367) > cwidth:
                ws.col(col_num).width = (len(columnas) * 367)
            ws.write(filas_resultante + 5, col_num, columnas[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        filasd=[]
        suma_credito=0
        suma_debito=0
        for d in response7:
            suma_credito = suma_credito + int(d['Credit'])
            suma_debito = suma_debito + int(d['Debit'])

            cuenta = d['Account'],
            descripcion = d['AcctName'],
            debito = d['Debit'],
            credito = d['Credit'],
            comentarios = d['LineMemo'],
            datosd = [(
                str(cuenta[0]),
                str(descripcion[0]),
                str(debito[0]),
                str(credito[0]),
                str(comentarios[0]),

            )]
            filasd.extend(datosd)
        dato_casilla=filas_resultante+6
        suma_credito = '$ ' + str(suma_credito)
        suma_debito = '$ ' + str(suma_debito)
        for row in filasd:
            dato_casilla += 1
            for col_num in range(len(row)):
                ws.write(dato_casilla, col_num, row[col_num], font_style)

        ws.write(dato_casilla + 2, 2, suma_debito, font_style)
        ws.write(dato_casilla + 2, 3, suma_credito, font_style)

        wb.save(response)
        return response





def tarea_correo_pedido_tres():

    try:
        if not SESSION:
            sap_login()
        cookie = "B1SESSION=" + SESSION
        if ROUTEID:  # por si tu SAP devuelve ROUTEID
            cookie += "; ROUTEID=" + ROUTEID
        estado = 'bost_Open'
        now = datetime.now(pytz.timezone('America/Bogota'))
        hoy=now.date()
        hora = now.time()
        errores = HistorialErrorTarea(
            accion='Inicio de tarea',
            fecha=hoy,
            hora=hora,
            empresa='No Corresponde',
            pedido='No Corresponde',
        )
        errores.save()

        url2 = IP_SAP + "PurchaseOrders?$orderby=DocDate desc&$select=DocNum,DocEntry,CardCode,CardName&$filter=DocDate eq '" \
               + str(hoy) + "'"

        headers = {
            'Prefer': 'odata.maxpagesize=999999',
            'Cookie': cookie
        }
        response = requests.request("GET", url2, headers=headers, verify=False)
        response = ast.literal_eval(response.text)
        response = response['value']



        now = datetime.now(pytz.timezone('America/Bogota'))
        hoy = now.date()
        hora = now.time()
        errores = HistorialErrorTarea(
            accion='Fin de tarea',
            fecha=hoy,
            hora=hora,
            empresa='No Corresponde',
            pedido='No Corresponde',
        )
        errores.save()
    except:
        error = str(sys.exc_info()[1])
        now = datetime.now(pytz.timezone('America/Bogota'))
        hoy = now.date()
        hora = now.time()
        errores = HistorialErrorTarea(
            accion='Error de conexion a: ' + error,
            fecha=hoy,
            hora=hora,
            empresa='No Corresponde',
            pedido='No Corresponde',
        )
        errores.save()
